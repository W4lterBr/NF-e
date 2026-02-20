from __future__ import annotations

import os
import sys
import json
import subprocess
import logging
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Callable
from datetime import datetime
import sqlite3
import ctypes
from ctypes import wintypes

# Logger
logger = logging.getLogger('nfe_search')

# Sistema de Temas
try:
    from themes import ThemeManager
    THEMES_AVAILABLE = True
except ImportError as e:
    print(f"[THEME] ⚠️ Sistema de temas não disponível: {e}")
    THEMES_AVAILABLE = False
    ThemeManager = None

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QLineEdit, QComboBox, QProgressBar, QTextEdit,
    QDialog, QMessageBox, QFileDialog, QInputDialog, QStatusBar,
    QTreeWidget, QTreeWidgetItem, QSplitter, QAction, QMenu, QSystemTrayIcon,
    QProgressDialog, QStyledItemDelegate, QStyleOptionViewItem, QScrollArea, QFrame,
    QGroupBox, QRadioButton, QDateEdit, QStyle, QCheckBox, QTabWidget, QListWidget,
    QListWidgetItem
)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QSettings, QSize
from PyQt5.QtGui import QIcon, QColor, QBrush, QFont, QCloseEvent

# Classe customizada para ordenação numérica
class NumericTableWidgetItem(QTableWidgetItem):
    """Item de tabela que ordena numericamente em vez de alfabeticamente"""
    def __init__(self, text: str, numeric_value: float = 0.0):
        super().__init__(text)
        self._numeric_value = numeric_value
        # Armazena o valor também no UserRole para debug
        self.setData(Qt.UserRole, numeric_value)
    
    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            return self._numeric_value < other._numeric_value
        # Fallback: tenta comparar pelo UserRole
        try:
            self_val = self.data(Qt.UserRole)
            other_val = other.data(Qt.UserRole)
            if self_val is not None and other_val is not None:
                return float(self_val) < float(other_val)
        except:
            pass
        return super().__lt__(other)

# Delegate para centralizar ícones na coluna XML e Status
class CenterIconDelegate(QStyledItemDelegate):
    """Delegate que centraliza ícones em células"""
    def paint(self, painter, option, index):
        if index.column() == 0:  # Apenas coluna XML
            # Desenha o fundo
            painter.fillRect(option.rect, option.backgroundBrush)
            
            # Tamanho do ícone (definido antes para ser usado depois)
            icon_size = 20
            
            # Pega o ícone
            icon = index.data(Qt.DecorationRole)
            if icon and not icon.isNull():
                # Calcula posição centralizada
                x = option.rect.x() + (option.rect.width() - icon_size) // 2
                y = option.rect.y() + (option.rect.height() - icon_size) // 2
                # Desenha o ícone centralizado
                icon.paint(painter, x, y, icon_size, icon_size)
        else:
            super().paint(painter, option, index)
    
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)

def get_data_dir():
    """Retorna o diretório de dados do aplicativo (AppData para compilados, local para dev)."""
    import sys
    
    # Se estiver executando como executável PyInstaller
    if getattr(sys, 'frozen', False):
        # Usa AppData do usuário para dados persistentes
        app_data = Path(os.environ.get('APPDATA', Path.home()))
        data_dir = app_data / "Busca XML"
    else:
        # Desenvolvimento: usa pasta local
        data_dir = Path(__file__).parent
    
    # Garante que o diretório existe
    try:
        data_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"AVISO: Não foi possível criar {data_dir}: {e}")
        # Fallback para pasta temporária
        data_dir = Path(os.environ.get('TEMP', Path.home())) / "Busca XML"
        data_dir.mkdir(parents=True, exist_ok=True)
    
    return data_dir

# Paths
# FIX: Corrige caminho dos recursos para PyInstaller onedir
if getattr(sys, 'frozen', False):
    # Modo executável: recursos estão em _internal (onedir) ou em sys._MEIPASS
    if hasattr(sys, '_MEIPASS'):
        BASE_DIR = Path(sys._MEIPASS)  # Onefile mode
    else:
        BASE_DIR = Path(sys.executable).parent / '_internal'  # Onedir mode
else:
    # Modo desenvolvimento
    BASE_DIR = Path(__file__).parent

DATA_DIR = get_data_dir()
DB_PATH = DATA_DIR / "notas.db"
LOGS_DIR = DATA_DIR / "logs"

# Backend
from modules.database import DatabaseManager as UIDB
from modules import sandbox_worker as sandbox


def ensure_logs_dir():
    try:
        LOGS_DIR.mkdir(exist_ok=True)
    except Exception:
        pass


def run_search(progress_cb: Optional[Callable[[str], None]] = None) -> Dict[str, Any]:
    """Executa a busca de NFe/CTe na SEFAZ."""
    # Usa sys.__stdout__ que é garantido ser o stdout original
    original_stdout = sys.__stdout__ if hasattr(sys, '__stdout__') else sys.stdout
    old_stdout = sys.stdout
    
    try:
        # Adiciona BASE_DIR ao sys.path para importação
        if str(BASE_DIR) not in sys.path:
            sys.path.insert(0, str(BASE_DIR))
        
        # Import estático (detectado pelo PyInstaller)
        # FIX: Troca import dinâmico por estático para funcionar no executável
        import nfe_search
        
        # Usa threading.Lock para proteção thread-safe contra recursão
        import threading
        _callback_lock = threading.Lock()
        
        # Classe para capturar e enviar progresso em tempo real
        class ProgressCapture:
            def write(self, text):
                try:
                    # Usa original_stdout que nunca é None
                    if original_stdout:
                        original_stdout.write(text)
                    
                    # PROTEÇÃO: Usa trylock para evitar recursão sem bloquear
                    if progress_cb and text.strip():
                        if _callback_lock.acquire(blocking=False):
                            try:
                                progress_cb(text.rstrip())
                            finally:
                                _callback_lock.release()
                except Exception as e:
                    # Fallback: tenta imprimir no console de qualquer forma
                    try:
                        if original_stdout:
                            original_stdout.write(f"[ERRO ProgressCapture] {e}\n")
                    except:
                        pass  # Silenciosamente ignora se nem isso funcionar
                    
            def flush(self):
                try:
                    if original_stdout:
                        original_stdout.flush()
                except Exception:
                    pass
        
        # Adiciona handler ao logger para capturar logs
        import logging
        if progress_cb:
            class ProgressHandler(logging.Handler):
                def emit(self, record):
                    try:
                        # PROTEÇÃO: Usa trylock para evitar recursão
                        if _callback_lock.acquire(blocking=False):
                            try:
                                msg = self.format(record)
                                progress_cb(msg)
                            finally:
                                _callback_lock.release()
                    except:
                        pass
            
            # Obtém o logger do nfe_search
            logger = logging.getLogger('nfe_search')
            
            # Remove apenas ProgressHandler antigos para evitar duplicação
            # MAS mantém StreamHandler (console) e FileHandler
            handlers_to_remove = [h for h in logger.handlers if isinstance(h, ProgressHandler)]
            for h in handlers_to_remove:
                logger.removeHandler(h)
            
            handler = ProgressHandler()
            handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        
        sys.stdout = ProgressCapture()
        
        # Executa apenas UMA iteração de busca (sem loop infinito)
        try:
            nfe_search.run_single_cycle()
        except SystemExit:
            # nfe_search pode chamar sys.exit() - ignorar
            pass
        except Exception as e:
            import traceback
            error_msg = f"Erro durante execução do nfe_search: {str(e)}\n{traceback.format_exc()}"
            sys.stdout = old_stdout if old_stdout else original_stdout
            return {"ok": False, "error": error_msg}
        
        # Restaura stdout
        sys.stdout = old_stdout if old_stdout else original_stdout
        
        return {"ok": True, "message": "Busca concluída"}
        
    except Exception as e:
        sys.stdout = old_stdout if old_stdout else original_stdout
        import traceback
        error_msg = f"Erro na busca: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)  # Log no console
        return {"ok": False, "error": error_msg}
    finally:
        # Garante que stdout sempre será restaurado
        try:
            sys.stdout = old_stdout if old_stdout else original_stdout
        except:
            sys.stdout = original_stdout
        sys.stdout = old_stdout


def resolve_xml_text(item: Dict[str, Any]) -> Optional[str]:
    try:
        chave = (item.get("chave") or "").strip()
        if not chave:
            return None
        
        import sqlite3
        import time
        
        with sqlite3.connect(str(DB_PATH)) as conn:
            # VERIFICAÇÃO RÁPIDA: Se a nota é RESUMO, não tem XML completo
            xml_status_row = conn.execute("SELECT xml_status FROM notas_detalhadas WHERE chave = ?", (chave,)).fetchone()
            if xml_status_row and xml_status_row[0] == 'RESUMO':
                print(f"[DEBUG XML] ⚠️ Nota marcada como RESUMO - sem XML completo disponível")
                return None
            
            # Tenta pegar xml_completo primeiro (mais rápido)
            row = conn.execute("SELECT xml_completo, caminho_arquivo FROM xmls_baixados WHERE chave = ?", (chave,)).fetchone()
            if row:
                # Se tem xml_completo no banco, usa ele
                if row[0]:
                    print(f"[DEBUG XML] ✅ XML encontrado no banco (xml_completo)")
                    return row[0]
                # Se não, tenta ler do arquivo
                if row[1] and os.path.exists(row[1]):
                    try:
                        print(f"[DEBUG XML] ✅ XML encontrado no arquivo: {row[1]}")
                        return Path(row[1]).read_text(encoding="utf-8", errors="ignore")
                    except Exception as e:
                        print(f"[DEBUG XML] ⚠️ Erro ao ler arquivo: {e}")
        
        print(f"[DEBUG XML] Buscando XML nas pastas locais para chave: {chave}")
        
        # 🆕 CORREÇÃO NFS-e: Detecta tipo de documento para busca correta
        tipo = str(item.get('tipo', '')).upper()
        is_nfse = 'NFS' in tipo
        
        # Para NFS-e, busca pelo NÚMERO ao invés da CHAVE
        if is_nfse:
            numero = item.get('nNF') or item.get('numero')  # Campo pode variar
            if numero:
                print(f"[DEBUG XML] NFS-e detectada - Buscando por número: {numero}")
                # Busca padrões múltiplos:
                # 1. NFSe_{numero}.xml (padrão antigo)
                # 2. {numero}-*.xml (padrão novo: {NUMERO}-{FORNECEDOR}.xml)
                # 3. *{numero}*.xml (fallback genérico)
                search_patterns = [
                    f"NFSe_{numero}.xml",
                    f"{numero}-*.xml",
                    f"*{numero}*.xml"
                ]
            else:
                print(f"[DEBUG XML] ⚠️ NFS-e sem número definido, tentando busca por chave")
                search_patterns = [f"*{chave}*.xml"]
        else:
            # NF-e e CT-e: busca por chave (padrão)
            search_patterns = [f"*{chave}*.xml"]
        
        roots = [
            DATA_DIR / "xmls",
            DATA_DIR / "xmls_chave",
            DATA_DIR / "xmls_nfce",
            DATA_DIR / "xml_NFs",
            DATA_DIR / "xml_envio",
            DATA_DIR / "xml_extraidos",
            BASE_DIR / "xmls",  # Fallback para desenvolvimento
            BASE_DIR / "xmls_chave",
            BASE_DIR / "xmls_nfce",
        ]
        # Pastas configuradas pelo usuário por CNPJ (emitidos)
        try:
            digits_inf = ''.join(ch for ch in str(item.get('informante') or '') if ch.isdigit())
            if digits_inf:
                s = QSettings('NFE_System', 'BOT_NFE')
                extra = str(s.value(f'emit_dirs/{digits_inf}', '') or '')
                for p in extra.split(';'):
                    p = p.strip()
                    if p:
                        roots.append(Path(p))
        except Exception:
            pass
        
        # Primeira tentativa: busca rápida pelo nome do arquivo
        search_start = time.time()
        for r in roots:
            if not r.exists():
                continue
            # Timeout de 5 segundos para evitar travamentos
            if time.time() - search_start > 5.0:
                print(f"[DEBUG XML] ⏱️ Timeout na busca de XML (5s)")
                break
            
            # Tenta cada padrão de busca
            for search_pattern in search_patterns:
                for f in r.rglob(search_pattern):
                    try:
                        print(f"[DEBUG XML] ✅ XML encontrado em: {f}")
                        xml_content = f.read_text(encoding="utf-8", errors="ignore")
                        
                        # 🆕 CORREÇÃO: Se é NFS-e e encontrou XML, marca como COMPLETO no banco
                        if is_nfse and xml_content:
                            try:
                                with sqlite3.connect(str(DB_PATH)) as conn_update:
                                    # Atualiza notas_detalhadas
                                    conn_update.execute(
                                        "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                        (chave,)
                                    )
                                    conn_update.commit()
                                    print(f"[DEBUG XML] ✅ NFS-e marcada como COMPLETO no banco")
                            except Exception as e_update:
                                print(f"[DEBUG XML] ⚠️ Erro ao atualizar status: {e_update}")
                        
                        return xml_content
                    except Exception:
                        continue
        
        # Segunda tentativa REMOVIDA (muito lenta e trava a interface)
        # A busca por conteúdo em TODOS os XMLs pode levar minutos
        print(f"[DEBUG XML] ❌ XML não encontrado em nenhuma pasta")
    except Exception as e:
        print(f"[DEBUG XML] ❌ Erro ao buscar XML: {e}")
    return None


class SearchDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Buscar na SEFAZ")
        self.resize(700, 420)
        v = QVBoxLayout(self)
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        v.addWidget(self.progress)
        v.addWidget(self.log)

    def append(self, text: str):
        self.log.append(text)
        self.log.ensureCursorVisible()

    def set_percent(self, p: int):
        self.progress.setValue(max(0, min(100, p)))


class BrasilNFeConfigDialog(QDialog):
    """Diálogo para configurar API BrasilNFe (manifestação de notas)."""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Configuração API BrasilNFe")
        self.setModal(True)
        self.resize(600, 400)
        
        layout = QVBoxLayout(self)
        
        # Título e descrição
        titulo = QLabel("<h2>🔌 API BrasilNFe</h2>")
        layout.addWidget(titulo)
        
        desc = QLabel(
            "<p>Configure o token da API BrasilNFe para manifestação automática de NF-e.</p>"
            "<p><b>Vantagens:</b></p>"
            "<ul>"
            "<li>✅ Assinatura garantida (sem erro 297)</li>"
            "<li>✅ Compatibilidade 100% com SEFAZ</li>"
            "<li>✅ Não depende de xmlsec local</li>"
            "</ul>"
            "<p><b>Como obter:</b> Crie conta em <a href='https://brasilnfe.com.br'>brasilnfe.com.br</a></p>"
        )
        desc.setWordWrap(True)
        desc.setOpenExternalLinks(True)
        layout.addWidget(desc)
        
        # Campo de token
        token_layout = QHBoxLayout()
        token_layout.addWidget(QLabel("Token:"))
        self.token_input = QLineEdit()
        self.token_input.setEchoMode(QLineEdit.Password)
        self.token_input.setPlaceholderText("Cole seu token da API BrasilNFe aqui...")
        token_atual = self.db.get_config('brasilnfe_token', '')
        self.token_input.setText(token_atual)
        token_layout.addWidget(self.token_input, 1)
        
        # Botão para mostrar/ocultar token
        self.btn_toggle = QPushButton("👁")
        self.btn_toggle.setFixedWidth(40)
        self.btn_toggle.clicked.connect(self._toggle_visibility)
        token_layout.addWidget(self.btn_toggle)
        
        layout.addLayout(token_layout)
        
        # Status
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)
        
        # Botões de ação
        btn_layout = QHBoxLayout()
        
        btn_testar = QPushButton("🧪 Testar Conexão")
        btn_testar.clicked.connect(self._testar_conexao)
        btn_layout.addWidget(btn_testar)
        
        btn_layout.addStretch()
        
        btn_salvar = QPushButton("💾 Salvar")
        btn_salvar.clicked.connect(self.accept)
        btn_layout.addWidget(btn_salvar)
        
        btn_cancelar = QPushButton("❌ Cancelar")
        btn_cancelar.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancelar)
        
        layout.addLayout(btn_layout)
        
        # Informações adicionais
        info = QLabel(
            "<hr><p style='color: gray; font-size: 10pt;'>"
            "<b>Nota:</b> Se deixar em branco, o sistema usará assinatura local (xmlsec) "
            "que pode ter problemas de compatibilidade (erro 297)."
            "</p>"
        )
        info.setWordWrap(True)
        layout.addWidget(info)
    
    def _toggle_visibility(self):
        """Alterna visibilidade do token."""
        if self.token_input.echoMode() == QLineEdit.Password:
            self.token_input.setEchoMode(QLineEdit.Normal)
            self.btn_toggle.setText("🙈")
        else:
            self.token_input.setEchoMode(QLineEdit.Password)
            self.btn_toggle.setText("👁")
    
    def _testar_conexao(self):
        """Testa conexão com API BrasilNFe."""
        token = self.token_input.text().strip()
        
        if not token:
            self.status_label.setText("❌ Token vazio")
            self.status_label.setStyleSheet("color: red;")
            return
        
        self.status_label.setText("🔄 Testando conexão...")
        self.status_label.setStyleSheet("color: blue;")
        QApplication.processEvents()
        
        try:
            # Tenta importar módulo BrasilNFe
            from modules.brasilnfe_api import BrasilNFeAPI
            
            # Cria instância e testa conexão
            api = BrasilNFeAPI(token)
            
            # Faz requisição de teste (pode ser um endpoint de status se existir)
            # Por ora, apenas verifica se o token está no formato correto
            if len(token) < 20:
                self.status_label.setText("⚠️ Token parece inválido (muito curto)")
                self.status_label.setStyleSheet("color: orange;")
                return
            
            self.status_label.setText(
                "✅ Token configurado! Teste completo requer manifestação real."
            )
            self.status_label.setStyleSheet("color: green;")
            
        except ImportError:
            self.status_label.setText("❌ Módulo BrasilNFe não encontrado")
            self.status_label.setStyleSheet("color: red;")
        except Exception as e:
            self.status_label.setText(f"❌ Erro: {str(e)}")
            self.status_label.setStyleSheet("color: red;")
    
    def accept(self):
        """Salva configuração ao aceitar."""
        token = self.token_input.text().strip()
        
        try:
            if token:
                self.db.set_config('brasilnfe_token', token)
                QMessageBox.information(
                    self,
                    "Sucesso",
                    "Token BrasilNFe salvo!\n\n"
                    "A partir de agora, manifestações de NF-e usarão a API BrasilNFe "
                    "(assinatura remota garantida)."
                )
            else:
                # Remove token
                self.db.set_config('brasilnfe_token', '')
                QMessageBox.information(
                    self,
                    "Token Removido",
                    "Token BrasilNFe removido.\n\n"
                    "Manifestações voltarão a usar assinatura local (xmlsec)."
                )
            
            super().accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar token: {e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # 🎨 Aplica o tema salvo (antes de criar elementos visuais)
        self._current_theme_colors = None  # Inicializa cache de cores do tema
        if THEMES_AVAILABLE:
            try:
                tema_atual = ThemeManager.load_theme_preference()
                app = QApplication.instance()
                ThemeManager.apply_theme(app, tema_atual)
                # Carrega cores do tema para uso na tabela
                self._current_theme_colors = ThemeManager.get_status_colors(tema_atual)
                print(f"[THEME] ✅ Tema '{tema_atual}' carregado na inicialização")
            except Exception as e:
                print(f"[THEME] ⚠️ Erro ao carregar tema: {e}")
        
        self._update_window_title()
        self.resize(1200, 720)
        
        # Define o ícone da janela principal (tenta .ico primeiro, depois .png)
        icon_path = BASE_DIR / 'Logo.ico'
        if not icon_path.exists():
            icon_path = BASE_DIR / 'Logo.png'
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        ensure_logs_dir()

        self.db = UIDB(DB_PATH)
        
        # Importa módulos de sistema
        from modules.startup_manager import StartupManager
        from modules.task_scheduler import TaskScheduler
        
        # Sistema de inicialização automática
        self.startup_manager = StartupManager("BOT Busca NFE")
        
        # Sistema de tarefas agendadas
        self.task_scheduler = TaskScheduler()
        
        # System Tray (ícone na bandeja do sistema)
        self.tray_icon = None
        self._setup_system_tray()
        
        # Cache de PDFs para abertura rápida {chave: pdf_path}
        self._pdf_cache = {}
        self._cache_building = False
        self._cache_worker = None  # Referência para a thread do cache
        self._refreshing_emitidos = False  # Flag para evitar múltiplos refreshes simultâneos
        
        # Sistema de trabalhos em background
        self._trabalhos_ativos = []
        self._sync_worker = None
        self._sync_thread = None
        self._sync_cancelada_pelo_usuario = False  # Flag para não reativar auto-sync
        self._ultimo_evento_usuario = datetime.now()
        self._inatividade_timer = QTimer()
        self._inatividade_timer.timeout.connect(self._check_inatividade)
        self._inatividade_timer.start(10000)  # Verifica a cada 10 segundos
        self._auto_update_executado = False  # Flag para não repetir auto-update

        central = QWidget()
        self.setCentralWidget(central)

        # Use a horizontal splitter: left = certificados tree, right = main area (toolbar + tabs)
        main_split = QSplitter()
        main_split.setChildrenCollapsible(False)

        # Left panel: certificados tree
        left_widget = QWidget()
        left_v = QVBoxLayout(left_widget)
        left_v.setContentsMargins(4, 4, 4, 4)
        # Left header with title and filter
        lh = QHBoxLayout()
        lbl_left = QLabel("Certificados")
        f = lbl_left.font(); f.setBold(True); lbl_left.setFont(f)
        self.cert_filter = QLineEdit(); self.cert_filter.setPlaceholderText("Filtrar…")
        self.cert_filter.textChanged.connect(self._filter_certs_tree)
        btn_clear_filter = QPushButton("Limpar")
        btn_clear_filter.setToolTip("Limpar filtro/seleção")
        btn_clear_filter.clicked.connect(self._clear_cert_filter)
        lh.addWidget(lbl_left)
        lh.addStretch(1)
        lh.addWidget(self.cert_filter)
        lh.addWidget(btn_clear_filter)
        left_v.addLayout(lh)
        self.tree_certs = QTreeWidget()
        self.tree_certs.setHeaderLabels(["Certificados"])
        self.tree_certs.setColumnCount(1)
        self.tree_certs.itemClicked.connect(self._on_tree_cert_clicked)
        left_v.addWidget(self.tree_certs)

        # Right panel: toolbar + content
        right_widget = QWidget()
        v = QVBoxLayout(right_widget)
        v.setContentsMargins(4, 4, 4, 4)

        # Toolbar
        t = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Buscar por emitente, número ou CNPJ…")
        self.search_edit.textChanged.connect(self._on_filter_changed)
        
        # Filtros de data
        from PyQt5.QtCore import QDate
        date_label = QLabel("Data:")
        self.date_inicio = QDateEdit()
        self.date_inicio.setCalendarPopup(True)
        self.date_inicio.setDisplayFormat("dd/MM/yyyy")
        self.date_inicio.setDate(QDate.currentDate().addMonths(-3))  # Padrão: 3 meses atrás
        self.date_inicio.dateChanged.connect(self._on_filter_changed)
        self.date_inicio.setToolTip("Data inicial do filtro")
        
        date_ate_label = QLabel("até:")
        self.date_fim = QDateEdit()
        self.date_fim.setCalendarPopup(True)
        self.date_fim.setDisplayFormat("dd/MM/yyyy")
        self.date_fim.setDate(QDate.currentDate())  # Padrão: hoje
        self.date_fim.dateChanged.connect(self._on_filter_changed)
        self.date_fim.setToolTip("Data final do filtro")
        
        self.status_dd = QComboBox(); self.status_dd.addItems(["Todos","Autorizado","Cancelado","Denegado"])
        self.status_dd.currentTextChanged.connect(self._on_filter_changed)
        self.tipo_dd = QComboBox(); self.tipo_dd.addItems(["Todos","NFe","CTe","NFS-e"])
        self.tipo_dd.currentTextChanged.connect(self._on_filter_changed)
        
        # Seletor de quantidade de linhas exibidas
        limit_label = QLabel("Exibir:")
        self.limit_dd = QComboBox()
        self.limit_dd.addItems(["50", "100", "500", "1000", "Todos"])
        
        # Restaura a seleção salva do usuário
        settings = QSettings('NFE_System', 'BOT_NFE')
        saved_limit = settings.value('display/limit', '100')  # Padrão: 100 linhas
        self.limit_dd.setCurrentText(str(saved_limit))
        
        self.limit_dd.currentTextChanged.connect(self._on_filter_changed)
        self.limit_dd.currentTextChanged.connect(self._save_limit_preference)
        self.limit_dd.setToolTip("Quantidade de documentos a exibir na tabela")
        
        self.btn_refresh = QPushButton("Atualizar"); self.btn_refresh.clicked.connect(self.refresh_all)
        btn_busca = QPushButton("Buscar na SEFAZ"); btn_busca.clicked.connect(self.do_search)
        btn_busca_completa = QPushButton("Busca Completa"); btn_busca_completa.clicked.connect(self.do_busca_completa)
        btn_manifestacao = QPushButton("Manifestação Manual"); btn_manifestacao.clicked.connect(lambda: self._manifestar_nota(None))
        btn_manifestacao.setToolTip("Manifestar um documento digitando a chave manualmente")
        btn_exportar = QPushButton("Exportar"); btn_exportar.clicked.connect(self.abrir_exportacao)
        btn_relatorio = QPushButton("Relatório IBS/CBS"); btn_relatorio.clicked.connect(self.abrir_relatorio)
        btn_relatorio.setToolTip("Relatório analítico com IBS e CBS")
        
        # Seletor de intervalo entre buscas
        from PyQt5.QtWidgets import QSpinBox
        intervalo_label = QLabel("Buscas Em:")
        self.spin_intervalo = QSpinBox()
        self.spin_intervalo.setMinimum(1)
        self.spin_intervalo.setMaximum(23)
        self.spin_intervalo.setSuffix(" horas")
        self.spin_intervalo.setValue(self._load_intervalo_config())
        self.spin_intervalo.valueChanged.connect(self._save_intervalo_config)
        self.spin_intervalo.setToolTip("Intervalo mínimo: 1 hora | Intervalo máximo: 23 horas")
        
        # Icons (fallback to standard icons if custom not found)
        def _icon(name: str, std=None) -> QIcon:
            p = BASE_DIR / 'Icone' / name
            if p.exists():
                return QIcon(str(p))
            try:
                return self.style().standardIcon(std) if std is not None else QIcon()
            except Exception:
                return QIcon()

        try:
            from PyQt5.QtWidgets import QStyle
            self.btn_refresh.setIcon(_icon('refresh.png', QStyle.SP_BrowserReload))
            btn_busca.setIcon(_icon('search.png', QStyle.SP_FileDialogContentsView))
            btn_busca_completa.setIcon(_icon('search.png', QStyle.SP_FileDialogContentsView))
            btn_manifestacao.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
            btn_exportar.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
            btn_relatorio.setIcon(self.style().standardIcon(QStyle.SP_FileDialogInfoView))
        except Exception:
            pass
        t.addWidget(self.search_edit)
        t.addWidget(date_label)
        t.addWidget(self.date_inicio)
        t.addWidget(date_ate_label)
        t.addWidget(self.date_fim)
        t.addWidget(self.status_dd)
        t.addWidget(self.tipo_dd)
        t.addWidget(limit_label)
        t.addWidget(self.limit_dd)
        t.addStretch(1)
        t.addWidget(intervalo_label)
        t.addWidget(self.spin_intervalo)
        t.addWidget(self.btn_refresh)
        t.addWidget(btn_busca)
        t.addWidget(btn_busca_completa)
        t.addWidget(btn_manifestacao)
        t.addWidget(btn_exportar)
        t.addWidget(btn_relatorio)
        v.addLayout(t)

        # Tabs: create a tab widget to host different views (emitidos por terceiros / pela empresa)
        from PyQt5.QtWidgets import QTabWidget
        self.tabs = QTabWidget()

        # Table (main) inside first tab
        tab1 = QWidget()
        tab1_layout = QVBoxLayout(tab1)
        self.table = QTableWidget()
        headers = [
            "XML","Num","D/Emit","Tipo","Valor","Venc.",
            "Emissor CNPJ","Emissor Nome","Natureza","UF","Base ICMS",
            "Valor ICMS","IBS","CBS","Status","CFOP","NCM","Tomador IE","Chave"
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # Adiciona tooltips aos headers IBS e CBS
        header_item_ibs = self.table.horizontalHeaderItem(12)
        if header_item_ibs:
            header_item_ibs.setToolTip("💰 IBS - Imposto sobre Bens e Serviços (Reforma Tributária)")
        header_item_cbs = self.table.horizontalHeaderItem(13)
        if header_item_cbs:
            header_item_cbs.setToolTip("💰 CBS - Contribuição sobre Bens e Serviços (Reforma Tributária)")
        
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # Habilita seleção múltipla (Ctrl+Click ou Shift+Click)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # Habilita ordenação clicável nos cabeçalhos
        self.table.setSortingEnabled(True)
        # Centraliza ícones na coluna XML (coluna 0)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        # Permite arrastar colunas para reordenar
        self.table.horizontalHeader().setSectionsMovable(True)
        # Conecta sinal para salvar ordem quando mudada
        self.table.horizontalHeader().sectionMoved.connect(lambda: self._save_column_order('table'))
        # Aplica delegate para centralizar ícones
        self.table.setItemDelegateForColumn(0, CenterIconDelegate(self.table))
        # Menu de contexto para buscar XML completo
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)
        # Use o sinal específico da tabela com linha/coluna
        self.table.cellDoubleClicked.connect(self._on_table_double_clicked)
        
        # Tooltip instantâneo (100ms de delay)
        QApplication.instance().setStyleSheet(QApplication.instance().styleSheet() + 
            "QToolTip { border: 1px solid #333; padding: 4px; border-radius: 3px; }")
        # Configura delay mínimo para tooltips aparecerem rapidamente
        self.table.setMouseTracking(True)
        hh = self.table.horizontalHeader()
        try:
            hh.setSectionResizeMode(QHeaderView.Interactive)
        except Exception:
            pass
        try:
            hh.setStretchLastSection(False)
        except Exception:
            pass
        # Colunas fixas para ícones, demais auto-ajustáveis
        try:
            self.table.setColumnWidth(0, 50)  # XML - ícone fixo
        except Exception:
            pass
        tab1_layout.addWidget(self.table)
        self.tabs.addTab(tab1, "Emitidos por terceiros")

        # Second tab: Emitidos pela empresa
        tab2 = QWidget()
        tab2_layout = QVBoxLayout(tab2)
        
        # Toolbar com botões
        toolbar_emitidos = QWidget()
        toolbar_layout_emitidos = QHBoxLayout(toolbar_emitidos)
        toolbar_layout_emitidos.setContentsMargins(0, 0, 0, 5)
        
        # Botão recarregar
        btn_reload_emitidos = QPushButton("🔄 Recarregar")
        btn_reload_emitidos.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 6px 15px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
        """)
        btn_reload_emitidos.clicked.connect(self.refresh_emitidos_table)
        toolbar_layout_emitidos.addWidget(btn_reload_emitidos)
        
        # Label informativo
        lbl_info_emitidos = QLabel("⚠️ NFS-e podem estar como RESUMO - clique 2x para tentar baixar XML completo")
        lbl_info_emitidos.setStyleSheet("color: #666; font-size: 10px; padding: 5px;")
        toolbar_layout_emitidos.addWidget(lbl_info_emitidos)
        
        toolbar_layout_emitidos.addStretch()
        tab2_layout.addWidget(toolbar_emitidos)
        
        # Cria tabela para notas emitidas pela empresa
        self.table_emitidos = QTableWidget()
        headers_emitidos = [
            "XML","Num","D/Emit","Tipo","Valor","Venc.",
            "Destinatário CNPJ","Destinatário Nome","Natureza","UF","Base ICMS",
            "Valor ICMS","IBS","CBS","Status","CFOP","NCM","Tomador IE","Chave"
        ]
        self.table_emitidos.setColumnCount(len(headers_emitidos))
        self.table_emitidos.setHorizontalHeaderLabels(headers_emitidos)
        
        # Adiciona tooltips aos headers IBS e CBS
        header_item_ibs_emit = self.table_emitidos.horizontalHeaderItem(12)
        if header_item_ibs_emit:
            header_item_ibs_emit.setToolTip("💰 IBS - Imposto sobre Bens e Serviços (Reforma Tributária)")
        header_item_cbs_emit = self.table_emitidos.horizontalHeaderItem(13)
        if header_item_cbs_emit:
            header_item_cbs_emit.setToolTip("💰 CBS - Contribuição sobre Bens e Serviços (Reforma Tributária)")
        
        self.table_emitidos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_emitidos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_emitidos.setSortingEnabled(True)
        self.table_emitidos.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        # Permite arrastar colunas para reordenar
        self.table_emitidos.horizontalHeader().setSectionsMovable(True)
        # Conecta sinal para salvar ordem quando mudada
        self.table_emitidos.horizontalHeader().sectionMoved.connect(lambda: self._save_column_order('table_emitidos'))
        self.table_emitidos.setItemDelegateForColumn(0, CenterIconDelegate(self.table_emitidos))
        self.table_emitidos.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_emitidos.customContextMenuRequested.connect(self._on_table_emitidos_context_menu)
        self.table_emitidos.cellDoubleClicked.connect(self._on_table_emitidos_double_clicked)
        self.table_emitidos.setMouseTracking(True)
        
        # Configura larguras das colunas
        try:
            hh_emitidos = self.table_emitidos.horizontalHeader()
            hh_emitidos.setSectionResizeMode(QHeaderView.Interactive)
            hh_emitidos.setStretchLastSection(False)
            # Apenas coluna XML fixa, demais auto-ajustáveis
            self.table_emitidos.setColumnWidth(0, 50)  # XML - ícone fixo
        except Exception:
            pass
        
        tab2_layout.addWidget(self.table_emitidos)
        self.tabs.addTab(tab2, "Emitidos pela empresa")

        # Restaura ordem personalizada das colunas
        self._restore_column_order('table')
        self._restore_column_order('table_emitidos')

        v.addWidget(self.tabs)

        # Add widgets to splitter
        main_split.addWidget(left_widget)
        main_split.addWidget(right_widget)
        main_split.setStretchFactor(0, 0)
        main_split.setStretchFactor(1, 1)

        # Place splitter in central layout
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.addWidget(main_split)

        # Status bar com resumo de busca
        self._statusbar = QStatusBar()
        self.setStatusBar(self._statusbar)
        
        # Inicializa status com última hora de busca
        last_search_text = self._get_last_search_status()
        self.status_label = QLabel(last_search_text)
        self._statusbar.addWidget(self.status_label, 1)  # stretch=1
        
        # Copyright - Direitos Reservados (primeiro, no canto direito)
        self.copyright_label = QLabel("© 2025 DWM System Developer")
        self.copyright_label.setStyleSheet("color: #555; font-size: 9px; padding: 0 5px;")
        self._statusbar.addPermanentWidget(self.copyright_label)
        
        # Separador visual
        separator = QLabel("|")
        separator.setStyleSheet("color: #ccc; padding: 0 5px;")
        self._statusbar.addPermanentWidget(separator)
        
        # Resumo de busca (sempre visível) - sem cor hardcoded para respeitar tema
        self.search_summary_label = QLabel("")
        self.search_summary_label.setStyleSheet("font-weight: bold; padding: 0 10px;")
        self._statusbar.addPermanentWidget(self.search_summary_label)
        
        # Progress bar compacta na status bar
        self.search_progress = QProgressBar()
        self.search_progress.setMaximumWidth(200)
        self.search_progress.setMaximumHeight(16)
        self.search_progress.setTextVisible(True)
        self.search_progress.setVisible(False)
        self._statusbar.addPermanentWidget(self.search_progress)

        # Menus (Central de Tarefas)
        print("DEBUG: Iniciando criação do menu Configurações...")
        self._setup_tasks_menu()
        print("DEBUG: Menu Configurações criado com sucesso!")

        # Data cache
        self.notes = []
        
        # Estatísticas de busca
        self._search_stats = {
            'nfes_found': 0,
            'ctes_found': 0,
            'nfses_found': 0,
            'start_time': None,
            'last_cert': ''
        }

        # Incremental table fill state
        self._table_fill_timer = QTimer(self)
        self._table_fill_timer.setInterval(0)
        self._table_fill_timer.timeout.connect(self._fill_table_step)
        self._table_fill_items = []
        self._table_fill_index = 0
        self._table_fill_chunk = 100
        self._table_filling = False

        # Loading flags/workers
        self._loading_notes = False
        self._load_worker = None
        self._search_worker = None
        self._search_in_progress = False
        self._next_search_time = None
        self._selected_cert_cnpj = None  # Inicializa seleção de certificado
        self._pdf_generator_worker = None  # Thread de geração de PDFs
        self._status_timer = QTimer(self)
        self._status_timer.timeout.connect(self._update_search_status)
        self._status_timer.start(1000)  # Atualiza a cada segundo

        # Initial load
        QTimer.singleShot(50, self.refresh_all)
        # Atualiza UFs dos certificados existentes
        QTimer.singleShot(100, self._atualizar_ufs_certificados)
        # Gera PDFs faltantes
        QTimer.singleShot(500, self._gerar_pdfs_faltantes)
        # Verifica se há sincronização pendente
        QTimer.singleShot(1500, self._verificar_sync_pendente)
        # ⛔ DESABILITADO: Consulta automática de status ao iniciar
        # A consulta de eventos só deve ocorrer:
        # 1. Após busca na SEFAZ (distribuição DFe)
        # 2. Ao clicar no botão "🔄 Atualizar Status"
        # 3. Ao clicar no botão "Sincronizar Agora"
        # QTimer.singleShot(3000, self._atualizar_status_background)
        # ✅ BUSCA AUTOMÁTICA HABILITADA - Inicia após 10 segundos da inicialização
        QTimer.singleShot(10000, self._auto_start_search)
        # ✅ AGENDADOR DE TAREFAS - Verifica tarefas agendadas ao iniciar
        QTimer.singleShot(15000, self._verificar_tarefas_agendadas_inicializacao)
        # Timer periódico para verificar tarefas por horário/intervalo
        self._agendador_timer = QTimer()
        self._agendador_timer.timeout.connect(self._verificar_tarefas_agendadas_periodico)
        self._agendador_timer.start(60000)  # Verifica a cada 1 minuto
        self._apply_theme()

    def _gerar_pdfs_faltantes(self):
        """Gera PDFs para XMLs que ainda não possuem PDF (em background)."""
        # Evita iniciar múltiplas threads
        if self._pdf_generator_worker and self._pdf_generator_worker.isRunning():
            return
        
        class PDFGeneratorWorker(QThread):
            finished_signal = pyqtSignal(int)
            
            def run(self):
                try:
                    xmls_dir = DATA_DIR / "xmls"
                    if not xmls_dir.exists():
                        self.finished_signal.emit(0)
                        return
                    
                    print("[VERIFICAÇÃO] Procurando XMLs sem PDF...")
                    count = 0
                    for xml_file in xmls_dir.rglob("*.xml"):
                        # Verifica se thread deve parar
                        if self.isInterruptionRequested():
                            print("[INFO] Geração de PDFs interrompida")
                            break
                        
                        # ⛔ PULA EVENTOS - Eventos NUNCA devem gerar PDF!
                        # Verifica pelo caminho (pasta "Eventos") OU pelo nome do arquivo
                        if "Eventos" in str(xml_file.parent) or "\\Eventos\\" in str(xml_file):
                            continue
                        
                        # Pula também por palavras-chave no nome
                        nome_arquivo = xml_file.stem.upper()
                        if any(keyword in nome_arquivo for keyword in ['EVENTO', 'CIENCIA', '-CANCELAMENTO', '-CORRECAO', 'RESUMO']):
                            continue
                        
                        pdf_file = xml_file.with_suffix('.pdf')
                        if not pdf_file.exists():
                            try:
                                xml_text = xml_file.read_text(encoding='utf-8')
                                
                                # ⛔ APENAS DOCUMENTOS COMPLETOS: Pula eventos, resumos, etc
                                # Só gera PDF se for nfeProc ou cteProc (documentos completos)
                                if '<nfeProc' not in xml_text and '<cteProc' not in xml_text:
                                    continue
                                
                                # Detecta tipo do documento
                                tipo = "CTe" if "<CTe" in xml_text or "<cte" in xml_text else "NFe"
                                
                                from modules.pdf_simple import generate_danfe_pdf
                                success = generate_danfe_pdf(xml_text, str(pdf_file), tipo)
                                if success:
                                    count += 1
                                    print(f"[PDF GERADO] {pdf_file.name}")
                            except Exception as e:
                                print(f"[ERRO PDF] {xml_file.name}: {e}")
                    
                    if count > 0:
                        print(f"[CONCLUÍDO] {count} PDFs gerados")
                    else:
                        print("[INFO] Todos os XMLs já possuem PDFs")
                    
                    self.finished_signal.emit(count)
                except Exception as e:
                    print(f"[ERRO] Falha ao gerar PDFs: {e}")
                    self.finished_signal.emit(0)
        
        # Cria e inicia worker
        self._pdf_generator_worker = PDFGeneratorWorker()
        self._pdf_generator_worker.start()

    def _atualizar_ufs_certificados(self):
        """Atualiza as UFs e razões sociais dos certificados existentes consultando a API Brasil."""
        try:
            print("[DEBUG] Atualizando UFs e razões sociais dos certificados existentes...")
            
            # Mapeia UF para código
            uf_to_codigo = {
                'AC': '12', 'AL': '27', 'AP': '16', 'AM': '13', 'BA': '29',
                'CE': '23', 'DF': '53', 'ES': '32', 'GO': '52', 'MA': '21',
                'MT': '51', 'MS': '50', 'MG': '31', 'PA': '15', 'PB': '25',
                'PR': '41', 'PE': '26', 'PI': '22', 'RJ': '33', 'RN': '24',
                'RS': '43', 'RO': '11', 'RR': '14', 'SC': '42', 'SP': '35',
                'SE': '28', 'TO': '17'
            }
            
            import requests
            certificados = self.db.load_certificates()
            
            for cert in certificados:
                cert_id = cert.get('id')
                cnpj = cert.get('cnpj_cpf', '')
                uf_atual = cert.get('cUF_autor', '')
                razao_atual = cert.get('razao_social', '')
                
                if len(cnpj) == 14:  # É CNPJ
                    try:
                        url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
                        response = requests.get(url, timeout=5)
                        
                        if response.status_code == 200:
                            data = response.json()
                            uf = data.get('uf')
                            codigo_uf = uf_to_codigo.get(uf)
                            razao_social = data.get('razao_social', '')
                            
                            # Flags para controlar o que precisa atualizar
                            atualizar_uf = codigo_uf and codigo_uf != uf_atual
                            atualizar_razao = razao_social and not razao_atual
                            
                            if atualizar_uf or atualizar_razao:
                                if atualizar_uf:
                                    print(f"[DEBUG] Atualizando UF do certificado {cnpj}: {uf_atual} -> {codigo_uf} ({uf})")
                                if atualizar_razao:
                                    print(f"[DEBUG] Adicionando razão social do certificado {cnpj}: {razao_social}")
                                
                                # Atualiza no banco
                                with sqlite3.connect(str(DB_PATH)) as conn:
                                    if atualizar_uf and atualizar_razao:
                                        conn.execute("UPDATE certificados SET cUF_autor = ?, razao_social = ? WHERE id = ?", 
                                                   (codigo_uf, razao_social, cert_id))
                                    elif atualizar_uf:
                                        conn.execute("UPDATE certificados SET cUF_autor = ? WHERE id = ?", 
                                                   (codigo_uf, cert_id))
                                    elif atualizar_razao:
                                        conn.execute("UPDATE certificados SET razao_social = ? WHERE id = ?", 
                                                   (razao_social, cert_id))
                                    conn.commit()
                            elif codigo_uf == uf_atual and razao_atual:
                                print(f"[DEBUG] Certificado {cnpj} já está completo (UF: {uf_atual}, Razão: {razao_atual[:40]}...)")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao atualizar certificado {cnpj}: {e}")
            
            print("[DEBUG] Atualização de certificados concluída")
            
            # Recarrega a árvore de certificados para exibir as razões sociais
            self._populate_certs_tree()
        except Exception as e:
            print(f"[ERRO] Erro ao atualizar certificados: {e}")
    
    def _verificar_sync_pendente(self):
        """Verifica se há sincronização pendente e pergunta se quer retomar."""
        try:
            estado = self.db.get_sync_state()
            if estado:
                from datetime import datetime
                
                # Formata a data de início
                try:
                    data_inicio = datetime.fromisoformat(estado['data_inicio'])
                    data_str = data_inicio.strftime("%d/%m/%Y às %H:%M:%S")
                except:
                    data_str = "data desconhecida"
                
                processados = estado.get('docs_processados', 0)
                total = estado.get('total_docs', 0)
                restantes = total - processados
                percentual = int((processados / total) * 100) if total > 0 else 0
                
                resposta = QMessageBox.question(
                    self,
                    "🔄 Sincronização Pendente",
                    f"<b>Foi detectada uma sincronização incompleta:</b><br><br>"
                    f"📅 <b>Iniciada em:</b> {data_str}<br>"
                    f"📊 <b>Progresso:</b> {processados}/{total} documentos ({percentual}%)<br>"
                    f"⏳ <b>Restantes:</b> {restantes} documentos<br><br>"
                    f"<b>Deseja retomar de onde parou?</b><br>"
                    f"<i>(Se escolher 'Não', a sincronização será reiniciada do zero)</i>",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if resposta == QMessageBox.Yes:
                    print(f"[SYNC] Retomando sincronização: {processados}/{total} documentos")
                    # Aguarda 1 segundo para a interface carregar completamente
                    QTimer.singleShot(1000, lambda: self._retomar_sync_background(estado))
                else:
                    print("[SYNC] Usuário optou por não retomar. Limpando estado...")
                    self.db.clear_sync_state()
                    self._sync_cancelada_pelo_usuario = True  # Marca que usuário cancelou
        except Exception as e:
            print(f"[SYNC] Erro ao verificar sincronização pendente: {e}")
    
    def _atualizar_status_background(self):
        """Atualiza status das notas silenciosamente em background na inicialização."""
        try:
            # Verifica se já foi executado (evita loop infinito)
            if self._auto_update_executado:
                print("[AUTO-UPDATE] Atualização automática já foi executada, pulando...")
                return
            
            self._auto_update_executado = True  # Marca como executado
            
            print("[AUTO-UPDATE] Iniciando atualização automática de status...")
            
            # Verifica se já há uma atualização em andamento
            if hasattr(self, '_auto_update_worker') and self._auto_update_worker and self._auto_update_worker.isRunning():
                print("[AUTO-UPDATE] Atualização já está em andamento, pulando...")
                return
            
            # Obtém certificados
            certs = self.db.load_certificates()
            if not certs:
                print("[AUTO-UPDATE] Nenhum certificado configurado")
                return
            
            # Obtém lista de chaves (apenas notas com status "Autorizado" para otimizar)
            chaves = []
            for nota in self.notes:
                status = (nota.get('status') or '').lower()  # CORRIGIDO: era 'status_nota'
                chave = nota.get('chave')
                # Consulta apenas notas "autorizadas" (não consulta já canceladas)
                if chave and len(chave) == 44 and 'autoriza' in status:
                    chaves.append(chave)
            
            if not chaves:
                print("[AUTO-UPDATE] Nenhuma nota autorizada para atualizar")
                return
            
            print(f"[AUTO-UPDATE] {len(chaves)} notas serão verificadas")
            
            # Atualiza status na barra (silencioso, sem diálogo)
            self.set_status(f"🔄 Atualizando status de {len(chaves)} notas...")
            
            # Executa em thread
            from PyQt5.QtCore import QThread, pyqtSignal
            
            class UpdateStatusWorker(QThread):
                finished = pyqtSignal(dict)
                error = pyqtSignal(str)
                
                def __init__(self, db, certs, chaves):
                    super().__init__()
                    self.db = db
                    self.certs = certs
                    self.chaves = chaves
                
                def run(self):
                    try:
                        from nfe_search import atualizar_status_notas_lote
                        stats = atualizar_status_notas_lote(
                            self.db,
                            self.certs,
                            self.chaves,
                            None  # Sem callback de progresso para ser silencioso
                        )
                        self.finished.emit(stats)
                    except Exception as e:
                        self.error.emit(str(e))
            
            def on_finished(stats):
                msg = f"✅ Status atualizado: {stats.get('atualizadas', 0)} alterações"
                if stats.get('canceladas', 0) > 0:
                    msg += f" ({stats.get('canceladas', 0)} canceladas)"
                
                print(f"[AUTO-UPDATE] {msg}")
                self.set_status(msg, 5000)
                
                # Limpa referência ao worker ANTES de recarregar (evita loop)
                self._auto_update_worker = None
                
                # FORÇA recarregar dados do banco antes de atualizar tabela
                print("[AUTO-UPDATE] Recarregando dados do banco...")
                old_count = len(self.notes)
                self.notes = self.db.load_notes(limit=5000)  # Aumenta limite para 5000
                print(f"[AUTO-UPDATE] {len(self.notes)} notas carregadas (antes: {old_count})")
                
                # Corrige xml_status baseado em arquivos existentes
                self._corrigir_xml_status_automatico()
                
                # Verifica quantas estão canceladas
                canceladas_count = sum(1 for n in self.notes if 'cancel' in (n.get('status') or '').lower())
                print(f"[AUTO-UPDATE] {canceladas_count} notas canceladas detectadas nos dados")
                
                # Atualiza interface SEM chamar refresh_all (evita loop)
                try:
                    self._refresh_table_only()  # Usa método específico que não recarrega dados
                except:
                    pass  # Fallback silencioso
            
            def on_error(error_msg):
                print(f"[AUTO-UPDATE] Erro: {error_msg}")
                self.set_status("Status atualizado com erros", 3000)
                
                # Limpa referência ao worker
                self._auto_update_worker = None
            
            worker = UpdateStatusWorker(self.db, certs, chaves)
            worker.finished.connect(on_finished)
            worker.error.connect(on_error)
            worker.start()
            
            # Mantém referência ao worker
            self._auto_update_worker = worker
            
        except Exception as e:
            print(f"[AUTO-UPDATE] Erro ao iniciar atualização: {e}")
    
    def _atualizar_status_apos_busca(self):
        """Atualiza status das notas e CT-es após busca na SEFAZ (somente documentos recentes)."""
        try:
            print("[PÓS-BUSCA] Verificando documentos recentes para consulta de eventos...")
            
            # Verifica se já há uma atualização em andamento
            if hasattr(self, '_auto_update_worker') and self._auto_update_worker and self._auto_update_worker.isRunning():
                print("[PÓS-BUSCA] Atualização já está em andamento, pulando...")
                return
            
            # Obtém certificados
            certs = self.db.load_certificates()
            if not certs:
                print("[PÓS-BUSCA] Nenhum certificado configurado")
                return
            
            # Obtém apenas documentos dos ÚLTIMOS 7 DIAS (otimização)
            from datetime import datetime, timedelta
            data_limite = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
            
            chaves_nfe = []
            chaves_cte = []
            
            for nota in self.notes:
                status = (nota.get('status') or '').lower()
                chave = nota.get('chave')
                data_emissao = (nota.get('data_emissao') or '')[:10]
                tipo = (nota.get('tipo') or '').upper()
                
                # Consulta apenas documentos RECENTES e AUTORIZADOS
                if (chave and len(chave) == 44 and 
                    'autoriza' in status and 
                    data_emissao >= data_limite):
                    
                    # Separa por tipo
                    if tipo == 'CTE':
                        chaves_cte.append(chave)
                    else:
                        chaves_nfe.append(chave)
            
            total_docs = len(chaves_nfe) + len(chaves_cte)
            
            if total_docs == 0:
                print("[PÓS-BUSCA] Nenhum documento recente para atualizar")
                return
            
            print(f"[PÓS-BUSCA] {len(chaves_nfe)} NF-es e {len(chaves_cte)} CT-es recentes serão verificados")
            self.set_status(f"🔄 Verificando eventos de {total_docs} documentos recentes (NF-e: {len(chaves_nfe)}, CT-e: {len(chaves_cte)})...")
            
            # Executa em thread
            from PyQt5.QtCore import QThread, pyqtSignal
            
            class UpdateStatusWorker(QThread):
                finished = pyqtSignal(dict)
                error = pyqtSignal(str)
                
                def __init__(self, db, certs, chaves_nfe, chaves_cte):
                    super().__init__()
                    self.db = db
                    self.certs = certs
                    self.chaves_nfe = chaves_nfe
                    self.chaves_cte = chaves_cte
                
                def run(self):
                    try:
                        from nfe_search import atualizar_status_notas_lote
                        
                        # Primeiro: NF-es
                        stats_nfe = {'consultadas': 0, 'canceladas': 0, 'atualizadas': 0}
                        if self.chaves_nfe:
                            print(f"[PÓS-BUSCA] Consultando eventos de {len(self.chaves_nfe)} NF-es...")
                            stats_nfe = atualizar_status_notas_lote(
                                self.db,
                                self.certs,
                                self.chaves_nfe,
                                None  # Sem callback de progresso
                            )
                        
                        # Segundo: CT-es
                        stats_cte = {'consultadas': 0, 'canceladas': 0, 'atualizadas': 0}
                        if self.chaves_cte:
                            print(f"[PÓS-BUSCA] Consultando eventos de {len(self.chaves_cte)} CT-es...")
                            stats_cte = atualizar_status_notas_lote(
                                self.db,
                                self.certs,
                                self.chaves_cte,
                                None  # Sem callback de progresso
                            )
                        
                        # Combina estatísticas
                        stats = {
                            'consultadas': stats_nfe.get('consultadas', 0) + stats_cte.get('consultadas', 0),
                            'canceladas': stats_nfe.get('canceladas', 0) + stats_cte.get('canceladas', 0),
                            'atualizadas': stats_nfe.get('atualizadas', 0) + stats_cte.get('atualizadas', 0),
                            'nfes': len(self.chaves_nfe),
                            'ctes': len(self.chaves_cte),
                            'canceladas_nfe': stats_nfe.get('canceladas', 0),
                            'canceladas_cte': stats_cte.get('canceladas', 0)
                        }
                        
                        self.finished.emit(stats)
                    except Exception as e:
                        self.error.emit(str(e))
            
            def on_finished(stats):
                nfes = stats.get('nfes', 0)
                ctes = stats.get('ctes', 0)
                canceladas_nfe = stats.get('canceladas_nfe', 0)
                canceladas_cte = stats.get('canceladas_cte', 0)
                total_canceladas = stats.get('canceladas', 0)
                
                msg = f"✅ Eventos verificados: NF-e ({nfes}, {canceladas_nfe} canceladas) | CT-e ({ctes}, {canceladas_cte} canceladas)"
                print(f"[PÓS-BUSCA] {msg}")
                self.set_status(msg, 5000)
                
                # Limpa worker
                self._auto_update_worker = None
                
                # Recarrega dados se houver alterações
                if stats.get('atualizadas', 0) > 0:
                    print("[PÓS-BUSCA] Recarregando dados...")
                    self.notes = self.db.load_notes(limit=5000)
                    self._refresh_table_only()
                
                # 🆕 SEMPRE executa correção de status (mesmo sem atualizações)
                print("[PÓS-BUSCA] Executando correção automática de status XML...")
                QTimer.singleShot(500, lambda: self._executar_correcao_status())
            
            def on_error(error_msg):
                print(f"[PÓS-BUSCA] Erro: {error_msg}")
                self._auto_update_worker = None
            
            worker = UpdateStatusWorker(self.db, certs, chaves_nfe, chaves_cte)
            worker.finished.connect(on_finished)
            worker.error.connect(on_error)
            worker.start()
            
            # Mantém referência
            self._auto_update_worker = worker
            
        except Exception as e:
            print(f"[PÓS-BUSCA] Erro geral: {e}")
            import traceback
            traceback.print_exc()
    
    def _executar_correcao_status(self):
        """
        Executa correção de xml_status em thread separada (background task)
        
        📋 MESMA LÓGICA da função _corrigir_xml_status_automatico()
        
        Diferença: Esta função roda em thread separada (não bloqueia a UI)
        Uso: Chamada após buscas na SEFAZ para atualizar status automaticamente
        
        Ver documentação completa em: _corrigir_xml_status_automatico()
        """
        try:
            # Evita múltiplas execuções simultâneas
            if hasattr(self, '_correcao_worker') and self._correcao_worker:
                print("[CORREÇÃO] Já existe uma correção em andamento, aguardando...")
                return
            
            print("[CORREÇÃO] Iniciando correção automática de xml_status em background...")
            
            from PyQt5.QtCore import QThread, pyqtSignal
            
            class CorrecaoStatusWorker(QThread):
                finished = pyqtSignal(int)  # Número de registros corrigidos
                error = pyqtSignal(str)
                
                def __init__(self, parent_window):
                    super().__init__()
                    self.parent_window = parent_window
                
                def run(self):
                    try:
                        from pathlib import Path
                        import sqlite3
                        corrigidos = 0
                        
                        print("[CORREÇÃO-THREAD] Verificando consistência de xml_status...")
                        
                        for nota in self.parent_window.notes:
                            chave = nota.get('chave')
                            xml_status_atual = (nota.get('xml_status') or 'RESUMO').upper()
                            informante = nota.get('informante', '')
                            tipo = (nota.get('tipo') or 'NFe').strip().upper().replace('-', '')
                            data_emissao = (nota.get('data_emissao') or '')[:10]
                            numero = nota.get('nNF') or nota.get('numero')
                            
                            if not chave or not informante or not data_emissao:
                                continue
                            
                            # ⚠️ NUNCA TOCAR EM REGISTROS EVENTO (são eventos, não notas)
                            if xml_status_atual == 'EVENTO':
                                continue  # Pula, não corrige
                            
                            # Extrai ano-mês
                            year_month = data_emissao[:7] if len(data_emissao) >= 7 else None
                            if not year_month:
                                continue
                            
                            # Debug específico para NFS-e
                            is_nfse = tipo == 'NFSE'
                            if is_nfse and numero:
                                print(f"[DEBUG-NFSE] Verificando NFS-e {numero} - Status atual: {xml_status_atual}")
                                print(f"[DEBUG-NFSE] Informante: {informante}, Ano-Mês: {year_month}")
                            
                            # Verifica se arquivo existe (múltiplas possibilidades)
                            xml_path = DATA_DIR / "xmls" / informante / year_month / tipo / f"{chave}.xml"
                            pdf_path = DATA_DIR / "xmls" / informante / year_month / tipo / f"{chave}.pdf"
                            
                            # Tenta estrutura antiga também
                            if not xml_path.exists():
                                xml_path = DATA_DIR / "xmls" / informante / year_month / f"{chave}.xml"
                                pdf_path = DATA_DIR / "xmls" / informante / year_month / f"{chave}.pdf"
                            
                            # 🆕 CORREÇÃO NFS-e (Thread): Busca por múltiplos padrões
                            # Mesma lógica da função principal _corrigir_xml_status_automatico()
                            # Suporta formatos: YYYY-MM e MM-YYYY
                            if not xml_path.exists() and is_nfse and numero:
                                # Tenta formato padrão: YYYY-MM
                                pasta_nfse = DATA_DIR / "xmls" / informante / year_month / tipo
                                print(f"[DEBUG-NFSE] Pasta NFS-e (YYYY-MM): {pasta_nfse}")
                                print(f"[DEBUG-NFSE] Pasta existe? {pasta_nfse.exists()}")
                                
                                # Se não existir, tenta formato antigo: MM-YYYY
                                if not pasta_nfse.exists() and len(year_month) == 7:
                                    try:
                                        year, month = year_month.split('-')
                                        year_month_old = f"{month}-{year}"  # 2024-09 -> 09-2024
                                        pasta_nfse = DATA_DIR / "xmls" / informante / year_month_old / tipo
                                        print(f"[DEBUG-NFSE] Tentando formato antigo (MM-YYYY): {pasta_nfse}")
                                        print(f"[DEBUG-NFSE] Pasta existe? {pasta_nfse.exists()}")
                                    except:
                                        pass
                                
                                if pasta_nfse.exists():
                                    # Busca por qualquer arquivo que comece com o número
                                    # Encontra: 8189-FORNECEDOR.xml, 8189-NFSe.xml, etc
                                    import glob
                                    pattern = str(pasta_nfse / f"{numero}-*.xml")
                                    print(f"[DEBUG-NFSE] Padrão de busca: {pattern}")
                                    arquivos = glob.glob(pattern)
                                    print(f"[DEBUG-NFSE] Arquivos encontrados: {arquivos}")
                                    if arquivos:
                                        xml_path = Path(arquivos[0])
                                        pdf_path = xml_path.with_suffix('.pdf')
                                        print(f"[DEBUG-NFSE] ✅ XML encontrado: {xml_path}")
                                        print(f"[DEBUG-NFSE] PDF correspondente: {pdf_path}")
                            
                            # Verifica também no banco xmls_baixados (NÃO aplicável para NFS-e)
                            arquivo_existe = xml_path.exists() or pdf_path.exists()
                            
                            if not arquivo_existe and not is_nfse:
                                # NFS-e não usa tabela xmls_baixados, só NF-e/CT-e
                                try:
                                    with self.parent_window.db._connect() as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT xml_completo FROM xmls_baixados WHERE chave = ?", (chave,))
                                        if cursor.fetchone():
                                            arquivo_existe = True
                                except Exception:
                                    pass
                            
                            # Corrige inconsistência
                            try:
                                if is_nfse:
                                    # 🆕 NFS-e: SEMPRE é COMPLETO (não tem conceito de RESUMO)
                                    # NFS-e sempre vem completa da prefeitura, exibe ícone verde sempre
                                    if xml_status_atual != 'COMPLETO':
                                        nota['xml_status'] = 'COMPLETO'
                                        with self.parent_window.db._connect() as conn:
                                            conn.execute(
                                                "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                                (chave,)
                                            )
                                        corrigidos += 1
                                        print(f"[DEBUG-NFSE] ✅ NFS-e marcada como COMPLETO (padrão)")
                                else:
                                    # NF-e/CT-e: Tem conceito de RESUMO vs COMPLETO
                                    if arquivo_existe and xml_status_atual == 'RESUMO':
                                        nota['xml_status'] = 'COMPLETO'
                                        with self.parent_window.db._connect() as conn:
                                            conn.execute(
                                                "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                                (chave,)
                                            )
                                        corrigidos += 1
                                    elif not arquivo_existe and xml_status_atual == 'COMPLETO':
                                        nota['xml_status'] = 'RESUMO'
                                        with self.parent_window.db._connect() as conn:
                                            conn.execute(
                                                "UPDATE notas_detalhadas SET xml_status = 'RESUMO' WHERE chave = ?",
                                                (chave,)
                                            )
                                        corrigidos += 1
                            except sqlite3.OperationalError as db_error:
                                # Ignora erros de lock do banco (thread concorrente)
                                if 'locked' in str(db_error).lower():
                                    print(f"[CORREÇÃO-THREAD] ⚠️ Banco travado ao corrigir {chave}, pulando...")
                                    continue
                                else:
                                    raise
                        
                        self.finished.emit(corrigidos)
                        
                    except Exception as e:
                        import traceback
                        error_msg = f"Erro na correção: {str(e)}\n{traceback.format_exc()}"
                        self.error.emit(error_msg)
            
            def on_finished(corrigidos):
                if corrigidos > 0:
                    print(f"[CORREÇÃO] ✅ {corrigidos} registros corrigidos")
                    # Atualiza visualização
                    self._refresh_table_only()
                    self.set_status(f"✅ {corrigidos} status XML corrigidos", 5000)
                else:
                    print(f"[CORREÇÃO] ✅ Todos os registros estão consistentes")
                
                # Limpa worker
                self._correcao_worker = None
            
            def on_error(error_msg):
                print(f"[CORREÇÃO] ❌ Erro: {error_msg}")
                self._correcao_worker = None
            
            worker = CorrecaoStatusWorker(self)
            worker.finished.connect(on_finished)
            worker.error.connect(on_error)
            worker.start()
            
            # Mantém referência
            self._correcao_worker = worker
            
        except Exception as e:
            print(f"[CORREÇÃO] Erro ao iniciar worker: {e}")
            import traceback
            traceback.print_exc()
    
    def set_status(self, msg: str, timeout_ms: int = 0):
        """
        Define mensagem na barra de status.
        A cor será determinada pelo tema, não por stylesheet inline.
        """
        self.status_label.setText(msg)
        # Não define cores inline - deixa o tema controlar
        if timeout_ms:
            QTimer.singleShot(timeout_ms, lambda: self.status_label.setText("Pronto"))
    
    def _quit_application(self):
        """Encerra a aplicação completamente."""
        reply = QMessageBox.question(
            self,
            "Confirmar saída",
            "Deseja realmente encerrar o Busca XML?\n\nA busca automática de documentos fiscais será interrompida.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            print("[DEBUG] Encerrando aplicação")
            
            # Finaliza threads ativas
            if self._cache_worker and self._cache_worker.isRunning():
                print("[DEBUG] Finalizando thread do cache...")
                self._cache_worker.wait(1000)  # Aguarda até 1 segundo
                if self._cache_worker.isRunning():
                    self._cache_worker.terminate()
            
            if hasattr(self, '_load_worker') and self._load_worker and self._load_worker.isRunning():
                print("[DEBUG] Finalizando thread de carregamento...")
                self._load_worker.wait(1000)
                if self._load_worker.isRunning():
                    self._load_worker.terminate()
            
            # Finaliza threads de geração de PDF
            if hasattr(self, '_pdf_workers'):
                for worker in self._pdf_workers[:]:  # Cópia da lista
                    if worker and worker.isRunning():
                        print(f"[DEBUG] Aguardando finalização de thread PDF...")
                        worker.wait(2000)  # Aguarda até 2 segundos
                        if worker.isRunning():
                            print(f"[DEBUG] Forçando término de thread PDF...")
                            worker.terminate()
                            worker.wait(500)
                self._pdf_workers.clear()
            
            if hasattr(self, 'tray_icon'):
                self.tray_icon.hide()
            QApplication.quit()
    
    def _setup_system_tray(self):
        """Configura o ícone na bandeja do sistema"""
        icon_path = BASE_DIR / 'Logo.ico'
        if not icon_path.exists():
            icon_path = BASE_DIR / 'Logo.png'
        
        if icon_path.exists():
            self.tray_icon = QSystemTrayIcon(QIcon(str(icon_path)), self)
            
            # Menu do tray
            tray_menu = QMenu()
            
            # Ação: Mostrar/Ocultar janela
            show_action = QAction("Mostrar/Ocultar", self)
            show_action.triggered.connect(self._toggle_window)
            tray_menu.addAction(show_action)
            
            tray_menu.addSeparator()
            
            # Ação: Buscar notas
            buscar_action = QAction("🔄 Buscar Notas Agora", self)
            buscar_action.triggered.connect(self._buscar_notas_manual)
            tray_menu.addAction(buscar_action)
            
            # Ação: Gerenciador de tarefas
            tasks_action = QAction("⏱️ Gerenciador de Tarefas", self)
            tasks_action.triggered.connect(self._show_task_manager)
            tray_menu.addAction(tasks_action)
            
            tray_menu.addSeparator()
            
            # Ação: Inicialização automática
            self.startup_action = QAction("", self)
            self.startup_action.setCheckable(True)
            self.startup_action.triggered.connect(self._toggle_startup)
            self._update_startup_action_text()
            tray_menu.addAction(self.startup_action)
            
            tray_menu.addSeparator()
            
            # Ação: Sair
            quit_action = QAction("Sair", self)
            quit_action.triggered.connect(self._quit_application)
            tray_menu.addAction(quit_action)
            
            self.tray_icon.setContextMenu(tray_menu)
            self.tray_icon.activated.connect(self._tray_icon_activated)
            self.tray_icon.show()
    
    def _toggle_window(self):
        """Alterna entre mostrar/ocultar janela"""
        if self.isVisible():
            self.hide()
        else:
            self.show()
            self.activateWindow()
            self.raise_()
    
    def _tray_icon_activated(self, reason):
        """Callback quando ícone do tray é clicado"""
        if reason == QSystemTrayIcon.DoubleClick:
            self._toggle_window()
    
    def _update_startup_action_text(self):
        """Atualiza texto da ação de startup"""
        if self.startup_manager.is_startup_enabled():
            self.startup_action.setText("✓ Iniciar com Windows (Ativado)")
            self.startup_action.setChecked(True)
        else:
            self.startup_action.setText("Iniciar com Windows")
            self.startup_action.setChecked(False)
    
    def _toggle_startup(self):
        """Habilita/desabilita inicialização automática (chamado do tray menu)"""
        success = self.startup_manager.toggle_startup()
        self._update_startup_action_text()
        
        # Atualiza também o menu principal se existir
        if hasattr(self, '_act_iniciar_windows'):
            self._act_iniciar_windows.setChecked(self.startup_manager.is_startup_enabled())
        
        if success:
            if self.startup_manager.is_startup_enabled():
                QMessageBox.information(
                    self,
                    "Inicialização Automática",
                    "✓ O aplicativo agora iniciará automaticamente com o Windows.\n\n"
                    "Você pode verificar isso em:\n"
                    "Configurações > Aplicativos > Inicialização"
                )
            else:
                QMessageBox.information(
                    self,
                    "Inicialização Automática",
                    "✗ Inicialização automática desabilitada."
                )
        else:
            QMessageBox.warning(
                self,
                "Erro",
                "Não foi possível alterar a configuração de inicialização automática."
            )
    
    def _toggle_startup_menu(self):
        """Habilita/desabilita inicialização automática (chamado do menu principal)"""
        success = self.startup_manager.toggle_startup()
        
        # Atualiza checkbox do menu
        if hasattr(self, '_act_iniciar_windows'):
            self._act_iniciar_windows.setChecked(self.startup_manager.is_startup_enabled())
        
        # Atualiza também o menu do tray
        self._update_startup_action_text()
        
        if success:
            status_msg = "habilitada" if self.startup_manager.is_startup_enabled() else "desabilitada"
            self.set_status(f"✓ Inicialização automática {status_msg}", 3000)
            
            if self.startup_manager.is_startup_enabled():
                QMessageBox.information(
                    self,
                    "Inicialização Automática Ativada",
                    "✓ O aplicativo agora iniciará automaticamente com o Windows.\n\n"
                    "• Aparecerá apenas na bandeja do sistema\n"
                    "• Busca automática será executada após 10 minutos\n"
                    "• Você pode desabilitar a qualquer momento\n\n"
                    "Verifique em: Configurações do Windows > Aplicativos > Inicialização"
                )
        else:
            QMessageBox.warning(
                self,
                "Erro",
                "Não foi possível alterar a configuração de inicialização automática.\n\n"
                "Verifique se você tem permissões administrativas."
            )
    
    def _buscar_notas_manual(self):
        """Inicia busca de notas manualmente"""
        try:
            self.task_scheduler.cancel_task("Busca Automática SEFAZ")
            self.refresh_all()
            if self.tray_icon:
                self.tray_icon.showMessage(
                    "Busca Iniciada",
                    "Buscando novas notas fiscais...",
                    QSystemTrayIcon.Information,
                    3000
                )
        except Exception as e:
            print(f"[TRAY] Erro ao buscar notas: {e}")
    
    def _show_task_manager(self):
        """Mostra janela do gerenciador de tarefas"""
        from modules.task_manager_dialog import TaskManagerDialog
        dialog = TaskManagerDialog(self.task_scheduler, self)
        dialog.exec_()
    
    def _quit_application(self):
        """Fecha completamente o aplicativo"""
        reply = QMessageBox.question(
            self,
            "Confirmar Saída",
            "Deseja realmente sair do aplicativo?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.tray_icon:
                self.tray_icon.hide()
            QApplication.quit()
    
    def closeEvent(self, event: QCloseEvent):
        """Intercepta o evento de fechar a janela."""
        # Finaliza threads ativas
        if self._cache_worker and self._cache_worker.isRunning():
            self._cache_worker.wait(1000)  # Aguarda até 1 segundo
            if self._cache_worker.isRunning():
                self._cache_worker.terminate()  # Força finalização se necessário
        
        if hasattr(self, '_load_worker') and self._load_worker and self._load_worker.isRunning():
            self._load_worker.wait(1000)
            if self._load_worker.isRunning():
                self._load_worker.terminate()
        
        # Finaliza thread de busca
        if hasattr(self, '_search_worker') and self._search_worker and self._search_worker.isRunning():
            print(f"[DEBUG] Aguardando finalização de thread de busca...")
            self._search_worker.wait(2000)
            if self._search_worker.isRunning():
                print(f"[DEBUG] Forçando término de thread de busca...")
                self._search_worker.terminate()
                self._search_worker.wait(500)
        
        # Finaliza threads de geração de PDF
        if hasattr(self, '_pdf_workers'):
            for worker in self._pdf_workers[:]:  # Cópia da lista
                if worker and worker.isRunning():
                    print(f"[DEBUG] Aguardando finalização de thread PDF...")
                    worker.wait(2000)  # Aguarda até 2 segundos
                    if worker.isRunning():
                        print(f"[DEBUG] Forçando término de thread PDF...")
                        worker.terminate()
                        worker.wait(500)
            self._pdf_workers.clear()
        
        # Finaliza thread de geração de PDFs em background
        if hasattr(self, '_pdf_generator_worker') and self._pdf_generator_worker and self._pdf_generator_worker.isRunning():
            print(f"[DEBUG] Aguardando finalização de thread de geração de PDFs...")
            self._pdf_generator_worker.requestInterruption()  # Sinal para parar graciosamente
            self._pdf_generator_worker.wait(3000)  # Aguarda até 3 segundos
            if self._pdf_generator_worker.isRunning():
                print(f"[DEBUG] Forçando término de thread de geração de PDFs...")
                self._pdf_generator_worker.terminate()
                self._pdf_generator_worker.wait(500)
        
        # Ao invés de fechar, minimiza para bandeja
        event.ignore()
        self.hide()
        
        # Mostra notificação
        if hasattr(self, 'tray_icon') and self.tray_icon.isVisible():
            self.tray_icon.showMessage(
                "Busca de Notas Fiscais",
                "O sistema continua rodando em segundo plano. Clique com botão direito no ícone da bandeja para acessar as opções.",
                QSystemTrayIcon.Information,
                3000
            )
    
    def _update_search_status(self):
        """Atualiza o status da busca no rodapé."""
        try:
            if self._search_in_progress:
                self.status_label.setText("🔄 Busca em andamento...")
            elif self._next_search_time:
                from datetime import datetime
                now = datetime.now()
                diff = (self._next_search_time - now).total_seconds()
                
                if diff > 0:
                    minutes = int(diff / 60)
                    seconds = int(diff % 60)
                    if minutes > 0:
                        self.status_label.setText(f"⏳ Próxima busca em {minutes}min {seconds}s")
                    else:
                        self.status_label.setText(f"⏳ Próxima busca em {seconds}s")
                else:
                    self.status_label.setText("⏳ Iniciando próxima busca...")
                    self._next_search_time = None
            else:
                # Atualiza com última busca se não estiver buscando
                last_search_text = self._get_last_search_status()
                self.status_label.setText(last_search_text)
        except Exception as e:
            print(f"[DEBUG] Erro em _update_search_status: {e}")
            import traceback
            traceback.print_exc()

    def _setup_tasks_menu(self):
        print("DEBUG: Dentro de _setup_tasks_menu()")
        # Cria um menu 'Configurações' no menu bar com as ações principais
        from PyQt5.QtWidgets import QActionGroup
        menubar = self.menuBar()
        tarefas = menubar.addMenu("Configurações")
        print(f"DEBUG: Menu 'Configurações' criado: {tarefas}")

        # Helper para criar ações com ícone opcional (QStyle ou arquivo)
        def add_action(menu: QMenu, text: str, slot, shortcut: Optional[str] = None, icon_name: Optional[str] = None, qstyle_icon=None):
            act = QAction(text, self)
            if shortcut:
                try:
                    act.setShortcut(shortcut)
                except Exception:
                    pass
            # Tenta QStyle icon primeiro, depois arquivo
            if qstyle_icon:
                try:
                    act.setIcon(self.style().standardIcon(qstyle_icon))
                except Exception:
                    pass
            elif icon_name:
                try:
                    icon_path = BASE_DIR / 'Icone' / icon_name
                    if icon_path.exists():
                        act.setIcon(QIcon(str(icon_path)))
                except Exception:
                    pass
            act.triggered.connect(slot)
            menu.addAction(act)
            return act

        # Ações principais já presentes na toolbar
        add_action(tarefas, "Atualizar", self.refresh_all, "F5", qstyle_icon=QStyle.SP_BrowserReload)
        add_action(tarefas, "🔄 Sincronizar XMLs", self.sincronizar_xmls_interface, "Ctrl+Shift+S", qstyle_icon=QStyle.SP_FileDialogDetailedView)
        add_action(tarefas, "📥 Baixar XMLs Faltantes", self.baixar_xmls_faltantes_por_chave, "Ctrl+Shift+D", qstyle_icon=QStyle.SP_ArrowDown)
        # Removido temporariamente: add_action(tarefas, "🔄 Atualizar Status das Notas", self._atualizar_status_lote, "Ctrl+Shift+R", qstyle_icon=QStyle.SP_BrowserReload)
        tarefas.addSeparator()
        add_action(tarefas, "Buscar na SEFAZ", self.do_search, "Ctrl+B", qstyle_icon=QStyle.SP_FileDialogContentsView)
        add_action(tarefas, "Busca Completa", self.do_busca_completa, "Ctrl+Shift+B", qstyle_icon=QStyle.SP_FileDialogDetailedView)
        # Removido: PDFs em lote (função não existe)
        tarefas.addSeparator()
        add_action(tarefas, "Busca por chave", self.buscar_por_chave, "Ctrl+K", qstyle_icon=QStyle.SP_FileDialogListView)
        add_action(tarefas, "📤 Exportar", self.abrir_exportacao, "Ctrl+E", qstyle_icon=QStyle.SP_DialogSaveButton)
        add_action(tarefas, "Certificados…", self.open_certificates, "Ctrl+Shift+C", qstyle_icon=QStyle.SP_DialogApplyButton)
        add_action(tarefas, "�📁 Importar XMLs", self.importar_xmls_pasta, "Ctrl+I", qstyle_icon=QStyle.SP_DialogOpenButton)
        tarefas.addSeparator()
        add_action(tarefas, "⚙️ Gerenciador de Trabalhos", self._abrir_gerenciador_trabalhos, "Ctrl+Shift+G", qstyle_icon=QStyle.SP_ComputerIcon)
        tarefas.addSeparator()
        add_action(tarefas, "💾 Armazenamento…", self.open_storage_config, "Ctrl+Shift+A", qstyle_icon=QStyle.SP_DriveFDIcon)
        add_action(tarefas, "🔄 Resetar Ordem das Colunas", self._resetar_ordem_colunas, None, qstyle_icon=QStyle.SP_BrowserReload)
        add_action(tarefas, "💰 Atualizar IBS/CBS das Notas", self._atualizar_ibs_cbs_notas, "Ctrl+Shift+U", qstyle_icon=QStyle.SP_FileDialogInfoView)
        tarefas.addSeparator()
        
        # Submenu: Intervalo de Busca Automática
        print("DEBUG: Criando submenu Intervalo de Busca Automática...")
        intervalo_submenu = tarefas.addMenu("⏱️ Intervalo de Busca Automática")
        try:
            intervalo_submenu.setIcon(self.style().standardIcon(QStyle.SP_DialogResetButton))
        except Exception as e:
            print(f"DEBUG: Erro ao definir ícone do submenu: {e}")
        
        # Cria ações para cada intervalo (1 a 23 horas)
        print("DEBUG: Criando grupo de ações...")
        intervalo_group = QActionGroup(self)
        intervalo_group.setExclusive(True)
        intervalo_atual = self._load_intervalo_config()
        print(f"DEBUG: Intervalo atual: {intervalo_atual} horas")
        
        for horas in [1, 2, 3, 4, 6, 8, 12, 16, 20, 23]:
            act_intervalo = QAction(f"{horas} {'hora' if horas == 1 else 'horas'}", self)
            act_intervalo.setCheckable(True)
            if horas == intervalo_atual:
                act_intervalo.setChecked(True)
                print(f"DEBUG: Marcando {horas} horas como selecionado")
            act_intervalo.triggered.connect(lambda checked, h=horas: self._set_intervalo_from_menu(h))
            intervalo_group.addAction(act_intervalo)
            intervalo_submenu.addAction(act_intervalo)
        
        print(f"DEBUG: Submenu criado com {len(intervalo_submenu.actions())} ações")
        
        # Submenu: Inicialização
        print("DEBUG: Criando submenu Inicialização...")
        tarefas.addSeparator()
        inicializacao_submenu = tarefas.addMenu("🚀 Inicialização")
        try:
            inicializacao_submenu.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        except Exception as e:
            print(f"DEBUG: Erro ao definir ícone do submenu de inicialização: {e}")
        
        # Ação: Iniciar com Windows
        self._act_iniciar_windows = QAction("Iniciar automaticamente com o Windows", self)
        self._act_iniciar_windows.setCheckable(True)
        self._act_iniciar_windows.setChecked(self.startup_manager.is_startup_enabled())
        self._act_iniciar_windows.triggered.connect(self._toggle_startup_menu)
        inicializacao_submenu.addAction(self._act_iniciar_windows)
        
        inicializacao_submenu.addSeparator()
        
        # Ação: Gerenciador de Tarefas Agendadas
        act_task_manager = QAction("⏱️ Gerenciador de Tarefas Agendadas", self)
        try:
            act_task_manager.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        except Exception:
            pass
        act_task_manager.triggered.connect(self._show_task_manager)
        inicializacao_submenu.addAction(act_task_manager)
        
        inicializacao_submenu.addSeparator()
        
        # Informação sobre busca automática
        info_action = QAction("ℹ️ Busca automática após 10 minutos (modo startup)", self)
        info_action.setEnabled(False)  # Apenas informativo
        inicializacao_submenu.addAction(info_action)
        
        print(f"DEBUG: Submenu de inicialização criado com {len(inicializacao_submenu.actions())} ações")
        
        # 🎨 Submenu: Temas
        if THEMES_AVAILABLE:
            print("DEBUG: Criando submenu Temas...")
            tarefas.addSeparator()
            temas_submenu = tarefas.addMenu("🎨 Temas")
            try:
                temas_submenu.setIcon(self.style().standardIcon(QStyle.SP_DesktopIcon))
            except Exception as e:
                print(f"DEBUG: Erro ao definir ícone do submenu de temas: {e}")
            
            # Cria grupo de ações para seleção exclusiva
            temas_group = QActionGroup(self)
            temas_group.setExclusive(True)
            
            # Obtém tema atual
            tema_atual = ThemeManager.load_theme_preference()
            print(f"DEBUG: Tema atual: {tema_atual}")
            
            # Adiciona ação para cada tema disponível
            for tema_nome in ThemeManager.get_theme_names():
                tema_info = ThemeManager.get_theme_info(tema_nome)
                act_tema = QAction(f"{tema_nome}", self)
                act_tema.setCheckable(True)
                act_tema.setToolTip(tema_info['description'])
                
                # Marca o tema atual
                if tema_nome == tema_atual:
                    act_tema.setChecked(True)
                    print(f"DEBUG: Marcando '{tema_nome}' como selecionado")
                
                # Conecta ao método de aplicação de tema
                act_tema.triggered.connect(lambda checked, nome=tema_nome: self._aplicar_tema(nome))
                temas_group.addAction(act_tema)
                temas_submenu.addAction(act_tema)
            
            print(f"DEBUG: Submenu de temas criado com {len(temas_submenu.actions())} temas")
        
        # Checkbox: Consultar Status na SEFAZ
        print("DEBUG: Criando checkbox Consultar Status...")
        self._act_consultar_status = QAction("✅ Consultar Status na SEFAZ", self)
        tarefas.addSeparator()
        add_action(tarefas, "🔄 Atualizações", self.check_updates, "Ctrl+U", qstyle_icon=QStyle.SP_BrowserReload)
        tarefas.addSeparator()
        add_action(tarefas, "Limpar", self.limpar_dados, "Ctrl+Shift+L", qstyle_icon=QStyle.SP_TrashIcon)
        tarefas.addSeparator()
        add_action(tarefas, "Abrir XMLs", self.open_xmls_folder, "Ctrl+Shift+X", qstyle_icon=QStyle.SP_DirIcon)
        add_action(tarefas, "Abrir logs", self.open_logs_folder, "Ctrl+L", qstyle_icon=QStyle.SP_FileDialogInfoView)

        # Alternativa: alternar 'PDF simples' (guarda em QSettings). Útil para modo seguro.
        try:
            self._act_pdf_simples = QAction("Usar PDF simples (modo seguro)", self)
            self._act_pdf_simples.setCheckable(True)
            s = QSettings('NFE_System', 'BOT_NFE')
            enabled = bool(s.value('pdf/force_simple_fallback', False, type=bool))
            self._act_pdf_simples.setChecked(enabled)
            def _toggle_pdf_simple(checked: bool):
                try:
                    QSettings('NFE_System', 'BOT_NFE').setValue('pdf/force_simple_fallback', bool(checked))
                except Exception:
                    pass
                self.set_status("PDF simples: " + ("ativado" if checked else "desativado"), 2500)
            self._act_pdf_simples.toggled.connect(_toggle_pdf_simple)
            tarefas.addSeparator()
            tarefas.addAction(self._act_pdf_simples)
        except Exception as e:
            print(f"DEBUG: Erro ao adicionar PDF simples: {e}")
        
        # Contar ações no menu
        total_acoes = len(tarefas.actions())
        print(f"DEBUG: Total de ações no menu 'Configurações': {total_acoes}")
        print("DEBUG: Listando todas as ações:")
        for i, action in enumerate(tarefas.actions(), 1):
            if action.isSeparator():
                print(f"  {i}. [SEPARADOR]")
            elif action.menu():
                print(f"  {i}. {action.text()} [SUBMENU com {len(action.menu().actions())} itens]")
            else:
                print(f"  {i}. {action.text()}")

    def sincronizar_xmls_interface(self):
        """Sincroniza dados da interface com XMLs físicos.
        Remove registros órfãos (sem XML correspondente)."""
        try:
            from pathlib import Path
            import os
            
            # Confirma operação
            reply = QMessageBox.question(
                self,
                "Sincronizar XMLs",
                "Esta operação irá:\n\n"
                "• Verificar todos os registros na interface\n"
                "• Remover registros sem XML físico correspondente\n"
                "• Verificar XMLs em pastas locais e banco de dados\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            self.set_status("Sincronizando XMLs com interface...", 0)
            QApplication.processEvents()
            
            # Estatísticas
            total_registros = 0
            registros_ok = 0
            registros_removidos = 0
            chaves_removidas = []
            
            # Carrega todos os registros do banco
            with sqlite3.connect(str(DB_PATH)) as conn:
                rows = conn.execute("SELECT chave FROM xmls_baixados").fetchall()
                total_registros = len(rows)
                
                # Cria diálogo de progresso
                from PyQt5.QtWidgets import QProgressDialog
                progress = QProgressDialog(
                    "Sincronizando XMLs...",
                    "Cancelar",
                    0,
                    total_registros,
                    self
                )
                progress.setWindowTitle("Sincronização")
                progress.setWindowModality(Qt.WindowModal)
                progress.setMinimumDuration(0)
                progress.setValue(0)
                
                cancelado = False
                
                for idx, (chave,) in enumerate(rows):
                    if progress.wasCanceled():
                        cancelado = True
                        break
                    
                    # Atualiza progresso
                    progress.setLabelText(
                        f"Verificando registro {idx + 1}/{total_registros}\n"
                        f"Chave: ...{chave[-12:]}\n\n"
                        f"✅ Encontrados: {registros_ok}\n"
                        f"❌ Removidos: {registros_removidos}"
                    )
                    progress.setValue(idx)
                    QApplication.processEvents()
                    
                    xml_existe = False
                    
                    # Busca em pastas locais
                    pastas = [
                        DATA_DIR / "xmls",
                        DATA_DIR / "xmls_chave",
                        DATA_DIR / "xml_NFs",
                        DATA_DIR / "xml_envio",
                        DATA_DIR / "xml_extraidos",
                        DATA_DIR / "xml_resposta_sefaz"
                    ]
                    
                    for pasta in pastas:
                        if not pasta.exists():
                            continue
                        
                        # Busca por nome de arquivo contendo a chave
                        for xml_file in pasta.rglob(f"*{chave}*.xml"):
                            xml_existe = True
                            break
                        
                        if xml_existe:
                            break
                    
                    if xml_existe:
                        registros_ok += 1
                    else:
                        # Remove registro órfão
                        conn.execute("DELETE FROM xmls_baixados WHERE chave = ?", (chave,))
                        registros_removidos += 1
                        chaves_removidas.append(chave)
                
                progress.setValue(total_registros)
                progress.close()
                
                if not cancelado:
                    conn.commit()
                else:
                    conn.rollback()
            
            # Mostra resumo
            if cancelado:
                mensagem = f"⏸️ Sincronização cancelada!\n\n"
            else:
                mensagem = f"✅ Sincronização concluída!\n\n"
            
            mensagem += f"Total de registros verificados: {idx + 1}/{total_registros}\n"
            mensagem += f"✅ XMLs encontrados: {registros_ok}\n"
            mensagem += f"❌ Registros removidos: {registros_removidos}\n"
            
            if chaves_removidas:
                mensagem += f"\n📋 Chaves removidas (primeiras 10):\n"
                for chave in chaves_removidas[:10]:
                    mensagem += f"  • {chave[:8]}...{chave[-8:]}\n"
                
                if len(chaves_removidas) > 10:
                    mensagem += f"  ... e mais {len(chaves_removidas) - 10} chaves\n"
            
            QMessageBox.information(self, "Sincronização Concluída", mensagem)
            
            # Atualiza interface
            self.refresh_all()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro na Sincronização", f"Erro ao sincronizar XMLs: {e}")
        finally:
            self.set_status("Sincronização concluída", 3000)

    def baixar_xmls_faltantes_por_chave(self):
        """Baixa XMLs completos usando consulta por chave (sem erro 656).
        Ideal para buscar XMLs que faltam quando tem bloqueio no NSU."""
        try:
            # Busca chaves sem arquivo XML
            with sqlite3.connect(str(DB_PATH)) as conn:
                # Busca chaves que não tem caminho_arquivo
                rows = conn.execute("""
                    SELECT xmls_baixados.chave, xmls_baixados.cnpj_cpf, 
                           certificados.cnpj_cpf, certificados.caminho, 
                           certificados.senha, certificados.cUF_autor
                    FROM xmls_baixados
                    JOIN certificados ON xmls_baixados.cnpj_cpf = certificados.informante
                    WHERE (xmls_baixados.caminho_arquivo IS NULL OR xmls_baixados.caminho_arquivo = '')
                    LIMIT 100
                """).fetchall()
                
                total_faltantes = len(rows)
            
            if total_faltantes == 0:
                QMessageBox.information(
                    self,
                    "XMLs Completos",
                    "✅ Todos os registros já possuem XML completo!"
                )
                return
            
            # Confirma operação
            reply = QMessageBox.question(
                self,
                "Baixar XMLs Faltantes",
                f"Encontradas {total_faltantes} chaves sem XML completo.\n\n"
                f"Esta operação irá:\n"
                f"• Consultar cada chave individualmente na SEFAZ\n"
                f"• Baixar XML completo usando consulta por chave\n"
                f"• NÃO gera erro 656 (pode usar a qualquer momento)\n"
                f"• Respeita limite de ~50 consultas/minuto\n\n"
                f"⏱️ Tempo estimado: ~{(total_faltantes / 50):.0f}-{(total_faltantes / 40):.0f} minutos\n\n"
                f"Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Importa função de consulta
            import sys
            sys.path.insert(0, str(BASE_DIR))
            from nfe_search import consultar_nfe_por_chave
            
            # Importa sistema de descriptografia
            try:
                from modules.crypto_portable import get_portable_crypto as get_crypto
                crypto = get_crypto()
                CRYPTO_AVAILABLE = True
            except ImportError:
                CRYPTO_AVAILABLE = False
                print("⚠️ Sistema de criptografia não disponível")
            
            # Dicionário para rastrear certificados com problemas
            certificados_com_erro = {}
            
            # Progresso
            from PyQt5.QtWidgets import QProgressDialog
            progress = QProgressDialog(
                "Baixando XMLs por chave...",
                "Cancelar",
                0,
                total_faltantes,
                self
            )
            progress.setWindowTitle("Baixando XMLs")
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            
            sucessos = 0
            falhas = 0
            cancelado = False
            
            import time
            
            for idx, (chave, cnpj_cpf, cnpj_cert, cert_path, senha_encriptada, cuf) in enumerate(rows):
                if progress.wasCanceled():
                    cancelado = True
                    break
                
                try:
                    # Descriptografa a senha se necessário
                    senha = senha_encriptada
                    if CRYPTO_AVAILABLE and senha_encriptada:
                        try:
                            senha = crypto.decrypt(senha_encriptada)
                        except Exception as e:
                            # Se falhar ao descriptografar, assume que é texto plano
                            print(f"⚠️ Usando senha em texto plano (não descriptografada): {e}")
                            senha = senha_encriptada
                    
                    progress.setLabelText(
                        f"Baixando XML {idx + 1}/{total_faltantes}\n"
                        f"Chave: ...{chave[-12:]}\n"
                        f"✅ Sucessos: {sucessos} | ❌ Falhas: {falhas}"
                    )
                    progress.setValue(idx)
                    QApplication.processEvents()
                    
                    # Consulta por chave (sem erro 656!)
                    xml = consultar_nfe_por_chave(chave, cert_path, senha, cnpj_cert, cuf)
                    
                    if xml:
                        # Salva XML em arquivo na pasta xmls_chave/cnpj/
                        pasta_cnpj = DATA_DIR / "xmls_chave" / cnpj_cpf
                        pasta_cnpj.mkdir(parents=True, exist_ok=True)
                        
                        caminho_xml = pasta_cnpj / f"{chave}.xml"
                        caminho_xml.write_text(xml, encoding='utf-8')
                        
                        # Atualiza caminho no banco
                        with sqlite3.connect(str(DB_PATH)) as conn:
                            conn.execute(
                                "UPDATE xmls_baixados SET caminho_arquivo = ? WHERE chave = ?",
                                (str(caminho_xml), chave)
                            )
                            conn.commit()
                        sucessos += 1
                    else:
                        falhas += 1
                    
                    # Rate limit: ~50 por minuto = 1.2 segundos entre requisições
                    if (idx + 1) % 50 == 0:
                        # A cada 50 requisições, aguarda 1 minuto
                        for segundo in range(60, 0, -1):
                            if progress.wasCanceled():
                                cancelado = True
                                break
                            progress.setLabelText(
                                f"⏸️ Rate limit: aguardando {segundo}s...\n"
                                f"Processados: {idx + 1}/{total_faltantes}\n"
                                f"✅ Sucessos: {sucessos} | ❌ Falhas: {falhas}"
                            )
                            QApplication.processEvents()
                            time.sleep(1)
                        
                        if cancelado:
                            break
                    else:
                        # Pequeno delay entre requisições
                        time.sleep(1.2)
                    
                except Exception as e:
                    erro_msg = str(e)
                    print(f"Erro ao consultar chave {chave}: {erro_msg}")
                    
                    # Rastreia certificados com problema
                    cert_info = f"{cert_cnpj} (UF {cuf}) - {cert_path}"
                    if cert_info not in certificados_com_erro:
                        certificados_com_erro[cert_info] = {
                            'count': 0,
                            'erro': erro_msg,
                            'chaves': []
                        }
                    certificados_com_erro[cert_info]['count'] += 1
                    if len(certificados_com_erro[cert_info]['chaves']) < 3:
                        certificados_com_erro[cert_info]['chaves'].append(chave[:20] + '...')
                    
                    falhas += 1
            
            progress.setValue(total_faltantes)
            progress.close()
            
            # Mostra resumo
            if cancelado:
                mensagem = f"⏸️ Operação cancelada!\n\n"
            else:
                mensagem = f"✅ Download concluído!\n\n"
            
            mensagem += f"Total processado: {idx + 1}/{total_faltantes}\n"
            mensagem += f"✅ Sucessos: {sucessos}\n"
            mensagem += f"❌ Falhas: {falhas}\n"
            
            if sucessos > 0:
                mensagem += f"\n💾 {sucessos} XMLs salvos no banco de dados!"
            
            # Adiciona relatório de certificados com problema
            if certificados_com_erro:
                mensagem += "\n\n⚠️ CERTIFICADOS COM PROBLEMAS:\n"
                mensagem += "=" * 50 + "\n"
                for cert_info, dados in certificados_com_erro.items():
                    mensagem += f"\n📜 {cert_info}\n"
                    mensagem += f"   Erro: {dados['erro']}\n"
                    mensagem += f"   Falhas: {dados['count']}\n"
                    if dados['chaves']:
                        mensagem += f"   Exemplos: {', '.join(dados['chaves'])}\n"
            
            QMessageBox.information(self, "Download Concluído", mensagem)
            
            # Atualiza interface
            if sucessos > 0:
                self.refresh_all()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao baixar XMLs: {e}")

    def _save_column_order(self, table_name: str):
        """Salva a ordem das colunas quando o usuário reorganiza"""
        try:
            settings = QSettings('NFE_System', 'BOT_NFE')
            table = self.table if table_name == 'table' else self.table_emitidos
            header = table.horizontalHeader()
            
            # Obtém a ordem visual das colunas
            order = []
            for i in range(header.count()):
                order.append(header.visualIndex(i))
            
            # Salva no QSettings
            settings.setValue(f'columns/{table_name}/order', order)
            print(f"✅ Ordem de colunas salva para {table_name}: {order}")
        except Exception as e:
            print(f"❌ Erro ao salvar ordem de colunas: {e}")
    
    def _atualizar_ibs_cbs_notas(self):
        """Atualiza IBS e CBS de todas as notas NFe existentes no banco"""
        try:
            from lxml import etree
            
            # Confirmação do usuário
            reply = QMessageBox.question(
                self,
                "Atualizar IBS/CBS",
                "Esta operação irá atualizar os valores de IBS e CBS de todas as notas NFe "
                "existentes no banco de dados, extraindo os valores dos XMLs salvos.\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Busca notas sem IBS/CBS
            with self.db._connect() as conn:
                cursor = conn.execute("""
                    SELECT chave, tipo
                    FROM notas_detalhadas
                    WHERE (v_ibs IS NULL OR v_ibs = '' OR v_ibs = '0' OR v_ibs = '0.00')
                      AND (v_cbs IS NULL OR v_cbs = '' OR v_cbs = '0' OR v_cbs = '0.00')
                      AND tipo = 'NFe'
                """)
                notas = cursor.fetchall()
            
            total = len(notas)
            
            if total == 0:
                QMessageBox.information(
                    self,
                    "IBS/CBS Atualizado",
                    "Todas as notas NFe já possuem IBS/CBS atualizados!"
                )
                return
            
            # Progress dialog
            progress = QProgressDialog(
                f"Atualizando IBS/CBS de {total} notas...",
                "Cancelar",
                0,
                total,
                self
            )
            progress.setWindowTitle("Atualizando IBS/CBS")
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            atualizadas = 0
            nao_encontradas = 0
            sem_valores = 0
            
            # Namespaces para XML
            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
            
            for idx, (chave, tipo) in enumerate(notas):
                if progress.wasCanceled():
                    break
                
                progress.setValue(idx)
                progress.setLabelText(
                    f"Processando nota {idx+1}/{total}...\n"
                    f"✅ Atualizadas: {atualizadas} | ⚠️ XML não encontrado: {nao_encontradas} | ℹ️ Sem valores: {sem_valores}"
                )
                QApplication.processEvents()
                
                try:
                    # Busca XML no banco primeiro
                    xml_path = None
                    with self.db._connect() as conn:
                        cursor = conn.execute(
                            "SELECT caminho_arquivo FROM xmls_baixados WHERE chave = ?",
                            (chave,)
                        )
                        row = cursor.fetchone()
                        if row and row[0]:
                            xml_path = Path(row[0])
                            if not xml_path.exists():
                                xml_path = None
                    
                    # Se não encontrou no banco, busca nos diretórios
                    if not xml_path:
                        # Busca recursiva em xmls/
                        xmls_dir = Path(self.BASE_DIR) / 'xmls'
                        if xmls_dir.exists():
                            matches = list(xmls_dir.rglob(f"*{chave}*.xml"))
                            if matches:
                                xml_path = matches[0]
                    
                    if not xml_path or not xml_path.exists():
                        nao_encontradas += 1
                        continue
                    
                    # Extrai IBS e CBS do XML
                    with open(xml_path, 'r', encoding='utf-8') as f:
                        xml_content = f.read()
                    
                    tree = etree.fromstring(xml_content.encode('utf-8'))
                    
                    v_ibs = ''
                    v_cbs = ''
                    
                    # Tenta extrair do grupo IBSCBSTot
                    ibs_cbs_tot = tree.find('.//{http://www.portalfiscal.inf.br/nfe}IBSCBSTot')
                    if ibs_cbs_tot is not None:
                        g_ibs = ibs_cbs_tot.find('{http://www.portalfiscal.inf.br/nfe}gIBS')
                        if g_ibs is not None:
                            v_ibs = g_ibs.findtext('{http://www.portalfiscal.inf.br/nfe}vIBS') or ''
                        g_cbs = ibs_cbs_tot.find('{http://www.portalfiscal.inf.br/nfe}gCBS')
                        if g_cbs is not None:
                            v_cbs = g_cbs.findtext('{http://www.portalfiscal.inf.br/nfe}vCBS') or ''
                    
                    # Se não encontrou, tenta buscar direto no ICMSTot
                    if not v_ibs:
                        ibs_tags = tree.xpath('.//nfe:ICMSTot/nfe:vIBS', namespaces=ns)
                        if ibs_tags and ibs_tags[0].text:
                            v_ibs = ibs_tags[0].text
                    
                    if not v_cbs:
                        cbs_tags = tree.xpath('.//nfe:ICMSTot/nfe:vCBS', namespaces=ns)
                        if cbs_tags and cbs_tags[0].text:
                            v_cbs = cbs_tags[0].text
                    
                    # Busca sem namespace (fallback)
                    if not v_ibs:
                        ibs_no_ns = tree.xpath(".//*[local-name()='vIBS']")
                        if ibs_no_ns and ibs_no_ns[0].text:
                            v_ibs = ibs_no_ns[0].text
                    
                    if not v_cbs:
                        cbs_no_ns = tree.xpath(".//*[local-name()='vCBS']")
                        if cbs_no_ns and cbs_no_ns[0].text:
                            v_cbs = cbs_no_ns[0].text
                    
                    if v_ibs or v_cbs:
                        # Atualiza no banco
                        with self.db._connect() as conn:
                            conn.execute("""
                                UPDATE notas_detalhadas 
                                SET v_ibs = ?, v_cbs = ?
                                WHERE chave = ?
                            """, (v_ibs or '0', v_cbs or '0', chave))
                            conn.commit()
                        atualizadas += 1
                    else:
                        sem_valores += 1
                
                except Exception as e:
                    print(f"Erro ao processar nota {chave}: {e}")
                    nao_encontradas += 1
            
            progress.setValue(total)
            
            # Mostra resultado
            mensagem = (
                f"Atualização de IBS/CBS concluída!\n\n"
                f"✅ Notas atualizadas: {atualizadas}\n"
                f"⚠️ XMLs não encontrados: {nao_encontradas}\n"
                f"ℹ️ Notas sem valores IBS/CBS: {sem_valores}\n"
                f"📋 Total processado: {total}"
            )
            
            QMessageBox.information(self, "Atualização Concluída", mensagem)
            
            # Atualiza a interface se houve alterações
            if atualizadas > 0:
                print(f"✅ {atualizadas} notas foram atualizadas com IBS/CBS")
                self.refresh_all()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar IBS/CBS: {e}")
            import traceback
            traceback.print_exc()
    
    def _resetar_ordem_colunas(self):
        """Reseta a ordem das colunas para o padrão"""
        try:
            reply = QMessageBox.question(
                self,
                "Resetar Ordem das Colunas",
                "Isso irá resetar a ordem das colunas para o padrão original.\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            settings = QSettings('NFE_System', 'BOT_NFE')
            
            # Remove configurações salvas
            settings.remove('columns/table/order')
            settings.remove('columns/table_emitidos/order')
            
            print("✅ Configuração de ordem de colunas resetada")
            
            # Informa ao usuário
            QMessageBox.information(
                self,
                "Ordem Resetada",
                "A ordem das colunas foi resetada para o padrão.\n\n"
                "Reinicie o aplicativo para aplicar as mudanças."
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao resetar ordem das colunas: {e}")
    
    def _restore_column_order(self, table_name: str):
        """Restaura a ordem das colunas salva pelo usuário"""
        try:
            settings = QSettings('NFE_System', 'BOT_NFE')
            saved_order = settings.value(f'columns/{table_name}/order', None)
            
            if not saved_order:
                return  # Sem preferência salva, usa ordem padrão
            
            table = self.table if table_name == 'table' else self.table_emitidos
            header = table.horizontalHeader()
            
            # ⚠️ VALIDAÇÃO: Verifica se a ordem salva é válida
            if not isinstance(saved_order, list):
                print(f"⚠️ Ordem salva inválida (não é lista) para {table_name}, usando padrão")
                settings.remove(f'columns/{table_name}/order')  # Remove configuração inválida
                return
            
            if len(saved_order) != header.count():
                print(f"⚠️ Ordem salva tem tamanho incorreto ({len(saved_order)} vs {header.count()}) para {table_name}, usando padrão")
                settings.remove(f'columns/{table_name}/order')  # Remove configuração inválida
                return
            
            # ⚠️ VALIDAÇÃO: Verifica se todos os índices são válidos (sem duplicatas)
            try:
                indices_unicos = set(int(v) for v in saved_order)
                if len(indices_unicos) != len(saved_order):
                    print(f"⚠️ Ordem salva contém duplicatas para {table_name}, usando padrão")
                    settings.remove(f'columns/{table_name}/order')
                    return
                
                # Verifica se todos os índices estão no range válido
                if not all(0 <= int(v) < header.count() for v in saved_order):
                    print(f"⚠️ Ordem salva contém índices fora do range para {table_name}, usando padrão")
                    settings.remove(f'columns/{table_name}/order')
                    return
            except (ValueError, TypeError) as e:
                print(f"⚠️ Erro ao validar ordem salva para {table_name}: {e}, usando padrão")
                settings.remove(f'columns/{table_name}/order')
                return
            
            # Restaura a ordem visual
            for logical_index, visual_index in enumerate(saved_order):
                visual_index = int(visual_index)
                current_visual = header.visualIndex(logical_index)
                header.moveSection(current_visual, visual_index)
            print(f"✅ Ordem de colunas restaurada para {table_name}: {saved_order}")
        except Exception as e:
            print(f"❌ Erro ao restaurar ordem de colunas: {e}")
            # Remove configuração problemática
            try:
                settings = QSettings('NFE_System', 'BOT_NFE')
                settings.remove(f'columns/{table_name}/order')
            except:
                pass

    def refresh_all(self):
        # Evita reentrância e trava de UI: carrega notas em thread
        if self._loading_notes:
            return
        self._loading_notes = True
        try:
            if self.btn_refresh:
                self.btn_refresh.setEnabled(False)
        except Exception:
            pass
        self.set_status("Carregando…")

        class LoadNotesWorker(QThread):
            finished_notes = pyqtSignal(list)
            def __init__(self, db: UIDB, limit: int = 5000):
                super().__init__()
                self.db = db
                self.limit = limit
            def run(self):
                try:
                    notes = self.db.load_notes(limit=self.limit)
                except Exception:
                    notes = []
                self.finished_notes.emit(notes)

        def on_loaded(notes: List[Dict[str, Any]]):
            try:
                self.notes = notes or []
                
                # Corrige xml_status baseado em arquivos existentes
                self._corrigir_xml_status_automatico()
                
                try:
                    self._populate_certs_tree()
                except Exception:
                    pass
                self.refresh_table()
                self.refresh_emitidos_table()  # Popula também a tabela de emitidos
                self.set_status(f"{len(self.notes)} registros carregados", 3000)
                
                # Constrói cache de PDFs em background (não bloqueia UI)
                if self.notes and not self._cache_building:
                    self._build_pdf_cache_async()
                
                # Gera PDFs faltantes em segundo plano
                QTimer.singleShot(1000, self._gerar_pdfs_faltantes)
                    
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao carregar dados: {e}")
            finally:
                self._loading_notes = False
                try:
                    if self.btn_refresh:
                        self.btn_refresh.setEnabled(True)
                except Exception:
                    pass

        def on_thread_finished():
            """Chamado quando a thread realmente terminou de executar"""
            if self._load_worker:
                self._load_worker.deleteLater()
                self._load_worker = None

        self._load_worker = LoadNotesWorker(self.db, limit=5000)
        self._load_worker.finished_notes.connect(on_loaded)
        self._load_worker.finished.connect(on_thread_finished)
        self._load_worker.start()
    
    def _refresh_table_only(self):
        """Atualiza apenas a visualização da tabela sem recarregar dados do banco"""
        try:
            print("[REFRESH] Atualizando visualização da tabela (sem recarregar dados)...")
            self.refresh_table()
            self.refresh_emitidos_table()
            print("[REFRESH] Tabelas atualizadas")
        except Exception as e:
            print(f"[REFRESH] Erro ao atualizar tabelas: {e}")
    
    def _corrigir_xml_status_automatico(self):
        """
        ⚡ CORREÇÃO AUTOMÁTICA DE STATUS XML (Chamado pelo botão "Atualizar")
        
        Verifica a existência física de arquivos XML/PDF no disco e corrige o status no banco.
        
        📋 PADRÕES DE NOMENCLATURA SUPORTADOS:
        
        NF-e e CT-e:
        ├─ {CHAVE}.xml (ex: 52026015045348000172570010014777191002562584.xml)
        └─ Estrutura: xmls/{CNPJ}/{ANO-MES}/{TIPO}/{CHAVE}.xml
        
        NFS-e (MÚLTIPLOS PADRÕES):
        ├─ {NUMERO}-{FORNECEDOR}.xml  (ex: 8189-AILTON MORAIS JARDIM.xml) ✅ NOVO
        ├─ {NUMERO}-NFSe.xml          (ex: 8189-NFSe.xml)
        ├─ NFSe_{NUMERO}.xml          (ex: NFSe_8189.xml)
        └─ Busca via glob: "{NUMERO}-*.xml" para encontrar qualquer variação
        
        🔍 LÓGICA DE VERIFICAÇÃO:
        
        **NF-e / CT-e:**
        - Arquivo existe → COMPLETO
        - Arquivo não existe → RESUMO
        
        **NFS-e:**
        - ✅ SEMPRE COMPLETO (não tem conceito de RESUMO)
        - NFS-e sempre vem completa da prefeitura
        - Ícone verde ✅ sempre visível quando NFS-e está no banco
        
        📊 CASOS DE USO:
        - XMLs salvos com novo padrão: {NUMERO}-{FORNECEDOR}.xml
        - XMLs antigos migrados: {NUMERO}-NFSe.xml
        - XMLs de diferentes fontes: NFSe_{NUMERO}.xml
        
        ✅ RESULTADO: Ícone verde aparece na interface (NF-e/CT-e quando arquivo existe, NFS-e sempre)
        """
        try:
            print("[CORREÇÃO] Verificando consistência de xml_status...")
            corrigidos = 0
            
            for nota in self.notes:
                chave = nota.get('chave')
                xml_status_atual = (nota.get('xml_status') or 'RESUMO').upper()
                informante = nota.get('informante', '')
                tipo = (nota.get('tipo') or 'NFe').strip().upper().replace('-', '')
                data_emissao = (nota.get('data_emissao') or '')[:10]
                numero = nota.get('nNF') or nota.get('numero')
                
                if not chave or not informante or not data_emissao:
                    continue
                
                # ⚠️ NUNCA TOCAR EM REGISTROS EVENTO (são eventos, não notas)
                if xml_status_atual == 'EVENTO':
                    continue
                
                # Extrai ano-mês
                year_month = data_emissao[:7] if len(data_emissao) >= 7 else None
                if not year_month:
                    continue
                
                # Identifica tipo de documento
                is_nfse = tipo == 'NFSE'
                
                # Verifica se arquivo existe (múltiplas possibilidades)
                xml_path = DATA_DIR / "xmls" / informante / year_month / tipo / f"{chave}.xml"
                pdf_path = DATA_DIR / "xmls" / informante / year_month / tipo / f"{chave}.pdf"
                
                # Tenta estrutura antiga também
                if not xml_path.exists():
                    xml_path = DATA_DIR / "xmls" / informante / year_month / f"{chave}.xml"
                    pdf_path = DATA_DIR / "xmls" / informante / year_month / f"{chave}.pdf"
                
                # 🆕 CORREÇÃO NFS-e: Busca por múltiplos padrões de nomenclatura
                # Problema resolvido: NFS-e tem nomes variados ({NUMERO}-{FORNECEDOR}.xml, {NUMERO}-NFSe.xml, etc)
                # Solução: Usa wildcard {NUMERO}-*.xml para encontrar qualquer variação
                # Também suporta múltiplos formatos de pasta: YYYY-MM e MM-YYYY
                if not xml_path.exists() and is_nfse and numero:
                    # Tenta formato padrão: YYYY-MM
                    pasta_nfse = DATA_DIR / "xmls" / informante / year_month / tipo
                    print(f"[CORREÇÃO-NFSE]   Pasta (YYYY-MM): {pasta_nfse}, Existe: {pasta_nfse.exists()}")
                    
                    # Se não existir, tenta formato antigo: MM-YYYY
                    if not pasta_nfse.exists() and len(year_month) == 7:
                        try:
                            year, month = year_month.split('-')
                            year_month_old = f"{month}-{year}"  # 2024-09 -> 09-2024
                            pasta_nfse = DATA_DIR / "xmls" / informante / year_month_old / tipo
                            print(f"[CORREÇÃO-NFSE]   Tentando formato antigo (MM-YYYY): {pasta_nfse}, Existe: {pasta_nfse.exists()}")
                        except:
                            pass
                    
                    if pasta_nfse.exists():
                        import glob
                        # Busca: 8189-*.xml encontra:
                        # - 8189-AILTON MORAIS JARDIM.xml
                        # - 8189-NFSe.xml
                        # - 8189-qualquer-nome.xml
                        pattern = str(pasta_nfse / f"{numero}-*.xml")
                        print(f"[CORREÇÃO-NFSE]   Padrão de busca: {pattern}")
                        arquivos = glob.glob(pattern)
                        print(f"[CORREÇÃO-NFSE]   Arquivos encontrados: {arquivos}")
                        if arquivos:
                            xml_path = Path(arquivos[0])
                            pdf_path = xml_path.with_suffix('.pdf')
                            print(f"[CORREÇÃO-NFSE]   ✅ XML encontrado: {xml_path}")
                
                # Verifica também no banco xmls_baixados (NÃO aplicável para NFS-e)
                arquivo_existe = xml_path.exists() or pdf_path.exists()
                
                if not arquivo_existe and not is_nfse:
                    # NFS-e não usa tabela xmls_baixados, só NF-e/CT-e
                    try:
                        with self.db._connect() as conn:
                            cursor = conn.cursor()
                            cursor.execute("SELECT xml_completo FROM xmls_baixados WHERE chave = ?", (chave,))
                            if cursor.fetchone():
                                arquivo_existe = True
                    except Exception:
                        pass
                
                # Corrige inconsistência
                try:
                    if is_nfse:
                        # 🆕 NFS-e: SEMPRE é COMPLETO (não tem conceito de RESUMO)
                        # NFS-e sempre vem completa da prefeitura, exibe ícone verde sempre
                        if xml_status_atual != 'COMPLETO':
                            nota['xml_status'] = 'COMPLETO'
                            with self.db._connect() as conn:
                                conn.execute(
                                    "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                    (chave,)
                                )
                                conn.commit()
                            corrigidos += 1
                            print(f"[CORREÇÃO-NFSE]   ✅ NFS-e marcada como COMPLETO (padrão)")
                    else:
                        # NF-e/CT-e: Tem conceito de RESUMO vs COMPLETO
                        if arquivo_existe and xml_status_atual == 'RESUMO':
                            nota['xml_status'] = 'COMPLETO'
                            with self.db._connect() as conn:
                                conn.execute(
                                    "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                    (chave,)
                                )
                                conn.commit()
                            corrigidos += 1
                        elif not arquivo_existe and xml_status_atual == 'COMPLETO':
                            nota['xml_status'] = 'RESUMO'
                            with self.db._connect() as conn:
                                conn.execute(
                                    "UPDATE notas_detalhadas SET xml_status = 'RESUMO' WHERE chave = ?",
                                    (chave,)
                                )
                                conn.commit()
                            corrigidos += 1
                except Exception as e:
                    print(f"[CORREÇÃO] Erro ao corrigir {chave}: {e}")
            
            if corrigidos > 0:
                print(f"[CORREÇÃO] ✅ {corrigidos} registros corrigidos")
            else:
                print(f"[CORREÇÃO] ✅ Todos os registros estão consistentes")
                
        except Exception as e:
            print(f"[CORREÇÃO] Erro: {e}")
            import traceback
            traceback.print_exc()
    
    def _buscar_nfse_automatico(self, busca_completa=False):
        """
        Executa busca automática de NFS-e em thread separada.
        Chamado automaticamente após buscas de NF-e/CT-e na SEFAZ.
        
        Args:
            busca_completa: Se True, executa busca completa (--completa). Se False, busca incremental.
        """
        try:
            from PyQt5.QtCore import QThread
            import subprocess
            
            class NFSeBuscaWorker(QThread):
                def __init__(self, busca_completa=False, parent=None):
                    super().__init__(parent)
                    self.busca_completa = busca_completa
                    
                def run(self):
                    try:
                        modo = "COMPLETA" if self.busca_completa else "INCREMENTAL"
                        print(f"\n{'='*70}")
                        print(f"[NFS-e] 🔄 INICIANDO BUSCA {modo} DE NFS-e")
                        print(f"[NFS-e] 📋 Executando script separado: buscar_nfse_auto.py")
                        print(f"[NFS-e] ⏰ Aguarde... (pode levar alguns minutos)")
                        print(f"{'='*70}\n")
                        
                        # Executa buscar_nfse_auto.py como subprocesso
                        script_path = BASE_DIR / 'buscar_nfse_auto.py'
                        if script_path.exists():
                            # Usa o mesmo interpretador Python da aplicação
                            import sys
                            
                            # Monta comando com ou sem --completa
                            cmd = [sys.executable, str(script_path)]
                            if self.busca_completa:
                                cmd.append('--completa')
                            
                            result = subprocess.run(
                                cmd,
                                capture_output=True,
                                text=True,
                                encoding='utf-8',  # Força UTF-8 para evitar erro de decode no Windows
                                errors='ignore',   # Ignora caracteres inválidos
                                timeout=600  # 10 minutos de timeout (busca completa pode demorar)
                            )
                            
                            print(f"\n{'='*70}")
                            if result.returncode == 0:
                                print(f"[NFS-e] ✅ Busca {modo} de NFS-e concluída com sucesso!")
                                print(f"[NFS-e] 📊 Clique em 'Atualizar' para visualizar as NFS-e")
                            else:
                                print(f"[NFS-e] ⚠️  Busca de NFS-e finalizada com código {result.returncode}")
                                if result.stderr:
                                    print(f"[NFS-e] Detalhes: {result.stderr[:200]}")
                            print(f"{'='*70}\n")
                        else:
                            print(f"[NFS-e] ⚠️  Script não encontrado: {script_path}")
                    except subprocess.TimeoutExpired:
                        print(f"[NFS-e] ⚠️  Timeout na busca de NFS-e ({10 if self.busca_completa else 5} minutos excedidos)")
                    except Exception as e:
                        print(f"[NFS-e] ❌ Erro ao executar busca de NFS-e: {e}")
            
            # Executa em thread separada (não bloqueia UI)
            if not hasattr(self, '_nfse_worker') or not self._nfse_worker.isRunning():
                self._nfse_worker = NFSeBuscaWorker(busca_completa=busca_completa, parent=self)
                self._nfse_worker.start()
                modo = "COMPLETA" if busca_completa else "INCREMENTAL"
                print(f"[NFS-e] Thread de busca {modo} NFS-e iniciada")
            else:
                print("[NFS-e] Busca NFS-e já em execução, pulando...")
                
        except Exception as e:
            print(f"[NFS-e] Erro ao iniciar busca automática: {e}")
    
    def _clear_date_filters(self):
        """Limpa os filtros de data (volta ao padrão)"""
        try:
            from PyQt5.QtCore import QDate
            # Remove valores dos campos de data (desabilita filtro)
            self.date_inicio.setDate(QDate.currentDate().addMonths(-3))
            self.date_fim.setDate(QDate.currentDate())
            self.set_status("Filtro de data limpo", 1500)
        except Exception as e:
            print(f"[DEBUG] Erro ao limpar filtros de data: {e}")
    
    def _save_limit_preference(self, limit_text: str):
        """Salva a preferência de limite de exibição do usuário"""
        try:
            settings = QSettings('NFE_System', 'BOT_NFE')
            settings.setValue('display/limit', limit_text)
            settings.sync()
        except Exception as e:
            print(f"[DEBUG] Erro ao salvar preferência de limite: {e}")
    
    def _on_filter_changed(self):
        """Atualiza ambas as abas quando qualquer filtro é alterado"""
        try:
            # Atualiza aba "Emitidos por terceiros"
            self.refresh_table()
            # Atualiza aba "Emitidos pela empresa"
            self.refresh_emitidos_table()
        except Exception as e:
            print(f"[DEBUG] Erro ao atualizar tabelas após filtro: {e}")

    def filtered(self) -> List[Dict[str, Any]]:
        q = (self.search_edit.text() or "").lower().strip()
        selected_cert = getattr(self, '_selected_cert_cnpj', None)
        st = (self.status_dd.currentText() or "Todos").lower()
        tp = (self.tipo_dd.currentText() or "Todos").lower().replace('-', '')
        
        # Filtro de data
        date_inicio_filter = None
        date_fim_filter = None
        if hasattr(self, 'date_inicio') and hasattr(self, 'date_fim'):
            try:
                from PyQt5.QtCore import QDate
                # Verifica se não é a data padrão (não aplicar filtro se usuário não alterou)
                date_inicio_qdate = self.date_inicio.date()
                date_fim_qdate = self.date_fim.date()
                
                # Só aplica filtro se as datas forem válidas
                if date_inicio_qdate.isValid() and date_fim_qdate.isValid():
                    date_inicio_filter = date_inicio_qdate.toString("yyyy-MM-dd")
                    date_fim_filter = date_fim_qdate.toString("yyyy-MM-dd")
            except Exception as e:
                print(f"[DEBUG] Erro ao processar filtro de data: {e}")
        
        # Limite de linhas
        limit_text = self.limit_dd.currentText()
        limit = None if limit_text == "Todos" else int(limit_text)
        
        # Função para normalizar CNPJ (remove pontuação)
        def normalizar_cnpj(cnpj: str) -> str:
            return ''.join(c for c in str(cnpj or '') if c.isdigit())
        
        # Carrega CNPJs dos certificados cadastrados (empresa)
        try:
            certs = self.db.load_certificates()
            company_cnpjs = {normalizar_cnpj(c.get('cnpj_cpf') or '') for c in certs}
            company_cnpjs.discard('')  # Remove string vazia se houver
        except Exception as e:
            print(f"[DEBUG] Erro ao carregar certificados para filtro: {e}")
            company_cnpjs = set()
        
        out: List[Dict[str, Any]] = []
        for it in (self.notes or []):
            # NÃO MOSTRAR eventos na interface (apenas armazenar em disco)
            xml_status = (it.get('xml_status') or '').upper()
            if xml_status == 'EVENTO':
                continue
            
            # FILTRO PRINCIPAL: Exclui notas emitidas pela própria empresa E destinadas à própria empresa
            # Esta aba deve mostrar "Emitidos por terceiros" OU "Emitidos para terceiros"
            cnpj_emitente_normalizado = normalizar_cnpj(it.get('cnpj_emitente') or '')
            nota_destinatario_norm = normalizar_cnpj(it.get('cnpj_destinatario') or '')
            
            # Só exclui se foi emitida pela empresa E destinada à própria empresa (operação interna)
            if cnpj_emitente_normalizado in company_cnpjs and nota_destinatario_norm in company_cnpjs:
                continue  # Pula notas de operação interna (emitida E destinada à própria empresa)
            
            # 🆕 FILTRO CRÍTICO: Mostra notas DESTINADAS à empresa OU EMITIDAS pela empresa para terceiros
            # Verifica informante (quem baixou) OU cnpj_destinatario (destinatário no XML)
            # Isso garante que NFS-e recebidas e emitidas para terceiros apareçam corretamente
            nota_informante_norm = normalizar_cnpj(it.get('informante') or '')
            
            # Se tem um certificado específico selecionado, filtra por ele
            if selected_cert:
                selected_cert_norm = normalizar_cnpj(str(selected_cert))
                # Nota deve ter sido baixada por este certificado OU emitida por ele OU destinada a ele
                nota_pertence = (
                    nota_informante_norm == selected_cert_norm or
                    cnpj_emitente_normalizado == selected_cert_norm or
                    nota_destinatario_norm == selected_cert_norm
                )
                if not nota_pertence:
                    continue
            else:
                # "Todos" selecionado: mostra notas relacionadas a QUALQUER certificado da empresa
                # Verifica se informante, emitente OU destinatário está nos certificados
                # (operações internas já foram filtradas acima)
                pertence_empresa = (
                    nota_informante_norm in company_cnpjs or
                    cnpj_emitente_normalizado in company_cnpjs or
                    nota_destinatario_norm in company_cnpjs
                )
                if not pertence_empresa:
                    continue  # Pula notas que não têm relação com a empresa
            if q:
                if q not in (it.get("nome_emitente", "").lower()) and q not in (str(it.get("numero", "")).lower()) and q not in (it.get("cnpj_emitente", "").lower()):
                    continue
            if st != "todos":
                try:
                    status_nota = (it.get("status") or "").lower()
                    # Usa raiz das palavras para filtrar corretamente
                    # "Cancelado" -> busca "cancel" (pega "Cancelamento de NF-e homologado")
                    # "Autorizado" -> busca "autor" (pega "Autorizado o uso da NF-e")
                    search_term = st
                    if st == "cancelado":
                        search_term = "cancel"
                    elif st == "autorizado":
                        search_term = "autor"
                    elif st == "denegado":
                        search_term = "denega"
                    
                    if search_term not in status_nota:
                        continue
                except Exception:
                    continue
            if tp != "todos":
                raw = (it.get("tipo", "") or "").strip().upper().replace('_','').replace(' ','')
                if tp == "nfe" and raw not in ("NFE", "NF-E"):
                    continue
                if tp == "cte" and raw not in ("CTE", "CT-E"):
                    continue
                if tp == "nfse" and raw not in ("NFSE", "NFS-E"):
                    continue
            
            # Filtro de data - permite NULL (RESUMO)
            if date_inicio_filter and date_fim_filter:
                data_emissao = (it.get("data_emissao") or "")[:10]  # YYYY-MM-DD
                # Permite NULL (RESUMO) ou dentro do range
                if data_emissao and not (date_inicio_filter <= data_emissao <= date_fim_filter):
                    continue
            
            out.append(it)
            
            # Aplica limite se definido
            if limit and len(out) >= limit:
                break
        
        return out
    
    def filtered_emitidos(self) -> List[Dict[str, Any]]:
        """Filtra notas emitidas pela empresa (onde cnpj_emitente pertence a um certificado)"""
        q = (self.search_edit.text() or "").lower().strip()
        selected_cert = getattr(self, '_selected_cert_cnpj', None)
        st = (self.status_dd.currentText() or "Todos").lower()
        tp = (self.tipo_dd.currentText() or "Todos").lower().replace('-', '')
        
        print(f"\n[FILTERED_EMITIDOS] Iniciando filtro...")
        print(f"[FILTERED_EMITIDOS] Certificado selecionado: {selected_cert if selected_cert else 'TODOS'}")
        
        # Filtro de data
        date_inicio_filter = None
        date_fim_filter = None
        if hasattr(self, 'date_inicio') and hasattr(self, 'date_fim'):
            try:
                from PyQt5.QtCore import QDate
                date_inicio_qdate = self.date_inicio.date()
                date_fim_qdate = self.date_fim.date()
                
                if date_inicio_qdate.isValid() and date_fim_qdate.isValid():
                    date_inicio_filter = date_inicio_qdate.toString("yyyy-MM-dd")
                    date_fim_filter = date_fim_qdate.toString("yyyy-MM-dd")
            except Exception as e:
                print(f"[DEBUG] Erro ao processar filtro de data: {e}")
        
        # Limite de linhas
        limit_text = self.limit_dd.currentText()
        limit = None if limit_text == "Todos" else int(limit_text)
        
        # Função para normalizar CNPJ (remove pontuação)
        def normalizar_cnpj(cnpj: str) -> str:
            return ''.join(c for c in str(cnpj or '') if c.isdigit())
        
        # Carrega CNPJs dos certificados cadastrados
        try:
            certs = self.db.load_certificates()
            company_cnpjs = {normalizar_cnpj(c.get('cnpj_cpf') or '') for c in certs}
            company_cnpjs.discard('')  # Remove string vazia se houver
            print(f"[DEBUG] Certificados encontrados: {len(certs)}")
            print(f"[DEBUG] CNPJs da empresa (normalizados): {company_cnpjs}")
        except Exception as e:
            print(f"[DEBUG] Erro ao carregar certificados: {e}")
            company_cnpjs = set()
        
        if not company_cnpjs:
            print(f"[DEBUG] Nenhum certificado encontrado, retornando lista vazia")
            return []
        
        # ALTERAÇÃO: Carrega DIRETAMENTE do banco com filtros SQL em vez de usar self.notes
        # Isso garante que todas as notas emitidas sejam encontradas, não apenas as primeiras 1000
        out: List[Dict[str, Any]] = []
        
        try:
            with self.db._connect() as conn:
                # DIAGNÓSTICO: Verifica o que existe no banco
                try:
                    total_notas = conn.execute("SELECT COUNT(*) FROM notas_detalhadas").fetchone()[0]
                    total_nao_eventos = conn.execute("SELECT COUNT(*) FROM notas_detalhadas WHERE xml_status != 'EVENTO'").fetchone()[0]
                    print(f"[DEBUG] Total de notas no banco: {total_notas}")
                    print(f"[DEBUG] Total de notas não-eventos: {total_nao_eventos}")
                    
                    # Verifica se cnpj_emitente tem valores
                    sample_cnpjs = conn.execute("""
                        SELECT DISTINCT cnpj_emitente 
                        FROM notas_detalhadas 
                        WHERE cnpj_emitente IS NOT NULL AND cnpj_emitente != '' 
                        LIMIT 10
                    """).fetchall()
                    print(f"[DEBUG] Exemplos de cnpj_emitente no banco:")
                    for (cnpj,) in sample_cnpjs[:5]:
                        normalized = ''.join(c for c in str(cnpj or '') if c.isdigit())
                        print(f"[DEBUG]   - {cnpj} (normalizado: {normalized})")
                    
                    # Testa a query de normalização diretamente
                    for test_cnpj in list(company_cnpjs)[:2]:
                        test_result = conn.execute(f"""
                            SELECT COUNT(*) 
                            FROM notas_detalhadas 
                            WHERE REPLACE(REPLACE(REPLACE(cnpj_emitente, '.', ''), '/', ''), '-', '') = ?
                            AND xml_status != 'EVENTO'
                        """, (test_cnpj,)).fetchone()[0]
                        print(f"[DEBUG] Teste CNPJ {test_cnpj}: {test_result} notas encontradas")
                except Exception as e:
                    print(f"[DEBUG] Erro no diagnóstico: {e}")
                
                # Constrói query SQL com filtros
                where_clauses = ["xml_status != 'EVENTO'"]
                params = []
                
                # 🆕 FILTRO PRINCIPAL: Mostra apenas notas onde EU SOU O EMITENTE
                # Aba "Emitidos pela empresa" = cnpj_emitente deve estar nos certificados
                if selected_cert:
                    # Filtro por certificado específico selecionado
                    print(f"[DEBUG] Aplicando filtro por certificado selecionado: {selected_cert}")
                    where_clauses.append("REPLACE(REPLACE(REPLACE(cnpj_emitente, '.', ''), '/', ''), '-', '') = ?")
                    params.append(normalizar_cnpj(str(selected_cert)))
                else:
                    # "Todos" selecionado - mostra notas de TODOS OS CERTIFICADOS (não todas as notas do banco!)
                    print(f"[DEBUG] 🏢 FILTRO EMITIDOS - Mostrando notas onde EU SOU O EMITENTE (cnpj_emitente nos certificados)")
                    
                    # Cria lista de placeholders para SQL IN clause
                    placeholders = ','.join(['?' for _ in company_cnpjs])
                    where_clauses.append(f"REPLACE(REPLACE(REPLACE(cnpj_emitente, '.', ''), '/', ''), '-', '') IN ({placeholders})")
                    params.extend(list(company_cnpjs))
                
                # Filtro por status
                if st != "todos":
                    try:
                        # Usa raiz das palavras para filtrar corretamente
                        search_term = st
                        if st == "cancelado":
                            search_term = "cancel"
                        elif st == "autorizado":
                            search_term = "autor"
                        elif st == "denegado":
                            search_term = "denega"
                        
                        where_clauses.append("LOWER(status) LIKE ?")
                        params.append(f"%{search_term}%")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao aplicar filtro de status: {e}")
                
                # Filtro por tipo
                if tp != "todos":
                    tipo_patterns = []
                    if tp == "nfe":
                        tipo_patterns = ["NFE", "NF-E"]
                    elif tp == "cte":
                        tipo_patterns = ["CTE", "CT-E"]
                    elif tp == "nfse":
                        tipo_patterns = ["NFSE", "NFS-E"]
                    
                    if tipo_patterns:
                        tipo_clauses = " OR ".join(["UPPER(REPLACE(REPLACE(tipo, '_', ''), ' ', '')) = ?" for _ in tipo_patterns])
                        where_clauses.append(f"({tipo_clauses})")
                        params.extend(tipo_patterns)
                
                # Filtro por data
                if date_inicio_filter and date_fim_filter:
                    # Permite NULL (RESUMO) OU dentro do range de datas
                    where_clauses.append("(data_emissao IS NULL OR SUBSTR(data_emissao, 1, 10) BETWEEN ? AND ?)")
                    params.extend([date_inicio_filter, date_fim_filter])
                
                # Busca por texto (nome, número, CNPJ)
                if q:
                    where_clauses.append("(LOWER(nome_emitente) LIKE ? OR CAST(numero AS TEXT) LIKE ? OR cnpj_emitente LIKE ?)")
                    params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
                
                # Monta query completa
                where_sql = " AND ".join(where_clauses)
                # COALESCE coloca RESUMO (NULL) no final, mas ainda os inclui
                query = f"SELECT * FROM notas_detalhadas WHERE {where_sql} ORDER BY COALESCE(data_emissao, '9999-12-31') DESC"
                
                # Aplica limite se definido
                if limit:
                    query += f" LIMIT {limit}"
                    print(f"[DEBUG] ⚠️ LIMITE DE EXIBIÇÃO ATIVO: {limit} notas")
                else:
                    print(f"[DEBUG] ✅ SEM LIMITE - Mostrando TODAS as notas do banco")
                
                print(f"[DEBUG] Query SQL para notas emitidas: {query[:250]}...")
                print(f"[DEBUG] Parâmetros: {params}")
                print(f"[DEBUG] Filtros ativos: Status={st}, Tipo={tp}, Data={date_inicio_filter} até {date_fim_filter}")
                
                cursor = conn.execute(query, params)
                rows = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                
                for row in rows:
                    out.append(dict(zip(columns, row)))
                
                print(f"[DEBUG] ✅ Total de notas carregadas do banco: {len(out)}")
                if len(out) > 0:
                    print(f"[DEBUG] Primeira nota: {out[0].get('data_emissao', 'N/A')} - {out[0].get('numero', 'N/A')}")
                    print(f"[DEBUG] Última nota: {out[-1].get('data_emissao', 'N/A')} - {out[-1].get('numero', 'N/A')}")
                
        except Exception as e:
            print(f"[DEBUG] Erro ao carregar notas emitidas do banco: {e}")
            import traceback
            traceback.print_exc()
        
        return out

    def _populate_certs_tree(self):
        # Preenche a árvore com certificados do banco (ativos)
        try:
            certs = self.db.load_certificates()
        except Exception:
            certs = []
        
        # Salva a seleção atual antes de limpar
        current_selection = getattr(self, '_selected_cert_cnpj', None)
        
        # Bloqueia sinais para evitar múltiplos triggers durante repopulação
        self.tree_certs.blockSignals(True)
        try:
            self.tree_certs.clear()
            # Add 'Todos' entry on top
            all_item = QTreeWidgetItem(["Todos"])
            all_item.setData(0, 32, None)
            self.tree_certs.addTopLevelItem(all_item)
            
            # Se não há seleção, seleciona "Todos" por padrão
            if current_selection is None:
                all_item.setSelected(True)
            
            # Ordena por razao_social ou informante
            def keyf(c: Dict[str, Any]):
                razao = (c.get('razao_social') or '').strip()
                informante = (c.get('informante') or '').strip()
                cnpj = (c.get('cnpj_cpf') or '').strip()
                return (razao or informante or cnpj).lower()
            
            for c in sorted(certs, key=keyf):
                informante = (c.get('informante') or '').strip()
                cnpj = (c.get('cnpj_cpf') or '').strip()
                razao_social = (c.get('razao_social') or '').strip()
                
                # Prioriza razão social, depois informante, depois CNPJ
                if razao_social:
                    label = f"{razao_social}"
                elif informante and informante != cnpj:
                    label = f"{informante}"
                elif cnpj:
                    label = cnpj
                else:
                    label = 'Sem nome'
                
                node = QTreeWidgetItem([label])
                node.setToolTip(0, f"CNPJ: {cnpj}\nCaminho: {c.get('caminho') or ''}")
                node.setData(0, 32, informante)  # Salva informante, não cnpj_cpf
                self.tree_certs.addTopLevelItem(node)
                
                # Restaura seleção se corresponder ao item atual
                if current_selection and str(informante).strip() == str(current_selection).strip():
                    node.setSelected(True)
        finally:
            # Reativa sinais
            self.tree_certs.blockSignals(False)

    def _on_tree_cert_clicked(self, item: QTreeWidgetItem, col: int):
        try:
            informante = item.data(0, 32)
            # Converte para string ou None
            if informante:
                new_selection = str(informante).strip()
            else:
                new_selection = None
            
            # Só atualiza se a seleção mudou
            if new_selection != self._selected_cert_cnpj:
                print(f"\n[CERTIFICADO] Seleção mudou de '{self._selected_cert_cnpj}' para '{new_selection}'")
                self._selected_cert_cnpj = new_selection
                self.search_edit.clear()
                # Atualiza AMBAS as abas
                print(f"[CERTIFICADO] Atualizando tabela 'Emitidos por terceiros'...")
                self.refresh_table()
                print(f"[CERTIFICADO] Atualizando tabela 'Emitidos pela empresa'...")
                self.refresh_emitidos_table()
                print(f"[CERTIFICADO] Ambas as tabelas atualizadas!")
        except Exception as e:
            print(f"[CERTIFICADO] Erro ao mudar seleção: {e}")

    def _format_date_br(self, date_str: str) -> str:
        """Converte data de AAAA-MM-DD para DD/MM/AAAA."""
        if not date_str:
            return ""
        try:
            # Se já está no formato DD/MM/AAAA, retorna como está
            if "/" in date_str:
                return date_str
            # Converte de AAAA-MM-DD para DD/MM/AAAA
            date_part = date_str[:10]  # Pega apenas a parte da data
            if len(date_part) == 10 and date_part[4] == '-' and date_part[7] == '-':
                ano, mes, dia = date_part.split('-')
                return f"{dia}/{mes}/{ano}"
            return date_str
        except Exception:
            return date_str
    
    def _parse_valor(self, valor_raw) -> tuple:
        """
        Converte valor de diferentes formatos para float e string formatada BR.
        
        Retorna: (valor_formatado: str, valor_num: float)
        
        Suporta formatos:
        - US: 1234.56 ou 1234567.89
        - BR: 1.234,56 ou 1.234.567,89
        - Já formatado: R$ 1.234,56
        """
        valor_formatado = ""
        valor_num = 0.0
        
        try:
            if valor_raw:
                # Limpa prefixos e espaços
                valor_str = str(valor_raw).replace("R$", "").strip()
                
                # Detecta formato: se tem vírgula E ponto, é formato BR (1.234,56)
                if "," in valor_str and "." in valor_str:
                    # Formato BR: 1.234,56 -> remove pontos, troca vírgula por ponto
                    valor_str = valor_str.replace(".", "").replace(",", ".")
                elif "," in valor_str:
                    # Apenas vírgula: pode ser BR (1234,56) ou separador errado
                    # Se vírgula está nos últimos 3 chars, é decimal BR
                    pos_virgula = valor_str.rfind(",")
                    if len(valor_str) - pos_virgula <= 3:  # ,XX ou ,X
                        valor_str = valor_str.replace(",", ".")
                    else:
                        # Vírgula como separador de milhar (raro) - remove
                        valor_str = valor_str.replace(",", "")
                elif "." in valor_str:
                    # Apenas ponto: pode ser US (1234.56) ou BR (1.234)
                    # Se ponto está nos últimos 3 chars, é decimal US
                    pos_ponto = valor_str.rfind(".")
                    if len(valor_str) - pos_ponto <= 3:  # .XX ou .X
                        # É decimal US - mantém como está
                        pass
                    else:
                        # Ponto como separador de milhar BR - remove
                        valor_str = valor_str.replace(".", "")
                
                valor_num = float(valor_str)
                # Formata no padrão brasileiro
                valor_formatado = f"R$ {valor_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            else:
                valor_formatado = ""
        except Exception:
            # Fallback: tenta conversão simples
            try:
                valor_num = float(str(valor_raw).replace(",", "."))
                valor_formatado = f"R$ {valor_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                valor_formatado = str(valor_raw or "")
                valor_num = 0.0
        
        return (valor_formatado, valor_num)
    
    def _codigo_uf_to_sigla(self, codigo: str) -> str:
        """Converte código UF para sigla."""
        if not codigo:
            return ""
        # Mapeamento de código para sigla UF
        codigo_to_uf = {
            '11': 'RO', '12': 'AC', '13': 'AM', '14': 'RR', '15': 'PA',
            '16': 'AP', '17': 'TO', '21': 'MA', '22': 'PI', '23': 'CE',
            '24': 'RN', '25': 'PB', '26': 'PE', '27': 'AL', '28': 'SE',
            '29': 'BA', '31': 'MG', '32': 'ES', '33': 'RJ', '35': 'SP',
            '41': 'PR', '42': 'SC', '43': 'RS', '50': 'MS', '51': 'MT',
            '52': 'GO', '53': 'DF'
        }
        return codigo_to_uf.get(str(codigo).strip(), codigo)
    
    def _verificar_notas_cinza(self):
        """
        [DEPRECATED] Auto-verificação foi movida para o Gerenciador de Trabalhos.
        
        Para usar a auto-verificação:
        1. Abra o Gerenciador de Trabalhos (Ctrl+Shift+G ou menu Configurações)
        2. Clique no botão "🔍 Auto-Verificação"
        3. A tarefa rodará em segundo plano sem travar a interface
        
        Esta função permanece aqui apenas para compatibilidade.
        """
        pass  # Não faz nada - funcionalidade movida para o Gerenciador de Trabalhos
    
    def _buscar_xml_completo_silencioso(self, item):
        """Busca XML completo em background sem mostrar diálogos."""
        chave = item.get('chave')
        if not chave:
            return
        
        try:
            print(f"[AUTO-VERIFICAÇÃO] Buscando XML completo para chave: {chave}")
            
            # Usa a mesma lógica de _buscar_xml_completo mas sem diálogos
            from modules.sandbox_worker import run_task as sandbox_run_task
            
            certs = self.db.load_certificates()
            if not certs:
                print(f"[AUTO-VERIFICAÇÃO] ⚠️ Nenhum certificado configurado")
                return
            
            xml_encontrado = False
            
            # Tenta com cada certificado
            for cert in certs:
                payload = {
                    'cert': {
                        'path': cert.get('caminho') or '',
                        'senha': cert.get('senha') or '',
                        'cnpj': cert.get('cnpj_cpf') or '',
                        'cuf': cert.get('cUF_autor') or ''
                    },
                    'chave': chave,
                    'prefer': ['nfeProc', 'NFe']
                }
                
                try:
                    result = sandbox_run_task('fetch_by_chave', payload, timeout=30)
                    
                    if result.get('ok') and result.get('data', {}).get('xml'):
                        xml_text = result['data']['xml']
                        
                        # Salva XML
                        informante = item.get('informante') or cert.get('cnpj_cpf')
                        tipo = (item.get('tipo') or 'NFe').strip().upper().replace('-', '')
                        data_emissao = (item.get('data_emissao') or '')[:10]
                        
                        if data_emissao and len(data_emissao) >= 7:
                            year_month = data_emissao[:7]
                        else:
                            from datetime import datetime
                            year_month = datetime.now().strftime("%Y-%m")
                        
                        xmls_root = DATA_DIR / "xmls" / informante / tipo / year_month
                        xmls_root.mkdir(parents=True, exist_ok=True)
                        xml_file = xmls_root / f"{chave}.xml"
                        xml_file.write_text(xml_text, encoding='utf-8')
                        
                        # Atualiza banco
                        self.db.register_xml_download(chave, str(xml_file), informante)
                        self.db.save_note({
                            'chave': chave,
                            'xml_status': 'COMPLETO',
                            'informante': informante
                        })
                        
                        print(f"[AUTO-VERIFICAÇÃO] ✅ XML completo salvo: {chave}")
                        xml_encontrado = True
                        
                        # Atualiza interface
                        QTimer.singleShot(100, self.refresh_all)
                        break  # Sucesso, não precisa tentar outros certificados
                        
                except Exception as e:
                    print(f"[AUTO-VERIFICAÇÃO] Erro ao buscar com certificado {cert.get('cnpj_cpf')}: {e}")
                    continue
            
            # Se não encontrou XML em nenhum certificado
            if not xml_encontrado:
                print(f"[AUTO-VERIFICAÇÃO] ⚠️ XML não encontrado: {chave}")
                    
        except Exception as e:
            print(f"[ERRO] Erro em _buscar_xml_completo_silencioso: {e}")
    
    def refresh_table(self):
        # Stop any ongoing fill
        try:
            if self._table_fill_timer.isActive():
                self._table_fill_timer.stop()
        except Exception:
            pass
        items = self.filtered()
        # Prepare table
        try:
            self._restore_sorting = self.table.isSortingEnabled()
            self.table.setSortingEnabled(False)
        except Exception:
            self._restore_sorting = False
        try:
            self.table.clearContents()
            self.table.setRowCount(len(items))
        except Exception:
            pass
        # Start incremental fill
        self._table_fill_items = items
        self._table_fill_index = 0
        self._table_filling = True
        self.set_status("Montando tabela…")
        try:
            self._table_fill_timer.start()
        except Exception:
            # Fallback: if timer fails, do synchronous (last resort)
            for r, it in enumerate(items):
                self._populate_row(r, it)
            try:
                self.table.setSortingEnabled(self._restore_sorting)
            except Exception:
                pass
            self._table_filling = False
            self.set_status(f"{len(items)} registros carregados", 2000)
    
    def refresh_emitidos_table(self):
        """Popula a tabela de notas emitidas pela empresa (usa mesma lógica de _populate_row)"""
        # Evitar múltiplas execuções simultâneas
        if self._refreshing_emitidos:
            print("[DEBUG] ⏭️ refresh_emitidos_table já está executando, pulando chamada duplicada")
            return
        
        self._refreshing_emitidos = True
        try:
            print("[DEBUG] ========== REFRESH_EMITIDOS_TABLE CHAMADO ==========")
            items = self.filtered_emitidos()
            print(f"[DEBUG] Populando tabela_emitidos com {len(items)} itens")
            
            # Mostra na UI quantas notas foram encontradas
            self.set_status(f"📤 Carregando {len(items)} notas emitidas...", 1000)
            
            try:
                sorting_enabled = self.table_emitidos.isSortingEnabled()
                self.table_emitidos.setSortingEnabled(False)
            except Exception:
                sorting_enabled = False
            
            try:
                self.table_emitidos.clearContents()
                self.table_emitidos.setRowCount(len(items))
            except Exception:
                pass
            
            # Popula diretamente (sem timer, pois geralmente há menos itens)
            print(f"[DEBUG] Iniciando população de {len(items)} linhas na table_emitidos...")
            for r, it in enumerate(items):
                print(f"[DEBUG] Populando linha {r}: nota {it.get('numero')}")
                self._populate_emitidos_row(r, it)
            print(f"[DEBUG] População concluída!")
            
            # Auto-ajusta largura das colunas ao conteúdo (exceto XML que é fixo)
            try:
                for col in range(1, self.table_emitidos.columnCount()):
                    self.table_emitidos.resizeColumnToContents(col)
            except Exception:
                pass
            
            try:
                self.table_emitidos.setSortingEnabled(sorting_enabled)
            except Exception:
                pass
            
            # Confirma na UI
            if len(items) > 0:
                # Verifica se há filtro de data ativo
                try:
                    from PyQt5.QtCore import QDate
                    date_inicio_qdate = self.date_inicio.date()
                    date_fim_qdate = self.date_fim.date()
                    if date_inicio_qdate.isValid() and date_fim_qdate.isValid():
                        date_inicio_str = date_inicio_qdate.toString("dd/MM/yyyy")
                        date_fim_str = date_fim_qdate.toString("dd/MM/yyyy")
                        self.set_status(f"✅ {len(items)} notas emitidas ({date_inicio_str} até {date_fim_str}) ⚠️ Filtro de data ativo!", 3000)
                    else:
                        self.set_status(f"✅ {len(items)} notas emitidas carregadas", 2000)
                except:
                    self.set_status(f"✅ {len(items)} notas emitidas carregadas", 2000)
            else:
                self.set_status("⚠️ Nenhuma nota emitida encontrada - Verifique o filtro de data!", 3000)
                
        finally:
            self._refreshing_emitidos = False

    def _populate_row(self, r: int, it: Dict[str, Any]):
        def cell(c: Any) -> QTableWidgetItem:
            return QTableWidgetItem(str(c or ""))
        
        def limpar_status(status: str) -> str:
            """Remove código '100 - ' do status para deixar mais limpo"""
            if status and status.startswith("100 - "):
                return status[6:]  # Remove '100 - '
            return status
        
        xml_status = (it.get("xml_status") or "RESUMO").upper()
        status_nota = (it.get("status") or "").lower()
        
        # Verifica se a nota está cancelada (NF-e ou CT-e)
        is_cancelada = 'cancelamento' in status_nota or 'cancel' in status_nota
        
        # Obtém cores do tema atual (se disponível)
        if hasattr(self, '_current_theme_colors') and self._current_theme_colors:
            cor_autorizada = self._current_theme_colors.get('autorizada', '#d6f5e0')
            cor_cancelada = self._current_theme_colors.get('cancelada', '#ffdcdc')
            cor_outros = self._current_theme_colors.get('outros', '#ebebeb')
        else:
            # Cores padrão se tema não estiver carregado
            cor_autorizada = '#d6f5e0'
            cor_cancelada = '#ffdcdc'
            cor_outros = '#ebebeb'
        
        # Define texto e cores baseado no tipo (eventos não aparecem aqui pois são filtrados)
        # Prioriza status de cancelamento
        if is_cancelada:
            status_text = ""  # Apenas ícone, sem texto
            bg_color = QColor(cor_cancelada)  # Cor de cancelada do tema
            icon_name = 'cancelado.png'
            # Tooltip diferente se tem XML completo ou só resumo
            if xml_status == "COMPLETO":
                tooltip_text = "❌ Nota Cancelada - XML Completo disponível"
            else:
                tooltip_text = "❌ Nota Cancelada - Apenas Resumo"
        elif xml_status == "COMPLETO":
            status_text = ""  # Apenas ícone, sem texto
            bg_color = QColor(cor_autorizada)  # Cor de autorizada do tema
            tooltip_text = "✅ XML Completo disponível"
            icon_name = 'xml.png'
        else:  # RESUMO
            status_text = ""  # Sem ícone para facilitar identificação
            bg_color = QColor(cor_outros)  # Cor de outros do tema
            tooltip_text = "⚠️ Apenas Resumo - clique para baixar XML completo"
            icon_name = None  # Resumo não mostra ícone
        
        c0 = cell(status_text)
        c0.setBackground(QBrush(bg_color))
        c0.setTextAlignment(Qt.AlignCenter)
        c0.setToolTip(tooltip_text)
        # Só adiciona ícone se definido (COMPLETO ou CANCELADO)
        if icon_name:
            try:
                icon_path = BASE_DIR / 'Icone' / icon_name
                if icon_path.exists():
                    icon = QIcon(str(icon_path))
                    c0.setIcon(icon)
                    # Define tamanho do ícone para melhor centralização
                    self.table.setIconSize(QSize(20, 20))
            except Exception:
                pass
        self.table.setItem(r, 0, c0)
        # Coluna Número - ordenação numérica
        numero = it.get("numero") or ""
        # Para RESUMO sem número, tenta extrair da chave (posição 25-34)
        if not numero and xml_status == "RESUMO":
            chave = it.get("chave") or ""
            if len(chave) >= 34:
                try:
                    numero = str(int(chave[25:34]))  # Remove zeros à esquerda
                except:
                    numero = "S/N"
        try:
            numero_int = int(str(numero)) if numero else 0
        except Exception:
            numero_int = 0
        self.table.setItem(r, 1, NumericTableWidgetItem(str(numero) if numero else "S/N", float(numero_int)))
        # Coluna Data Emissão - ordenação por timestamp
        data_emissao_raw = it.get("data_emissao") or ""
        # Para RESUMO sem data, tenta extrair da chave (posição 2-5: AAMM - chave não contém dia)
        if not data_emissao_raw and xml_status == "RESUMO":
            chave = it.get("chave") or ""
            if len(chave) >= 6:
                try:
                    aa = chave[2:4]  # Ano (2 dígitos)
                    mm = chave[4:6]  # Mês (2 dígitos)
                    # Chave não contém dia - usa dia 01 como padrão
                    data_emissao_raw = f"20{aa}-{mm}-01"
                except:
                    data_emissao_raw = ""
        
        data_emissao_br = self._format_date_br(data_emissao_raw) if data_emissao_raw else "(Resumo)"
        # Converte data para timestamp para ordenação correta
        try:
            if data_emissao_raw and len(data_emissao_raw) >= 10:
                from datetime import datetime
                dt = datetime.strptime(data_emissao_raw[:10], "%Y-%m-%d")
                timestamp = dt.timestamp()
            else:
                timestamp = 9999999999.0  # Coloca RESUMO no final ao ordenar por data
        except Exception:
            timestamp = 9999999999.0
        self.table.setItem(r, 2, NumericTableWidgetItem(data_emissao_br, timestamp))
        self.table.setItem(r, 3, cell(it.get("tipo")))
        # Coluna Valor - ordenação numérica com exibição formatada
        valor_raw = it.get("valor")
        valor_formatado, valor_num = self._parse_valor(valor_raw)
        c_val = NumericTableWidgetItem(valor_formatado, valor_num)
        c_val.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(r, 4, c_val)
        # Coluna Vencimento - ordenação por timestamp
        vencimento_raw = it.get("vencimento") or ""
        vencimento_br = self._format_date_br(vencimento_raw)
        try:
            if vencimento_raw and len(vencimento_raw) >= 10:
                from datetime import datetime
                dt = datetime.strptime(vencimento_raw[:10], "%Y-%m-%d")
                timestamp = dt.timestamp()
            else:
                timestamp = 0.0
        except Exception:
            timestamp = 0.0
        self.table.setItem(r, 5, NumericTableWidgetItem(vencimento_br, timestamp))
        
        # Colunas de dados (ajustados após remover coluna Status)
        self.table.setItem(r, 6, cell(it.get("cnpj_emitente")))
        # Nome do emitente - para RESUMO pode estar vazio
        nome_emitente = it.get("nome_emitente") or ""
        if not nome_emitente and xml_status == "RESUMO":
            nome_emitente = "(Emitente não informado)"
        self.table.setItem(r, 7, cell(nome_emitente))
        self.table.setItem(r, 8, cell(it.get("natureza")))
        self.table.setItem(r, 9, cell(self._codigo_uf_to_sigla(it.get("uf") or "")))
        
        # Coluna Base ICMS - ordenação numérica com formatação BR
        base_icms_raw = it.get("base_icms")
        base_icms_formatado, base_icms_num = self._parse_valor(base_icms_raw)
        c_base = NumericTableWidgetItem(base_icms_formatado, base_icms_num)
        c_base.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(r, 10, c_base)
        
        # Coluna Valor ICMS - ordenação numérica com formatação BR
        valor_icms_raw = it.get("valor_icms")
        valor_icms_formatado, valor_icms_num = self._parse_valor(valor_icms_raw)
        c_icms = NumericTableWidgetItem(valor_icms_formatado, valor_icms_num)
        c_icms.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setItem(r, 11, c_icms)
        
        # Coluna IBS (Reforma Tributária) - ordenação numérica com formatação BR
        v_ibs_raw = it.get("v_ibs") or ""
        v_ibs_formatado, v_ibs_num = self._parse_valor(v_ibs_raw)
        c_ibs = NumericTableWidgetItem(v_ibs_formatado, v_ibs_num)
        c_ibs.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        c_ibs.setToolTip("IBS - Imposto sobre Bens e Serviços (Reforma Tributária)")
        self.table.setItem(r, 12, c_ibs)
        
        # Coluna CBS (Reforma Tributária) - ordenação numérica com formatação BR
        v_cbs_raw = it.get("v_cbs") or ""
        v_cbs_formatado, v_cbs_num = self._parse_valor(v_cbs_raw)
        c_cbs = NumericTableWidgetItem(v_cbs_formatado, v_cbs_num)
        c_cbs.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        c_cbs.setToolTip("CBS - Contribuição sobre Bens e Serviços (Reforma Tributária)")
        self.table.setItem(r, 13, c_cbs)
        
        # Coluna Status - remove código "100 - " para deixar mais limpo
        status_original = it.get("status") or ""
        status_limpo = limpar_status(status_original)
        self.table.setItem(r, 14, cell(status_limpo))
        
        self.table.setItem(r, 15, cell(it.get("cfop")))
        self.table.setItem(r, 16, cell(it.get("ncm")))
        self.table.setItem(r, 17, cell(it.get("ie_tomador")))
        self.table.setItem(r, 18, cell(it.get("chave")))
    
    def _populate_emitidos_row(self, r: int, it: Dict[str, Any]):
        """Popula uma linha da tabela de emitidos (mesma estrutura que _populate_row)"""
        def cell(c: Any) -> QTableWidgetItem:
            return QTableWidgetItem(str(c or ""))
        
        def limpar_status(status: str) -> str:
            """Remove código '100 - ' do status para deixar mais limpo"""
            if status and status.startswith("100 - "):
                return status[6:]  # Remove '100 - '
            return status
        
        # DEBUG: Log SEMPRE no início para confirmar que está sendo chamada
        numero_nota = str(it.get('numero') or '')
        tipo_nota = (it.get('tipo') or '').upper().replace('-', '')
        
        # 🆕 NFS-e NÃO TEM RESUMO - Sempre é COMPLETO
        # Diferente de NF-e/CT-e que vêm como resumo do DistDFe, NFS-e sempre vem completa
        if tipo_nota == 'NFSE':
            xml_status = "COMPLETO"
        else:
            xml_status = (it.get("xml_status") or "RESUMO").upper()
        status_nota = (it.get("status") or "").lower()
        
        # Verifica se a nota está cancelada (NF-e ou CT-e)
        is_cancelada = 'cancelamento' in status_nota or 'cancel' in status_nota
        
        # Obtém cores do tema atual (se disponível)
        if hasattr(self, '_current_theme_colors') and self._current_theme_colors:
            cor_autorizada = self._current_theme_colors.get('autorizada', '#d6f5e0')
            cor_cancelada = self._current_theme_colors.get('cancelada', '#ffdcdc')
            cor_outros = self._current_theme_colors.get('outros', '#ebebeb')
        else:
            # Cores padrão se tema não estiver carregado
            cor_autorizada = '#d6f5e0'
            cor_cancelada = '#ffdcdc'
            cor_outros = '#ebebeb'
        
        # Define texto e cores baseado no tipo
        # Prioriza status de cancelamento sobre xml_status
        if is_cancelada:
            status_text = ""
            bg_color = QColor(cor_cancelada)  # Cor de cancelada do tema
            icon_name = 'cancelado.png'
            # Tooltip diferente se tem XML completo ou só resumo
            if xml_status == "COMPLETO":
                tooltip_text = "❌ Nota Cancelada - XML Completo disponível"
            else:
                tooltip_text = "❌ Nota Cancelada - Apenas Resumo"
        elif xml_status == "COMPLETO":
            status_text = ""
            bg_color = QColor(cor_autorizada)  # Cor de autorizada do tema
            tooltip_text = "✅ XML Completo disponível"
            icon_name = 'xml.png'
        else:  # RESUMO
            status_text = ""  # Sem ícone para facilitar identificação
            bg_color = QColor(cor_outros)  # Cor de outros do tema
            tooltip_text = "⚠️ Apenas Resumo - clique para baixar XML completo"
            icon_name = None  # Resumo não mostra ícone
        
        # DEBUG: Log do ícone escolhido
        if numero_nota in ['29511', '5629031']:
            print(f"  icon_name escolhido: {icon_name}")
            print(f"  bg_color: {bg_color.name()}")
            print(f"  tooltip: {tooltip_text}")
        
        c0 = cell(status_text)
        c0.setBackground(QBrush(bg_color))
        c0.setTextAlignment(Qt.AlignCenter)
        c0.setToolTip(tooltip_text)
        
        # Só adiciona ícone se definido (COMPLETO ou CANCELADO)
        if icon_name:
            try:
                icon_path = BASE_DIR / 'Icone' / icon_name
                if icon_path.exists():
                    icon = QIcon(str(icon_path))
                    c0.setIcon(icon)
                    # Define tamanho do ícone para melhor centralização
                    self.table_emitidos.setIconSize(QSize(20, 20))
            except Exception:
                pass
        
        self.table_emitidos.setItem(r, 0, c0)
        
        # Coluna Número - ordenação numérica
        numero = it.get("numero") or ""
        # Para RESUMO sem número, tenta extrair da chave (posição 25-34)
        if not numero and xml_status == "RESUMO":
            chave = it.get("chave") or ""
            if len(chave) >= 34:
                try:
                    numero = str(int(chave[25:34]))  # Remove zeros à esquerda
                except:
                    numero = "S/N"
        try:
            numero_int = int(str(numero)) if numero else 0
        except Exception:
            numero_int = 0
        self.table_emitidos.setItem(r, 1, NumericTableWidgetItem(str(numero) if numero else "S/N", float(numero_int)))
        
        # Coluna Data Emissão - ordenação por timestamp
        data_emissao_raw = it.get("data_emissao") or ""
        # Para RESUMO sem data, tenta extrair da chave (posição 2-5: AAMM - chave não contém dia)
        if not data_emissao_raw and xml_status == "RESUMO":
            chave = it.get("chave") or ""
            if len(chave) >= 6:
                try:
                    aa = chave[2:4]  # Ano (2 dígitos)
                    mm = chave[4:6]  # Mês (2 dígitos)
                    # Chave não contém dia - usa dia 01 como padrão
                    data_emissao_raw = f"20{aa}-{mm}-01"
                except:
                    data_emissao_raw = ""
        
        data_emissao_br = self._format_date_br(data_emissao_raw) if data_emissao_raw else "(Resumo)"
        try:
            if data_emissao_raw and len(data_emissao_raw) >= 10:
                from datetime import datetime
                dt = datetime.strptime(data_emissao_raw[:10], "%Y-%m-%d")
                timestamp = dt.timestamp()
            else:
                timestamp = 9999999999.0  # Coloca RESUMO no final
        except Exception:
            timestamp = 9999999999.0
        self.table_emitidos.setItem(r, 2, NumericTableWidgetItem(data_emissao_br, timestamp))
        
        self.table_emitidos.setItem(r, 3, cell(it.get("tipo")))
        
        # Coluna Valor - ordenação numérica com exibição formatada
        valor_raw = it.get("valor")
        valor_formatado, valor_num = self._parse_valor(valor_raw)
        c_val = NumericTableWidgetItem(valor_formatado, valor_num)
        c_val.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table_emitidos.setItem(r, 4, c_val)
        
        # Coluna Vencimento - ordenação por timestamp
        vencimento_raw = it.get("vencimento") or ""
        vencimento_br = self._format_date_br(vencimento_raw)
        try:
            if vencimento_raw and len(vencimento_raw) >= 10:
                from datetime import datetime
                dt = datetime.strptime(vencimento_raw[:10], "%Y-%m-%d")
                timestamp = dt.timestamp()
            else:
                timestamp = 0.0
        except Exception:
            timestamp = 0.0
        self.table_emitidos.setItem(r, 5, NumericTableWidgetItem(vencimento_br, timestamp))
        
        # Colunas de dados (ajustados após remover coluna Status)
        # 🔧 IMPORTANTE: Para "Emitidos pela empresa", mostramos DESTINATÁRIO (para quem a empresa prestou serviço)
        # Esta aba mostra notas onde SUA EMPRESA É O EMITENTE (cnpj_emitente = certificado)
        # Então exibimos: DESTINATÁRIO (quem recebeu o serviço/produto)
        self.table_emitidos.setItem(r, 6, cell(it.get("cnpj_destinatario") or ""))
        
        # Nome do destinatário (quem recebeu o serviço/produto)
        nome_destinatario = it.get("nome_destinatario") or ""
        if not nome_destinatario and xml_status == "RESUMO":
            nome_destinatario = "(Destinatário não informado)"
        self.table_emitidos.setItem(r, 7, cell(nome_destinatario))
        
        self.table_emitidos.setItem(r, 8, cell(it.get("natureza")))
        self.table_emitidos.setItem(r, 9, cell(self._codigo_uf_to_sigla(it.get("uf") or "")))
        
        # Coluna Base ICMS - ordenação numérica com formatação BR
        base_icms_raw = it.get("base_icms")
        base_icms_formatado, base_icms_num = self._parse_valor(base_icms_raw)
        c_base = NumericTableWidgetItem(base_icms_formatado, base_icms_num)
        c_base.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table_emitidos.setItem(r, 10, c_base)
        
        # Coluna Valor ICMS - ordenação numérica com formatação BR
        valor_icms_raw = it.get("valor_icms")
        valor_icms_formatado, valor_icms_num = self._parse_valor(valor_icms_raw)
        c_icms = NumericTableWidgetItem(valor_icms_formatado, valor_icms_num)
        c_icms.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table_emitidos.setItem(r, 11, c_icms)
        
        # Coluna IBS (Reforma Tributária) - ordenação numérica com formatação BR
        v_ibs_raw = it.get("v_ibs") or ""
        v_ibs_formatado, v_ibs_num = self._parse_valor(v_ibs_raw)
        c_ibs = NumericTableWidgetItem(v_ibs_formatado, v_ibs_num)
        c_ibs.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        c_ibs.setToolTip("IBS - Imposto sobre Bens e Serviços (Reforma Tributária)")
        self.table_emitidos.setItem(r, 12, c_ibs)
        
        # Coluna CBS (Reforma Tributária) - ordenação numérica com formatação BR
        v_cbs_raw = it.get("v_cbs") or ""
        v_cbs_formatado, v_cbs_num = self._parse_valor(v_cbs_raw)
        c_cbs = NumericTableWidgetItem(v_cbs_formatado, v_cbs_num)
        c_cbs.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        c_cbs.setToolTip("CBS - Contribuição sobre Bens e Serviços (Reforma Tributária)")
        self.table_emitidos.setItem(r, 13, c_cbs)
        
        # Coluna Status - remove código "100 - " para deixar mais limpo
        status_original = it.get("status") or ""
        status_limpo = limpar_status(status_original)
        self.table_emitidos.setItem(r, 14, cell(status_limpo))
        
        self.table_emitidos.setItem(r, 15, cell(it.get("cfop")))
        self.table_emitidos.setItem(r, 16, cell(it.get("ncm")))
        self.table_emitidos.setItem(r, 17, cell(it.get("ie_tomador")))
        self.table_emitidos.setItem(r, 18, cell(it.get("chave")))

    def _fill_table_step(self):
        try:
            total = len(self._table_fill_items)
            if self._table_fill_index >= total:
                # Done
                try:
                    self._table_fill_timer.stop()
                except Exception:
                    pass
                self._table_filling = False
                # Sempre habilita sorting após preenchimento
                try:
                    self.table.setSortingEnabled(True)
                except Exception:
                    pass
                # Auto-ajusta largura das colunas ao conteúdo (exceto XML que é fixo)
                try:
                    for col in range(1, self.table.columnCount()):
                        self.table.resizeColumnToContents(col)
                except Exception:
                    pass
                self.set_status(f"{total} registros carregados", 2000)
                
                # ⚠️ AUTO-VERIFICAÇÃO REMOVIDA DAQUI
                # Agora está disponível no Gerenciador de Trabalhos (Ctrl+Shift+G)
                # O usuário pode iniciar quando quiser sem travar a interface
                return
            start = self._table_fill_index
            end = min(start + self._table_fill_chunk, total)
            for r in range(start, end):
                it = self._table_fill_items[r]
                self._populate_row(r, it)
            self._table_fill_index = end
            # Update status lightly
            try:
                pct = int((end / max(1, total)) * 100)
                self.set_status(f"Montando tabela… {pct}%")
            except Exception:
                pass
        except Exception:
            # On any error, stop to avoid tight loop
            try:
                self._table_fill_timer.stop()
            except Exception:
                pass
            self._table_filling = False

        # Table tweaks
        try:
            self.table.setAlternatingRowColors(True)
            self.table.verticalHeader().setDefaultSectionSize(24)
        except Exception:
            pass

    def _update_window_title(self):
        """Atualiza o título da janela com a versão atual."""
        try:
            version_file = BASE_DIR / "version.txt"
            if version_file.exists():
                version = version_file.read_text(encoding='utf-8').strip()
            else:
                version = "1.0.0"
        except Exception:
            version = "1.0.0"
        
        self.setWindowTitle(f"Busca XML - v{version}")
    
    def _center_window(self):
        """Centraliza a janela na tela."""
        from PyQt5.QtWidgets import QApplication
        from PyQt5.QtCore import QRect
        screen = QApplication.desktop().screenGeometry()
        size = self.geometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )
    
    def _update_search_summary(self):
        """Atualiza resumo de busca em tempo real."""
        from datetime import datetime
        try:
            elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
            cert_info = f"Cert: ...{self._search_stats['last_cert']}" if self._search_stats['last_cert'] else ""
            
            summary = f"🔍 NFes: {self._search_stats['nfes_found']} | CTes: {self._search_stats['ctes_found']} | NFSes: {self._search_stats['nfses_found']}"
            if cert_info:
                summary += f" | {cert_info}"
            summary += f" | {elapsed:.0f}s"
            
            self.search_summary_label.setText(summary)
            # REMOVIDO: QApplication.processEvents() causava reentrância e travamento
        except Exception:
            pass  # Silencioso para evitar recursão
    
    def _get_last_search_status(self):
        """Retorna texto com status da última busca."""
        from datetime import datetime
        try:
            last_search = self.db.get_last_search_time()
            if last_search:
                last_dt = datetime.fromisoformat(last_search)
                now = datetime.now()
                diff_minutes = (now - last_dt).total_seconds() / 60
                
                # Formata tempo decorrido
                if diff_minutes < 60:
                    tempo = f"{diff_minutes:.0f}min"
                elif diff_minutes < 1440:  # menos de 24h
                    horas = diff_minutes / 60
                    tempo = f"{horas:.1f}h"
                else:
                    dias = diff_minutes / 1440
                    tempo = f"{dias:.1f}d"
                
                # Formata hora
                hora = last_dt.strftime("%H:%M")
                return f"Última busca: {hora} (há {tempo})"
            else:
                return "Pronto - Nenhuma busca realizada"
        except Exception:
            return "Pronto"  # Silencioso para evitar recursão
    
    def _load_intervalo_config(self) -> int:
        """Carrega intervalo de busca configurado (em horas). Padrão: 1 hora."""
        try:
            intervalo_minutos = self.db.get_next_search_interval()
            # Converte de minutos para horas
            return max(1, min(23, intervalo_minutos // 60))
        except Exception:
            return 1  # Padrão: 1 hora
    
    def _save_intervalo_config(self, horas: int):
        """Salva intervalo de busca configurado (converte horas para minutos)."""
        try:
            minutos = horas * 60
            self.db.set_next_search_interval(minutos)
            self.set_status(f"Intervalo de busca atualizado: {horas} hora(s)", 3000)
            
            # 🔄 RECALCULA PRÓXIMA BUSCA SE JÁ HOUVER UMA AGENDADA
            if hasattr(self, '_next_search_time') and self._next_search_time:
                from datetime import datetime, timedelta
                
                # ❌ Marca o timer antigo como cancelado
                self._timer_cancelled = True
                
                # ⏰ Calcula novo horário e atualiza contador
                self._next_search_time = datetime.now() + timedelta(minutes=minutos)
                self._update_search_status()  # Atualiza contador imediatamente
                
                # 📅 Agenda nova busca com novo intervalo
                delay_ms = int(minutos * 60 * 1000)
                self._timer_cancelled = False
                self._scheduled_timer_id = QTimer.singleShot(delay_ms, self._executar_busca_agendada)
                
                print(f"[DEBUG] ✅ Timer reagendado! Próxima busca: {self._next_search_time.strftime('%H:%M:%S')}")
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Não foi possível salvar intervalo: {e}")
    
    def _set_intervalo_from_menu(self, horas: int):
        """Define intervalo de busca a partir do menu (sincroniza com SpinBox)."""
        try:
            self.spin_intervalo.setValue(horas)  # Dispara _save_intervalo_config automaticamente
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Não foi possível definir intervalo: {e}")
    
    def _aplicar_tema(self, tema_nome: str):
        """
        Aplica um tema à aplicação e salva a preferência.
        
        Args:
            tema_nome: Nome do tema a ser aplicado
        """
        if not THEMES_AVAILABLE:
            QMessageBox.warning(self, "Erro", "Sistema de temas não disponível!")
            return
        
        try:
            # Aplica o tema
            app = QApplication.instance()
            success = ThemeManager.apply_theme(app, tema_nome)
            
            if success:
                # Salva preferência
                ThemeManager.save_theme_preference(tema_nome)
                
                # Atualiza cores da tabela baseado no novo tema
                self._atualizar_cores_tabela_tema(tema_nome)
                
                # Limpa stylesheets inline que possam sobrescrever o tema
                self._limpar_cores_inline()
                
                # Força atualização completa da interface
                self._forcar_atualizacao_interface()
                
                # Mensagem de sucesso
                self.set_status(f"✅ Tema '{tema_nome}' aplicado com sucesso!", 3000)
                
            else:
                QMessageBox.warning(
                    self,
                    "Erro ao Aplicar Tema",
                    f"Não foi possível aplicar o tema '{tema_nome}'.\n"
                    "Verifique os logs para mais detalhes."
                )
        
        except Exception as e:
            print(f"[THEME] Erro ao aplicar tema: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "Erro",
                f"Ocorreu um erro ao aplicar o tema:\n{e}"
            )
    
    def _atualizar_cores_tabela_tema(self, tema_nome: str):
        """
        Atualiza as cores da tabela baseado no tema selecionado.
        
        Args:
            tema_nome: Nome do tema atual
        """
        try:
            # Obtém cores do tema
            cores = ThemeManager.get_status_colors(tema_nome)
            
            # Armazena cores para uso posterior
            self._current_theme_colors = cores
            
            # Força atualização da tabela
            self.refresh_all()
            
            print(f"[THEME] Cores da tabela atualizadas para tema '{tema_nome}'")
        
        except Exception as e:
            print(f"[THEME] Erro ao atualizar cores da tabela: {e}")
    
    def _limpar_cores_inline(self):
        """Remove stylesheets inline que podem sobrescrever o tema."""
        try:
            # Remove cores hardcoded do label de resumo de busca
            if hasattr(self, 'search_summary_label'):
                # Mantém apenas formatação, remove cor específica
                self.search_summary_label.setStyleSheet("font-weight: bold; padding: 0 10px;")
            
            # Remove cor inline do status_label para usar cor do tema
            if hasattr(self, 'status_label'):
                self.status_label.setStyleSheet("")
            
            print("[THEME] Cores inline limpas")
        
        except Exception as e:
            print(f"[THEME] Erro ao limpar cores inline: {e}")
    
    def _forcar_atualizacao_interface(self):
        """Força atualização completa de todos os elementos da interface."""
        try:
            # Atualiza widgets principais
            if hasattr(self, 'centralWidget'):
                self.centralWidget().update()
            
            # Atualiza tabelas
            if hasattr(self, 'table'):
                self.table.viewport().update()
            if hasattr(self, 'table_emitidos'):
                self.table_emitidos.viewport().update()
            
            # Atualiza árvore de certificados
            if hasattr(self, 'tree_certs'):
                self.tree_certs.viewport().update()
            
            # Atualiza janela principal
            self.update()
            
            # Processa eventos pendentes
            QApplication.processEvents()
            
            print("[THEME] Interface atualizada")
        
        except Exception as e:
            print(f"[THEME] Erro ao atualizar interface: {e}")
    
    def _get_status_color(self, tipo: str) -> str:
        """
        Obtém a cor de status baseada no tema atual.
        
        Args:
            tipo: Tipo de status ('success', 'error', 'warning', 'info')
            
        Returns:
            str: Cor em formato hexadecimal
        """
        if THEMES_AVAILABLE and hasattr(self, '_current_theme_colors'):
            # Tenta obter tema atual
            tema_atual = ThemeManager.load_theme_preference()
            cores_msg = ThemeManager.get_message_colors(tema_atual)
            return cores_msg.get(tipo, '#000000')
        
        # Cores padrão se tema não disponível
        cores_padrao = {
            'success': '#28a745',
            'error': '#dc3545',
            'warning': '#ffc107',
            'info': '#17a2b8'
        }
        return cores_padrao.get(tipo, '#000000')
    
    def _set_intervalo_from_menu(self, horas: int):
        """Define intervalo de busca a partir do menu (sincroniza com SpinBox)."""
        try:
            self._save_intervalo_config(horas)
            if hasattr(self, 'spin_intervalo'):
                self.spin_intervalo.setValue(horas)
        except Exception as e:
            print(f"Erro ao definir intervalo do menu: {e}")


    def _apply_theme(self):
        # Global stylesheet for a clean modern look
        try:
            self.setStyleSheet(
                """
                QWidget { font-size: 11px; }
                QLineEdit, QComboBox { padding: 4px 6px; }
                QPushButton { padding: 4px 10px; }
                QTreeWidget { border: 1px solid #d0d0d0; border-radius: 4px; }
                QTableWidget { border: 1px solid #d0d0d0; border-radius: 4px; gridline-color: #ededed; }
                QHeaderView::section { background: #fafafa; padding: 6px; border: 0px; border-bottom: 1px solid #e0e0e0; font-weight: 600; }
                QStatusBar { background: #f6f6f6; }
                """
            )
            # Slightly larger headers
            hfont = self.table.horizontalHeader().font()
            current_size = hfont.pointSize()
            if current_size > 0:
                hfont.setPointSize(current_size + 1)
            else:
                hfont.setPointSize(10)  # Tamanho padrão se não conseguir obter
            hfont.setBold(True)
            self.table.horizontalHeader().setFont(hfont)
        except Exception:
            pass

    def _filter_certs_tree(self, text: str):
        try:
            text = (text or '').lower().strip()
            for i in range(self.tree_certs.topLevelItemCount()):
                it = self.tree_certs.topLevelItem(i)
                visible = (text in (it.text(0) or '').lower()) if text else True
                it.setHidden(not visible)
        except Exception:
            pass

    def _clear_cert_filter(self):
        try:
            self.cert_filter.clear()
            self._selected_cert_cnpj = None
            # Limpa seleção visual na árvore
            self.tree_certs.clearSelection()
            # Seleciona "Todos"
            if self.tree_certs.topLevelItemCount() > 0:
                todos_item = self.tree_certs.topLevelItem(0)
                todos_item.setSelected(True)
            self.refresh_table()
        except Exception:
            pass

    def get_selected_item(self) -> Optional[Dict[str, Any]]:
        sel = self.table.selectionModel()
        if sel is None:
            return None
        idxs = sel.selectedRows()
        if not idxs:
            return None
        row = idxs[0].row()
        # We need to map back to filtered list order
        flt = self.filtered()
        if 0 <= row < len(flt):
            return flt[row]
        return None

    def _on_table_context_menu(self, pos):
        """Menu de contexto com opções para a nota/CT-e selecionada"""
        # Pega o item clicado
        item_at_pos = self.table.itemAt(pos)
        if not item_at_pos:
            return
        
        # Obtém todas as linhas selecionadas
        selected_rows = list(set(index.row() for index in self.table.selectedIndexes()))
        if not selected_rows:
            return
        
        # Usa a linha clicada como referência para o menu
        row = item_at_pos.row()
        
        # ⚠️ IMPORTANTE: Não usar filtered()[row] porque a ordem muda após sorting!
        # Busca pela chave que está na célula da tabela
        
        # Encontra o índice da coluna "Chave"
        chave_col_index = None
        for col in range(self.table.columnCount()):
            header_text = self.table.horizontalHeaderItem(col).text()
            if header_text == "Chave":
                chave_col_index = col
                break
        
        if chave_col_index is None:
            print(f"[DEBUG] Erro: Coluna 'Chave' não encontrada na table!")
            return
        
        # Pega a chave da linha clicada
        chave_item = self.table.item(row, chave_col_index)
        if not chave_item:
            print(f"[DEBUG] Erro: Não encontrou chave na linha {row}")
            return
        
        chave = chave_item.text()
        
        # Busca o item completo pela chave no banco
        try:
            import sqlite3
            conn = sqlite3.connect(str(DATA_DIR / 'notas.db'))
            conn.row_factory = sqlite3.Row
            nota = conn.execute('SELECT * FROM notas_detalhadas WHERE chave = ?', (chave,)).fetchone()
            conn.close()
            
            if not nota:
                QMessageBox.warning(self, "Erro", "Documento não encontrado no banco de dados!")
                return
            
            # Converte de Row para Dict
            item = dict(nota)
            
        except Exception as e:
            print(f"[DEBUG] Erro ao buscar nota: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao buscar documento: {e}")
            return
        
        xml_status = (item.get('xml_status') or '').upper()
        status_nota = (item.get('status') or '').lower()
        
        # 🔍 Verifica se é REALMENTE um resumo (sem dados completos)
        # Resumo = sem número OU sem data de emissão OU sem emitente
        numero = item.get('numero') or item.get('nNF') or ''
        data_emissao = item.get('data_emissao') or item.get('dhEmi') or ''
        emitente = item.get('nome_emitente') or item.get('xNome') or ''
        
        is_resumo = (not numero or not data_emissao or not emitente or 
                     xml_status == 'RESUMO')
        
        # Cria menu
        menu = QMenu(self)
        
        # ⭐ OPÇÃO NO TOPO: XML Completo (só para RESUMO ou dados incompletos)
        if is_resumo:
            # Verifica se há múltiplas seleções
            if len(selected_rows) > 1:
                action_xml_completo = menu.addAction(f"✅ Baixar XML Completo ({len(selected_rows)} notas)")
                action_xml_completo.setToolTip(f"Manifestar e baixar {len(selected_rows)} XMLs completos da SEFAZ")
            else:
                action_xml_completo = menu.addAction("✅ Baixar XML Completo")
                action_xml_completo.setToolTip("Manifestar, baixar XML completo da SEFAZ e gerar PDF")
            menu.addSeparator()
            print(f"[DEBUG MENU] ✅ Botão 'Baixar XML Completo' adicionado ({len(selected_rows)} selecionadas)")
        else:
            action_xml_completo = None
            print(f"[DEBUG MENU] ⚠️ Botão 'Baixar XML Completo' NÃO adicionado (nota completa)")
        
        # Opção: Ver Detalhes Completos (sempre disponível)
        action_detalhes = menu.addAction("📄 Ver Detalhes Completos")
        
        # Opção: Eventos (sempre disponível)
        menu.addSeparator()
        action_eventos = menu.addAction("📋 Ver Eventos")
        
        # Opção: Manifestar (só para notas RECEBIDAS - Emitidos por Terceiros)
        # NF-e e CT-e permitem manifestação do destinatário
        tipo_doc = (item.get('tipo') or '').upper()
        if tipo_doc in ['NFE', 'NF-E', 'CTE', 'CT-E']:
            menu.addSeparator()
            if tipo_doc in ['NFE', 'NF-E']:
                action_manifestar = menu.addAction("✉️ Manifestar Destinatário")
            else:  # CTE
                action_manifestar = menu.addAction("✉️ Manifestar CT-e")
        else:
            action_manifestar = None
        
        # Mostra menu e pega ação
        action = menu.exec_(self.table.viewport().mapToGlobal(pos))
        
        if action == action_xml_completo:
            # Se múltiplas notas selecionadas, baixa todas
            if len(selected_rows) > 1:
                self._baixar_xml_e_pdf_multiplos(selected_rows)
            else:
                self._baixar_xml_e_pdf(item)  # Método direto para uma nota
        elif action == action_detalhes:
            self._mostrar_detalhes_nota(item)
        elif action == action_eventos:
            self._mostrar_eventos(item)
        elif action == action_manifestar:
            self._manifestar_nota(item)
    
    def _on_table_emitidos_context_menu(self, pos):
        """Menu de contexto para a tabela de notas emitidas pela empresa"""
        # Pega o item clicado
        item_at_pos = self.table_emitidos.itemAt(pos)
        if not item_at_pos:
            return
        
        row = item_at_pos.row()
        
        # ⚠️ IMPORTANTE: Precisa usar os dados armazenados na própria linha da tabela!
        # filtered_emitidos() pode estar em ordem diferente por causa de filtros/ordenação
        
        # Encontra o índice da coluna "Chave" (última coluna, independente de reordenação visual)
        chave_col_index = None
        for col in range(self.table_emitidos.columnCount()):
            header_text = self.table_emitidos.horizontalHeaderItem(col).text()
            if header_text == "Chave":
                chave_col_index = col
                break
        
        if chave_col_index is None:
            print(f"[DEBUG] Erro: Coluna 'Chave' não encontrada!")
            return
        
        # Pega o número visível na tela para comparação
        numero_item = self.table_emitidos.item(row, 1)  # Coluna 1 = número
        numero_tela = numero_item.text() if numero_item else "???"
        
        # Pega a chave da coluna correta (pelo índice lógico encontrado)
        chave_item = self.table_emitidos.item(row, chave_col_index)
        if not chave_item:
            print(f"[DEBUG] Erro: Não encontrou chave na coluna {chave_col_index} da linha {row}")
            return
        
        chave = chave_item.text()
        
        # Busca o item completo pelo chave no banco
        try:
            import sqlite3
            conn = sqlite3.connect(str(DATA_DIR / 'notas.db'))
            conn.row_factory = sqlite3.Row
            nota = conn.execute('SELECT * FROM notas_detalhadas WHERE chave = ?', (chave,)).fetchone()
            conn.close()
            
            if not nota:
                QMessageBox.warning(self, "Erro", "Documento não encontrado no banco de dados!")
                return
            
            # Converte de Row para Dict
            item = dict(nota)
            
            print(f"\n[DEBUG MENU EMITIDOS] ========== CLIQUE NO MENU ==========")
            print(f"[DEBUG MENU EMITIDOS] Row clicada: {row}")
            print(f"[DEBUG MENU EMITIDOS] Número visível na tela (coluna 1): {numero_tela}")
            print(f"[DEBUG MENU EMITIDOS] Chave lida da coluna 16: {chave}")
            print(f"[DEBUG MENU EMITIDOS] Número do banco pela chave: {item.get('numero')}")
            print(f"[DEBUG MENU EMITIDOS] Tipo do banco: {item.get('tipo')}")
            print(f"[DEBUG MENU EMITIDOS] ⚠️ COMPARAÇÃO: Número tela={numero_tela} vs Número banco={item.get('numero')}")
            
        except Exception as e:
            print(f"[DEBUG] Erro ao buscar nota: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao buscar documento: {e}")
            return
        xml_status = (item.get('xml_status') or '').upper()
        status_nota = (item.get('status') or '').lower()
        
        # Cria menu
        menu = QMenu(self)
        
        # Opção: Ver Detalhes Completos (sempre disponível)
        action_detalhes = menu.addAction("📄 Ver Detalhes Completos")
        
        # Opção: Buscar XML Completo (só para RESUMO)
        menu.addSeparator()
        if xml_status == 'RESUMO':
            action_buscar = menu.addAction("🔍 Buscar XML Completo na SEFAZ")
        else:
            action_buscar = None
        
        # Opção: Eventos (sempre disponível)
        menu.addSeparator()
        action_eventos = menu.addAction("📋 Ver Eventos")
        
        # Mostra menu e pega ação
        action = menu.exec_(self.table_emitidos.viewport().mapToGlobal(pos))
        
        if action == action_xml_completo:
            self._baixar_xml_e_pdf(item)  # Novo método direto
        elif action == action_detalhes:
            self._mostrar_detalhes_nota(item)
        elif action == action_buscar:
            self._buscar_xml_completo(item)
        elif action == action_eventos:
            print(f"\n[DEBUG ANTES MOSTRAR EVENTOS] ========== ANTES DE CHAMAR _mostrar_eventos ==========")
            print(f"[DEBUG ANTES MOSTRAR EVENTOS] item['numero']: {item.get('numero')}")
            print(f"[DEBUG ANTES MOSTRAR EVENTOS] item['tipo']: {item.get('tipo')}")
            print(f"[DEBUG ANTES MOSTRAR EVENTOS] item['chave']: {item.get('chave')}")
            self._mostrar_eventos(item)
    
    def _mostrar_detalhes_nota(self, item: Dict[str, Any]):
        """Exibe uma janela com todos os detalhes da nota fiscal"""
        if not item:
            return
        
        # Cria janela de diálogo
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Detalhes - Nota {item.get('numero', 'N/A')}")
        dialog.setMinimumWidth(700)
        dialog.setMinimumHeight(600)
        
        # Layout principal
        layout = QVBoxLayout(dialog)
        
        # Área de scroll para o conteúdo
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # Função helper para adicionar campo
        def add_field(label: str, value: Any, is_header: bool = False):
            container = QWidget()
            h_layout = QHBoxLayout(container)
            h_layout.setContentsMargins(5, 5, 5, 5)
            
            label_widget = QLabel(f"<b>{label}:</b>")
            label_widget.setMinimumWidth(180)
            
            value_str = str(value or "")
            value_widget = QLabel(value_str)
            value_widget.setWordWrap(True)
            value_widget.setTextInteractionFlags(Qt.TextSelectableByMouse)
            
            if is_header:
                font = value_widget.font()
                font.setPointSize(font.pointSize() + 2)
                font.setBold(True)
                value_widget.setFont(font)
                label_widget.setFont(font)
            
            h_layout.addWidget(label_widget)
            h_layout.addWidget(value_widget, 1)
            
            scroll_layout.addWidget(container)
        
        # Função para adicionar separador
        def add_separator(title: str = ""):
            line = QFrame()
            line.setFrameShape(QFrame.HLine)
            line.setFrameShadow(QFrame.Sunken)
            scroll_layout.addWidget(line)
            
            if title:
                title_label = QLabel(f"<h3>{title}</h3>")
                scroll_layout.addWidget(title_label)
        
        # === INFORMAÇÕES PRINCIPAIS ===
        add_field("Número da Nota", item.get('numero', 'N/A'), is_header=True)
        add_field("Tipo", item.get('tipo', 'N/A'))
        add_field("Status", item.get('status', 'N/A'))
        add_field("Status XML", item.get('xml_status', 'N/A'))
        
        add_separator("Emissor")
        add_field("Nome do Emitente", item.get('nome_emitente', 'N/A'))
        add_field("CNPJ Emitente", item.get('cnpj_emitente', 'N/A'))
        add_field("UF", self._codigo_uf_to_sigla(item.get('uf', '')))
        
        add_separator("Valores")
        add_field("Valor Total", item.get('valor', 'N/A'))
        add_field("Base ICMS", item.get('base_icms', 'N/A'))
        add_field("Valor ICMS", item.get('valor_icms', 'N/A'))
        
        add_separator("Datas")
        add_field("Data de Emissão", self._format_date_br(item.get('data_emissao', '')))
        add_field("Data de Vencimento", self._format_date_br(item.get('vencimento', '')))
        add_field("Atualizado em", item.get('atualizado_em', 'N/A'))
        
        add_separator("Informações Fiscais")
        add_field("CFOP", item.get('cfop', 'N/A'))
        add_field("NCM", item.get('ncm', 'N/A'))
        add_field("Natureza da Operação", item.get('natureza', 'N/A'))
        add_field("IE Tomador", item.get('ie_tomador', 'N/A'))
        
        add_separator("Informações do Sistema")
        add_field("Informante (CNPJ Consulta)", item.get('informante', 'N/A'))
        add_field("Chave de Acesso", item.get('chave', 'N/A'))
        
        # CNPJ Destinatário (se disponível)
        if item.get('cnpj_destinatario'):
            add_field("CNPJ Destinatário", item.get('cnpj_destinatario', 'N/A'))
        
        # Finaliza scroll
        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        # Botões
        button_box = QWidget()
        button_layout = QHBoxLayout(button_box)
        
        # Botão Copiar Chave
        btn_copy_chave = QPushButton("📋 Copiar Chave")
        btn_copy_chave.clicked.connect(lambda: self._copiar_para_clipboard(item.get('chave', '')))
        button_layout.addWidget(btn_copy_chave)
        
        # Botão Copiar CNPJ
        btn_copy_cnpj = QPushButton("📋 Copiar CNPJ Emitente")
        btn_copy_cnpj.clicked.connect(lambda: self._copiar_para_clipboard(item.get('cnpj_emitente', '')))
        button_layout.addWidget(btn_copy_cnpj)
        
        button_layout.addStretch()
        
        # Botão Fechar
        btn_close = QPushButton("✖ Fechar")
        btn_close.clicked.connect(dialog.accept)
        button_layout.addWidget(btn_close)
        
        layout.addWidget(button_box)
        
        # Exibe o diálogo
        dialog.exec_()
    
    def _copiar_para_clipboard(self, texto: str):
        """Copia texto para a área de transferência"""
        if texto:
            clipboard = QApplication.clipboard()
            clipboard.setText(texto)
            self.set_status(f"✅ Copiado: {texto}", 2000)
    
    def _buscar_xml_completo(self, item: Dict[str, Any]):
        """Busca o XML completo de um resumo na SEFAZ"""
        chave = item.get('chave')
        if not chave or len(chave) != 44:
            QMessageBox.warning(self, "Erro", "Chave de acesso inválida!")
            return
        
        # Determina qual certificado usar
        informante = item.get('informante')
        certs = self.db.load_certificates()
        
        cert_to_use = None
        if informante:
            # Tenta usar o certificado que baixou esse resumo
            for c in certs:
                if c.get('informante') == informante:
                    cert_to_use = c
                    break
        
        if not cert_to_use and certs:
            # Usa o primeiro certificado ativo
            cert_to_use = certs[0]
        
        if not cert_to_use:
            QMessageBox.warning(self, "Erro", "Nenhum certificado disponível!")
            return
        
        # Busca na SEFAZ
        try:
            from nfe_search import consultar_nfe_por_chave
            
            self.set_status(f"Buscando XML completo da chave {chave[:10]}...")
            QApplication.processEvents()
            
            xml_completo = consultar_nfe_por_chave(
                chave=chave,
                certificado_path=cert_to_use.get('caminho'),
                senha=cert_to_use.get('senha'),
                cnpj=cert_to_use.get('cnpj_cpf'),
                cuf=cert_to_use.get('cUF_autor')
            )
            
            if xml_completo:
                # Verifica se é XML completo (com dados da nota) ou apenas protocolo
                xml_lower = xml_completo.lower()
                is_only_protocol = (
                    '<retconssit' in xml_lower and 
                    '<protnfe' in xml_lower and
                    '<nfeproc' not in xml_lower and
                    '<nfe' not in xml_lower.replace('nferesultmsg', '').replace('protnfe', '')
                )
                
                if is_only_protocol:
                    # Apenas protocolo disponível, XML completo não está mais acessível
                    self.set_status("⚠ Apenas protocolo disponível (XML completo não acessível)", 5000)
                    
                    # Remove da interface (deleta do banco)
                    self.db.deletar_nota_detalhada(chave)
                    self.refresh_table()
                    
                    QMessageBox.warning(
                        self, 
                        "XML Completo Não Disponível", 
                        f"A SEFAZ retornou apenas o protocolo de autorização.\n\n"
                        f"O XML completo da NFe não está mais disponível para consulta.\n\n"
                        f"Possíveis causas:\n"
                        f"• NFe muito antiga (fora do período de disponibilidade)\n"
                        f"• Nota cancelada\n"
                        f"• Acesso negado\n\n"
                        f"O resumo foi removido da listagem."
                    )
                else:
                    # XML completo obtido com sucesso
                    from nfe_search import salvar_xml_por_certificado
                    
                    # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local + TODOS os perfis ativos
                    cnpj_informante = informante or cert_to_use.get('cnpj_cpf')
                    nome_cert = cert_to_use.get('nome_certificado')
                    
                    # 1. Salva em backup local (xmls/)
                    salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base="xmls")
                    
                    # 2. Salva em TODOS os perfis ativos (pasta_base=None)
                    salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base=None, nome_certificado=nome_cert)
                    
                    # Atualiza no banco
                    nota = item.copy()
                    nota['xml_status'] = 'COMPLETO'
                    self.db.salvar_nota_detalhada(nota)
                    
                    self.set_status(f"✓ XML completo baixado e salvo!", 3000)
                    self.refresh_table()
                    
                    QMessageBox.information(
                        self, 
                        "Sucesso", 
                        f"XML completo da nota {item.get('numero')} baixado com sucesso!\n\n"
                        f"O registro foi atualizado de 'Resumo' para 'Completo'."
                    )
            else:
                self.set_status("Erro ao buscar XML completo", 3000)
                QMessageBox.warning(
                    self, 
                    "Erro", 
                    "Não foi possível obter o XML completo da SEFAZ.\n\n"
                    "Possíveis causas:\n"
                    "- Nota não encontrada\n"
                    "- Certificado sem permissão de acesso\n"
                    "- Problema de conexão com SEFAZ"
                )
        except Exception as e:
            self.set_status(f"Erro: {str(e)}", 5000)
            QMessageBox.critical(self, "Erro", f"Erro ao buscar XML completo:\n\n{str(e)}")
    
    def _baixar_xml_e_pdf_multiplos(self, selected_rows: list):
        """
        Baixa XMLs completos para múltiplas notas selecionadas.
        Processa cada nota sequencialmente com barra de progresso.
        """
        total = len(selected_rows)
        
        # Busca todas as notas selecionadas no banco
        notas = []
        chave_col_index = None
        
        # Encontra índice da coluna Chave
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            if header_item and header_item.text() == "Chave":
                chave_col_index = col
                break
        
        if chave_col_index is None:
            QMessageBox.warning(self, "Erro", "Coluna 'Chave' não encontrada!")
            return
        
        # Coleta chaves de todas as linhas selecionadas
        for row in selected_rows:
            chave_item = self.table.item(row, chave_col_index)
            if chave_item:
                chave = chave_item.text()
                try:
                    import sqlite3
                    conn = sqlite3.connect(str(DATA_DIR / 'notas.db'))
                    conn.row_factory = sqlite3.Row
                    nota = conn.execute('SELECT * FROM notas_detalhadas WHERE chave = ?', (chave,)).fetchone()
                    conn.close()
                    
                    if nota:
                        notas.append(dict(nota))
                except Exception as e:
                    print(f"[ERRO] Falha ao buscar nota {chave}: {e}")
        
        if not notas:
            QMessageBox.warning(self, "Erro", "Nenhuma nota válida selecionada!")
            return
        
        # Confirma ação
        reply = QMessageBox.question(
            self,
            "Confirmar Download",
            f"Deseja baixar XML completo de {len(notas)} nota(s)?\n\n"
            f"Esta operação irá:\n"
            f"• Manifestar ciência (para NF-e)\n"
            f"• Baixar XMLs completos da SEFAZ\n"
            f"• Gerar PDFs automaticamente\n"
            f"• Atualizar interface",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # Processa cada nota
        sucessos = []
        falhas = []
        
        for i, nota in enumerate(notas, 1):
            chave = nota.get('chave')
            numero = nota.get('numero') or nota.get('nNF') or chave[:10]
            
            self.set_status(f"📥 Processando {i}/{len(notas)}: Nota {numero}...", 0)
            QApplication.processEvents()
            
            try:
                # Chama função existente em modo silencioso (sem diálogos)
                resultado = self._baixar_xml_e_pdf(nota, silent_mode=True)
                
                if resultado.get('sucesso'):
                    sucessos.append({
                        'chave': chave,
                        'numero': numero,
                        'nota': nota.get('nota_fiscal', numero)
                    })
                else:
                    falhas.append({
                        'chave': chave,
                        'numero': numero,
                        'nota': nota.get('nota_fiscal', numero),
                        'erro': resultado.get('erro', 'Erro desconhecido')
                    })
                
            except Exception as e:
                print(f"[ERRO] Falha ao processar nota {chave}: {e}")
                falhas.append({
                    'chave': chave,
                    'numero': numero,
                    'nota': nota.get('nota_fiscal', numero),
                    'erro': str(e)
                })
        
        # Atualiza interface ao final
        self.refresh_table()
        self.refresh_emitidos_table()
        
        # Monta mensagem de resumo detalhada
        msg = f"📊 Resultado do Download em Lote\n\n"
        msg += f"Total processado: {len(notas)} nota(s)\n\n"
        
        if sucessos:
            msg += f"✅ Sucesso: {len(sucessos)} nota(s)\n"
            msg += f"   • XMLs completos baixados\n"
            msg += f"   • PDFs gerados automaticamente\n"
            msg += f"   • Interface atualizada\n\n"
        
        if falhas:
            msg += f"❌ Falhas: {len(falhas)} nota(s)\n"
            # Mostra primeiras 3 falhas
            max_mostrar = min(3, len(falhas))
            for i, f in enumerate(falhas[:max_mostrar]):
                msg += f"   • Nota {f['numero']}: {f['erro'][:60]}...\n"
            
            if len(falhas) > max_mostrar:
                msg += f"   • ... e mais {len(falhas) - max_mostrar} erro(s)\n"
        
        # Mostra diálogo com resumo
        if falhas:
            QMessageBox.warning(self, "Download Concluído com Avisos", msg)
        else:
            QMessageBox.information(self, "Download Concluído", msg)
        
        self.set_status(f"✅ {len(sucessos)}/{len(notas)} XMLs baixados com sucesso", 5000)
    
    def _baixar_xml_e_pdf(self, item: Dict[str, Any], silent_mode: bool = False):
        """
        Manifesta Ciência da Operação, baixa XML completo da SEFAZ, 
        atualiza interface para verde e gera PDF automaticamente.
        Método otimizado para ação direta sem diálogos intermediários.
        
        Args:
            item: Dicionário com dados da nota
            silent_mode: Se True, suprime diálogos de erro e retorna dict com resultado
        
        Returns:
            dict: {'sucesso': bool, 'erro': str} quando silent_mode=True
        """
        chave = item.get('chave')
        if not chave or len(chave) != 44:
            if silent_mode:
                return {'sucesso': False, 'erro': 'Chave de acesso inválida'}
            QMessageBox.warning(self, "Erro", "Chave de acesso inválida!")
            return
        
        # Determina certificado
        informante = item.get('informante')
        certs = self.db.load_certificates()
        
        cert_to_use = None
        if informante:
            for c in certs:
                if c.get('informante') == informante:
                    cert_to_use = c
                    break
        
        if not cert_to_use and certs:
            cert_to_use = certs[0]
        
        if not cert_to_use:
            if silent_mode:
                return {'sucesso': False, 'erro': 'Nenhum certificado disponível'}
            QMessageBox.warning(self, "Erro", "Nenhum certificado disponível!")
            return
        
        cert_path = cert_to_use.get('caminho')
        cert_senha = cert_to_use.get('senha')
        cert_cnpj = cert_to_use.get('cnpj_cpf')
        
        # Detecta tipo de documento (modelo)
        modelo = chave[20:22]
        is_nfe = modelo == '55'
        is_cte = modelo == '57'
        
        try:
            from nfe_search import NFeService, salvar_xml_por_certificado, extrair_nota_detalhada
            from modules.manifestacao_service import ManifestacaoService
            import time
            
            # 0️⃣ MANIFESTAR CIÊNCIA (SOMENTE PARA NF-e)
            if is_nfe:
                self.set_status(f"📝 Manifestando ciência da operação...")
                QApplication.processEvents()
                
                # Verifica se já foi manifestado
                eventos_existentes = self.db.get_manifestacoes_by_chave(chave)
                ja_manifestado = any(e.get('tipo_evento') == '210200' for e in eventos_existentes)
                
                if not ja_manifestado:
                    # 🔔 MANIFESTAR CIÊNCIA DA OPERAÇÃO (evento 210200)
                    try:
                        manifesta_service = ManifestacaoService(cert_path, cert_senha)
                        
                        sucesso, protocolo, mensagem, xml_resposta = manifesta_service.enviar_manifestacao(
                            chave=chave,
                            tipo_evento='210200',  # Ciência da Operação
                            cnpj_destinatario=cert_cnpj,
                            justificativa=None  # Evento 210200 não requer justificativa
                        )
                        
                        if not sucesso:
                            print(f"[INFO] Manifestação não realizada: {mensagem}")
                            self.set_status("ℹ️ Pulando manifestação (continuando com download)", 2000)
                            # ⚠️ NÃO mostra popup - continua silenciosamente com download
                            # Continua tentando baixar mesmo sem manifestar
                        else:
                            # ✅ NÃO salva retEnvEvento (apenas confirmação, não contém nota)
                            # Removido: salvar_xml_por_certificado(xml_resposta, ...) 
                            
                            # Registra no banco
                            self.db.register_manifestacao(
                                chave=chave,
                                tipo_evento='210200',
                                informante=informante or cert_cnpj,
                                status="REGISTRADA",
                                protocolo=protocolo
                            )
                            
                            self.set_status("⏱️ Aguardando processamento SEFAZ (3s)...", 0)
                            QApplication.processEvents()
                            time.sleep(3)  # Aguarda SEFAZ processar a manifestação
                            
                    except Exception as e:
                        print(f"[INFO] Erro ao manifestar: {e}")
                        self.set_status("ℹ️ Pulando manifestação (continuando com download)", 2000)
                        # ⚠️ NÃO mostra popup - continua silenciosamente
                else:
                    self.set_status("✅ Já manifestado anteriormente", 1000)
                    QApplication.processEvents()
            elif is_cte:
                self.set_status("ℹ️ CT-e não requer manifestação prévia", 1000)
                QApplication.processEvents()
            
            # 1️⃣ BUSCAR XML NO SEFAZ
            self.set_status(f"🔄 Baixando XML completo da chave {chave[:10]}...")
            QApplication.processEvents()
            
            svc = NFeService(
                cert_path,
                cert_senha,
                cert_cnpj,
                cert_to_use.get('cUF_autor')
            )
            
            # Busca XML via Distribuição DFe (método mais confiável)
            xml_completo = None
            resposta_sefaz = None
            cstat_sefaz = None
            motivo_sefaz = None
            
            try:
                resposta_sefaz = svc.fetch_by_chave_dist(chave)
                if resposta_sefaz:
                    # Extrai cStat e xMotivo da resposta SEFAZ
                    from lxml import etree
                    import base64
                    import gzip
                    
                    try:
                        root = etree.fromstring(resposta_sefaz.encode('utf-8'))
                        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                        
                        # Busca cStat/xMotivo do retDistDFeInt
                        cstat_sefaz = root.findtext('.//nfe:cStat', namespaces=ns) or root.findtext('.//cStat')
                        motivo_sefaz = root.findtext('.//nfe:xMotivo', namespaces=ns) or root.findtext('.//xMotivo')
                        
                        logger.info(f"📋 Resposta SEFAZ - cStat: {cstat_sefaz} | xMotivo: {motivo_sefaz}")
                    except Exception as e:
                        logger.warning(f"⚠️ Erro ao extrair cStat/xMotivo: {e}")
                    
                    # Verifica se tem o XML completo em texto claro
                    if '<nfeProc' in resposta_sefaz or '<procNFe' in resposta_sefaz:
                        xml_completo = resposta_sefaz
                        logger.info(f"✅ XML completo encontrado na resposta (texto claro)")
                    else:
                        # Tenta descompactar documentos zipados (docZip)
                        try:
                            root = etree.fromstring(resposta_sefaz.encode('utf-8'))
                            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                            
                            # Procura por loteDistDFeInt com docZip
                            docZips = root.findall('.//nfe:docZip', namespaces=ns) or root.findall('.//docZip')
                            
                            if docZips:
                                logger.info(f"📦 Encontrados {len(docZips)} documento(s) compactado(s)")
                                
                                for docZip in docZips:
                                    try:
                                        # Decodifica base64 e descompacta gzip
                                        zip_b64 = docZip.text
                                        if zip_b64:
                                            zip_bytes = base64.b64decode(zip_b64)
                                            xml_bytes = gzip.decompress(zip_bytes)
                                            xml_descompactado = xml_bytes.decode('utf-8')
                                            
                                            # Verifica se é procNFe/nfeProc
                                            if '<procNFe' in xml_descompactado or '<nfeProc' in xml_descompactado:
                                                xml_completo = xml_descompactado
                                                logger.info(f"✅ XML completo descompactado com sucesso")
                                                break
                                    except Exception as e:
                                        logger.warning(f"⚠️ Erro ao descompactar docZip: {e}")
                                        continue
                            else:
                                logger.warning(f"⚠️ Resposta não contém XML completo nem docZip")
                        except Exception as e:
                            logger.warning(f"⚠️ Erro ao processar documentos zipados: {e}")
            except Exception as e:
                logger.error(f"❌ Erro ao buscar por distribuição: {e}")
            
            if not xml_completo:
                erro_msg = "XML não disponível no SEFAZ"
                if cstat_sefaz and motivo_sefaz:
                    erro_msg = f"{cstat_sefaz}: {motivo_sefaz}"
                
                self.set_status("❌ XML não disponível no SEFAZ", 3000)
                
                if silent_mode:
                    return {'sucesso': False, 'erro': erro_msg}
                
                # Monta mensagem com detalhes da SEFAZ
                msg_detalhes = "Não foi possível obter o XML completo da SEFAZ.\n\n"
                
                if cstat_sefaz and motivo_sefaz:
                    msg_detalhes += f"📋 Resposta SEFAZ:\n"
                    msg_detalhes += f"   • cStat: {cstat_sefaz}\n"
                    msg_detalhes += f"   • Motivo: {motivo_sefaz}\n\n"
                
                msg_detalhes += "💡 Possíveis causas:\n"
                msg_detalhes += "   • Nota muito antiga (fora do prazo de disponibilidade)\n"
                msg_detalhes += "   • Nota cancelada sem XML completo\n"
                msg_detalhes += "   • Acesso negado pelo certificado\n"
                msg_detalhes += "   • Problema de conexão"
                
                QMessageBox.warning(
                    self,
                    "XML Não Disponível",
                    msg_detalhes
                )
                return
            
            self.set_status("💾 Salvando XML...", 0)
            QApplication.processEvents()
            
            # 2️⃣ SALVAR XML
            cnpj_informante = informante or cert_to_use.get('cnpj_cpf')
            nome_cert = self.db.get_cert_nome_by_informante(cnpj_informante)
            
            # 1. Salva em backup local (xmls/)
            salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base="xmls")
            
            # 2. Salva em TODOS os perfis ativos (pasta_base=None)
            salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base=None, nome_certificado=nome_cert)
            
            # 3️⃣ ATUALIZAR BANCO (xml_status = COMPLETO)
            # Carrega dados existentes
            logger.info(f"🔍 Buscando nota existente no banco com chave: {chave}")
            with self.db._connect() as conn:
                existing = conn.execute(
                    "SELECT * FROM notas_detalhadas WHERE chave = ?",
                    (chave,)
                ).fetchone()
                
                logger.info(f"📊 Nota {'ENCONTRADA' if existing else 'NÃO ENCONTRADA'} no banco")
                
                if existing:
                    columns = [desc[0] for desc in conn.execute("SELECT * FROM notas_detalhadas LIMIT 0").description]
                    nota_update = dict(zip(columns, existing))
                    
                    # 🔒 PROTEÇÃO: Nunca sobrescreve EVENTO para COMPLETO
                    old_xml_status = nota_update.get('xml_status', 'RESUMO').upper()
                    if old_xml_status == 'EVENTO':
                        self.set_status("ℹ️ Registro é EVENTO, não será atualizado", 2000)
                        if silent_mode:
                            return {'sucesso': False, 'erro': 'Registro é EVENTO, não pode ser atualizado'}
                        return  # Não atualiza eventos
                    
                    nota_update['xml_status'] = 'COMPLETO'
                    
                    # Extrai dados completos do XML
                    from nfe_search import extrair_nota_detalhada, XMLProcessor
                    parser = XMLProcessor()
                    nota_detalhada = extrair_nota_detalhada(
                        xml_txt=xml_completo,
                        parser=parser,
                        db=self.db,
                        chave=chave,
                        informante=informante or cert_cnpj,
                        nsu_documento=None  # Não temos NSU ao baixar por chave
                    )
                    if nota_detalhada:
                        # Atualiza campos com dados do XML completo
                        for key, value in nota_detalhada.items():
                            if value and value != '':  # Só sobrescreve se tem valor
                                nota_update[key] = value
                    
                    # Salva no banco
                    self.db.save_note(nota_update)
                else:
                    # Se não existir, cria novo registro
                    from nfe_search import extrair_nota_detalhada, XMLProcessor
                    parser = XMLProcessor()
                    nota_detalhada = extrair_nota_detalhada(
                        xml_txt=xml_completo,
                        parser=parser,
                        db=self.db,
                        chave=chave,
                        informante=informante or cert_cnpj,
                        nsu_documento=None
                    )
                    if nota_detalhada:
                        self.db.save_note(nota_detalhada)
            
            self.set_status("📄 Gerando PDF...", 0)
            QApplication.processEvents()
            
            # Determina pasta de destino (mesmo local do XML)
            tipo = (item.get('tipo') or 'NFe').strip().upper().replace('-', '')
            data_emissao = (item.get('data_emissao') or '')[:10]
            
            if data_emissao and len(data_emissao) >= 7:
                year_month = data_emissao[:7]
            else:
                from datetime import datetime
                year_month = datetime.now().strftime("%Y-%m")
            
            xmls_root = DATA_DIR / "xmls" / (informante or cert_to_use.get('cnpj_cpf')) / tipo / year_month
            xml_file = xmls_root / f"{chave}.xml"
            
            # 4️⃣ GERAR PDF AUTOMATICAMENTE (se XML existe)
            if xml_file.exists():
                try:
                    # PDF já é gerado automaticamente pelo salvar_xml_por_certificado
                    pdf_file = xml_file.with_suffix('.pdf')
                    if pdf_file.exists():
                        logger.info(f"[XML COMPLETO] ✅ PDF já gerado: {pdf_file}")
                    else:
                        logger.warning(f"[XML COMPLETO] ⚠️ PDF não foi gerado automaticamente")
                except Exception as e:
                    logger.warning(f"[XML COMPLETO] ⚠️ Erro ao verificar PDF: {e}")
            
            # 5️⃣ ATUALIZAR INTERFACE (CINZA → VERDE)
            self.set_status("✅ XML completo baixado e PDF gerado!", 3000)
            self.refresh_table()
            self.refresh_emitidos_table()
            
            # Retorna sucesso em silent_mode ou mostra diálogo
            if silent_mode:
                return {'sucesso': True}
            
            QMessageBox.information(
                self,
                "Sucesso!",
                f"✅ XML completo baixado com sucesso!\n"
                f"📄 PDF gerado automaticamente\n"
                f"🟢 Interface atualizada\n\n"
                f"Nota: {item.get('numero')}\n"
                f"Pasta: {xmls_root.name if xmls_root.exists() else 'xmls'}"
            )
            
        except Exception as e:
            self.set_status(f"❌ Erro: {str(e)}", 5000)
            
            if silent_mode:
                return {'sucesso': False, 'erro': str(e)}
            
            QMessageBox.critical(
                self,
                "Erro",
                f"Erro ao baixar XML completo:\n\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def _baixar_xml_completo_apos_manifestacao(self, chave: str, informante: str, cert_path: str, cert_senha: str, cert_cnpj: str):
        """
        Baixa XML completo da SEFAZ após manifestação bem-sucedida.
        NÃO manifesta novamente - assume que manifestação já foi feita.
        
        Args:
            chave: Chave de acesso da nota (44 dígitos)
            informante: CNPJ/CPF do informante
            cert_path: Caminho do certificado
            cert_senha: Senha do certificado
            cert_cnpj: CNPJ do certificado
        """
        try:
            from nfe_search import NFeService, salvar_xml_por_certificado, extrair_nota_detalhada, XMLProcessor
            
            print(f"[AUTO-DOWNLOAD] 🔄 Buscando XML completo da chave {chave[:25]}...")
            self.set_status(f"🔄 Baixando XML completo automaticamente...", 0)
            QApplication.processEvents()
            
            # Obtém cUF do certificado
            certs = self.db.load_certificates()
            cuf = None
            for c in certs:
                if c.get('informante') == informante or c.get('cnpj_cpf') == cert_cnpj:
                    cuf = c.get('cUF_autor')
                    break
            
            # Cria serviço NFeService
            svc = NFeService(cert_path, cert_senha, cert_cnpj, cuf)
            
            # Tenta buscar XML pela distribuição DFe
            xml_completo = None
            try:
                xml_completo = svc.fetch_by_chave_dist(chave)
                if xml_completo and (('<nfeProc' in xml_completo) or ('<procNFe' in xml_completo)):
                    print(f"[AUTO-DOWNLOAD] ✅ XML obtido via DistribuicaoDFe")
                else:
                    xml_completo = None
            except Exception as e:
                print(f"[AUTO-DOWNLOAD] ⚠️ Erro fetch_by_chave_dist: {e}")
            
            # Fallback para método alternativo
            if not xml_completo:
                try:
                    xml_completo = svc.fetch_by_key(chave)
                    if xml_completo:
                        print(f"[AUTO-DOWNLOAD] ✅ XML obtido via fetch_by_key")
                except Exception as e:
                    print(f"[AUTO-DOWNLOAD] ⚠️ Erro fetch_by_key: {e}")
            
            if not xml_completo or (('<nfeProc' not in xml_completo) and ('<procNFe' not in xml_completo)):
                print(f"[AUTO-DOWNLOAD] ❌ XML não disponível")
                self.set_status("⚠️ XML completo não disponível ainda", 3000)
                return False
            
            print(f"[AUTO-DOWNLOAD] 💾 Salvando XML...")
            self.set_status("💾 Salvando XML...", 0)
            QApplication.processEvents()
            
            # Salva XML no disco
            cnpj_informante = informante or cert_cnpj
            nome_cert = self.db.get_cert_nome_by_informante(cnpj_informante)
            
            # 1. Salva em backup local (xmls/)
            salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base="xmls")
            
            # 2. Salva em TODOS os perfis ativos (pasta_base=None)
            salvar_xml_por_certificado(xml_completo, cnpj_informante, pasta_base=None, nome_certificado=nome_cert)
            
            # Atualiza banco de dados
            print(f"[AUTO-DOWNLOAD] 📝 Atualizando banco de dados...")
            with self.db._connect() as conn:
                existing = conn.execute(
                    "SELECT * FROM notas_detalhadas WHERE chave = ?",
                    (chave,)
                ).fetchone()
                
                if existing:
                    columns = [desc[0] for desc in conn.execute("SELECT * FROM notas_detalhadas LIMIT 0").description]
                    nota_update = dict(zip(columns, existing))
                    
                    # Verifica se não é EVENTO
                    old_xml_status = nota_update.get('xml_status', 'RESUMO').upper()
                    if old_xml_status == 'EVENTO':
                        print(f"[AUTO-DOWNLOAD] ℹ️ Registro é EVENTO, não será atualizado")
                        return True  # XML foi salvo, mas não atualiza registro de evento
                    
                    nota_update['xml_status'] = 'COMPLETO'
                    
                    # Extrai dados completos do XML
                    parser = XMLProcessor()
                    nota_detalhada = extrair_nota_detalhada(
                        xml_txt=xml_completo,
                        parser=parser,
                        db=self.db,
                        chave=chave,
                        informante=informante or cert_cnpj,
                        nsu_documento=None  # Não temos NSU ao baixar por chave
                    )
                    
                    if nota_detalhada:
                        # Atualiza campos com dados do XML completo
                        for key, value in nota_detalhada.items():
                            if value and value != '':
                                nota_update[key] = value
                    
                    # Salva no banco
                    self.db.save_note(nota_update)
                    print(f"[AUTO-DOWNLOAD] ✅ Banco de dados atualizado")
                else:
                    # Se não existe, cria novo registro
                    parser = XMLProcessor()
                    nota_detalhada = extrair_nota_detalhada(
                        xml_txt=xml_completo,
                        parser=parser,
                        db=self.db,
                        chave=chave,
                        informante=informante or cert_cnpj,
                        nsu_documento=None
                    )
                    if nota_detalhada:
                        self.db.save_note(nota_detalhada)
                        print(f"[AUTO-DOWNLOAD] ✅ Nova nota criada no banco")
            
            # Tenta gerar PDF
            print(f"[AUTO-DOWNLOAD] 📄 Gerando PDF...")
            self.set_status("📄 Gerando PDF...", 0)
            QApplication.processEvents()
            
            try:
                from modules.pdf_generator import generate_pdf_from_xml
                from pathlib import Path
                
                # Busca arquivo XML salvo
                xmls_base = DATA_DIR / "xmls" / (informante or cert_cnpj)
                
                # Procura recursivamente pelo XML
                xml_files = list(xmls_base.rglob(f"{chave}.xml"))
                if xml_files:
                    xml_file = xml_files[0]
                    pdf_file = xml_file.with_suffix('.pdf')
                    generate_pdf_from_xml(str(xml_file), str(pdf_file))
                    print(f"[AUTO-DOWNLOAD] ✅ PDF gerado: {pdf_file.name}")
            except Exception as e:
                print(f"[AUTO-DOWNLOAD] ⚠️ Erro ao gerar PDF: {e}")
            
            # Atualiza interface
            print(f"[AUTO-DOWNLOAD] 🔄 Atualizando interface...")
            self.set_status("✅ XML completo baixado e interface atualizada!", 3000)
            self.refresh_table()
            self.refresh_emitidos_table()
            
            print(f"[AUTO-DOWNLOAD] ✅ Processo concluído com sucesso!")
            return True
            
        except Exception as e:
            print(f"[AUTO-DOWNLOAD] ❌ Erro: {e}")
            import traceback
            traceback.print_exc()
            self.set_status(f"❌ Erro no download automático", 3000)
            return False
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.set_status(f"❌ Erro: {str(e)}", 5000)
            QMessageBox.critical(
                self,
                "Erro",
                f"Erro ao processar XML:\n\n{str(e)}\n\nVerifique os logs para mais detalhes."
            )
    
    def _consultar_status_nota(self, item: Dict[str, Any]):
        """Consulta o status atual de uma nota na SEFAZ"""
        chave = item.get('chave')
        if not chave or len(chave) != 44:
            QMessageBox.warning(self, "Erro", "Chave de acesso inválida!")
            return
        
        # Determina qual certificado usar
        informante = item.get('informante')
        certs = self.db.load_certificates()
        
        cert_to_use = None
        if informante:
            # Tenta usar o certificado que baixou esse resumo
            for c in certs:
                if c.get('informante') == informante:
                    cert_to_use = c
                    break
        
        if not cert_to_use and certs:
            # Usa o primeiro certificado ativo
            cert_to_use = certs[0]
        
        if not cert_to_use:
            QMessageBox.warning(self, "Erro", "Nenhum certificado disponível!")
            return
        
        # Consulta na SEFAZ
        try:
            from nfe_search import consultar_nfe_por_chave
            
            self.set_status(f"🔄 Consultando status da nota {item.get('numero', '')}...")
            QApplication.processEvents()
            
            xml_resposta = consultar_nfe_por_chave(
                chave=chave,
                certificado_path=cert_to_use.get('caminho'),
                senha=cert_to_use.get('senha'),
                cnpj=cert_to_use.get('cnpj_cpf'),
                cuf=cert_to_use.get('cUF_autor')
            )
            
            if xml_resposta:
                # Extrai o status da resposta
                import re
                from xml.etree import ElementTree as ET
                
                try:
                    root = ET.fromstring(xml_resposta)
                    
                    # Procura pelo cStat (código de status)
                    cstat = None
                    xmotivo = None
                    
                    # Namespace comum em respostas NFe
                    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                    
                    # Tenta encontrar cStat e xMotivo
                    for cstat_elem in root.iter():
                        if 'cStat' in cstat_elem.tag:
                            cstat = cstat_elem.text
                        if 'xMotivo' in cstat_elem.tag:
                            xmotivo = cstat_elem.text
                    
                    # Interpreta o status
                    if cstat:
                        status_msg = f"Status: {cstat}"
                        if xmotivo:
                            status_msg += f"\n{xmotivo}"
                        
                        # Determina o novo status da nota
                        novo_status = None
                        if cstat in ['100', '150']:  # Autorizada
                            novo_status = "Autorizado o uso da NF-e"
                            icone = "✅"
                            cor = "green"
                        elif cstat in ['101', '151', '135', '155']:  # Cancelada
                            novo_status = "Cancelamento homologado"
                            icone = "❌"
                            cor = "red"
                        elif cstat in ['110', '301', '302']:  # Denegada
                            novo_status = "Uso Denegado"
                            icone = "⚠️"
                            cor = "orange"
                        elif cstat == '217':  # Nota não existe
                            novo_status = "Nota não consta na base da SEFAZ"
                            icone = "❓"
                            cor = "gray"
                        else:
                            novo_status = xmotivo or f"Status {cstat}"
                            icone = "ℹ️"
                            cor = "blue"
                        
                        # Atualiza no banco de dados
                        if novo_status and novo_status != item.get('status'):
                            print(f"[DEBUG STATUS] Atualizando status da nota:")
                            print(f"  Chave: {chave}")
                            print(f"  Status antigo: {item.get('status')}")
                            print(f"  Status novo: {novo_status}")
                            
                            self.db.atualizar_status_nota(chave, novo_status)
                            
                            # Recarrega os dados do banco para garantir sincronização
                            print(f"[DEBUG STATUS] Recarregando notas do banco...")
                            self.notes = self.db.load_notes(limit=5000)
                            
                            # Verifica se o status foi atualizado na memória
                            nota_atualizada = next((n for n in self.notes if n.get('chave') == chave), None)
                            if nota_atualizada:
                                print(f"[DEBUG STATUS] Nota encontrada após reload:")
                                print(f"  Status: {nota_atualizada.get('status')}")
                                print(f"  xml_status: {nota_atualizada.get('xml_status')}")
                            else:
                                print(f"[DEBUG STATUS] ⚠️ Nota não encontrada após reload!")
                            
                            # Recarrega as tabelas para atualizar os ícones
                            print(f"[DEBUG STATUS] Atualizando tabelas...")
                            self.refresh_table()
                            self.refresh_emitidos_table()
                            
                            self.set_status(f"✓ Status atualizado: {novo_status}", 5000)
                            
                            # Se foi cancelada, tenta buscar o evento de cancelamento automaticamente
                            if cstat in ['101', '151', '135', '155']:  # Cancelada
                                QTimer.singleShot(500, lambda: self._buscar_evento_cancelamento(item))
                        else:
                            self.set_status(f"Status consultado: {novo_status}", 5000)
                        
                        # Mostra resultado
                        QMessageBox.information(
                            self,
                            f"{icone} Status da Nota",
                            f"<h3>Consulta realizada com sucesso!</h3>"
                            f"<p><b>Nota:</b> {item.get('numero', 'N/A')}</p>"
                            f"<p><b>Código:</b> {cstat}</p>"
                            f"<p><b>Status:</b> <span style='color: {cor};'>{novo_status}</span></p>"
                            f"<p><b>Chave:</b> {chave}</p>"
                            f"{'<p><i>Buscando evento de cancelamento...</i></p>' if cstat in ['101', '151', '135', '155'] else ''}"
                        )
                    else:
                        # Não encontrou cStat na resposta
                        self.set_status("⚠ Resposta da SEFAZ sem código de status", 3000)
                        QMessageBox.warning(
                            self,
                            "Aviso",
                            "A SEFAZ retornou uma resposta, mas não foi possível extrair o código de status.\n\n"
                            "Tente novamente mais tarde."
                        )
                        
                except ET.ParseError as e:
                    self.set_status(f"Erro ao processar resposta XML", 3000)
                    QMessageBox.warning(
                        self,
                        "Erro",
                        f"Não foi possível processar a resposta da SEFAZ:\n\n{str(e)}"
                    )
            else:
                self.set_status("Erro ao consultar status", 3000)
                QMessageBox.warning(
                    self,
                    "Erro",
                    "Não foi possível consultar o status na SEFAZ.\n\n"
                    "Possíveis causas:\n"
                    "- Problema de conexão com SEFAZ\n"
                    "- Certificado sem permissão\n"
                    "- Serviço temporariamente indisponível"
                )
                
        except Exception as e:
            self.set_status(f"Erro: {str(e)}", 5000)
            QMessageBox.critical(self, "Erro", f"Erro ao consultar status:\n\n{str(e)}")
    
    def _buscar_evento_cancelamento(self, item: Dict[str, Any]):
        """Busca automaticamente o evento de cancelamento na SEFAZ"""
        chave = item.get('chave')
        if not chave or len(chave) != 44:
            return
        
        # Verifica se já tem o evento localmente
        xmls_root = DATA_DIR / "xmls"
        evento_encontrado = False
        
        try:
            if xmls_root.exists():
                for eventos_folder in xmls_root.rglob("Eventos"):
                    for xml_file in eventos_folder.glob("*.xml"):
                        try:
                            xml_content = xml_file.read_text(encoding='utf-8')
                            if chave in xml_content and '110111' in xml_content:  # 110111 = evento de cancelamento
                                evento_encontrado = True
                                self.set_status("✅ Evento de cancelamento já está salvo localmente", 3000)
                                return
                        except Exception:
                            continue
        except Exception:
            pass
        
        # Se não encontrou, tenta buscar na SEFAZ
        if not evento_encontrado:
            informante = item.get('informante')
            certs = self.db.load_certificates()
            
            cert_to_use = None
            if informante:
                for c in certs:
                    if c.get('informante') == informante:
                        cert_to_use = c
                        break
            
            if not cert_to_use and certs:
                cert_to_use = certs[0]
            
            if not cert_to_use:
                self.set_status("⚠️ Nenhum certificado para buscar eventos", 3000)
                return
            
            try:
                # Busca o XML completo que deve conter o protocolo de cancelamento
                from nfe_search import consultar_nfe_por_chave
                
                self.set_status("🔍 Buscando evento de cancelamento...", 0)
                QApplication.processEvents()
                
                xml_resposta = consultar_nfe_por_chave(
                    chave=chave,
                    certificado_path=cert_to_use.get('caminho'),
                    senha=cert_to_use.get('senha'),
                    cnpj=cert_to_use.get('cnpj_cpf'),
                    cuf=cert_to_use.get('cUF_autor')
                )
                
                if xml_resposta and ('retCancNFe' in xml_resposta or 'procEventoNFe' in xml_resposta):
                    # Salva o evento
                    from nfe_search import salvar_xml_por_certificado
                    
                    # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local + TODOS os perfis ativos
                    cnpj_informante = informante or cert_to_use.get('cnpj_cpf')
                    nome_cert = cert_to_use.get('nome_certificado')
                    
                    # 1. Salva em backup local (xmls/)
                    salvar_xml_por_certificado(xml_resposta, cnpj_informante, pasta_base="xmls")
                    
                    # 2. Salva em TODOS os perfis ativos (pasta_base=None)
                    salvar_xml_por_certificado(xml_resposta, cnpj_informante, pasta_base=None, nome_certificado=nome_cert)
                    
                    self.set_status("✅ Evento de cancelamento baixado e salvo!", 3000)
                else:
                    self.set_status("ℹ️ Evento de cancelamento não disponível na SEFAZ", 3000)
                    
            except Exception as e:
                self.set_status(f"⚠️ Erro ao buscar evento: {str(e)}", 3000)
    
    def _abrir_pdf_evento(self, evento: Dict[str, Any], dialog_parent: QDialog):
        """Abre o PDF de um documento quando clicado 2x na lista de eventos"""
        try:
            # Verifica se é um documento vinculado (tem relacao 'Vinculado')
            relacao = evento.get('relacao', '')
            if 'Vinculado' not in relacao:
                QMessageBox.information(dialog_parent, "Info", 
                    "Este é um evento deste documento.\n\n"
                    "Duplo-clique funciona apenas em documentos VINCULADOS (aba Vínculos).")
                return
            
            # Para documentos vinculados, precisa extrair a chave do caminho do arquivo
            caminho = evento.get('caminho', '')
            if not caminho:
                QMessageBox.information(dialog_parent, "Info", "Documento vinculado não possui caminho de arquivo.")
                return
            
            # Extrai a chave do nome do arquivo (44 dígitos)
            from pathlib import Path
            arquivo = Path(caminho).stem
            
            # Procura por 44 dígitos consecutivos no nome do arquivo
            import re
            match = re.search(r'\d{44}', arquivo)
            if not match:
                # Tenta ler a chave do próprio XML
                try:
                    xml_content = Path(caminho).read_text(encoding='utf-8')
                    # Procura por diferentes tags de chave
                    chave_patterns = [
                        r'<chCTe>(\d{44})</chCTe>',
                        r'<chNFe>(\d{44})</chNFe>',
                        r'<chMDFe>(\d{44})</chMDFe>',
                    ]
                    for pattern in chave_patterns:
                        match = re.search(pattern, xml_content)
                        if match:
                            break
                except Exception:
                    pass
                
                if not match:
                    QMessageBox.warning(dialog_parent, "Erro", 
                        "Não foi possível identificar a chave do documento.\n\n"
                        f"Arquivo: {Path(caminho).name}")
                    return
            
            chave_vinculada = match.group(1) if match.lastindex else match.group(0)
            
            print(f"\n[DEBUG ABRIR PDF EVENTO] ========== DUPLO CLIQUE EM EVENTO ==========")
            print(f"[DEBUG ABRIR PDF EVENTO] Arquivo: {Path(caminho).name}")
            print(f"[DEBUG ABRIR PDF EVENTO] Chave extraída: {chave_vinculada}")
            
            # Busca o documento no banco pela chave
            import sqlite3
            conn = sqlite3.connect(str(DATA_DIR / 'notas.db'))
            conn.row_factory = sqlite3.Row
            nota = conn.execute('SELECT * FROM notas_detalhadas WHERE chave = ?', (chave_vinculada,)).fetchone()
            
            if not nota:
                print(f"[DEBUG ABRIR PDF EVENTO] ❌ Não encontrou no banco com chave exata")
                # Tenta buscar por LIKE (caso tenha espaços ou caracteres extras)
                nota = conn.execute('SELECT * FROM notas_detalhadas WHERE REPLACE(chave, " ", "") = ?', (chave_vinculada.replace(" ", ""),)).fetchone()
                if nota:
                    print(f"[DEBUG ABRIR PDF EVENTO] ✅ Encontrou com busca LIKE")
            else:
                print(f"[DEBUG ABRIR PDF EVENTO] ✅ Encontrou no banco")
                print(f"[DEBUG ABRIR PDF EVENTO] Tipo: {nota['tipo']}, Número: {nota['numero']}")
            
            conn.close()
            
            if not nota:
                QMessageBox.warning(dialog_parent, "Erro", f"Documento não encontrado no banco.\nChave: {chave_vinculada[:10]}...")
                return
            
            # Converte para dict
            nota_dict = dict(nota)
            informante = nota_dict.get('informante', '')
            
            # OTIMIZAÇÃO 0: Verifica pdf_path do banco primeiro (SUPER RÁPIDO)
            pdf_path_db = nota_dict.get('pdf_path')
            if pdf_path_db:
                print(f"[DEBUG ABRIR PDF EVENTO] ⚡ PDF path do banco: {pdf_path_db}")
                pdf_file_db = Path(pdf_path_db)
                if pdf_file_db.exists():
                    print(f"[DEBUG ABRIR PDF EVENTO] ✅ Abrindo PDF do banco...")
                    import subprocess, sys
                    pdf_str = str(pdf_file_db.absolute())
                    if sys.platform == "win32":
                        subprocess.Popen(["cmd", "/c", "start", "", pdf_str], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)
                    else:
                        subprocess.Popen(["xdg-open", pdf_str])
                    self.set_status("✅ PDF aberto (cache DB)", 1000)
                    return
                else:
                    print(f"[DEBUG ABRIR PDF EVENTO] ⚠️ PDF path do banco inválido (arquivo não existe)")
            else:
                print(f"[DEBUG ABRIR PDF EVENTO] PDF path não está no banco, iniciando busca...")
            
            # Busca PDF já existente (MESMA LÓGICA DA TABELA PRINCIPAL)
            pdf_encontrado = None
            xmls_root = DATA_DIR / "xmls"
            tipo_doc = nota_dict.get('tipo', 'NFe')
            tipo_normalized = tipo_doc.strip().upper().replace('-', '') if tipo_doc else 'NFe'
            data_emissao = nota_dict.get('data_emissao', '')[:10] if nota_dict.get('data_emissao') else ''
            
            print(f"[DEBUG ABRIR PDF EVENTO] 🔍 Procurando PDF...")
            print(f"[DEBUG ABRIR PDF EVENTO] Chave: {chave_vinculada}")
            print(f"[DEBUG ABRIR PDF EVENTO] Informante: {informante}")
            print(f"[DEBUG ABRIR PDF EVENTO] Tipo: {tipo_normalized}")
            print(f"[DEBUG ABRIR PDF EVENTO] Data emissão: {data_emissao}")
            
            # ETAPA 1: Busca direta na pasta do mês (ESTRUTURA NOVA)
            if chave_vinculada and informante and data_emissao:
                try:
                    year_month = data_emissao[:7] if len(data_emissao) >= 7 else None
                    if year_month:
                        # Estrutura: xmls/{CNPJ}/{ANO-MES}/{TIPO}/{CHAVE}.pdf
                        specific_path = xmls_root / informante / year_month / tipo_normalized / f"{chave_vinculada}.pdf"
                        print(f"[DEBUG ABRIR PDF EVENTO] 📁 Estrutura nova: {specific_path}")
                        if specific_path.exists():
                            print(f"[DEBUG ABRIR PDF EVENTO] ✅ Encontrado (estrutura nova)!")
                            pdf_encontrado = specific_path
                            # AUTO-CURA: Salva no banco
                            self.db.atualizar_pdf_path(chave_vinculada, str(pdf_encontrado.absolute()))
                            print(f"[DEBUG ABRIR PDF EVENTO] 🔄 Auto-cura: PDF path salvo no banco")
                        else:
                            # Estrutura antiga: xmls/{CNPJ}/{ANO-MES}/{CHAVE}.pdf
                            old_path = xmls_root / informante / year_month / f"{chave_vinculada}.pdf"
                            print(f"[DEBUG ABRIR PDF EVENTO] 📁 Estrutura antiga: {old_path}")
                            if old_path.exists():
                                print(f"[DEBUG ABRIR PDF EVENTO] ✅ Encontrado (estrutura antiga)!")
                                pdf_encontrado = old_path
                                # AUTO-CURA: Salva no banco
                                self.db.atualizar_pdf_path(chave_vinculada, str(pdf_encontrado.absolute()))
                                print(f"[DEBUG ABRIR PDF EVENTO] 🔄 Auto-cura: PDF path salvo no banco")
                            else:
                                print(f"[DEBUG ABRIR PDF EVENTO] ❌ Não encontrado nas estruturas diretas")
                except Exception as e:
                    print(f"[DEBUG ABRIR PDF EVENTO] ⚠️ Erro na busca direta: {e}")
            
            # ETAPA 2: Busca recursiva (último recurso)
            if not pdf_encontrado and chave_vinculada and informante:
                print(f"[DEBUG ABRIR PDF EVENTO] 🔄 Iniciando busca recursiva...")
                pasta_informante = xmls_root / informante
                if pasta_informante.exists():
                    folders = list(sorted(pasta_informante.glob("20*"), reverse=True))
                    folders.extend(sorted(pasta_informante.glob("*/20*"), reverse=True))
                    print(f"[DEBUG ABRIR PDF EVENTO] 📂 Verificando {len(folders)} pastas...")
                    for idx, year_month_folder in enumerate(folders[:20], 1):  # Limita a 20 primeiras pastas
                        potential_pdf = year_month_folder / f"{chave_vinculada}.pdf"
                        if potential_pdf.exists():
                            print(f"[DEBUG ABRIR PDF EVENTO] ✅ Encontrado (busca recursiva [{idx}]): {year_month_folder.name}")
                            pdf_encontrado = potential_pdf
                            # AUTO-CURA: Salva no banco
                            self.db.atualizar_pdf_path(chave_vinculada, str(pdf_encontrado.absolute()))
                            print(f"[DEBUG ABRIR PDF EVENTO] 🔄 Auto-cura: PDF path salvo no banco")
                            break
                    if not pdf_encontrado:
                        print(f"[DEBUG ABRIR PDF EVENTO] ❌ Não encontrado em {min(len(folders), 20)} pastas verificadas")
                else:
                    print(f"[DEBUG ABRIR PDF EVENTO] ❌ Pasta do informante não existe: {pasta_informante}")
            
            if not pdf_encontrado:
                print(f"[DEBUG ABRIR PDF EVENTO] ❌ PDF não encontrado em nenhuma localização")
            
            # Abre PDF se encontrado
            if pdf_encontrado and pdf_encontrado.exists():
                import subprocess, sys
                pdf_str = str(pdf_encontrado.absolute())
                if sys.platform == "win32":
                    subprocess.Popen(["cmd", "/c", "start", "", pdf_str], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)
                else:
                    subprocess.Popen(["xdg-open", pdf_str])
                self.set_status("✅ PDF aberto", 1000)
            else:
                # PDF não encontrado, oferece gerar
                resposta = QMessageBox.question(
                    dialog_parent, 
                    "PDF não encontrado",
                    f"PDF não encontrado para este documento.\n\n"
                    f"Tipo: {nota_dict.get('tipo', 'N/A')}\n"
                    f"Número: {nota_dict.get('numero', 'N/A')}\n"
                    f"Chave: {chave_vinculada[:10]}...\n\n"
                    f"Deseja gerar o PDF agora?",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if resposta == QMessageBox.Yes:
                    # Procura o XML do documento
                    xml_path = None
                    for pasta_cert in xmls_root.iterdir():
                        if not pasta_cert.is_dir():
                            continue
                        for xml_file in pasta_cert.rglob("*.xml"):
                            if chave_vinculada in xml_file.stem:
                                xml_path = xml_file
                                break
                        if xml_path:
                            break
                    
                    if xml_path and xml_path.exists():
                        self.set_status("🔄 Gerando PDF...", 0)
                        QApplication.processEvents()
                        
                        from nfe_search import gerar_pdf_nfe
                        pdf_gerado = gerar_pdf_nfe(str(xml_path), informante or '')
                        
                        if pdf_gerado:
                            import subprocess, sys
                            if sys.platform == "win32":
                                subprocess.Popen(["cmd", "/c", "start", "", pdf_gerado], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)
                            else:
                                subprocess.Popen(["xdg-open", pdf_gerado])
                            self.set_status("✅ PDF gerado e aberto", 2000)
                        else:
                            QMessageBox.warning(dialog_parent, "Erro", "Erro ao gerar PDF. Verifique os logs.")
                            self.set_status("❌ Erro ao gerar PDF", 2000)
                    else:
                        QMessageBox.warning(dialog_parent, "Erro", 
                            f"XML do documento não encontrado.\n\n"
                            f"Chave: {chave_vinculada}")
                        self.set_status("❌ XML não encontrado", 2000)
                        
        except Exception as e:
            print(f"[DEBUG] Erro ao abrir PDF do evento: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(dialog_parent, "Erro", f"Erro ao abrir PDF:\n{e}")
    
    def _mostrar_eventos(self, item: Dict[str, Any]):
        """Mostra os eventos vinculados a uma NFe/CT-e"""
        
        def formatar_data_evento(data_str: str) -> str:
            """Converte data de aaaa-mm-ddThh:mm:ss para dd/mm/aaaa - hh:mm:ss"""
            if not data_str or data_str == 'N/A':
                return 'N/A'
            try:
                # Remove possível timezone (ex: 2025-11-27T13:03:38-03:00)
                data_limpa = data_str.split('-03:00')[0].split('+')[0]
                
                # Se tem 'T', é formato ISO
                if 'T' in data_limpa:
                    from datetime import datetime
                    dt = datetime.fromisoformat(data_limpa[:19])
                    return dt.strftime('%d/%m/%Y - %H:%M:%S')
                
                # Se já está no formato dd/mm/yyyy, retorna como está
                if '/' in data_limpa:
                    return data_limpa
                
                # Outros casos, retorna original
                return data_str[:19] if len(data_str) >= 19 else data_str
            except Exception:
                return data_str[:19] if len(data_str) >= 19 else data_str
        
        chave = item.get('chave', '')
        if not chave or len(chave) != 44:
            QMessageBox.warning(self, "Eventos", "Chave de acesso inválida!")
            return
        
        informante = item.get('informante', '')
        
        # Detecta o tipo pela chave (posições 20-21 = modelo)
        # Modelo 55 = NFe, Modelo 57 = CTe, Modelo 58 = MDFe
        modelo_chave = chave[20:22] if len(chave) >= 22 else '55'
        if modelo_chave == '57':
            tipo = 'CTE'
        elif modelo_chave == '58':
            tipo = 'MDFE'
        else:
            tipo = 'NFE'  # Padrão = NFe (modelo 55)
        
        numero = item.get('numero', chave[:10])
        
        print(f"\n[DEBUG] ========== BUSCANDO EVENTOS ==========")
        print(f"[DEBUG] Chave: {chave}")
        print(f"[DEBUG] Modelo da chave: {modelo_chave}")
        print(f"[DEBUG] Tipo detectado: {tipo}")
        print(f"[DEBUG] Número: {numero}")
        
        # Mostra indicador de busca
        self.set_status("🔍 Procurando eventos...")
        QApplication.processEvents()  # Atualiza UI
        
        # Busca eventos nos XMLs locais
        eventos_encontrados = []
        eventos_unicos = set()  # Para evitar duplicatas (usa nome do arquivo como chave)
        
        try:
            # 1️⃣ Busca manifestações registradas no banco
            print(f"[DEBUG] 1️⃣ Buscando manifestações no banco...")
            try:
                manifestacoes = self.db.get_manifestacoes_by_chave(chave)
                print(f"[DEBUG] Encontradas {len(manifestacoes)} manifestações no banco")
                for manif in manifestacoes:
                    tipo_evento = manif.get('tipo_evento', '')
                    protocolo = manif.get('protocolo', 'N/A')
                    data_envio = manif.get('enviado_em', 'N/A')
                    status = manif.get('status', 'N/A')
                    
                    # Mapeia tipo de evento para descrição amigável
                    tipos_eventos = {
                        '210200': '📬 Confirmação da Operação',
                        '210210': '❓ Ciência da Operação',
                        '210220': '⛔ Desconhecimento da Operação',
                        '210240': '🚫 Operação não Realizada',
                    }
                    
                    evento_desc = tipos_eventos.get(tipo_evento, f"Manifestação {tipo_evento}")
                    
                    # Cria chave única para evitar duplicatas
                    chave_unica = f"MANIF_{tipo_evento}_{protocolo}_{data_envio}"
                    if chave_unica not in eventos_unicos:
                        eventos_unicos.add(chave_unica)
                        eventos_encontrados.append({
                            'arquivo': f'Manifestação {tipo_evento}',
                            'tipo': evento_desc,
                            'descricao': f"Protocolo: {protocolo}",
                            'data': formatar_data_evento(data_envio),
                            'status': status,
                            'caminho': None  # Manifestação registrada no banco
                        })
            except Exception as e:
                print(f"[DEBUG] Erro ao buscar manifestações: {e}")
            
            # 2️⃣ Procura EVENTOS diretos do documento
            print(f"[DEBUG] 2️⃣ Buscando eventos em pastas Eventos...")
            xmls_root = DATA_DIR / "xmls"
            if xmls_root.exists():
                eventos_folders = list(xmls_root.rglob("Eventos"))
                print(f"[DEBUG] Encontradas {len(eventos_folders)} pastas de Eventos")
                # Busca eventos diretos (cancelamento, manifestação, etc)
                for eventos_folder in eventos_folders:
                    xml_files = list(eventos_folder.glob("*.xml"))
                    if xml_files:
                        print(f"[DEBUG] Verificando {len(xml_files)} arquivos em {eventos_folder}")
                    for xml_file in xml_files:
                        try:
                            xml_content = xml_file.read_text(encoding='utf-8')
                            
                            # Verifica se a chave do documento está neste evento
                            if chave not in xml_content:
                                continue
                            
                            print(f"[DEBUG] ✅ Evento encontrado: {xml_file.name}")
                                
                            # Extrai informações do evento
                            from lxml import etree
                            tree = etree.fromstring(xml_content.encode('utf-8'))
                            
                            # Tenta diferentes estruturas de evento
                            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                            
                            # Tipo de evento
                            tp_evento = tree.findtext('.//nfe:tpEvento', namespaces=ns)
                            if not tp_evento:
                                # Tenta sem namespace
                                tp_evento = tree.findtext('.//tpEvento') or 'N/A'
                            
                            # Descrição do evento
                            desc_evento = tree.findtext('.//nfe:descEvento', namespaces=ns) or tree.findtext('.//nfe:xEvento', namespaces=ns)
                            if not desc_evento:
                                desc_evento = tree.findtext('.//descEvento') or tree.findtext('.//xEvento') or 'N/A'
                            
                            # Data/hora do evento
                            dh_evento = tree.findtext('.//nfe:dhEvento', namespaces=ns) or tree.findtext('.//nfe:dhRegEvento', namespaces=ns) or tree.findtext('.//nfe:dhRecbto', namespaces=ns)
                            if not dh_evento:
                                dh_evento = tree.findtext('.//dhEvento') or tree.findtext('.//dhRegEvento') or tree.findtext('.//dhRecbto') or 'N/A'
                            
                            # Status
                            cstat = tree.findtext('.//nfe:cStat', namespaces=ns)
                            if not cstat:
                                cstat = tree.findtext('.//cStat') or 'N/A'
                            
                            xmotivo = tree.findtext('.//nfe:xMotivo', namespaces=ns)
                            if not xmotivo:
                                xmotivo = tree.findtext('.//xMotivo') or 'N/A'
                            
                            # Mapeia tipo de evento para descrição amigável
                            tipos_eventos = {
                                '110111': '❌ Cancelamento',
                                '110110': '✏️ Carta de Correção',
                                '210200': '📬 Confirmação da Operação',
                                '210210': '❓ Ciência da Operação',
                                '210220': '⛔ Desconhecimento da Operação',
                                '210240': '🚫 Operação não Realizada',
                                '110140': '🔒 EPEC (Contingência)',
                                # Eventos de CTe vinculados a NFe
                                '610130': '🚛 CTe Autorizado',
                                '610131': '🚛 CTe Cancelado',
                                '610500': '📦 MDFe Autorizado',
                                '610510': '📦 MDFe Cancelado',
                                '610514': '📦 MDFe com CTe',
                                '610600': '🚛 CTe Vinculado à NFe',
                                '610601': '🚛 CTe Desvinculado da NFe',
                                '610610': '📦 MDFe Vinculado à NFe',
                                '610611': '📦 MDFe Desvinculado da NFe',
                                '610614': '📦 MDFe Autorizado com CTe',
                                '610615': '📦 MDFe Cancelado com CTe',
                            }
                            
                            evento_desc = tipos_eventos.get(tp_evento, f"Evento {tp_evento}")
                            
                            # Cria chave única para evitar duplicatas (usa DADOS do evento, não nome do arquivo)
                            chave_unica = f"EVENTO_{tp_evento}_{dh_evento}_{desc_evento}"
                            if chave_unica not in eventos_unicos:
                                eventos_unicos.add(chave_unica)
                                eventos_encontrados.append({
                                    'arquivo': xml_file.name,
                                    'tipo': evento_desc,
                                    'descricao': desc_evento,
                                    'data': formatar_data_evento(dh_evento),
                                    'status': f"{cstat} - {xmotivo}",
                                    'caminho': str(xml_file),
                                    'relacao': 'Evento Direto'
                                })
                        except Exception:
                            continue
        except Exception as e:
            print(f"[DEBUG] Erro ao buscar eventos em pastas: {e}")
        
        # 3️⃣ Busca DOCUMENTOS que referenciam este documento
        # Para NFe: busca CTes e MDFes que mencionem essa chave
        # Para CTe: busca NFes e MDFes que mencionem essa chave
        try:
            xmls_root = DATA_DIR / "xmls"
            if xmls_root.exists():
                pastas_busca = []
                if tipo == "NFE":
                    pastas_busca = ["CTe", "MDFe"]
                elif tipo == "CTE":
                    pastas_busca = ["NFe", "MDFe"]
                elif tipo == "MDFE":
                    pastas_busca = ["NFe", "CTe"]
                else:
                    pastas_busca = []
                
                for pasta_tipo in pastas_busca:
                    for pasta_doc in xmls_root.rglob(pasta_tipo):
                        if not pasta_doc.is_dir():
                            continue
                        
                        for xml_file in pasta_doc.glob("*.xml"):
                            try:
                                xml_content = xml_file.read_text(encoding='utf-8')
                                
                                # Verifica se a chave do documento original está referenciada neste XML
                                if chave not in xml_content:
                                    continue
                                
                                # Parse do XML para extrair informações
                                from lxml import etree
                                tree = etree.fromstring(xml_content.encode('utf-8'))
                                
                                # Extrai chave e número do documento vinculado
                                ns_cte = {'cte': 'http://www.portalfiscal.inf.br/cte'}
                                ns_nfe = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                                ns_mdfe = {'mdfe': 'http://www.portalfiscal.inf.br/mdfe'}
                                
                                chave_vinculada = None
                                numero_vinculado = None
                                emitente_vinculado = None
                                data_vinculada = None
                                
                                if pasta_tipo == "CTe":
                                    chave_vinculada = tree.findtext('.//cte:chCTe', namespaces=ns_cte) or tree.findtext('.//chCTe')
                                    numero_vinculado = tree.findtext('.//cte:nCT', namespaces=ns_cte) or tree.findtext('.//nCT')
                                    emitente_vinculado = tree.findtext('.//cte:xNome', namespaces=ns_cte) or tree.findtext('.//xNome')
                                    data_vinculada = tree.findtext('.//cte:dhEmi', namespaces=ns_cte) or tree.findtext('.//dhEmi')
                                elif pasta_tipo == "NFe":
                                    chave_vinculada = tree.findtext('.//nfe:chNFe', namespaces=ns_nfe) or tree.findtext('.//chNFe')
                                    numero_vinculado = tree.findtext('.//nfe:nNF', namespaces=ns_nfe) or tree.findtext('.//nNF')
                                    emitente_vinculado = tree.findtext('.//nfe:xNome', namespaces=ns_nfe) or tree.findtext('.//xNome')
                                    data_vinculada = tree.findtext('.//nfe:dhEmi', namespaces=ns_nfe) or tree.findtext('.//dhEmi')
                                elif pasta_tipo == "MDFe":
                                    chave_vinculada = tree.findtext('.//mdfe:chMDFe', namespaces=ns_mdfe) or tree.findtext('.//chMDFe')
                                    numero_vinculado = tree.findtext('.//mdfe:nMDF', namespaces=ns_mdfe) or tree.findtext('.//nMDF')
                                    emitente_vinculado = tree.findtext('.//mdfe:xNome', namespaces=ns_mdfe) or tree.findtext('.//xNome')
                                    data_vinculada = tree.findtext('.//mdfe:dhEmi', namespaces=ns_mdfe) or tree.findtext('.//dhEmi')
                                
                                if not numero_vinculado:
                                    numero_vinculado = xml_file.stem[:10]  # Pega do nome do arquivo
                                
                                # Cria chave única para evitar duplicatas (usa chave do documento vinculado)
                                chave_unica = f"VINCULO_{chave_vinculada or xml_file.name}_{numero_vinculado}_{data_vinculada}"
                                if chave_unica not in eventos_unicos:
                                    eventos_unicos.add(chave_unica)
                                    eventos_encontrados.append({
                                        'arquivo': xml_file.name,
                                        'tipo': f'🔗 {pasta_tipo} Vinculado',
                                        'descricao': f"{pasta_tipo} Nº {numero_vinculado} - {emitente_vinculado or 'N/A'}",
                                        'data': formatar_data_evento(data_vinculada) if data_vinculada else 'N/A',
                                        'status': f"Chave: {chave_vinculada[:10]}..." if chave_vinculada else xml_file.name[:30],
                                        'caminho': str(xml_file),
                                        'relacao': f'{pasta_tipo} Vinculado'
                                    })
                            except Exception as ex:
                                print(f"[DEBUG] Erro ao processar {xml_file.name}: {ex}")
                                continue
        except Exception as e:
            print(f"[DEBUG] Erro ao buscar documentos vinculados: {e}")
        
        # Limpa status após busca
        if eventos_encontrados:
            self.set_status(f"✅ {len(eventos_encontrados)} evento(s) encontrado(s)", 2000)
        else:
            self.set_status("ℹ️ Nenhum evento encontrado", 2000)
        
        # Cria dialog para mostrar eventos
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Eventos e Vínculos - {tipo} {numero}")
        dialog.resize(1000, 600)
        
        layout = QVBoxLayout(dialog)
        
        # ===== CABEÇALHO COM INFORMAÇÕES DO DOCUMENTO =====
        header_frame = QFrame()
        header_frame.setStyleSheet("QFrame { background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 4px; padding: 10px; }")
        header_layout = QVBoxLayout(header_frame)
        
        titulo = QLabel(f"<h3>📄 {tipo} Nº {numero}</h3>")
        titulo.setStyleSheet("font-weight: bold;")
        header_layout.addWidget(titulo)
        
        chave_label = QLabel(f"<b>Chave de Acesso:</b> {chave}")
        chave_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        header_layout.addWidget(chave_label)
        
        layout.addWidget(header_frame)
        
        if not eventos_encontrados:
            no_eventos_label = QLabel("ℹ️ Nenhum evento ou documento vinculado encontrado.")
            no_eventos_label.setStyleSheet("padding: 20px; color: #666; font-size: 12pt;")
            no_eventos_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(no_eventos_label)
        else:
            # ===== SEPARAR EVENTOS POR TIPO =====
            eventos_proprios = []
            documentos_vinculados = []
            
            for evento in eventos_encontrados:
                relacao = evento.get('relacao', '')
                if 'Vinculado' in relacao:
                    documentos_vinculados.append(evento)
                else:
                    eventos_proprios.append(evento)
            
            # ===== ORDENA EVENTOS POR DATA (DO MAIS ANTIGO PARA O MAIS NOVO) =====
            def parse_data_br(data_str: str):
                """Converte dd/mm/yyyy - hh:mm:ss para timestamp para ordenação"""
                try:
                    if not data_str or data_str == 'N/A':
                        return 0
                    # Remove possível texto extra e pega só a data/hora
                    data_limpa = data_str.split(' - ')
                    if len(data_limpa) == 2:
                        from datetime import datetime
                        dt = datetime.strptime(f"{data_limpa[0]} {data_limpa[1]}", "%d/%m/%Y %H:%M:%S")
                        return dt.timestamp()
                    return 0
                except Exception:
                    return 0
            
            eventos_proprios.sort(key=lambda x: parse_data_br(x.get('data', '')))
            documentos_vinculados.sort(key=lambda x: parse_data_br(x.get('data', '')))
            
            # Cria abas
            tabs = QTabWidget()
            
            # ===== ABA 1: EVENTOS PRÓPRIOS =====
            if eventos_proprios:
                tab_eventos = QWidget()
                tab_eventos_layout = QVBoxLayout(tab_eventos)
                
                info_label = QLabel(f"<b>Eventos registrados para este documento ({len(eventos_proprios)})</b>")
                info_label.setStyleSheet("padding: 5px; color: #0066cc;")
                tab_eventos_layout.addWidget(info_label)
                
                eventos_table = QTableWidget()
                eventos_table.setColumnCount(5)
                eventos_table.setHorizontalHeaderLabels(["Data/Hora", "Tipo de Evento", "Descrição", "Status", "Protocolo"])
                eventos_table.setRowCount(len(eventos_proprios))
                eventos_table.setEditTriggers(QTableWidget.NoEditTriggers)
                eventos_table.setSelectionBehavior(QTableWidget.SelectRows)
                eventos_table.setAlternatingRowColors(True)
                
                for i, evento in enumerate(eventos_proprios):
                    # Data/Hora
                    data_item = QTableWidgetItem(evento['data'])
                    data_item.setTextAlignment(Qt.AlignCenter)
                    eventos_table.setItem(i, 0, data_item)
                    
                    # Tipo
                    tipo_item = QTableWidgetItem(evento['tipo'])
                    eventos_table.setItem(i, 1, tipo_item)
                    
                    # Descrição
                    eventos_table.setItem(i, 2, QTableWidgetItem(evento['descricao']))
                    
                    # Status
                    status_parts = evento['status'].split(' - ', 1)
                    status_item = QTableWidgetItem(status_parts[1] if len(status_parts) > 1 else evento['status'])
                    eventos_table.setItem(i, 3, status_item)
                    
                    # Protocolo
                    protocolo = status_parts[0] if len(status_parts) > 1 else 'N/A'
                    eventos_table.setItem(i, 4, QTableWidgetItem(protocolo))
                
                eventos_table.resizeColumnsToContents()
                eventos_table.horizontalHeader().setStretchLastSection(False)
                eventos_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
                
                # Adiciona handler de duplo-clique para abrir PDF
                eventos_table.cellDoubleClicked.connect(
                    lambda row, col: self._abrir_pdf_evento(eventos_proprios[row], dialog)
                )
                
                tab_eventos_layout.addWidget(eventos_table)
                tabs.addTab(tab_eventos, f"📋 Eventos ({len(eventos_proprios)})")
            
            # ===== ABA 2: DOCUMENTOS VINCULADOS =====
            if documentos_vinculados:
                tab_vinculos = QWidget()
                tab_vinculos_layout = QVBoxLayout(tab_vinculos)
                
                if tipo == "NFE":
                    info_texto = f"<b>Documentos de transporte vinculados a esta NF-e ({len(documentos_vinculados)})</b>"
                elif tipo == "CTE":
                    info_texto = f"<b>Documentos fiscais relacionados a este CT-e ({len(documentos_vinculados)})</b>"
                else:
                    info_texto = f"<b>Documentos vinculados ({len(documentos_vinculados)})</b>"
                
                info_label = QLabel(info_texto)
                info_label.setStyleSheet("padding: 5px; color: #0066cc;")
                tab_vinculos_layout.addWidget(info_label)
                
                vinculos_table = QTableWidget()
                vinculos_table.setColumnCount(5)
                vinculos_table.setHorizontalHeaderLabels(["Tipo", "Número", "Emitente", "Data Emissão", "Chave de Acesso"])
                vinculos_table.setRowCount(len(documentos_vinculados))
                vinculos_table.setEditTriggers(QTableWidget.NoEditTriggers)
                vinculos_table.setSelectionBehavior(QTableWidget.SelectRows)
                vinculos_table.setAlternatingRowColors(True)
                
                for i, doc in enumerate(documentos_vinculados):
                    # Tipo
                    tipo_doc = doc['tipo'].replace('🔗 ', '')
                    vinculos_table.setItem(i, 0, QTableWidgetItem(tipo_doc))
                    
                    # Extrair número e emitente da descrição
                    desc_parts = doc['descricao'].split(' - ', 1)
                    numero_doc = desc_parts[0].replace('CTe Nº ', '').replace('NFe Nº ', '').replace('MDFe Nº ', '')
                    emitente = desc_parts[1] if len(desc_parts) > 1 else 'N/A'
                    
                    vinculos_table.setItem(i, 1, QTableWidgetItem(numero_doc))
                    vinculos_table.setItem(i, 2, QTableWidgetItem(emitente))
                    
                    # Data
                    data_item = QTableWidgetItem(doc['data'])
                    data_item.setTextAlignment(Qt.AlignCenter)
                    vinculos_table.setItem(i, 3, data_item)
                    
                    # Chave (resumida)
                    chave_doc = doc['status'].replace('Chave: ', '') if doc['status'].startswith('Chave:') else doc['arquivo'][:44]
                    vinculos_table.setItem(i, 4, QTableWidgetItem(chave_doc))
                
                vinculos_table.resizeColumnsToContents()
                vinculos_table.horizontalHeader().setStretchLastSection(False)
                vinculos_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
                
                # Adiciona handler de duplo-clique para abrir PDF do documento vinculado
                vinculos_table.cellDoubleClicked.connect(
                    lambda row, col: self._abrir_pdf_evento(documentos_vinculados[row], dialog)
                )
                
                tab_vinculos_layout.addWidget(vinculos_table)
                
                # Legenda
                legenda = QLabel("💡 <i>Estes documentos fazem referência à chave de acesso deste documento.</i>")
                legenda.setStyleSheet("padding: 5px; color: #666; font-size: 9pt;")
                tab_vinculos_layout.addWidget(legenda)
                
                tabs.addTab(tab_vinculos, f"🔗 Vínculos ({len(documentos_vinculados)})")
            
            layout.addWidget(tabs)
            
            # Botão para abrir pasta de eventos
            btn_abrir_pasta = QPushButton("📁 Abrir pasta de eventos")
            btn_abrir_pasta.clicked.connect(lambda: self._abrir_pasta_eventos(informante))
            layout.addWidget(btn_abrir_pasta)
        
        # Botão fechar
        btn_fechar = QPushButton("Fechar")
        btn_fechar.clicked.connect(dialog.close)
        layout.addWidget(btn_fechar)
        
        dialog.exec_()
    
    def _abrir_pasta_eventos(self, informante: str):
        """Abre a pasta de eventos do informante no Windows Explorer"""
        try:
            eventos_path = DATA_DIR / "xmls" / informante
            if eventos_path.exists():
                if sys.platform == "win32":
                    os.startfile(str(eventos_path))  # type: ignore[attr-defined]
                else:
                    subprocess.Popen(["xdg-open", str(eventos_path)])
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir pasta: {e}")
    
    def _manifestar_nota(self, item: Dict[str, Any] = None):
        """
        Exibe dialog moderna para manifestar NF-e ou CT-e.
        NF-e: eventos 210200, 210210, 210220, 210240
        CT-e: eventos 610110, 610112
        
        CASO 1: item=None → Pede certificado, senha e chave (manifestação manual)
        CASO 2: item fornecido → Usa chave e certificado da seleção
        """
        # Debug para verificar se item está sendo passado
        print(f"[DEBUG MANIFESTAÇÃO] item recebido: {item is not None}")
        if item:
            print(f"[DEBUG MANIFESTAÇÃO] Chave: {item.get('chave', 'N/A')}")
            print(f"[DEBUG MANIFESTAÇÃO] Informante: {item.get('informante', 'N/A')}")
        
        # CASO 1: Manifestação manual (sem documento selecionado)
        if item is None:
            dialog_input = QDialog(self)
            dialog_input.setWindowTitle("✉️ Manifestar Documento (Manual)")
            dialog_input.setMinimumWidth(500)
            dialog_input.setMaximumWidth(550)
            
            layout = QVBoxLayout(dialog_input)
            layout.setSpacing(10)
            layout.setContentsMargins(15, 15, 15, 15)
            
            # Título
            title = QLabel("<h3 style='color: #2c3e50;'>📝 Manifestação Manual</h3>")
            layout.addWidget(title)
            
            info = QLabel("<span style='color: #666; font-size: 9pt;'>Preencha os dados abaixo para manifestar um documento.</span>")
            info.setWordWrap(True)
            layout.addWidget(info)
            
            # Seleção de certificado
            cert_label = QLabel("<b>Certificado Digital:</b>")
            cert_label.setStyleSheet("font-size: 9pt;")
            layout.addWidget(cert_label)
            
            cert_combo = QComboBox()
            cert_combo.setMinimumHeight(28)
            certs = self.db.load_certificates()
            for cert in certs:
                nome_cert = cert.get('nome_certificado', 'Sem Nome')
                cnpj = cert.get('cnpj_cpf', '')
                cert_combo.addItem(f"{nome_cert} - {cnpj}", cert)
            
            if cert_combo.count() == 0:
                QMessageBox.warning(self, "Erro", "Nenhum certificado cadastrado!\n\nCadastre um certificado primeiro.")
                return
            
            layout.addWidget(cert_combo)
            
            # Senha do certificado
            senha_label = QLabel("<b>Senha do Certificado:</b>")
            senha_label.setStyleSheet("font-size: 9pt;")
            layout.addWidget(senha_label)
            
            senha_input = QLineEdit()
            senha_input.setMinimumHeight(28)
            senha_input.setEchoMode(QLineEdit.Password)
            senha_input.setPlaceholderText("Digite a senha do certificado")
            layout.addWidget(senha_input)
            
            # Chave de acesso
            chave_label = QLabel("<b>Chave de Acesso (44 dígitos):</b>")
            chave_label.setStyleSheet("font-size: 9pt;")
            layout.addWidget(chave_label)
            
            chave_input = QLineEdit()
            chave_input.setMinimumHeight(28)
            chave_input.setPlaceholderText("Digite a chave de acesso do documento")
            chave_input.setMaxLength(44)
            layout.addWidget(chave_input)
            
            # Botões
            buttons_layout = QHBoxLayout()
            buttons_layout.setSpacing(10)
            
            btn_cancelar = QPushButton("❌ Cancelar")
            btn_cancelar.setMinimumHeight(32)
            btn_cancelar.setStyleSheet("""
                QPushButton {
                    background-color: #95a5a6;
                    color: white;
                    border: none;
                    padding: 6px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 9pt;
                }
                QPushButton:hover {
                    background-color: #7f8c8d;
                }
            """)
            btn_cancelar.clicked.connect(dialog_input.reject)
            buttons_layout.addWidget(btn_cancelar)
            
            btn_continuar = QPushButton("➡️ Continuar")
            btn_continuar.setMinimumHeight(32)
            btn_continuar.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    padding: 6px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 9pt;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """)
            
            def continuar_manifestacao():
                chave = chave_input.text().strip()
                senha = senha_input.text().strip()
                
                if len(chave) != 44:
                    QMessageBox.warning(dialog_input, "Erro", "A chave deve ter exatamente 44 dígitos!")
                    return
                
                if not senha:
                    QMessageBox.warning(dialog_input, "Erro", "Digite a senha do certificado!")
                    return
                
                cert_data = cert_combo.currentData()
                if not cert_data:
                    QMessageBox.warning(dialog_input, "Erro", "Selecione um certificado!")
                    return
                
                dialog_input.accept()
                
                # Cria item virtual para prosseguir
                item_virtual = {
                    'chave': chave,
                    'informante': cert_data.get('cnpj_cpf', ''),
                    'tipo': 'NFE',  # Detecta automaticamente depois
                    'numero': chave[-9:],
                    'nome_emitente': 'Desconhecido',
                    '_manual': True,
                    '_cert_path': cert_data.get('caminho'),
                    '_cert_senha': senha,
                    '_cert_cnpj': cert_data.get('cnpj_cpf')
                }
                
                # Continua com o fluxo normal
                self._manifestar_nota(item_virtual)
            
            btn_continuar.clicked.connect(continuar_manifestacao)
            buttons_layout.addWidget(btn_continuar)
            
            layout.addLayout(buttons_layout)
            dialog_input.exec_()
            return
        
        # CASO 2: Manifestação com documento selecionado
        chave = item.get('chave', '')
        if not chave or len(chave) != 44:
            QMessageBox.warning(self, "Manifestação", "Chave de acesso inválida!")
            return
        
        informante = item.get('informante', '')
        if not informante:
            QMessageBox.warning(self, "Manifestação", "Informante não identificado!")
            return
        
        tipo_doc = (item.get('tipo') or '').upper()
        is_cte = tipo_doc in ['CTE', 'CT-E']
        
        numero = item.get('numero', chave[-9:])
        emitente = item.get('nome_emitente', 'N/A')
        
        # Cria dialog moderna
        dialog = QDialog(self)
        dialog.setWindowTitle(f"✉️ Manifestar {'CT-e' if is_cte else 'NF-e'}")
        dialog.setMinimumWidth(450)
        dialog.setMaximumWidth(500)
        dialog.setMinimumHeight(400)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Header com informações da nota
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background-color: #f0f4f8;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        header_layout = QVBoxLayout(header_frame)
        
        doc_type = "CT-e" if is_cte else "NF-e"
        title_label = QLabel(f"<h3 style='color: #2c3e50; margin: 0;'>📄 {doc_type} {numero}</h3>")
        header_layout.addWidget(title_label)
        
        info_label = QLabel(f"<b>Emitente:</b> {emitente}<br><b>Chave:</b> {chave[:10]}...{chave[-10:]}")
        info_label.setStyleSheet("color: #555; font-size: 9pt;")
        header_layout.addWidget(info_label)
        
        layout.addWidget(header_frame)
        
        # Título da seção
        section_label = QLabel(f"<b style='color: #2c3e50; font-size: 10pt;'>Selecione o tipo de manifestação:</b>")
        layout.addWidget(section_label)
        
        # Tipos de manifestação com botões estilizados
        if is_cte:
            eventos = [
                {
                    'codigo': '610110',
                    'nome': 'Desacordo do Serviço',
                    'icone': '🛑',
                    'descricao': 'Declara desacordo com o serviço de transporte prestado',
                    'cor': '#e74c3c'
                },
                {
                    'codigo': '610112',
                    'nome': 'Cancelar Desacordo',
                    'icone': '↩️',
                    'descricao': 'Cancela declaração de desacordo anterior',
                    'cor': '#f39c12'
                }
            ]
        else:  # NF-e
            eventos = [
                {
                    'codigo': '210210',
                    'nome': 'Ciência da Operação',
                    'icone': '👁️',
                    'descricao': 'Registra que você tomou conhecimento da NF-e',
                    'cor': '#3498db'
                },
                {
                    'codigo': '210200',
                    'nome': 'Confirmação da Operação',
                    'icone': '✅',
                    'descricao': 'Confirma o recebimento da mercadoria/serviço',
                    'cor': '#27ae60'
                },
                {
                    'codigo': '210220',
                    'nome': 'Desconhecimento da Operação',
                    'icone': '🛑',
                    'descricao': 'Informa que você não reconhece esta operação',
                    'cor': '#e74c3c'
                },
                {
                    'codigo': '210240',
                    'nome': 'Operação não Realizada',
                    'icone': '⭕',
                    'descricao': 'Informa que a operação não foi realizada',
                    'cor': '#f39c12'
                }
            ]
        
        selected_evento = {'codigo': None}
        
        # Lista simples com radio buttons
        from PyQt5.QtWidgets import QRadioButton, QButtonGroup
        
        button_group = QButtonGroup(dialog)
        radio_buttons = []
        
        for idx, evento_data in enumerate(eventos):
            # Layout horizontal simples
            item_layout = QHBoxLayout()
            item_layout.setSpacing(12)
            
            # Radio button
            radio = QRadioButton()
            radio.codigo = evento_data['codigo']
            
            # Ícone
            icone = QLabel(evento_data['icone'])
            icone.setStyleSheet(f"font-size: 24pt; color: {evento_data['cor']};")
            icone.setFixedWidth(40)
            
            # Texto (nome + descrição)
            texto = QLabel(f"<b style='color: {evento_data['cor']};'>{evento_data['nome']}</b><br>"
                          f"<span style='color: #666; font-size: 9pt;'>{evento_data['descricao']}</span>")
            texto.setWordWrap(True)
            
            item_layout.addWidget(radio)
            item_layout.addWidget(icone)
            item_layout.addWidget(texto, 1)
            
            layout.addLayout(item_layout)
            
            button_group.addButton(radio, idx)
            radio_buttons.append(radio)
        
        # Seleciona o primeiro por padrão
        if radio_buttons:
            radio_buttons[0].setChecked(True)
        
        # Função para pegar o código selecionado
        def get_selected_codigo():
            for radio in radio_buttons:
                if radio.isChecked():
                    return radio.codigo
            return None
        
        layout.addStretch()
        
        # Botões de ação
        buttons_layout = QHBoxLayout()
        
        btn_cancelar = QPushButton("❌ Cancelar")
        btn_cancelar.setMinimumHeight(32)
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 9pt;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_cancelar.clicked.connect(dialog.reject)
        buttons_layout.addWidget(btn_cancelar)
        
        btn_enviar = QPushButton("📤 Enviar")
        btn_enviar.setMinimumHeight(32)
        btn_enviar.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 9pt;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        btn_enviar.clicked.connect(lambda: self._enviar_manifestacao(dialog, chave, informante, get_selected_codigo(), item))
        buttons_layout.addWidget(btn_enviar)
        
        layout.addLayout(buttons_layout)
        
        dialog.exec_()
    
    def _enviar_manifestacao(self, dialog, chave, informante, tipo_evento, item=None):
        """Envia manifestação para SEFAZ"""
        if not tipo_evento:
            QMessageBox.warning(dialog, "Atenção", "Selecione um tipo de manifestação!")
            return
        
        # Eventos que exigem justificativa
        eventos_com_justificativa = ['210220', '210240', '110111']
        justificativa = None
        
        # Solicita justificativa se necessário
        if tipo_evento in eventos_com_justificativa:
            dialog_just = QDialog(dialog)
            dialog_just.setWindowTitle("✍️ Justificativa Obrigatória")
            dialog_just.setMinimumWidth(500)
            
            layout_just = QVBoxLayout(dialog_just)
            layout_just.setSpacing(15)
            layout_just.setContentsMargins(20, 20, 20, 20)
            
            # Título
            tipo_names = {
                '210220': 'Desconhecimento da Operação',
                '210240': 'Operação não Realizada',
                '110111': 'Cancelamento'
            }
            tipo_nome = tipo_names.get(tipo_evento, 'Manifestação')
            
            title = QLabel(f"<h3 style='color: #e74c3c;'>⚠️ {tipo_nome}</h3>")
            layout_just.addWidget(title)
            
            info = QLabel("<span style='color: #555;'>Digite uma justificativa para esta manifestação (mínimo 15 caracteres):</span>")
            info.setWordWrap(True)
            layout_just.addWidget(info)
            
            # Campo de texto
            text_edit = QTextEdit()
            text_edit.setPlaceholderText("Exemplo: Mercadoria não foi recebida no endereço indicado...")
            text_edit.setMinimumHeight(120)
            layout_just.addWidget(text_edit)
            
            # Contador de caracteres
            char_label = QLabel("<span style='color: #999;'>0 caracteres</span>")
            def atualizar_contador():
                texto = text_edit.toPlainText()
                qtd = len(texto)
                cor = '#27ae60' if qtd >= 15 else '#e74c3c'
                char_label.setText(f"<span style='color: {cor};'>{qtd} caracteres</span>")
            
            text_edit.textChanged.connect(atualizar_contador)
            layout_just.addWidget(char_label)
            
            # Botões
            buttons_just = QHBoxLayout()
            
            btn_cancelar_just = QPushButton("❌ Cancelar")
            btn_cancelar_just.clicked.connect(dialog_just.reject)
            buttons_just.addWidget(btn_cancelar_just)
            
            btn_confirmar_just = QPushButton("✅ Confirmar")
            btn_confirmar_just.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 6px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            
            def confirmar_justificativa():
                texto = text_edit.toPlainText().strip()
                if len(texto) < 15:
                    QMessageBox.warning(dialog_just, "Atenção", "A justificativa deve ter no mínimo 15 caracteres!")
                    return
                dialog_just.accept()
            
            btn_confirmar_just.clicked.connect(confirmar_justificativa)
            buttons_just.addWidget(btn_confirmar_just)
            
            layout_just.addLayout(buttons_just)
            
            # Mostra dialog de justificativa
            if dialog_just.exec_() != QDialog.Accepted:
                return  # Usuário cancelou
            
            justificativa = text_edit.toPlainText().strip()
        
        # Verifica se já foi manifestada antes
        try:
            ja_manifestada = self.db.check_manifestacao_exists(
                chave=chave,
                tipo_evento=tipo_evento,
                informante=informante
            )
            
            if ja_manifestada:
                reply = QMessageBox.question(
                    dialog,
                    "Confirmação",
                    f"Esta manifestação já foi enviada anteriormente.\n\n"
                    f"Deseja enviar novamente?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return
        except Exception as e:
            print(f"[WARN] Erro ao verificar manifestação: {e}")
        
        # Fecha o dialog e mostra progresso
        dialog.accept()
        
        progress = QProgressDialog("Enviando manifestação para SEFAZ...", "Cancelar", 0, 0, self)
        progress.setWindowTitle("Manifestação")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        QApplication.processEvents()
        
        try:
            # Carrega certificado (manual ou da seleção)
            if item and item.get('_manual'):
                # Manifestação manual - usa dados fornecidos
                cert_path = item.get('_cert_path')
                cert_senha = item.get('_cert_senha')
                cert_cnpj = item.get('_cert_cnpj')
            else:
                # Manifestação normal - busca certificado do informante
                certs = self.db.load_certificates()
                cert_info = None
                for cert in certs:
                    if cert.get('informante') == informante:
                        cert_info = cert
                        break
                
                if not cert_info:
                    progress.close()
                    QMessageBox.critical(self, "Erro", f"Certificado do informante {informante} não encontrado!")
                    return
                
                cert_path = cert_info.get('caminho')
                cert_senha = cert_info.get('senha')
                cert_cnpj = cert_info.get('cnpj_cpf')
            
            # Prepara dados do evento
            from modules.manifestacao_service import ManifestacaoService
            from nfe_search import salvar_xml_por_certificado
            import sys
            sys.path.insert(0, str(BASE_DIR))
            
            # Envia manifestação REAL via SEFAZ
            print(f"[MANIFESTAÇÃO] Enviando {tipo_evento} para chave {chave}...")
            
            # Justificativa só para eventos que exigem (já foi solicitada antes se necessário)
            # Eventos 210210 e 210200 NÃO devem ter justificativa
            # Justificativa já foi capturada antes para 210220, 210240, 110111
            
            # Cria serviço de manifestação
            manifesta_service = ManifestacaoService(cert_path, cert_senha)
            
            # Envia para SEFAZ (justificativa=None para eventos que não precisam)
            sucesso, protocolo, mensagem, xml_resposta = manifesta_service.enviar_manifestacao(
                chave=chave,
                tipo_evento=tipo_evento,
                cnpj_destinatario=cert_cnpj,
                justificativa=justificativa  # Já é None para 210210/210200
            )
            
            if not sucesso:
                progress.close()
                QMessageBox.critical(
                    self,
                    "❌ Erro SEFAZ",
                    f"A SEFAZ rejeitou a manifestação:\n\n{mensagem}"
                )
                return
            
            # ✅ NÃO salva XML de retorno da manifestação (retEnvEvento)
            # O retEnvEvento é apenas a confirmação do protocolo, não contém a nota fiscal
            # O XML completo da nota será baixado separadamente se necessário
            # Removido: salvar_xml_por_certificado(xml_resposta, ...) - causava SEM_NUMERO-SEM_NOME.xml
            
            # Registra no banco
            self.db.register_manifestacao(
                chave=chave,
                tipo_evento=tipo_evento,
                informante=informante,
                status="REGISTRADA",
                protocolo=protocolo
            )
            
            # 🔄 AUTO-DOWNLOAD: Baixa XML completo automaticamente após manifestação
            should_download_xml = False
            
            # Verifica se é evento 210200 ou 210210 (Ciência/Confirmação)
            # Apenas para NF-e (modelo 55)
            modelo = chave[20:22] if len(chave) >= 22 else ''
            is_nfe = modelo == '55'
            
            if is_nfe and tipo_evento in ['210200', '210210']:
                # Verifica se é nota RESUMO (precisa baixar completa)
                try:
                    with self.db._connect() as conn:
                        nota_db = conn.execute(
                            "SELECT xml_status FROM notas_detalhadas WHERE chave = ?",
                            (chave,)
                        ).fetchone()
                        
                        if nota_db:
                            xml_status = (nota_db[0] or 'RESUMO').upper()
                            if xml_status == 'RESUMO':
                                should_download_xml = True
                                print(f"[MANIFESTAÇÃO] ✅ Nota é RESUMO - XML completo será baixado automaticamente")
                            else:
                                print(f"[MANIFESTAÇÃO] ℹ️ Nota já é {xml_status} - download não necessário")
                except Exception as e:
                    print(f"[MANIFESTAÇÃO] ⚠️ Erro ao verificar xml_status: {e}")
            
            progress.close()
            
            # Mensagem de sucesso com informação sobre download
            msg_texto = f"Manifestação enviada com sucesso!\n\nProtocolo: {protocolo}\n\n"
            if should_download_xml:
                msg_texto += "⏳ Aguardando 3s para SEFAZ processar...\nXML completo será baixado automaticamente."
            else:
                msg_texto += "A tabela será atualizada automaticamente."
            
            QMessageBox.information(
                self,
                "✅ Sucesso",
                msg_texto
            )
            
            # 📥 DOWNLOAD AUTOMÁTICO do XML completo (se necessário)
            if should_download_xml:
                print(f"[MANIFESTAÇÃO] ⏱️ Aguardando 3s para SEFAZ processar...")
                import time
                time.sleep(3)
                
                print(f"[MANIFESTAÇÃO] 🔄 Iniciando download automático do XML completo...")
                try:
                    # Chama o mesmo método usado pelo botão "Baixar XML Completo"
                    # mas sem manifestar novamente (já foi feito acima)
                    self._baixar_xml_completo_apos_manifestacao(chave, informante, cert_path, cert_senha, cert_cnpj)
                except Exception as e:
                    print(f"[MANIFESTAÇÃO] ❌ Erro no download automático: {e}")
                    QMessageBox.warning(
                        self,
                        "Aviso",
                        f"Manifestação registrada, mas erro ao baixar XML:\n\n{str(e)}\n\n"
                        f"Use o menu 'Baixar XML Completo' para tentar novamente."
                    )
            
            # Atualiza a tabela
            self.refresh_table()
            
        except Exception as e:
            progress.close()
            QMessageBox.critical(
                self,
                "❌ Erro",
                f"Erro ao enviar manifestação:\n\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def _build_pdf_cache_async(self):
        """Constrói cache de PDFs em background para abertura rápida"""
        if self._cache_building:
            return
        
        self._cache_building = True
        
        class CacheBuilder(QThread):
            cache_ready = pyqtSignal(dict)  # {chave: pdf_path}
            
            def __init__(self, notes, base_dir):
                super().__init__()
                self.notes = notes
                self.base_dir = base_dir
            
            def run(self):
                cache = {}
                try:
                    for note in self.notes:
                        chave = note.get('chave', '')
                        informante = note.get('informante', '')
                        data_emissao = (note.get('data_emissao') or '')[:10]
                        tipo = (note.get('tipo') or 'NFe').strip().upper().replace('-', '')
                        
                        if not (chave and informante and data_emissao):
                            continue
                        
                        # Extrai ano-mês
                        try:
                            year_month = data_emissao[:7] if len(data_emissao) >= 7 else None
                            if year_month:
                                # Verifica caminho direto (com tipo)
                                pdf_path = DATA_DIR / "xmls" / informante / tipo / year_month / f"{chave}.pdf"
                                if pdf_path.exists():
                                    cache[chave] = str(pdf_path)
                                    continue
                                
                                # Verifica caminho antigo (sem tipo)
                                pdf_path = DATA_DIR / "xmls" / informante / year_month / f"{chave}.pdf"
                                if pdf_path.exists():
                                    cache[chave] = str(pdf_path)
                                    continue
                        except Exception:
                            pass
                except Exception as e:
                    print(f"[DEBUG] Erro ao construir cache de PDFs: {e}")
                
                self.cache_ready.emit(cache)
        
        def on_cache_ready(cache: dict):
            self._pdf_cache = cache
            self._cache_building = False
            self._cache_worker = None
            print(f"[DEBUG] Cache de PDFs construído: {len(cache)} arquivos indexados")
        
        self._cache_worker = CacheBuilder(self.notes, BASE_DIR)
        self._cache_worker.cache_ready.connect(on_cache_ready)
        self._cache_worker.start()

    def _on_table_double_clicked(self, row: int, col: int):
        """Abre PDF (verifica existência primeiro, só gera se necessário) - OTIMIZADO"""
        import time
        start_time = time.time()
        print(f"\n[DEBUG PDF] ========== DUPLO CLIQUE ===========")
        print(f"[DEBUG PDF] Linha: {row}, Coluna: {col}")
        print(f"[DEBUG PDF] Aba ativa: {self.tabs.currentIndex()} (0=Recebidas, 1=Emitidas)")
        
        # Encontra o índice da coluna "Chave" (independente de reordenação visual)
        chave_col_index = None
        for c in range(self.table.columnCount()):
            header_text = self.table.horizontalHeaderItem(c).text()
            if header_text == "Chave":
                chave_col_index = c
                break
        
        if chave_col_index is None:
            print(f"[DEBUG PDF] ❌ Coluna 'Chave' não encontrada!")
            return
        
        # CORREÇÃO: Pega a chave diretamente da célula da tabela (coluna "Chave")
        chave_item = self.table.item(row, chave_col_index)
        if not chave_item:
            print(f"[DEBUG PDF] ❌ Célula de chave vazia na linha {row}, coluna {chave_col_index}")
            return
        
        chave = chave_item.text().strip()
        if not chave:
            print(f"[DEBUG PDF] ❌ Chave vazia")
            return
        
        print(f"[DEBUG PDF] Chave da célula: {chave}")
        
        # Busca o item correto em self.notes usando a chave
        item = None
        for note in self.notes:
            if note.get('chave') == chave:
                item = note
                break
        
        if not item:
            print(f"[DEBUG PDF] ❌ Documento não encontrado no banco com chave: {chave}")
            return
        
        print(f"[DEBUG PDF] Informante: {item.get('informante', 'N/A')}")
        print(f"[DEBUG PDF] Tipo: {item.get('tipo', 'N/A')}")
        # NFe usa nNF/xNome, CTe usa nCT/xNome (remetente), NFS-e usa numero/nome_emitente
        numero = item.get('nNF') or item.get('nCT') or item.get('numero') or 'N/A'
        emitente = item.get('xNome') or item.get('nome_emitente') or 'N/A'
        print(f"[DEBUG PDF] Número: {numero}")
        print(f"[DEBUG PDF] Emitente: {emitente}")
        
        # 🆕 DETECÇÃO DE TIPO DE DOCUMENTO (para lógica específica)
        tipo = str(item.get('tipo', '')).upper()
        is_nfse = 'NFS' in tipo
        
        # OTIMIZAÇÃO 0: Verifica pdf_path do banco (SUPER RÁPIDO - PRIORITÁRIO)
        # ⚠️ EXCEÇÃO: Para NFS-e, ignora cache em pasta TEMP (usa pasta definitiva)
        pdf_path_db = item.get('pdf_path')
        should_use_db_cache = False
        
        if pdf_path_db:
            print(f"[DEBUG PDF] Etapa 0: PDF path do banco encontrado: {pdf_path_db}")
            pdf_file_db = Path(pdf_path_db)
            
            # 🔍 NFS-e: Ignora cache em pasta temporária (prioriza pasta xmls)
            if is_nfse and ('Temp' in pdf_path_db or 'temp' in pdf_path_db):
                print(f"[DEBUG PDF] ⚠️ NFS-e com cache temporário - ignorando para buscar PDF definitivo")
                should_use_db_cache = False
            elif pdf_file_db.exists():
                should_use_db_cache = True
            else:
                print(f"[DEBUG PDF] ⚠️ PDF path do banco inválido (arquivo não existe mais)")
        else:
            print(f"[DEBUG PDF] Etapa 0: PDF path não está no banco")
        
        # Se cache do banco é válido e não deve ser ignorado, usa ele
        if should_use_db_cache:
            try:
                print(f"[DEBUG PDF] ⚡⚡ Database hit! Abrindo PDF direto do banco...")
                pdf_str = str(pdf_file_db.absolute())
                if sys.platform == "win32":
                    subprocess.Popen(["cmd", "/c", "start", "", pdf_str], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)  # type: ignore[attr-defined]
                else:
                    subprocess.Popen(["xdg-open", pdf_str])
                total_time = time.time() - start_time
                print(f"[DEBUG PDF] ✅ PDF aberto (banco) - Tempo total: {total_time:.3f}s")
                self.set_status("✅ PDF aberto (cache DB)", 1000)
                return
            except Exception as e:
                print(f"[DEBUG PDF] ❌ Erro ao abrir PDF do banco: {e}")
                QMessageBox.warning(self, "Erro ao abrir PDF", f"Erro: {e}")
                return
        
        # 🆕 BUSCA DE PDF EXISTENTE NA PASTA (antes de gerar novo)
        print(f"[DEBUG PDF] Etapa 1: Buscando PDF existente na pasta...")
        pdf_path = None
        
        # 🎯 OTIMIZAÇÃO: Busca DIRETA na pasta com MÚLTIPLOS PADRÕES
        # Padrões possíveis de nome de PDF
        search_patterns = []
        
        # Extrai informante do item
        informante = item.get('informante', '')
        
        # Para NFS-e, busca pelo padrão NFSe_{numero}.pdf
        if is_nfse:
            numero = item.get('nNF') or item.get('numero')
            if numero:
                print(f"[DEBUG PDF] NFS-e detectada - Buscando NFSe_{numero}.pdf")
                # Padrões múltiplos para NFS-e:
                # 1. NFSe_{numero}.pdf (padrão antigo)
                # 2. {numero}-*.pdf (padrão novo: {NUMERO}-{FORNECEDOR}.pdf)
                # 3. DANFSe_{numero}.pdf (padrão alternativo)
                search_patterns.extend([
                    f"NFSe_{numero}.pdf",
                    f"{numero}-*.pdf",
                    f"DANFSe_{numero}.pdf"
                ])
                print(f"[DEBUG PDF] Padrões de busca: {search_patterns}...")
        else:
            # Para NF-e e CT-e, busca por VÁRIOS padrões possíveis
            numero = item.get('numero') or item.get('nNF')
            emitente = item.get('emit_nome') or item.get('nome_emitente') or ''
            emitente_clean = emitente[:50].strip() if emitente else ''
            
            # Padrão 1: Chave completa (50260107398110000100550020000259161000199154.pdf)
            search_patterns.append(f"{chave}.pdf")
            
            # Padrão 2: Número-Emitente (25916-EDC AUTO PECAS LTDA.pdf)
            if numero and emitente_clean:
                search_patterns.append(f"{numero}-{emitente_clean}.pdf")
            
            # Padrão 3: Padrão timestamp SEFAZ (20260128_155606_416_01773924000193_cte_dist_xml_extraido_cte_completo_NSU*.pdf)
            if informante:
                search_patterns.append(f"*{informante}*{chave[:14]}*.pdf")
            
            # Padrão 4: Número do documento (000123.pdf, 1234.pdf)
            if numero:
                search_patterns.append(f"{numero}.pdf")
                search_patterns.append(f"{str(numero).zfill(6)}.pdf")
        
        print(f"[DEBUG PDF] Padrões de busca: {search_patterns[:3]}...")  # Mostra primeiros 3
        
        # Busca em todas as pastas de XMLs
        roots = [
            DATA_DIR / "xmls",
            BASE_DIR / "xmls"
        ]
        
        search_start = time.time()
        for r in roots:
            if not r.exists():
                continue
            
            # Timeout de 3 segundos para toda a busca
            if time.time() - search_start > 3.0:
                print(f"[DEBUG PDF] ⏱️ Timeout na busca de PDF (3s)")
                break
            
            # Tenta cada padrão
            for pattern in search_patterns:
                if time.time() - search_start > 3.0:
                    break
                
                # Se tem wildcard (*), usa glob
                if '*' in pattern:
                    matches = list(r.rglob(pattern))
                    if matches:
                        # Pega o mais recente se houver múltiplos
                        pdf_file = max(matches, key=lambda p: p.stat().st_mtime if p.exists() else 0)
                        if pdf_file.exists() and pdf_file.suffix.lower() == '.pdf':
                            print(f"[DEBUG PDF] ✅ PDF encontrado (padrão: {pattern}): {pdf_file}")
                            pdf_path = str(pdf_file.absolute())
                            break
                else:
                    # Busca exata
                    matches = list(r.rglob(pattern))
                    if matches and matches[0].exists():
                        print(f"[DEBUG PDF] ✅ PDF encontrado (padrão: {pattern}): {matches[0]}")
                        pdf_path = str(matches[0].absolute())
                        break
            
            if pdf_path:
                break
        
        # Se encontrou PDF existente, abre direto
        if pdf_path:
            try:
                # 🆕 CORREÇÃO: Para NFS-e, verifica XML e atualiza status no banco ANTES de abrir PDF
                if is_nfse:
                    print(f"[DEBUG PDF] 🔍 NFS-e: Verificando status do XML antes de abrir PDF...")
                    xml_text = resolve_xml_text(item)
                    if xml_text:
                        print(f"[DEBUG PDF] ✅ XML da NFS-e encontrado - atualizando status no banco...")
                        try:
                            with sqlite3.connect(str(DB_PATH)) as conn_update:
                                # Atualiza notas_detalhadas
                                conn_update.execute(
                                    "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                                    (chave,)
                                )
                                conn_update.commit()
                                print(f"[DEBUG PDF] ✅ NFS-e marcada como COMPLETO no banco")
                                
                                # 🔥 FORÇA REFRESH DA TABELA NA INTERFACE
                                if self.tab_widget.currentIndex() == 0:  # Aba Recebidas
                                    QTimer.singleShot(100, self.refresh_table)
                                else:  # Aba Emitidas
                                    QTimer.singleShot(100, self.refresh_emitidos_table)
                        except Exception as e_update:
                            print(f"[DEBUG PDF] ⚠️ Erro ao atualizar status: {e_update}")
                    else:
                        print(f"[DEBUG PDF] ⚠️ XML da NFS-e não encontrado")
                
                print(f"[DEBUG PDF] ⚡ Abrindo PDF existente sem gerar novo...")
                if sys.platform == "win32":
                    subprocess.Popen(["cmd", "/c", "start", "", pdf_path], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)  # type: ignore[attr-defined]
                else:
                    subprocess.Popen(["xdg-open", pdf_path])
                
                # Atualiza cache no banco
                try:
                    self.db.atualizar_pdf_path(chave, pdf_path)
                    print(f"[DEBUG PDF] ✅ Cache atualizado no banco")
                except:
                    pass
                
                total_time = time.time() - start_time
                print(f"[DEBUG PDF] ✅ PDF aberto (pasta) - Tempo total: {total_time:.3f}s")
                self.set_status("✅ PDF aberto (arquivo existente)", 1500)
                return
            except Exception as e:
                print(f"[DEBUG PDF] ❌ Erro ao abrir PDF: {e}")
                QMessageBox.warning(self, "Erro ao abrir PDF", f"Erro: {e}")
                return
        else:
            print(f"[DEBUG PDF] ⚠️ PDF não encontrado na pasta - será gerado novo")
        
        # Pula direto para verificação de XML e geração de PDF
        print(f"[DEBUG PDF] Etapa 5: Verificando XML antes de gerar PDF...")
        xml_check_start = time.time()
        
        # Busca o XML localmente primeiro
        xml_text = resolve_xml_text(item)
        if not xml_text:
            print(f"[DEBUG PDF] ❌ XML não encontrado localmente")
            QMessageBox.warning(
                self, 
                "XML não encontrado", 
                f"Não foi possível encontrar o XML para a chave {chave}.\n\n"
                "O PDF só pode ser gerado se o XML estiver disponível."
            )
            return
        
        print(f"[DEBUG PDF] ✅ XML encontrado, iniciando geração de PDF...")
        print(f"[DEBUG PDF] Etapa 5 concluída em {time.time() - xml_check_start:.3f}s")
        
        # Tem XML, pode gerar PDF (LENTO) - executa em thread separada
        print(f"[DEBUG PDF] Etapa 6: Geração de PDF necessária...")
        generation_start = time.time()
        self.set_status("⏳ Gerando PDF... Por favor aguarde...")
        QApplication.processEvents()
        
        # Cria worker thread para não travar a interface
        print(f"[DEBUG PDF] Criando worker thread para geração assíncrona...")
        self._process_pdf_async(item)
        print(f"[DEBUG PDF] Worker criado - aguardando conclusão em background")
        print(f"[DEBUG PDF] ========================================\n")
    
    def _on_table_emitidos_double_clicked(self, row: int, col: int):
        """Abre PDF da tabela de notas emitidas (mesma lógica que _on_table_double_clicked)"""
        import time
        start_time = time.time()
        
        print(f"\n[DEBUG PDF EMITIDOS] ========== DUPLO CLIQUE ===========")
        print(f"[DEBUG PDF EMITIDOS] Linha: {row}, Coluna: {col}")
        
        # Encontra o índice da coluna "Chave" (independente de reordenação visual)
        chave_col_index = None
        for c in range(self.table_emitidos.columnCount()):
            header_text = self.table_emitidos.horizontalHeaderItem(c).text()
            if header_text == "Chave":
                chave_col_index = c
                break
        
        if chave_col_index is None:
            print(f"[DEBUG PDF EMITIDOS] ❌ Coluna 'Chave' não encontrada!")
            return
        
        # CORREÇÃO: Pega a chave diretamente da célula da tabela (coluna "Chave")
        chave_item = self.table_emitidos.item(row, chave_col_index)
        if not chave_item:
            print(f"[DEBUG PDF EMITIDOS] ❌ Célula de chave vazia na linha {row}, coluna {chave_col_index}")
            return
        
        chave = chave_item.text().strip()
        if not chave:
            print(f"[DEBUG PDF EMITIDOS] ❌ Chave vazia")
            return
        
        print(f"[DEBUG PDF EMITIDOS] Chave da célula: {chave}")
        
        # Busca o item correto na lista de emitidos usando a chave
        flt = self.filtered_emitidos()
        item = None
        for note in flt:
            if note.get('chave') == chave:
                item = note
                break
        
        if not item:
            print(f"[DEBUG PDF EMITIDOS] ❌ Documento não encontrado com chave: {chave}")
            return
        
        print(f"[DEBUG PDF EMITIDOS] Informante: {item.get('informante', 'N/A')}")
        print(f"[DEBUG PDF EMITIDOS] Tipo: {item.get('tipo', 'N/A')}")
        
        # ⛔ CACHE DESATIVADO - Sempre gera PDF pelo BrazilFiscalReport
        print(f"[DEBUG PDF EMITIDOS] ⚠️ Cache desativado - Sempre usa BrazilFiscalReport para gerar PDF")
        
        # ⛔ BUSCA DE PDF DESATIVADA - Sempre gera pelo BrazilFiscalReport
        print(f"[DEBUG PDF EMITIDOS] ⚠️ Busca de PDF simplificado desativada - Sempre gera pelo BrazilFiscalReport")
        
        # Pula direto para verificação de XML e geração de PDF
        # Para notas emitidas, o cnpj_emitente é da empresa (quem emitiu)
        # e o informante é quem recebeu (destinatário)
        # O XML está salvo pelo informante (quem baixou)
        
        # Se não tem PDF, verifica se tem XML antes de gerar
        print(f"[DEBUG PDF EMITIDOS] Etapa 5: Verificando XML antes de gerar PDF...")
        xml_check_start = time.time()
        
        # 🆕 Detecta se é NFS-e
        tipo = str(item.get('tipo', '')).upper()
        is_nfse = 'NFS' in tipo
        
        # Busca o XML localmente primeiro
        xml_text = resolve_xml_text(item)
        if not xml_text:
            print(f"[DEBUG PDF EMITIDOS] ❌ XML não encontrado localmente")
            
            # Mensagem específica para NFS-e RESUMO
            if is_nfse:
                xml_status_atual = str(item.get('xml_status', 'RESUMO')).upper()
                QMessageBox.warning(
                    self, 
                    "NFS-e Incompleta", 
                    f"Esta NFS-e está marcada como {xml_status_atual} - apenas metadados estão disponíveis.\n\n"
                    f"📄 Chave: {chave}\n"
                    f"📋 Número: {item.get('numero', 'N/A')}\n"
                    f"📅 Data: {item.get('data_emissao', 'N/A')[:10]}\n"
                    f"💰 Valor: R$ {float(item.get('valor', 0)):.2f}\n\n"
                    "⚠️ O XML completo não foi baixado da SEFAZ Nacional.\n\n"
                    "💡 Para obter o XML completo e gerar o PDF:\n"
                    "1. Aguarde a próxima busca automática de NFS-e\n"
                    "2. Ou execute uma busca manual de NFS-e"
                )
            else:
                QMessageBox.warning(
                    self, 
                    "XML não encontrado", 
                    f"Não foi possível encontrar o XML para a chave {chave}.\n\n"
                    "O PDF só pode ser gerado se o XML estiver disponível."
                )
            return
        
        # 🆕 CORREÇÃO: Se é NFS-e e encontrou XML, atualiza status no banco
        if is_nfse and xml_text:
            print(f"[DEBUG PDF EMITIDOS] ✅ XML da NFS-e encontrado - atualizando status no banco...")
            try:
                with sqlite3.connect(str(DB_PATH)) as conn_update:
                    conn_update.execute(
                        "UPDATE notas_detalhadas SET xml_status = 'COMPLETO' WHERE chave = ?",
                        (chave,)
                    )
                    conn_update.commit()
                    print(f"[DEBUG PDF EMITIDOS] ✅ NFS-e marcada como COMPLETO no banco")
            except Exception as e_update:
                print(f"[DEBUG PDF EMITIDOS] ⚠️ Erro ao atualizar status: {e_update}")
        
        print(f"[DEBUG PDF EMITIDOS] ✅ XML encontrado, iniciando geração de PDF...")
        print(f"[DEBUG PDF EMITIDOS] Etapa 5 concluída em {time.time() - xml_check_start:.3f}s")
        
        # Tem XML, pode gerar PDF (LENTO) - executa em thread separada
        print(f"[DEBUG PDF EMITIDOS] Etapa 6: Geração de PDF necessária...")
        generation_start = time.time()
        self.set_status("⏳ Gerando PDF... Por favor aguarde...")
        QApplication.processEvents()
        
        # Cria worker thread para não travar a interface
        print(f"[DEBUG PDF EMITIDOS] Criando worker thread para geração assíncrona...")
        self._process_pdf_async(item)
        print(f"[DEBUG PDF EMITIDOS] Worker criado - aguardando conclusão em background")
        print(f"[DEBUG PDF EMITIDOS] ========================================\n")
    
    def _process_pdf_async(self, item: Dict[str, Any]):
        """Processa PDF em thread separada para não travar a UI"""
        
        # ⛔ NUNCA GERAR PDF PARA EVENTOS!
        xml_status = (item.get('xml_status') or '').upper()
        if xml_status == 'EVENTO':
            print("[PDF] ⛔ Eventos não geram PDF - pulando...")
            QMessageBox.information(self, "PDF não disponível", 
                "Eventos não geram PDF.\nApenas o arquivo XML está disponível.")
            return
        
        class PDFWorker(QThread):
            finished = pyqtSignal(dict)  # {ok, pdf_path, error}
            status_update = pyqtSignal(str)  # Para atualizar status na UI
            
            def __init__(self, parent_window, item_data):
                super().__init__()
                self.parent_window = parent_window
                self.item = item_data
            
            def run(self):
                try:
                    chave = self.item.get('chave', '')
                    informante = self.item.get('informante', '')
                    tipo = (self.item.get('tipo') or 'NFe').strip().upper().replace('-', '')
                    
                    # Resolve XML
                    self.status_update.emit("🔍 Buscando XML...")
                    xml_text = resolve_xml_text(self.item)
                    
                    # Verifica se precisa buscar XML completo
                    def _is_nfe_full(x: str) -> bool:
                        xl = (x or '').lower()
                        return ('<nfeproc' in xl) or ('<protnfe' in xl)
                    def _is_cte_full(x: str) -> bool:
                        xl = (x or '').lower()
                        return ('<proccte' in xl) or ('<protcte' in xl)
                    
                    need_fetch = False
                    downloaded_from_sefaz = False
                    
                    if not xml_text:
                        need_fetch = True
                    else:
                        if tipo == 'NFE' and not _is_nfe_full(xml_text):
                            need_fetch = True
                        if tipo == 'CTE' and not _is_cte_full(xml_text):
                            need_fetch = True
                    
                    if need_fetch:
                        self.status_update.emit("⏳ Baixando XML completo da SEFAZ...")
                        try:
                            certs = self.parent_window.db.load_certificates()
                            selected = getattr(self.parent_window, '_selected_cert_cnpj', None)
                            if selected:
                                try:
                                    certs.sort(key=lambda c: 0 if (str(c.get('cnpj_cpf') or '') == selected or str(c.get('informante') or '') == selected) else 1)
                                except Exception:
                                    pass
                            prefer = ('nfeProc', 'NFe') if tipo == 'NFE' else ('procCTe', 'CTe')
                            for c in certs:
                                payload = {
                                    'cert': {
                                        'path': c.get('caminho') or '',
                                        'senha': c.get('senha') or '',
                                        'cnpj': c.get('cnpj_cpf') or '',
                                        'cuf': c.get('cUF_autor') or ''
                                    },
                                    'chave': chave,
                                    'prefer': list(prefer)
                                }
                                res = sandbox.run_task('fetch_by_chave', payload, timeout=240)
                                if res.get('ok') and res.get('data', {}).get('xml'):
                                    xml_text = res['data']['xml']
                                    downloaded_from_sefaz = True
                                    break
                        except Exception:
                            pass
                    
                    if not xml_text:
                        self.finished.emit({"ok": False, "error": "XML completo não encontrado (local/SEFAZ)."})
                        return
                    
                    # Save downloaded XML
                    saved_xml_path = None
                    if downloaded_from_sefaz:
                        self.status_update.emit("💾 Salvando XML...")
                        try:
                            data_emissao = (self.item.get('data_emissao') or '')[:10]
                            if chave and informante:
                                year_month = data_emissao[:7] if len(data_emissao) >= 7 else datetime.now().strftime("%Y-%m")
                                xmls_root = DATA_DIR / "xmls" / informante / tipo / year_month
                                xmls_root.mkdir(parents=True, exist_ok=True)
                                xml_file = xmls_root / f"{chave}.xml"
                                xml_file.write_text(xml_text, encoding='utf-8')
                                saved_xml_path = str(xml_file)
                                self.parent_window.db.register_xml_download(chave, saved_xml_path, informante)
                                upd = {'chave': chave, 'xml_status': 'COMPLETO', 'informante': informante}
                                for k in ['ie_tomador', 'nome_emitente', 'cnpj_emitente', 'numero',
                                          'data_emissao', 'tipo', 'valor', 'cfop', 'vencimento',
                                          'ncm', 'status', 'natureza', 'uf', 'base_icms', 'valor_icms']:
                                    if k in self.item:
                                        upd[k] = self.item[k]
                                self.parent_window.db.save_note(upd)
                        except Exception:
                            pass
                    
                    # Determine PDF path - OTIMIZADO para buscar o XML na estrutura organizada
                    if saved_xml_path:
                        pdf_path = Path(saved_xml_path).with_suffix('.pdf')
                    else:
                        if chave and informante:
                            # Busca o XML APENAS na estrutura organizada por CNPJ (xmls/{CNPJ}/...)
                            xmls_root = DATA_DIR / "xmls"
                            found_xml = None
                            
                            # Busca 1: Na pasta do informante por nome da chave (PADRÃO v1.0.86+)
                            informante_folder = xmls_root / informante
                            if informante_folder.exists():
                                print(f"[DEBUG XML] Buscando {chave}.xml em: {informante_folder}")
                                for xml_file in informante_folder.rglob(f"{chave}.xml"):
                                    # Ignora pastas de debug/backup
                                    if not any(x in str(xml_file).lower() for x in ['debug', 'backup', 'request', 'response']):
                                        found_xml = xml_file
                                        print(f"[DEBUG XML] ✅ XML encontrado: {xml_file}")
                                        break
                            
                            # Busca 2: Busca por padrão antigo ou conteúdo (FALLBACK)
                            if not found_xml and informante_folder.exists():
                                print(f"[DEBUG XML] Buscando XMLs legados com chave no nome...")
                                for xml_file in informante_folder.rglob("*.xml"):
                                    # Ignora arquivos de sistema
                                    if any(x in str(xml_file).lower() for x in ['debug', 'backup', 'request', 'response', 'protocolo']):
                                        continue
                                    if chave in xml_file.name:
                                        found_xml = xml_file
                                        print(f"[DEBUG XML] ✅ XML legado encontrado: {xml_file}")
                                        break
                            
                            if found_xml:
                                # Salva PDF junto com o XML encontrado
                                pdf_path = found_xml.with_suffix('.pdf')
                            else:
                                # Último recurso: pasta temporária
                                import tempfile
                                tmp = Path(tempfile.gettempdir()) / "BOT_Busca_NFE_PDFs"
                                tmp.mkdir(parents=True, exist_ok=True)
                                pdf_path = tmp / f"{tipo}-{chave}.pdf"
                        else:
                            import tempfile
                            tmp = Path(tempfile.gettempdir()) / "BOT_Busca_NFE_PDFs"
                            tmp.mkdir(parents=True, exist_ok=True)
                            pdf_path = tmp / f"{tipo}-{chave}.pdf"
                    
                    # Check if PDF exists (pode ter sido salvo junto com o XML)
                    if pdf_path.exists():
                        self.status_update.emit("✅ PDF encontrado!")
                        self.finished.emit({"ok": True, "pdf_path": str(pdf_path)})
                        return
                    
                    # Generate PDF
                    self.status_update.emit("📄 Gerando PDF do XML...")
                    payload: Dict[str, Any] = {
                        "xml": xml_text,
                        "tipo": (self.item.get("tipo") or "NFe"),
                        "out_path": str(pdf_path),
                        "force_simple_fallback": False,
                    }
                    res = sandbox.run_task("generate_pdf", payload, timeout=240)
                    if res.get("ok"):
                        if pdf_path.exists():
                            self.finished.emit({"ok": True, "pdf_path": str(pdf_path)})
                        else:
                            self.finished.emit({"ok": False, "error": "PDF não foi gerado"})
                    else:
                        self.finished.emit({"ok": False, "error": f"Falha ao gerar PDF: {res.get('error')}"})
                
                except Exception as e:
                    import traceback
                    self.finished.emit({"ok": False, "error": f"Erro: {str(e)}\n{traceback.format_exc()}"})
        
        # Cria e inicia worker
        worker = PDFWorker(self, item)
        
        def on_status(msg: str):
            self.set_status(msg)
        
        def on_finished(result: dict):
            if result.get("ok"):
                pdf_path = result.get("pdf_path")
                chave = item.get('chave')
                # AUTO-CURA: Atualiza PDF path no banco após gerar
                if pdf_path and chave:
                    self.db.atualizar_pdf_path(chave, pdf_path)
                    print(f"[DEBUG PDF] 🔄 PDF gerado - path salvo no banco: {chave}")
                try:
                    if sys.platform == "win32":
                        # Abre PDF com visualizador padrão do Windows (evita abrir interface se PDF estiver associado incorretamente)
                        subprocess.Popen(["cmd", "/c", "start", "", pdf_path], shell=False, creationflags=subprocess.CREATE_NO_WINDOW)  # type: ignore[attr-defined]
                    else:
                        subprocess.Popen(["xdg-open", pdf_path])
                    self.set_status("✅ PDF gerado e aberto com sucesso!", 2000)
                except Exception as e:
                    self.set_status("")
                    QMessageBox.warning(self, "Erro ao abrir PDF", f"Erro: {e}")
            else:
                self.set_status("")
                error_msg = result.get("error", "Erro desconhecido")
                QMessageBox.critical(self, "Erro", error_msg)
        
        worker.status_update.connect(on_status)
        worker.finished.connect(on_finished)
        worker.start()
        
        # Mantém referência para evitar garbage collection
        if not hasattr(self, '_pdf_workers'):
            self._pdf_workers = []
        self._pdf_workers.append(worker)
        worker.finished.connect(lambda: self._pdf_workers.remove(worker) if worker in self._pdf_workers else None)

    def _auto_start_search(self):
        """Inicia busca automaticamente ao iniciar o sistema."""
        from datetime import datetime, timedelta
        
        try:
            # Usa o intervalo configurado pelo usuário (em horas)
            intervalo_horas = self.spin_intervalo.value()
            intervalo_minutos = intervalo_horas * 60
            
            # Verificar última execução
            last_search = self.db.get_last_search_time()
            
            if last_search:
                # Converter para datetime
                try:
                    last_dt = datetime.fromisoformat(last_search)
                    now = datetime.now()
                    diff_minutes = (now - last_dt).total_seconds() / 60
                    
                    # Se já passou o intervalo, inicia busca imediatamente
                    if diff_minutes >= intervalo_minutos:
                        self._search_in_progress = True
                        self.set_status(f"Última busca: {diff_minutes:.0f} minutos atrás. Iniciando busca automática...", 5000)
                        # Registra o horário da busca
                        self.db.set_last_search_time(datetime.now().isoformat())
                        # Inicia a busca
                        QTimer.singleShot(500, self.do_search)
                    else:
                        # Ainda está no intervalo de espera, calcula próxima busca
                        minutos_restantes = intervalo_minutos - diff_minutes
                        self._next_search_time = now + timedelta(minutes=minutos_restantes)
                        self.set_status(f"Última busca há {diff_minutes:.0f} minutos. Próxima em {minutos_restantes:.0f} minutos.", 5000)
                        
                        # Agenda a próxima busca
                        delay_ms = int(minutos_restantes * 60 * 1000)
                        QTimer.singleShot(delay_ms, self._auto_start_search)
                        
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    self._search_in_progress = True
                    self.set_status("Iniciando busca automática...", 3000)
                    self.db.set_last_search_time(datetime.now().isoformat())
                    QTimer.singleShot(500, self.do_search)
            else:
                # Primeira execução
                self._search_in_progress = True
                self.set_status("Primeira execução. Iniciando busca automática...", 3000)
                # Registra o horário da busca
                self.db.set_last_search_time(datetime.now().isoformat())
                # Inicia a busca
                QTimer.singleShot(500, self.do_search)
            
        except Exception as e:
            print(f"[DEBUG] Erro em _auto_start_search: {e}")
            import traceback
            traceback.print_exc()
            self._search_in_progress = False
            self.set_status(f"Erro ao iniciar busca automática: {e}", 5000)
    
    def _executar_busca_agendada(self):
        """Executa busca agendada diretamente (já passou o intervalo configurado)."""
        from datetime import datetime
        
        try:
            # ⛔ Se o timer foi cancelado manualmente (mudança de intervalo), ignora
            if hasattr(self, '_timer_cancelled') and self._timer_cancelled:
                print("[DEBUG] ⏹️ Timer cancelado - ignorando execução agendada")
                return
            
            print("[AUTO-SEARCH] Executando busca agendada")
            
            # Marca busca em andamento
            self._search_in_progress = True
            
            # Registra o horário da busca
            self.db.set_last_search_time(datetime.now().isoformat())
            
            # Atualiza status
            self.set_status("Iniciando busca automática...", 3000)
            
            # Inicia a busca
            QTimer.singleShot(500, self.do_search)
            
        except Exception as e:
            print(f"[DEBUG] Erro em _executar_busca_agendada: {e}")
            import traceback
            traceback.print_exc()
            self._search_in_progress = False
            self.set_status(f"Erro ao executar busca agendada: {e}", 5000)
    
    def _verificar_tarefas_agendadas_inicializacao(self):
        """Verifica e executa tarefas configuradas para executar ao iniciar"""
        from PyQt5.QtCore import QSettings
        
        try:
            settings = QSettings('NFE_System', 'BOT_NFE')
            
            # Verifica se está configurado para executar ao iniciar
            ao_iniciar = settings.value('agendador/ao_iniciar', False, type=bool)
            
            if not ao_iniciar:
                print("[AGENDADOR] Nenhuma tarefa configurada para executar ao iniciar")
                return
            
            # Obtém qual tarefa executar
            tarefa_index = settings.value('agendador/tarefa_index', 0, type=int)
            tarefa_nome = settings.value('agendador/tarefa_nome', 'Buscar Notas na SEFAZ')
            
            print(f"[AGENDADOR] Executando tarefa ao iniciar: {tarefa_nome}")
            
            # Executa a tarefa correspondente
            self._executar_tarefa_agendada(tarefa_index, tarefa_nome)
            
        except Exception as e:
            print(f"[AGENDADOR] Erro ao verificar tarefas de inicialização: {e}")
    
    def _verificar_tarefas_agendadas_periodico(self):
        """Verifica e executa tarefas agendadas por horário ou intervalo"""
        from PyQt5.QtCore import QSettings, QTime
        from datetime import datetime, timedelta
        
        try:
            settings = QSettings('NFE_System', 'BOT_NFE')
            
            # Verifica tarefa por horário específico
            horario_ativo = settings.value('agendador/horario_ativo', False, type=bool)
            if horario_ativo:
                horario_config = settings.value('agendador/horario', '08:00')
                horario = QTime.fromString(horario_config, 'HH:mm')
                hora_atual = QTime.currentTime()
                
                # Verifica se está na hora (com margem de 1 minuto)
                if abs(hora_atual.secsTo(horario)) <= 60:
                    ultima_exec_horario = settings.value('agendador/ultima_exec_horario', '')
                    hoje = datetime.now().date().isoformat()
                    
                    # Executa apenas uma vez por dia
                    if ultima_exec_horario != hoje:
                        tarefa_index = settings.value('agendador/tarefa_index', 0, type=int)
                        tarefa_nome = settings.value('agendador/tarefa_nome', 'Buscar Notas na SEFAZ')
                        
                        print(f"[AGENDADOR] Executando tarefa por horário ({horario_config}): {tarefa_nome}")
                        
                        self._executar_tarefa_agendada(tarefa_index, tarefa_nome)
                        
                        # Marca como executado hoje
                        settings.setValue('agendador/ultima_exec_horario', hoje)
            
            # Verifica tarefa por intervalo
            intervalo_ativo = settings.value('agendador/intervalo_ativo', False, type=bool)
            if intervalo_ativo:
                intervalo_horas = settings.value('agendador/intervalo_horas', 2, type=int)
                ultima_exec_intervalo = settings.value('agendador/ultima_exec_intervalo', '')
                
                if ultima_exec_intervalo:
                    ultima = datetime.fromisoformat(ultima_exec_intervalo)
                    agora = datetime.now()
                    diff = (agora - ultima).total_seconds() / 3600  # em horas
                    
                    if diff >= intervalo_horas:
                        tarefa_index = settings.value('agendador/tarefa_index', 0, type=int)
                        tarefa_nome = settings.value('agendador/tarefa_nome', 'Buscar Notas na SEFAZ')
                        
                        print(f"[AGENDADOR] Executando tarefa por intervalo ({intervalo_horas}h): {tarefa_nome}")
                        
                        self._executar_tarefa_agendada(tarefa_index, tarefa_nome)
                        
                        # Atualiza timestamp
                        settings.setValue('agendador/ultima_exec_intervalo', datetime.now().isoformat())
                else:
                    # Primeira execução do intervalo
                    settings.setValue('agendador/ultima_exec_intervalo', datetime.now().isoformat())
            
        except Exception as e:
            print(f"[AGENDADOR] Erro ao verificar tarefas periódicas: {e}")
    
    def _executar_tarefa_agendada(self, tarefa_index: int, tarefa_nome: str):
        """Executa a tarefa agendada conforme o índice"""
        try:
            print(f"[AGENDADOR] Iniciando execução: {tarefa_nome}")
            self.set_status(f"⏰ Executando tarefa agendada: {tarefa_nome}")
            
            # Mapeamento de tarefas
            if tarefa_index == 0:  # Buscar Notas na SEFAZ
                self.do_search()
            elif tarefa_index == 1:  # Busca Completa (NSU)
                self.do_busca_completa()
            elif tarefa_index == 2:  # Atualizar Status de Notas
                self._atualizar_status_background()
            elif tarefa_index == 3:  # Baixar XMLs Faltantes
                self.baixar_xmls_faltantes_por_chave()
            elif tarefa_index == 4:  # Gerar PDFs Pendentes
                self._gerar_pdfs_faltantes()
            elif tarefa_index == 5:  # Manifestação Automática
                self._manifestar_nota(None)
            elif tarefa_index == 6:  # Sincronizar Documentos
                self.sincronizar_xmls_interface()
            
            print(f"[AGENDADOR] Tarefa '{tarefa_nome}' executada com sucesso")
            
        except Exception as e:
            print(f"[AGENDADOR] Erro ao executar tarefa '{tarefa_nome}': {e}")
            self.set_status(f"❌ Erro ao executar tarefa agendada: {e}")
    
    def do_search(self):
        from datetime import datetime, timedelta
        
        # Marca busca em andamento
        self._search_in_progress = True
        self._next_search_time = None
        
        # Reseta estatísticas
        self._search_stats = {
            'nfes_found': 0,
            'ctes_found': 0,
            'nfses_found': 0,
            'start_time': datetime.now(),
            'last_cert': '',
            'total_docs': 0
        }
        
        # Mostra progress bar
        self.search_progress.setVisible(True)
        self.search_progress.setRange(0, 0)  # Modo indeterminado
        self.search_summary_label.setText("🔍 Iniciando busca...")
        
        # Atualiza timestamp da última busca
        try:
            self.db.set_last_search_time(datetime.now().isoformat())
        except Exception:
            pass  # Silencioso para evitar recursão
        
        # Janela de debug desabilitada - usando apenas status bar
        # dlg = SearchDialog(self)
        # dlg.show()
        # QApplication.processEvents()

        def on_progress(line: str):
            if not line:
                return
            
            # Atualiza resumo baseado nos logs
            try:
                # Detecta processamento de certificado
                if "Processando certificado" in line:
                    import re
                    match = re.search(r'CNPJ=(\d+)', line)
                    if match:
                        cnpj = match.group(1)
                        self._search_stats['last_cert'] = cnpj[-4:]  # Últimos 4 dígitos
                        self._update_search_summary()
                
                # Detecta NFe encontrada
                if "registrar_xml" in line.lower() or "infnfe" in line.lower():
                    self._search_stats['nfes_found'] += 1
                    self._update_search_summary()
                
                # Detecta CTe encontrado
                if "processar_cte" in line.lower() or "🚛" in line:
                    self._search_stats['ctes_found'] += 1
                    self._update_search_summary()
                
                # Detecta NFS-e processada
                if "nfs-e processada com sucesso" in line.lower() or "📋" in line:
                    self._search_stats['nfses_found'] += 1
                    self._update_search_summary()
                
                # Detecta documentos processados
                if "docZip" in line or "NSU" in line:
                    self._search_stats['total_docs'] += 1
                    self._update_search_summary()
                    
            except Exception:
                pass  # REMOVIDO print() para evitar recursão via ProgressCapture
            
            # Detecta se a busca foi finalizada
            if "Busca de NSU finalizada" in line or "Próxima busca será agendada" in line or "Busca concluída" in line:
                # Marca que a busca finalizou
                self._search_in_progress = False
                
                # Oculta progress bar
                self.search_progress.setVisible(False)
                
                # Mostra resumo final
                elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                self.search_summary_label.setText(
                    f"✅ NFes: {self._search_stats['nfes_found']} | "
                    f"CTes: {self._search_stats['ctes_found']} | "
                    f"NFSes: {self._search_stats['nfses_found']} | "
                    f"Tempo: {elapsed:.0f}s"
                )
                
                # Usa o intervalo configurado pelo usuário (em horas)
                intervalo_horas = self.spin_intervalo.value()
                intervalo_minutos = intervalo_horas * 60
                
                # Calcula próxima busca baseado no intervalo configurado
                self._next_search_time = datetime.now() + timedelta(minutes=intervalo_minutos)
                
                # Atualiza status
                if intervalo_horas == 1:
                    self.set_status(f"Próxima busca em {intervalo_horas} hora", 0)
                else:
                    self.set_status(f"Próxima busca em {intervalo_horas} horas", 0)
                
                # Agenda a próxima busca automaticamente (executa diretamente sem verificação)
                delay_ms = int(intervalo_minutos * 60 * 1000)
                
                # 🔄 Armazena o timer para poder cancelá-lo se necessário
                self._scheduled_timer_id = QTimer.singleShot(delay_ms, self._executar_busca_agendada)
                
                return
            
            # Linha de progresso é exibida apenas nos logs (não mais em janela)
            # Mantém apenas para compatibilidade com código existente

        # Worker thread para não travar a interface
        class SearchWorker(QThread):
            finished_search = pyqtSignal(dict)
            progress_line = pyqtSignal(str)
            error_occurred = pyqtSignal(str)
            
            def run(self):
                try:
                    res = run_search(progress_cb=lambda line: self.progress_line.emit(line))
                    self.finished_search.emit(res)
                except Exception as e:
                    import traceback
                    error_msg = f"Erro fatal na thread de busca: {str(e)}\n{traceback.format_exc()}"
                    print(error_msg)
                    self.error_occurred.emit(error_msg)
                    self.finished_search.emit({"ok": False, "error": error_msg})
        
        def on_finished(res: Dict[str, Any]):
            try:
                if not res.get("ok"):
                    error = res.get('error') or res.get('message')
                    print(f"\nErro na busca: {error}")
                    self._search_in_progress = False
                    self.search_summary_label.setText(f"❌ Erro: {error[:50]}...")
                self.refresh_all()
                self._search_worker = None
                
                # Gera PDFs dos novos XMLs em background
                print("[INFO] Iniciando geração de PDFs dos novos XMLs...")
                QTimer.singleShot(1000, self._gerar_pdfs_faltantes)
                
                # 🆕 CONSULTA DE EVENTOS após busca SEFAZ (se busca foi bem-sucedida)
                if res.get("ok"):
                    print("[PÓS-BUSCA] Iniciando consulta de eventos dos documentos baixados...")
                    QTimer.singleShot(3000, lambda: self._atualizar_status_apos_busca())
                    
                    # 🆕 CORREÇÃO DE STATUS após busca (executa após eventos)
                    print("[PÓS-BUSCA] Agendando correção automática de status XML...")
                    QTimer.singleShot(10000, lambda: self._executar_correcao_status())
                    
                    # 🆕 BUSCA DE NFS-e após busca SEFAZ concluir com sucesso
                    print("\n" + "="*70)
                    print("[PÓS-BUSCA] ✅ NF-e e CT-e concluídos!")
                    print("[PÓS-BUSCA] 📋 NFS-e será processada em 5 segundos...")
                    print("[PÓS-BUSCA] ℹ️  NFS-e roda separadamente para evitar duplicação")
                    print("="*70 + "\n")
                    QTimer.singleShot(5000, self._buscar_nfse_automatico)
            except Exception as e:
                import traceback
                error_msg = f"Erro em on_finished: {str(e)}\n{traceback.format_exc()}"
                print(error_msg)
                QMessageBox.critical(self, "Erro", error_msg)
        
        def on_error(error_msg: str):
            try:
                print(f"[ERRO] {error_msg}")
                self._search_in_progress = False
                self.search_summary_label.setText(f"❌ Erro fatal")
                QMessageBox.critical(self, "Erro Fatal na Busca", error_msg)
            except Exception as e:
                print(f"Erro ao processar on_error: {e}")
        
        # Conecta sinais
        self._search_worker = SearchWorker()
        self._search_worker.progress_line.connect(on_progress)
        self._search_worker.finished_search.connect(on_finished)
        self._search_worker.error_occurred.connect(on_error)
        self._search_worker.start()

    # ==========================
    # Novas funções implementadas
    # ==========================
    def buscar_por_chave(self):
        """
        Busca NF-e/CT-e por chave de acesso (individual ou arquivo TXT).
        
        🎯 FUNCIONALIDADE:
        - Consulta a SEFAZ diretamente pela chave de 44 dígitos
        - Salva XMLs automaticamente (backup local + perfis ativos)
        - Extrai e salva dados completos no banco
        - Detecta automaticamente se é nota de ENTRADA ou SAÍDA
        
        📂 ONDE AS NOTAS APARECEM:
        - Notas de SAÍDA (emitidas pela empresa) → aba "Emitidos pela empresa"
        - Notas de ENTRADA (recebidas) → aba principal
        
        💡 USO RECOMENDADO:
        - Importar NF-e de saída que não vêm pela distribuição DFe
        - Recuperar notas antigas ou faltantes
        - Importar notas sem precisar dos XMLs salvos
        """
        try:
            # Dialog para escolher método de entrada
            reply = QMessageBox.question(
                self,
                "Busca por Chave - NF-e/CT-e entrada e saída",
                "🔍 BUSCAR NOTAS PELA CHAVE DE ACESSO\n\n"
                "Esta função busca notas diretamente na SEFAZ e salva automaticamente:\n"
                "  📥 Notas de ENTRADA (recebidas)\n"
                "  📤 Notas de SAÍDA (emitidas pela empresa)\n\n"
                "Como deseja informar a(s) chave(s)?\n\n"
                "• YES = Digitar chave única (44 dígitos)\n"
                "• NO = Importar arquivo .txt com múltiplas chaves",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Cancel:
                return
            
            chaves = []
            
            if reply == QMessageBox.Yes:
                # Entrada manual de chave única
                chave, ok = QInputDialog.getText(
                    self,
                    "Busca por Chave",
                    "Informe a chave de acesso (44 dígitos):"
                )
                if not ok or not chave:
                    return
                # Remove espaços e caracteres não numéricos
                chave_limpa = ''.join(c for c in chave if c.isdigit())
                if len(chave_limpa) != 44:
                    QMessageBox.warning(self, "Busca por Chave", "Chave inválida! Deve conter exatamente 44 dígitos.")
                    return
                chaves.append(chave_limpa)
            else:
                # Importar arquivo TXT
                arquivo, _ = QFileDialog.getOpenFileName(
                    self,
                    "Selecionar arquivo TXT com chaves",
                    "",
                    "Arquivos de texto (*.txt);;Todos os arquivos (*.*)"
                )
                if not arquivo:
                    return
                
                try:
                    with open(arquivo, 'r', encoding='utf-8') as f:
                        for linha in f:
                            chave_limpa = ''.join(c for c in linha if c.isdigit())
                            if len(chave_limpa) == 44:
                                chaves.append(chave_limpa)
                    
                    if not chaves:
                        QMessageBox.warning(self, "Busca por Chave", "Nenhuma chave válida encontrada no arquivo.")
                        return
                except Exception as e:
                    QMessageBox.critical(self, "Busca por Chave", f"Erro ao ler arquivo: {e}")
                    return
            
            # Confirma busca
            total_chaves = len(chaves)
            reply = QMessageBox.question(
                self,
                "Busca por Chave",
                f"Buscar {total_chaves} chave(s) de acesso na SEFAZ?\n\n"
                f"✅ As notas encontradas serão salvas automaticamente.\n"
                f"📂 Notas de SAÍDA → aba 'Emitidos pela empresa'\n"
                f"📂 Notas de ENTRADA → aba principal",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Executa busca em thread separada
            self._executar_busca_por_chaves(chaves)
            
        except Exception as e:
            QMessageBox.critical(self, "Busca por Chave", f"Erro: {e}")
    
    def _executar_busca_por_chaves(self, chaves):
        """Executa a busca das chaves em background."""
        import sys
        from lxml import etree
        sys.path.insert(0, str(BASE_DIR))
        from nfe_search import NFeService, XMLProcessor, DatabaseManager, salvar_xml_por_certificado, extrair_nota_detalhada
        
        # Cria dialog de progresso
        progress = QProgressDialog("Buscando chaves na SEFAZ...", "Cancelar", 0, len(chaves), self)
        progress.setWindowTitle("Busca por Chave")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        db = DatabaseManager(DB_PATH)
        parser = XMLProcessor()
        encontradas = 0
        nao_encontradas = 0
        erros = []
        
        # Carrega todos os certificados
        certificados = db.get_certificados()
        if not certificados:
            QMessageBox.warning(self, "Busca por Chave", "Nenhum certificado cadastrado!")
            return
        
        # Índice do certificado que teve sucesso (para priorizar nos próximos)
        ultimo_cert_sucesso = 0
        
        print(f"\n[BUSCA POR CHAVE] Iniciando busca de {len(chaves)} chaves")
        print(f"[BUSCA POR CHAVE] Certificados disponíveis: {len(certificados)}")
        
        for idx, chave in enumerate(chaves):
            if progress.wasCanceled():
                print(f"[BUSCA POR CHAVE] Busca cancelada pelo usuário na chave {idx+1}")
                break
            
            print(f"\n{'='*80}")
            print(f"[BUSCA POR CHAVE] Processando {idx+1}/{len(chaves)}: {chave}")
            print(f"{'='*80}")
            
            progress.setValue(idx)
            progress.setLabelText(f"Buscando chave {idx+1}/{len(chaves)}...\n{chave}")
            QApplication.processEvents()
            
            resp_xml = None
            cert_encontrado = None
            
            # Detecta tipo do documento pela chave (posição 20-21: modelo)
            # Modelo 55 = NF-e, Modelo 57 = CT-e
            modelo = chave[20:22] if len(chave) >= 22 else '55'
            is_cte = modelo == '57'
            tipo_doc = 'CT-e' if is_cte else 'NF-e'
            print(f"[DEBUG] Tipo detectado: {tipo_doc} (modelo={modelo})")
            
            # Extrai UF da chave (primeiros 2 dígitos)
            uf_chave = chave[:2] if len(chave) >= 2 else None
            print(f"[DEBUG] UF da chave: {uf_chave}")
            
            # Ordena certificados: prioriza UF da chave, depois último sucesso, depois resto
            ordem_certs = []
            
            # 1. Certificados da mesma UF da chave
            for i, cert in enumerate(certificados):
                if cert[4] == uf_chave:  # cuf == uf_chave
                    ordem_certs.append(i)
            
            # 2. Último certificado com sucesso (se não estiver na lista)
            if ultimo_cert_sucesso not in ordem_certs:
                ordem_certs.insert(0, ultimo_cert_sucesso)
            
            # 3. Resto dos certificados
            for i in range(len(certificados)):
                if i not in ordem_certs:
                    ordem_certs.append(i)
            
            print(f"[DEBUG] Ordem de tentativa dos certificados: {[i+1 for i in ordem_certs]}")
            
            for cert_idx in ordem_certs:
                try:
                    cnpj, path, senha, inf, cuf = certificados[cert_idx]
                    print(f"[DEBUG] Tentando certificado {cert_idx+1}/{len(certificados)}: CNPJ={cnpj}, Informante={inf}, UF={cuf}")
                    
                    svc = NFeService(path, senha, cnpj, cuf)
                    
                    # Busca o XML do documento (NF-e ou CT-e)
                    if is_cte:
                        print(f"[DEBUG] Chamando fetch_prot_cte para chave: {chave}")
                        resp_xml = svc.fetch_prot_cte(chave)
                    else:
                        print(f"[DEBUG] Chamando fetch_prot_nfe para chave: {chave}")
                        resp_xml = svc.fetch_prot_nfe(chave)
                    print(f"[DEBUG] Resposta recebida: {resp_xml[:200] if resp_xml else 'None'}...")
                    
                    if resp_xml:
                        # Verifica se é erro que indica "tente outro certificado"
                        # 217 = Nota não consta na base, 226 = UF divergente, 404 = namespace
                        erros_tentar_outro = ['217', '226', '404']
                        tem_erro_cert = any(f'<cStat>{cod}</cStat>' in resp_xml for cod in erros_tentar_outro)
                        
                        if not tem_erro_cert:
                            cert_encontrado = (cnpj, path, senha, inf, cuf)
                            ultimo_cert_sucesso = cert_idx
                            print(f"[SUCCESS] Nota encontrada com certificado {cert_idx+1}!")
                            break
                        else:
                            # Identifica qual erro
                            if '<cStat>217</cStat>' in resp_xml:
                                print(f"[INFO] Nota não consta na base deste certificado (217), tentando próximo...")
                            elif '<cStat>226</cStat>' in resp_xml:
                                print(f"[INFO] UF divergente (226), tentando próximo...")
                            elif '<cStat>404</cStat>' in resp_xml:
                                print(f"[INFO] Erro de namespace (404), tentando próximo...")
                            else:
                                print(f"[INFO] Erro ao buscar, tentando próximo certificado...")
                    
                except Exception as e:
                    print(f"[ERRO] Erro ao tentar certificado {cert_idx+1}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # Processa resultado
            try:
                if not resp_xml or not cert_encontrado:
                    nao_encontradas += 1
                    erros.append(f"{chave}: Não encontrada em nenhum certificado")
                    print(f"[ERRO] Chave não encontrada em nenhum certificado")
                    continue
                
                if resp_xml:
                    print(f"[DEBUG] Processando resposta XML...")
                    # Parse da resposta
                    try:
                        tree = etree.fromstring(resp_xml.encode('utf-8') if isinstance(resp_xml, str) else resp_xml)
                        print(f"[DEBUG] XML parseado com sucesso")
                        
                        # Extrai informações do protocolo (NF-e ou CT-e)
                        if is_cte:
                            # Namespace CT-e
                            NS = {'cte': 'http://www.portalfiscal.inf.br/cte'}
                            prot = tree.find('.//cte:protCTe', namespaces=NS)
                            print(f"[DEBUG] protCTe encontrado: {prot is not None}")
                            
                            if prot is not None:
                                ch = prot.findtext('cte:infProt/cte:chCTe', namespaces=NS) or ''
                                cStat = prot.findtext('cte:infProt/cte:cStat', namespaces=NS) or ''
                                xMotivo = prot.findtext('cte:infProt/cte:xMotivo', namespaces=NS) or ''
                                print(f"[DEBUG] Protocolo CT-e: chave={ch}, cStat={cStat}, xMotivo={xMotivo}")
                                
                                # Salva status
                                if ch and cStat:
                                    db.set_nf_status(ch, cStat, xMotivo)
                                    print(f"[DEBUG] Status CT-e salvo no banco")
                                
                                # Se autorizado (código 100 = Autorizado)
                                if cStat in ['100', '101', '110', '150', '301', '302']:
                                    # Salva XML completo do CT-e
                                    xml_completo = etree.tostring(tree, encoding='utf-8').decode('utf-8')
                                    cnpj_cert, _, _, inf_correto, _ = cert_encontrado
                                    
                                    # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local (xmls/) + TODOS os perfis ativos
                                    nome_cert = db.get_cert_nome_by_informante(inf_correto)
                                    
                                    # Salva em backup local (xmls/)
                                    resultado_local = salvar_xml_por_certificado(xml_completo, cnpj_cert, pasta_base="xmls")
                                    caminho_xml = resultado_local[0] if isinstance(resultado_local, tuple) else resultado_local
                                    
                                    # Salva em TODOS os perfis ativos configurados (pasta_base=None)
                                    salvar_xml_por_certificado(xml_completo, cnpj_cert, pasta_base=None, nome_certificado=nome_cert)
                                    
                                    # Registra no banco COM o caminho
                                    if caminho_xml:
                                        db.registrar_xml(chave, inf_correto, caminho_xml)
                                    else:
                                        db.registrar_xml(chave, inf_correto)
                                    
                                    encontradas += 1
                                    print(f"[SUCCESS] CT-e autorizado e registrado com certificado {inf_correto}!")
                                    
                                    # Processa e salva dados detalhados do CT-e
                                    try:
                                        extrair_nota_detalhada(xml_completo, parser, db, chave, inf_correto)
                                        print(f"[DEBUG] Dados detalhados do CT-e salvos no banco")
                                    except Exception as e_extract:
                                        print(f"[AVISO] Erro ao extrair dados detalhados do CT-e: {e_extract}")
                                else:
                                    nao_encontradas += 1
                                    erros.append(f"{chave}: {cStat} - {xMotivo}")
                                    print(f"[ERRO] CT-e não autorizado: {cStat} - {xMotivo}")
                            else:
                                # Tenta extrair erro da consulta
                                cStat = tree.findtext('.//cte:cStat', namespaces=NS) or ''
                                xMotivo = tree.findtext('.//cte:xMotivo', namespaces=NS) or 'Protocolo CT-e não encontrado'
                                print(f"[ERRO] protCTe não encontrado. cStat={cStat}, xMotivo={xMotivo}")
                                nao_encontradas += 1
                                erros.append(f"{chave}: {cStat} - {xMotivo}")
                        else:
                            # Namespace NF-e
                            NS = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                            prot = tree.find('.//nfe:protNFe', namespaces=NS)
                            print(f"[DEBUG] protNFe encontrado: {prot is not None}")
                            
                            if prot is not None:
                                ch = prot.findtext('nfe:infProt/nfe:chNFe', namespaces=NS) or ''
                                cStat = prot.findtext('nfe:infProt/nfe:cStat', namespaces=NS) or ''
                                xMotivo = prot.findtext('nfe:infProt/nfe:xMotivo', namespaces=NS) or ''
                                print(f"[DEBUG] Protocolo: chave={ch}, cStat={cStat}, xMotivo={xMotivo}")
                                
                                # Salva status
                                if ch and cStat:
                                    db.set_nf_status(ch, cStat, xMotivo)
                                    print(f"[DEBUG] Status salvo no banco")
                                
                                # Se autorizada
                                if cStat in ['100', '101', '110', '150', '301', '302']:
                                    # Salva XML da NFe
                                    xml_completo = etree.tostring(tree, encoding='utf-8').decode('utf-8')
                                    cnpj_cert, _, _, inf_correto, _ = cert_encontrado
                                    
                                    # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local (xmls/) + TODOS os perfis ativos
                                    nome_cert = db.get_cert_nome_by_informante(inf_correto)
                                    
                                    # Salva em backup local (xmls/)
                                    resultado_local = salvar_xml_por_certificado(xml_completo, cnpj_cert, pasta_base="xmls")
                                    caminho_xml = resultado_local[0] if isinstance(resultado_local, tuple) else resultado_local
                                    
                                    # Salva em TODOS os perfis ativos configurados (pasta_base=None)
                                    salvar_xml_por_certificado(xml_completo, cnpj_cert, pasta_base=None, nome_certificado=nome_cert)
                                    
                                    # Registra no banco COM o caminho
                                    if caminho_xml:
                                        db.registrar_xml(chave, inf_correto, caminho_xml)
                                    else:
                                        db.registrar_xml(chave, inf_correto)
                                    
                                    encontradas += 1
                                    print(f"[SUCCESS] Nota autorizada e registrada com certificado {inf_correto}!")
                                    
                                    # 🆕 EXTRAÇÃO COMPLETA DE DADOS (tanto de entrada quanto saída)
                                    try:
                                        print(f"[DEBUG] Extraindo dados completos da NF-e...")
                                        nota_dados = extrair_nota_detalhada(
                                            xml_completo, 
                                            db, 
                                            informante=inf_correto,
                                            nsu_documento=""  # Notas consultadas por chave não têm NSU
                                        )
                                        
                                        # Salva dados completos no banco
                                        db.salvar_nota_detalhada(nota_dados)
                                        
                                        # Mostra informações extraídas
                                        num = nota_dados.get('numero', 'N/A')
                                        emit = nota_dados.get('nome_emitente', 'N/A')
                                        cnpj_emit = nota_dados.get('cnpj_emitente', 'N/A')
                                        valor = nota_dados.get('valor', 'N/A')
                                        data = nota_dados.get('data_emissao', 'N/A')
                                        
                                        print(f"[DEBUG] Dados completos salvos:")
                                        print(f"  • Número: {num}")
                                        print(f"  • Emitente: {emit} (CNPJ: {cnpj_emit})")
                                        print(f"  • Valor: R$ {valor}")
                                        print(f"  • Data: {data}")
                                        print(f"  • Status: COMPLETO (XML salvo)")
                                        
                                        # 🎯 IDENTIFICA SE É DE SAÍDA OU ENTRADA
                                        # Verifica se o emitente é um dos certificados cadastrados
                                        certs_cnpjs = [c[0] for c in certificados]  # Lista de CNPJs dos certificados
                                        cnpj_emit_limpo = ''.join(c for c in cnpj_emit if c.isdigit())
                                        
                                        if cnpj_emit_limpo in certs_cnpjs:
                                            print(f"[INFO] ✅ NF-e de SAÍDA detectada (emitente = certificado)")
                                            print(f"[INFO] 📂 Aparecerá na aba 'Emitidos pela empresa'")
                                        else:
                                            print(f"[INFO] ✅ NF-e de ENTRADA detectada (emitente ≠ certificado)")
                                            print(f"[INFO] 📂 Aparecerá na aba principal")
                                        
                                    except Exception as e_extract:
                                        print(f"[AVISO] Erro ao extrair dados completos da NF-e: {e_extract}")
                                        import traceback
                                        traceback.print_exc()
                                else:
                                    nao_encontradas += 1
                                    erros.append(f"{chave}: {cStat} - {xMotivo}")
                                    print(f"[ERRO] Nota não autorizada: {cStat} - {xMotivo}")
                            else:
                                # Tenta extrair erro da consulta
                                cStat = tree.findtext('.//nfe:cStat', namespaces=NS) or ''
                                xMotivo = tree.findtext('.//nfe:xMotivo', namespaces=NS) or 'Protocolo não encontrado'
                                print(f"[ERRO] protNFe não encontrado. cStat={cStat}, xMotivo={xMotivo}")
                                nao_encontradas += 1
                                erros.append(f"{chave}: {cStat} - {xMotivo}")
                    
                    except Exception as e:
                        print(f"[EXCEPTION] Erro ao processar XML: {e}")
                        import traceback
                        traceback.print_exc()
                        nao_encontradas += 1
                        erros.append(f"{chave}: Erro ao processar XML - {str(e)}")
                else:
                    nao_encontradas += 1
                    erros.append(f"{chave}: Falha na comunicação")
                    
            except Exception as e:
                nao_encontradas += 1
                erros.append(f"{chave}: {str(e)}")
        
        print(f"\n[BUSCA POR CHAVE] Loop finalizado")
        print(f"[BUSCA POR CHAVE] Encontradas: {encontradas}, Não encontradas: {nao_encontradas}")
        print(f"[BUSCA POR CHAVE] Total processado: {encontradas + nao_encontradas} de {len(chaves)}")
        
        progress.setValue(len(chaves))
        
        # Atualiza tabela
        self.refresh_all()
        
        # Mostra resultado
        mensagem = f"✅ Busca concluída!\n\n"
        mensagem += f"📥 Encontradas e salvas: {encontradas}\n"
        mensagem += f"❌ Não encontradas/erro: {nao_encontradas}\n"
        mensagem += f"📊 Total processado: {encontradas + nao_encontradas} de {len(chaves)} chaves\n\n"
        mensagem += f"━━━━━━━━━━━━━━━━━━━━━━\n"
        mensagem += f"📂 Onde encontrar as notas:\n"
        mensagem += f"  • Notas de SAÍDA → aba 'Emitidos pela empresa'\n"
        mensagem += f"  • Notas de ENTRADA → aba principal\n"
        mensagem += f"  • XMLs salvos → pasta xmls/ e perfis ativos"
        
        if erros and len(erros) <= 10:
            mensagem += "\n\n⚠️ Erros:\n" + "\n".join(erros[:10])
        elif erros:
            mensagem += f"\n\n⚠️ {len(erros)} erros encontrados (veja o log para detalhes)"
        
        QMessageBox.information(self, "Busca por Chave - Resultado", mensagem)

    def _listar_certificados_windows(self):
        """Lista certificados instalados no Windows (DEPRECADO - usar seleção de .pfx)."""
        # Função mantida por compatibilidade mas não mais utilizada
        return []
    
    def _selecionar_certificado_pfx(self):
        """Abre dialog para selecionar arquivo .pfx do certificado."""
        from PyQt5.QtWidgets import QFileDialog
        
        arquivo, _ = QFileDialog.getOpenFileName(
            self,
            "Selecione o Certificado Digital",
            "",
            "Arquivos de Certificado (*.pfx *.p12);;Todos os Arquivos (*.*)"
        )
        
        if arquivo:
            self.manifestacao_pfx_path.setText(arquivo)
            self.manifestacao_pfx_path.setStyleSheet("background-color: #d4edda; border: 2px solid #28a745;")

    def abrir_manifestacao(self):
        """Lista certificados instalados no Windows Certificate Store."""
        certificados = []
        
        try:
            import sys
            if sys.platform != 'win32':
                print("[CERTIFICADOS] Sistema não é Windows")
                return certificados
            
            print("\n[CERTIFICADOS] Iniciando listagem de certificados do Windows...")
            
            # Tenta usar wincertstore + cryptography
            try:
                import wincertstore
                from cryptography import x509
                from cryptography.hazmat.backends import default_backend
                import re
                
                print("[CERTIFICADOS] Bibliotecas wincertstore e cryptography carregadas")
                
                # Tenta acessar certificados de CURRENT_USER e LOCAL_MACHINE
                import ctypes
                from ctypes import wintypes
                
                # Lista de stores e locais para verificar
                stores_to_check = [
                    ("MY", 0x00010000, "CURRENT_USER"),  # CERT_SYSTEM_STORE_CURRENT_USER
                    ("MY", 0x00020000, "LOCAL_MACHINE"), # CERT_SYSTEM_STORE_LOCAL_MACHINE
                ]
                
                for storename, location_flag, location_name in stores_to_check:
                    print(f"\n[CERTIFICADOS] Verificando store: {storename} em {location_name}")
                    print(f"[CERTIFICADOS] IMPORTANTE: Procurando certificados A1 (PFX instalados)")
                    
                    try:
                        # Para CURRENT_USER, usa wincertstore (mais simples)
                        if location_name == "CURRENT_USER":
                            # Conta TODOS os certificados antes de filtrar
                            total_certs_in_store = 0
                            with wincertstore.CertSystemStore(storename) as store_counter:
                                for _ in store_counter.itercerts():
                                    total_certs_in_store += 1
                            
                            print(f"[CERTIFICADOS] Total de certificados na store {storename} ({location_name}): {total_certs_in_store}")
                            
                            with wincertstore.CertSystemStore(storename) as store:
                                cert_count = 0
                                for cert_context in store.itercerts():
                                    cert_count += 1
                                    print(f"\n[CERTIFICADOS] === Processando certificado {cert_count} da store {storename} ===")
                                    try:
                                        # Converte CERT_CONTEXT para bytes
                                        cert_bytes = cert_context.get_encoded()
                                        cert = x509.load_der_x509_certificate(cert_bytes, default_backend())
                                        
                                        # LOG DETALHADO: Mostra TODOS os campos do Subject
                                        print(f"[CERTIFICADOS]   === SUBJECT COMPLETO ===")
                                        for attr in cert.subject:
                                            print(f"[CERTIFICADOS]     {attr.oid._name}: {attr.value}")
                                        print(f"[CERTIFICADOS]   === FIM SUBJECT ===")
                                        
                                        # Extrai CN (Common Name)
                                        cn = None
                                        try:
                                            cn_list = cert.subject.get_attributes_for_oid(x509.oid.NameOID.COMMON_NAME)
                                            if cn_list:
                                                cn = cn_list[0].value
                                                print(f"[CERTIFICADOS]   CN encontrado: {cn}")
                                        except Exception as e:
                                            print(f"[CERTIFICADOS]   Erro ao extrair CN: {e}")
                                        
                                        # Extrai Organization (O)
                                        org = None
                                        try:
                                            org_list = cert.subject.get_attributes_for_oid(x509.oid.NameOID.ORGANIZATION_NAME)
                                            if org_list:
                                                org = org_list[0].value
                                                print(f"[CERTIFICADOS]   Organization encontrada: {org}")
                                        except Exception as e:
                                            print(f"[CERTIFICADOS]   Erro ao extrair Organization: {e}")
                                        
                                        # Extrai Serial Number do subject
                                        serial_number = None
                                        try:
                                            sn_list = cert.subject.get_attributes_for_oid(x509.oid.NameOID.SERIAL_NUMBER)
                                            if sn_list:
                                                serial_number = sn_list[0].value
                                                print(f"[CERTIFICADOS]   Serial Number encontrado: {serial_number}")
                                        except Exception as e:
                                            print(f"[CERTIFICADOS]   Erro ao extrair Serial Number: {e}")
                                        
                                        # Tenta extrair CNPJ/CPF de várias fontes
                                        cnpj = None
                                        
                                        print(f"[CERTIFICADOS]   Tentando extrair CNPJ/CPF...")
                                        
                                        # 1. Do serialNumber
                                        if serial_number:
                                            nums = ''.join(c for c in str(serial_number) if c.isdigit())
                                            print(f"[CERTIFICADOS]   - SerialNumber (números): {nums}")
                                            if len(nums) >= 14:
                                                cnpj = nums[:14]
                                                print(f"[CERTIFICADOS]   - CNPJ extraído do SerialNumber: {cnpj}")
                                            elif len(nums) == 11:
                                                cnpj = nums  # CPF
                                                print(f"[CERTIFICADOS]   - CPF extraído do SerialNumber: {cnpj}")
                                        
                                        # 2. Do CN
                                        if not cnpj and cn:
                                            # Procura padrão de CNPJ (14 dígitos)
                                            cnpj_match = re.search(r'\d{14}', cn)
                                            if cnpj_match:
                                                cnpj = cnpj_match.group()
                                                print(f"[CERTIFICADOS]   - CNPJ extraído do CN: {cnpj}")
                                            else:
                                                # Procura CPF (11 dígitos)
                                                cpf_match = re.search(r'\d{11}', cn)
                                                if cpf_match:
                                                    cnpj = cpf_match.group()
                                                    print(f"[CERTIFICADOS]   - CPF extraído do CN: {cnpj}")
                                        
                                        # 3. Da organização
                                        if not cnpj and org:
                                            nums = ''.join(c for c in str(org) if c.isdigit())
                                            print(f"[CERTIFICADOS]   - Organization (números): {nums}")
                                            if len(nums) >= 14:
                                                cnpj = nums[:14]
                                                print(f"[CERTIFICADOS]   - CNPJ extraído da Organization: {cnpj}")
                                        
                                        # Formata CNPJ/CPF
                                        cnpj_formatado = "N/A"
                                        if cnpj:
                                            if len(cnpj) == 14:
                                                cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
                                            elif len(cnpj) == 11:
                                                cnpj_formatado = f"{cnpj[:3]}.{cnpj[3:6]}.{cnpj[6:9]}-{cnpj[9:11]}"
                                            else:
                                                cnpj_formatado = cnpj
                                        
                                        # Verifica validade
                                        import datetime
                                        hoje = datetime.datetime.now()
                                        esta_valido = cert.not_valid_before <= hoje <= cert.not_valid_after
                                        
                                        validade = cert.not_valid_after.strftime("%d/%m/%Y")
                                        
                                        # Emissor
                                        emissor = None
                                        try:
                                            emissor_list = cert.issuer.get_attributes_for_oid(x509.oid.NameOID.COMMON_NAME)
                                            if emissor_list:
                                                emissor = emissor_list[0].value
                                        except:
                                            pass
                                        
                                        # Verifica se tem uso para assinatura digital
                                        tem_assinatura = False
                                        try:
                                            key_usage = cert.extensions.get_extension_for_oid(x509.oid.ExtensionOID.KEY_USAGE)
                                            tem_assinatura = key_usage.value.digital_signature
                                            print(f"[CERTIFICADOS]   Key Usage: digital_signature={tem_assinatura}")
                                        except Exception as ku_err:
                                            # Se não tem extensão Key Usage, assume que pode assinar
                                            # Muitos certificados A1 não têm essa extensão explícita
                                            tem_assinatura = True
                                            print(f"[CERTIFICADOS]   Key Usage não encontrado (assumindo True): {ku_err}")
                                        
                                        # Nome para exibição
                                        nome_exibicao = cn or org or "Certificado sem nome"
                                        
                                        print(f"[CERTIFICADOS]   Resumo: {nome_exibicao[:50]}")
                                        print(f"[CERTIFICADOS]   - CNPJ/CPF: {cnpj_formatado}")
                                        print(f"[CERTIFICADOS]   - Validade: {validade} (Válido: {esta_valido})")
                                        print(f"[CERTIFICADOS]   - Emissor: {emissor or 'N/A'}")
                                        print(f"[CERTIFICADOS]   - Store: {storename}")
                                        print(f"[CERTIFICADOS]   - Pode assinar: {tem_assinatura}")
                                        
                                        # AJUSTE: Aceita certificados da store MY que estejam válidos
                                        # Remove a exigência de ter extensão de assinatura digital explícita
                                        if esta_valido and storename == "MY":
                                            cert_info = {
                                                'cn': nome_exibicao,
                                                'org': org or 'N/A',
                                                'cnpj': cnpj_formatado,
                                                'cnpj_raw': cnpj or '',
                                                'validade': validade,
                                                'emissor': emissor or 'Desconhecido',
                                                'cert_bytes': cert_bytes,
                                                'thumbprint': cert.fingerprint(cert.signature_hash_algorithm).hex(),
                                                'store': storename
                                            }
                                            certificados.append(cert_info)
                                            print(f"[CERTIFICADOS] ✓✓✓ ADICIONADO: {nome_exibicao[:50]} - CNPJ: {cnpj_formatado}")
                                        else:
                                            motivo = []
                                            if not esta_valido:
                                                motivo.append(f"vencido (expira em {validade})")
                                            if storename != "MY":
                                                motivo.append(f"store {storename} (precisa ser MY)")
                                            print(f"[CERTIFICADOS] ✗✗✗ IGNORADO ({', '.join(motivo)}): {nome_exibicao[:50]}")
                                    
                                    except Exception as e:
                                        print(f"[CERTIFICADOS] Erro ao processar certificado: {e}")
                                        import traceback
                                        traceback.print_exc()
                                        continue
                            
                            print(f"[CERTIFICADOS] Store {storename} ({location_name}): processados {cert_count} certificados")
                        
                        else:
                            # LOCAL_MACHINE - usa ctypes para acessar
                            print(f"[CERTIFICADOS] Processando {location_name} com ctypes...")
                            
                            # Abre store LOCAL_MACHINE
                            crypt32 = ctypes.windll.crypt32
                            CERT_STORE_PROV_SYSTEM = 10
                            CERT_STORE_OPEN_EXISTING_FLAG = 0x00004000
                            
                            store_handle = crypt32.CertOpenStore(
                                CERT_STORE_PROV_SYSTEM,
                                0,
                                None,
                                location_flag | CERT_STORE_OPEN_EXISTING_FLAG,
                                storename
                            )
                            
                            if not store_handle:
                                print(f"[CERTIFICADOS] Falha ao abrir {location_name}\\{storename}")
                                continue
                            
                            # Enumera certificados
                            cert_count = 0
                            cert_context_ptr = None
                            
                            # Define a estrutura CERT_CONTEXT
                            class CERT_CONTEXT(ctypes.Structure):
                                _fields_ = [
                                    ("dwCertEncodingType", wintypes.DWORD),
                                    ("pbCertEncoded", ctypes.POINTER(ctypes.c_byte)),
                                    ("cbCertEncoded", wintypes.DWORD),
                                    ("pCertInfo", ctypes.c_void_p),
                                    ("hCertStore", ctypes.c_void_p),
                                ]
                            
                            while True:
                                cert_context_ptr = crypt32.CertEnumCertificatesInStore(store_handle, cert_context_ptr)
                                if not cert_context_ptr:
                                    break
                                
                                cert_count += 1
                                print(f"\n[CERTIFICADOS] === Processando certificado {cert_count} da store {storename} ({location_name}) ===")
                                
                                try:
                                    # Lê a estrutura CERT_CONTEXT
                                    cert_context = ctypes.cast(cert_context_ptr, ctypes.POINTER(CERT_CONTEXT)).contents
                                    
                                    # Extrai os bytes do certificado
                                    cert_bytes = ctypes.string_at(cert_context.pbCertEncoded, cert_context.cbCertEncoded)
                                    cert = x509.load_der_x509_certificate(cert_bytes, default_backend())
                                    
                                    # LOG DETALHADO: Mostra TODOS os campos do Subject
                                    print(f"[CERTIFICADOS]   === SUBJECT COMPLETO ===")
                                    for attr in cert.subject:
                                        print(f"[CERTIFICADOS]     {attr.oid._name}: {attr.value}")
                                    print(f"[CERTIFICADOS]   === FIM SUBJECT ===")
                                    
                                    # Processa igual ao CURRENT_USER (mesmo código de extração)
                                    # Extrai CN
                                    cn = None
                                    try:
                                        cn_list = cert.subject.get_attributes_for_oid(x509.oid.NameOID.COMMON_NAME)
                                        if cn_list:
                                            cn = cn_list[0].value
                                            print(f"[CERTIFICADOS]   CN encontrado: {cn}")
                                    except Exception as e:
                                        print(f"[CERTIFICADOS]   Erro ao extrair CN: {e}")
                                    
                                    # Verifica validade
                                    import datetime
                                    hoje = datetime.datetime.now()
                                    esta_valido = cert.not_valid_before <= hoje <= cert.not_valid_after
                                    validade = cert.not_valid_after.strftime("%d/%m/%Y")
                                    
                                    # Nome para exibição
                                    nome_exibicao = cn or "Certificado sem nome"
                                    
                                    print(f"[CERTIFICADOS]   Resumo: {nome_exibicao[:50]}")
                                    print(f"[CERTIFICADOS]   - Validade: {validade} (Válido: {esta_valido})")
                                    print(f"[CERTIFICADOS]   - Store: {storename} ({location_name})")
                                    
                                    if esta_valido:
                                        cert_info = {
                                            'cn': nome_exibicao,
                                            'org': 'N/A',
                                            'cnpj': 'N/A',
                                            'cnpj_raw': '',
                                            'validade': validade,
                                            'emissor': 'Desconhecido',
                                            'cert_bytes': cert_bytes,
                                            'thumbprint': cert.fingerprint(cert.signature_hash_algorithm).hex(),
                                            'store': f"{storename} ({location_name})"
                                        }
                                        certificados.append(cert_info)
                                        print(f"[CERTIFICADOS] ✓✓✓ ADICIONADO: {nome_exibicao[:50]}")
                                    
                                except Exception as e:
                                    print(f"[CERTIFICADOS] Erro ao processar certificado LOCAL_MACHINE: {e}")
                                    import traceback
                                    traceback.print_exc()
                                    continue
                            
                            crypt32.CertCloseStore(store_handle, 0)
                            print(f"[CERTIFICADOS] Store {storename} ({location_name}): processados {cert_count} certificados")
                            
                    except Exception as e:
                        print(f"[CERTIFICADOS] Erro ao acessar store {storename} em {location_name}: {e}")
                        import traceback
                        traceback.print_exc()
                
                print(f"\n[CERTIFICADOS] Total de certificados válidos adicionados: {len(certificados)}")
                        
            except ImportError as ie:
                print(f"[CERTIFICADOS] Bibliotecas não disponíveis: {ie}")
                print("[CERTIFICADOS] Instale: pip install wincertstore cryptography")
                    
        except Exception as e:
            print(f"[CERTIFICADOS] Erro geral ao listar certificados: {e}")
            import traceback
            traceback.print_exc()
        
        return certificados

    def abrir_manifestacao(self):
        """Abre janela standalone para manifestação de documentos (NF-e/CT-e)."""
        dialog = QDialog(self)
        dialog.setWindowTitle("📨 Manifestação de Documentos")
        dialog.setMinimumWidth(700)
        dialog.setMinimumHeight(600)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        
        # === SEÇÃO: Certificado Digital ===
        cert_group = QGroupBox("🔐 Certificado Digital (.pfx)")
        cert_layout = QVBoxLayout()
        
        # Seleção de arquivo .pfx
        pfx_label = QLabel("Arquivo do Certificado:")
        pfx_label.setStyleSheet("font-weight: bold;")
        cert_layout.addWidget(pfx_label)
        
        pfx_h_layout = QHBoxLayout()
        self.manifestacao_pfx_path = QLineEdit()
        self.manifestacao_pfx_path.setPlaceholderText("Selecione o arquivo .pfx ou .p12 do certificado digital...")
        self.manifestacao_pfx_path.setReadOnly(True)
        self.manifestacao_pfx_path.setMinimumHeight(35)
        
        btn_selecionar_pfx = QPushButton("📁 Selecionar")
        btn_selecionar_pfx.setMinimumHeight(35)
        btn_selecionar_pfx.setMinimumWidth(120)
        btn_selecionar_pfx.clicked.connect(lambda: self._selecionar_certificado_pfx())
        
        pfx_h_layout.addWidget(self.manifestacao_pfx_path, 3)
        pfx_h_layout.addWidget(btn_selecionar_pfx, 1)
        cert_layout.addLayout(pfx_h_layout)
        
        # Campo de senha
        senha_label = QLabel("🔑 Senha do Certificado:")
        senha_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        cert_layout.addWidget(senha_label)
        
        self.manifestacao_senha = QLineEdit()
        self.manifestacao_senha.setPlaceholderText("Digite a senha do certificado...")
        self.manifestacao_senha.setEchoMode(QLineEdit.Password)
        self.manifestacao_senha.setMinimumHeight(35)
        cert_layout.addWidget(self.manifestacao_senha)
        
        # Checkbox para mostrar senha
        self.manifestacao_mostrar_senha = QCheckBox("Mostrar senha")
        self.manifestacao_mostrar_senha.stateChanged.connect(
            lambda state: self.manifestacao_senha.setEchoMode(
                QLineEdit.Normal if state else QLineEdit.Password
            )
        )
        cert_layout.addWidget(self.manifestacao_mostrar_senha)
        
        cert_group.setLayout(cert_layout)
        layout.addWidget(cert_group)
        
        # === SEÇÃO: Chave de Acesso ===
        chave_group = QGroupBox("🔑 Chave de Acesso do Documento")
        chave_layout = QVBoxLayout()
        
        self.manifestacao_chave_edit = QLineEdit()
        self.manifestacao_chave_edit.setPlaceholderText("Digite a chave de acesso (44 dígitos)")
        self.manifestacao_chave_edit.setMaxLength(44)
        self.manifestacao_chave_edit.setMinimumHeight(35)
        
        # Validação e detecção de tipo
        def validar_chave():
            chave = self.manifestacao_chave_edit.text().strip()
            if len(chave) == 44 and chave.isdigit():
                modelo = chave[20:22]
                if modelo == '55':
                    self.manifestacao_tipo_combo.setCurrentText("NF-e (Nota Fiscal Eletrônica)")
                    self.manifestacao_chave_edit.setStyleSheet("background-color: #d4edda; border: 2px solid #28a745;")
                elif modelo == '57':
                    self.manifestacao_tipo_combo.setCurrentText("CT-e (Conhecimento de Transporte)")
                    self.manifestacao_chave_edit.setStyleSheet("background-color: #d1ecf1; border: 2px solid #17a2b8;")
                else:
                    self.manifestacao_chave_edit.setStyleSheet("background-color: #fff3cd; border: 2px solid #ffc107;")
            elif len(chave) > 0:
                self.manifestacao_chave_edit.setStyleSheet("background-color: #f8d7da; border: 2px solid #dc3545;")
            else:
                self.manifestacao_chave_edit.setStyleSheet("")
        
        self.manifestacao_chave_edit.textChanged.connect(validar_chave)
        
        chave_layout.addWidget(QLabel("Digite a chave de acesso:"))
        chave_layout.addWidget(self.manifestacao_chave_edit)
        chave_group.setLayout(chave_layout)
        layout.addWidget(chave_group)
        
        # === SEÇÃO: Justificativa ===
        justificativa_group = QGroupBox("📝 Justificativa")
        justificativa_layout = QVBoxLayout()
        
        justificativa_info = QLabel(
            "ℹ️ Obrigatória para eventos: Desconhecimento da Operação e Operação não Realizada.\n"
            "Mínimo de 15 caracteres."
        )
        justificativa_info.setStyleSheet("color: #7f8c8d; font-size: 9pt; font-style: italic;")
        justificativa_layout.addWidget(justificativa_info)
        
        self.manifestacao_justificativa = QTextEdit()
        self.manifestacao_justificativa.setPlaceholderText(
            "Digite a justificativa para a manifestação (ex: mercadoria não solicitada, "
            "dados divergentes, operação cancelada, etc.)"
        )
        self.manifestacao_justificativa.setMinimumHeight(80)
        self.manifestacao_justificativa.setMaximumHeight(120)
        
        # Contador de caracteres
        self.manifestacao_justificativa_contador = QLabel("0 caracteres")
        self.manifestacao_justificativa_contador.setStyleSheet("color: #95a5a6; font-size: 9pt;")
        
        def atualizar_contador():
            texto = self.manifestacao_justificativa.toPlainText()
            count = len(texto)
            self.manifestacao_justificativa_contador.setText(f"{count} caracteres")
            
            if count >= 15:
                self.manifestacao_justificativa_contador.setStyleSheet("color: #27ae60; font-weight: bold; font-size: 9pt;")
            elif count > 0:
                self.manifestacao_justificativa_contador.setStyleSheet("color: #e67e22; font-size: 9pt;")
            else:
                self.manifestacao_justificativa_contador.setStyleSheet("color: #95a5a6; font-size: 9pt;")
        
        self.manifestacao_justificativa.textChanged.connect(atualizar_contador)
        
        justificativa_layout.addWidget(self.manifestacao_justificativa)
        justificativa_layout.addWidget(self.manifestacao_justificativa_contador)
        justificativa_group.setLayout(justificativa_layout)
        layout.addWidget(justificativa_group)
        
        # === SEÇÃO: Tipo de Documento ===
        tipo_group = QGroupBox("📄 Tipo de Documento")
        tipo_layout = QVBoxLayout()
        
        self.manifestacao_tipo_combo = QComboBox()
        self.manifestacao_tipo_combo.setMinimumHeight(35)
        self.manifestacao_tipo_combo.addItem("NF-e (Nota Fiscal Eletrônica)", "NFE")
        self.manifestacao_tipo_combo.addItem("CT-e (Conhecimento de Transporte)", "CTE")
        
        # Atualiza eventos ao mudar tipo
        def atualizar_eventos():
            # Remove eventos antigos
            for i in reversed(range(eventos_layout.count())):
                widget = eventos_layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()
            
            tipo = self.manifestacao_tipo_combo.currentData()
            if tipo == "NFE":
                eventos = [
                    {"codigo": "210210", "nome": "Ciência da Operação", "cor": "#3498db", "icon": "👁️", 
                     "desc": "Declara que tomou conhecimento da operação"},
                    {"codigo": "210200", "nome": "Confirmação da Operação", "cor": "#27ae60", "icon": "💡", 
                     "desc": "Confirma que a operação foi realizada"},
                    {"codigo": "210220", "nome": "Desconhecimento da Operação", "cor": "#e74c3c", "icon": "🛑", 
                     "desc": "Declara que não reconhece a operação"},
                    {"codigo": "210240", "nome": "Operação não Realizada", "cor": "#f39c12", "icon": "⭕", 
                     "desc": "Declara que a operação não ocorreu"}
                ]
            else:  # CTE
                eventos = [
                    {"codigo": "610110", "nome": "Desacordo do Serviço", "cor": "#e74c3c", "icon": "🛑", 
                     "desc": "Declara desacordo com o serviço prestado"},
                    {"codigo": "610112", "nome": "Cancelar Desacordo", "cor": "#f39c12", "icon": "↩️", 
                     "desc": "Cancela declaração de desacordo anterior"}
                ]
            
            for evento in eventos:
                btn = QPushButton(f"{evento['icon']} {evento['nome']}")
                btn.setMinimumHeight(60)
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {evento['cor']};
                        color: white;
                        border: 2px solid {evento['cor']};
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        padding: 10px;
                    }}
                    QPushButton:hover {{
                        background-color: {self._lighten_color(evento['cor'])};
                    }}
                    QPushButton:pressed {{
                        background-color: {self._darken_color(evento['cor'])};
                    }}
                """)
                btn.setToolTip(evento['desc'])
                btn.clicked.connect(lambda checked, e=evento: self._manifestar_standalone(e))
                eventos_layout.addWidget(btn)
        
        self.manifestacao_tipo_combo.currentIndexChanged.connect(atualizar_eventos)
        
        tipo_layout.addWidget(QLabel("Selecione o tipo de documento:"))
        tipo_layout.addWidget(self.manifestacao_tipo_combo)
        tipo_group.setLayout(tipo_layout)
        layout.addWidget(tipo_group)
        
        # === SEÇÃO: Eventos de Manifestação ===
        eventos_group = QGroupBox("📨 Eventos de Manifestação")
        eventos_layout = QVBoxLayout()
        eventos_group.setLayout(eventos_layout)
        layout.addWidget(eventos_group)
        
        # Popula eventos iniciais (NF-e)
        atualizar_eventos()
        
        # === BOTÕES DE AÇÃO ===
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_fechar = QPushButton("❌ Fechar")
        btn_fechar.setMinimumHeight(40)
        btn_fechar.setMinimumWidth(120)
        btn_fechar.clicked.connect(dialog.close)
        btn_fechar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border-radius: 5px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        
        btn_layout.addWidget(btn_fechar)
        layout.addLayout(btn_layout)
        
        # Salva referência do dialog para usar em _manifestar_standalone
        self.manifestacao_dialog = dialog
        
        dialog.exec_()

    def _manifestar_standalone(self, evento):
        """Envia manifestação a partir da janela standalone."""
        chave = self.manifestacao_chave_edit.text().strip()
        
        # Validações
        if not chave or len(chave) != 44 or not chave.isdigit():
            QMessageBox.warning(self.manifestacao_dialog, "Manifestação", 
                              "Digite uma chave de acesso válida (44 dígitos numéricos)!")
            return
        
        # Valida seleção de certificado e senha
        pfx_path = self.manifestacao_pfx_path.text().strip()
        senha = self.manifestacao_senha.text()
        
        if not pfx_path:
            QMessageBox.warning(self.manifestacao_dialog, "Manifestação", 
                              "Selecione o arquivo .pfx do certificado!")
            return
        
        if not senha:
            QMessageBox.warning(self.manifestacao_dialog, "Manifestação", 
                              "Digite a senha do certificado!")
            return
        
        # Carrega certificado do arquivo .pfx
        try:
            from cryptography.hazmat.primitives.serialization import pkcs12
            from cryptography.hazmat.backends import default_backend
            from cryptography.x509.oid import NameOID
            import os
            
            if not os.path.exists(pfx_path):
                QMessageBox.critical(self.manifestacao_dialog, "Erro", 
                                   f"Arquivo não encontrado:\n{pfx_path}")
                return
            
            with open(pfx_path, 'rb') as f:
                pfx_data = f.read()
            
            # Carrega o certificado com a senha
            try:
                private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                    pfx_data, senha.encode(), default_backend()
                )
            except Exception as e:
                QMessageBox.critical(
                    self.manifestacao_dialog, 
                    "Erro no Certificado",
                    f"Não foi possível carregar o certificado.\n\n"
                    f"Verifique se a senha está correta.\n\n"
                    f"Erro: {str(e)}"
                )
                return
            
            if not certificate:
                QMessageBox.critical(self.manifestacao_dialog, "Erro", 
                                   "Nenhum certificado encontrado no arquivo .pfx")
                return
            
            # Extrai informações do certificado
            cn = certificate.subject.get_attributes_for_oid(NameOID.COMMON_NAME)[0].value
            
            # Extrai CNPJ do certificado
            informante = ''
            for attr in certificate.subject:
                # Procura por CNPJ em diversos OIDs possíveis
                if 'CNPJ' in str(attr.oid) or attr.oid.dotted_string == '2.5.4.97':
                    informante = ''.join(filter(str.isdigit, attr.value))
                    break
            
            if not informante:
                # Tenta extrair do CN ou serialNumber
                try:
                    cn_nums = ''.join(filter(str.isdigit, cn))
                    if len(cn_nums) >= 11:  # CPF ou CNPJ
                        informante = cn_nums[:14] if len(cn_nums) >= 14 else cn_nums
                except:
                    pass
            
            if not informante:
                QMessageBox.warning(
                    self.manifestacao_dialog,
                    "CNPJ/CPF não encontrado",
                    f"Não foi possível extrair CNPJ/CPF do certificado.\n\n"
                    f"CN: {cn}\n\n"
                    f"O certificado pode não ser válido para manifestação."
                )
                return
            
            print(f"[MANIFESTAÇÃO] Certificado carregado: {cn}")
            print(f"[MANIFESTAÇÃO] Informante extraído: {informante}")
            
        except Exception as e:
            QMessageBox.critical(
                self.manifestacao_dialog, 
                "Erro ao Carregar Certificado",
                f"Erro inesperado ao carregar o certificado:\n\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
            return
        
        tipo_evento = evento['codigo']
        nome_evento = evento['nome']
        
        # Valida justificativa para eventos que exigem
        justificativa = self.manifestacao_justificativa.toPlainText().strip()
        eventos_requerem_justificativa = ['210220', '210240']  # Desconhecimento e Operação não Realizada
        
        if tipo_evento in eventos_requerem_justificativa:
            if not justificativa or len(justificativa) < 15:
                QMessageBox.warning(
                    self.manifestacao_dialog,
                    "Justificativa Obrigatória",
                    f"O evento '{nome_evento}' requer uma justificativa com no mínimo 15 caracteres.\n\n"
                    f"Caracteres informados: {len(justificativa)}\n"
                    f"Mínimo necessário: 15"
                )
                self.manifestacao_justificativa.setFocus()
                return
        
        print(f"[MANIFESTAÇÃO] Justificativa: {justificativa if justificativa else '(não informada)'}")
        
        # Verifica se já foi manifestada
        ja_manifestada = self.db.check_manifestacao_exists(chave, tipo_evento, informante)
        if ja_manifestada:
            reply = QMessageBox.question(
                self.manifestacao_dialog,
                "Manifestação Duplicada",
                f"Já existe uma manifestação '{nome_evento}' para esta chave.\n\n"
                f"Deseja enviar novamente?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
        
        # Confirmação
        import os
        cert_filename = os.path.basename(pfx_path)
        
        msg_confirmacao = (
            f"Deseja enviar a manifestação?\n\n"
            f"🔑 Chave: {chave}\n"
            f"📨 Evento: {nome_evento}\n"
            f"🔐 Certificado: {cert_filename}\n"
            f"📄 CN: {cn}\n"
            f"🏢 CNPJ/CPF: {informante}"
        )
        
        if justificativa:
            msg_confirmacao += f"\n📝 Justificativa: {justificativa[:50]}{'...' if len(justificativa) > 50 else ''}"
        
        reply = QMessageBox.question(
            self.manifestacao_dialog,
            "Confirmar Manifestação",
            msg_confirmacao,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.No:
            return
        
        # Progress dialog
        progress = QProgressDialog("Enviando manifestação...", "Cancelar", 0, 0, self.manifestacao_dialog)
        progress.setWindowTitle("Manifestação")
        progress.setWindowModality(Qt.WindowModal)
        progress.setCancelButton(None)
        progress.show()
        QApplication.processEvents()
        
        try:
            # Envia manifestação REAL para SEFAZ
            from modules.manifestacao_service import ManifestacaoService
            
            logger.info(f"[MANIFESTAÇÃO] Iniciando envio real para SEFAZ")
            logger.info(f"[MANIFESTAÇÃO] Chave: {chave}")
            logger.info(f"[MANIFESTAÇÃO] Evento: {tipo_evento} - {nome_evento}")
            logger.info(f"[MANIFESTAÇÃO] Informante: {informante}")
            
            try:
                # Cria serviço de manifestação
                manifest_service = ManifestacaoService(pfx_path, senha)
                
                # Envia para SEFAZ
                sucesso, protocolo, mensagem, xml_resposta = manifest_service.enviar_manifestacao(
                    chave=chave,
                    tipo_evento=tipo_evento,
                    cnpj_destinatario=informante,
                    justificativa=justificativa if justificativa else None
                )
                
                progress.close()
                
                if not sucesso:
                    QMessageBox.critical(
                        self.manifestacao_dialog,
                        "Erro na Manifestação",
                        f"❌ A SEFAZ rejeitou a manifestação:\n\n{mensagem}"
                    )
                    return
                
                logger.info(f"[MANIFESTAÇÃO] ✅ Sucesso! Protocolo: {protocolo}")
                
            except Exception as e:
                progress.close()
                QMessageBox.critical(
                    self.manifestacao_dialog,
                    "Erro ao Enviar",
                    f"❌ Erro ao comunicar com SEFAZ:\n\n{str(e)}\n\nVerifique:\n"
                    f"- Certificado digital válido\n"
                    f"- Conexão com internet\n"
                    f"- Chave de acesso correta"
                )
                import traceback
                traceback.print_exc()
                return
            
            status = "SUCESSO"
            
            # Registra no banco
            self.db.register_manifestacao(chave, tipo_evento, informante, status, protocolo)
            
            progress.close()
            
            # Salva automaticamente os arquivos XML e PDF
            import os
            pasta_base = os.path.join(os.getcwd(), "xmls", "Manifestação manual")
            os.makedirs(pasta_base, exist_ok=True)
            
            # ✅ NÃO salva retEnvEvento (apenas confirmação, não contém nota)
            # Manifestação registrada no banco, XML completo será baixado separadamente
            arquivos_salvos = True  # Considera salvo pois foi registrado no banco
            # Removido: _salvar_arquivos_manifestacao_automatico(..., xml_resposta) - causava SEM_NUMERO-SEM_NOME.xml
            
            # Atualiza tabelas
            self.refresh_table()
            self.refresh_emitidos_table()
            
            # Mensagem de sucesso
            msg_sucesso = f"✅ Manifestação '{nome_evento}' enviada com sucesso!\n\n"
            msg_sucesso += f"📋 Protocolo: {protocolo}\n"
            msg_sucesso += f"💬 {mensagem}\n\n"
            
            if arquivos_salvos:
                msg_sucesso += f"📁 Arquivos salvos em:\n{pasta_base}"
            
            QMessageBox.information(
                self.manifestacao_dialog,
                "Manifestação Enviada",
                msg_sucesso
            )
            
        except Exception as e:
            progress.close()
            QMessageBox.critical(
                self.manifestacao_dialog,
                "Erro",
                f"❌ Erro ao enviar manifestação:\n\n{str(e)}"
            )

    def _salvar_arquivos_manifestacao_automatico(self, pasta, chave, tipo_evento, protocolo, nome_evento, justificativa=""):
        """✅ OBSOLETO: Manifestação agora é apenas registrada no banco, XML completo será baixado separadamente.
        
        Mantido apenas para compatibilidade, mas não salva mais retEnvEvento.
        O XML da nota fiscal (não o retEnvEvento) será baixado via auto-download se configurado.
        """
        # ✅ NÃO salva retEnvEvento (apenas confirmação, não contém nota)
        # Método mantido vazio para compatibilidade com código legado
        logger.debug(f"[MANIFESTAÇÃO] Método _salvar_arquivos_manifestacao_automatico chamado mas não executa mais (retEnvEvento não deve ser salvo)")
        return

    def _salvar_arquivos_manifestacao(self, chave, tipo_evento, protocolo, nome_evento, justificativa=""):
        """Salva XML e PDF da manifestação."""
        from PyQt5.QtWidgets import QFileDialog
        import os
        
        # Pergunta onde salvar
        pasta = QFileDialog.getExistingDirectory(
            self.manifestacao_dialog, 
            "Selecione a pasta para salvar os arquivos",
            os.path.expanduser("~")
        )
        
        if not pasta:
            return
        
        # Nome base do arquivo
        nome_base = f"manifestacao_{tipo_evento}_{chave}"
        
        try:
            # Monta XML com ou sem justificativa
            justificativa_xml = ""
            if justificativa:
                justificativa_xml = f"            <xJust>{justificativa}</xJust>\n"
            
            # Salva XML (simulado - TODO: usar XML real da SEFAZ)
            xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<procEventoNFe versao="1.00">
    <evento versao="1.00">
        <infEvento>
            <chNFe>{chave}</chNFe>
            <tpEvento>{tipo_evento}</tpEvento>
            <nSeqEvento>1</nSeqEvento>
            <descEvento>{nome_evento}</descEvento>
{justificativa_xml}        </infEvento>
    </evento>
    <retEvento versao="1.00">
        <infEvento>
            <nProt>{protocolo}</nProt>
            <chNFe>{chave}</chNFe>
            <tpEvento>{tipo_evento}</tpEvento>
            <cStat>135</cStat>
            <xMotivo>Evento registrado e vinculado a NF-e</xMotivo>
        </infEvento>
    </retEvento>
</procEventoNFe>"""
            
            xml_path = os.path.join(pasta, f"{nome_base}.xml")
            with open(xml_path, 'w', encoding='utf-8') as f:
                f.write(xml_content)
            
            # Gera PDF da manifestação
            pdf_path = os.path.join(pasta, f"{nome_base}.pdf")
            self._gerar_pdf_manifestacao(pdf_path, chave, tipo_evento, protocolo, nome_evento, justificativa)
            
            QMessageBox.information(
                self.manifestacao_dialog,
                "Arquivos Salvos",
                f"📁 Arquivos salvos com sucesso!\n\n"
                f"📄 XML: {xml_path}\n"
                f"📄 PDF: {pdf_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self.manifestacao_dialog,
                "Erro ao Salvar",
                f"❌ Erro ao salvar arquivos:\n\n{str(e)}"
            )

    def _gerar_pdf_manifestacao(self, pdf_path, chave, tipo_evento, protocolo, nome_evento, justificativa=""):
        """Gera PDF da manifestação do documento."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from datetime import datetime
            
            # Cria o canvas do PDF
            c = canvas.Canvas(pdf_path, pagesize=A4)
            largura, altura = A4
            
            # Título
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(largura/2, altura - 2*cm, "COMPROVANTE DE MANIFESTAÇÃO")
            
            # Subtítulo com tipo de evento
            c.setFont("Helvetica-Bold", 14)
            c.setFillColor(colors.HexColor("#2c3e50"))
            c.drawCentredString(largura/2, altura - 3*cm, nome_evento)
            
            # Linha horizontal
            c.setStrokeColor(colors.HexColor("#3498db"))
            c.setLineWidth(2)
            c.line(2*cm, altura - 3.5*cm, largura - 2*cm, altura - 3.5*cm)
            
            # Informações principais
            y = altura - 5*cm
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 12)
            
            # Chave de Acesso
            c.drawString(2*cm, y, "Chave de Acesso:")
            c.setFont("Helvetica", 10)
            # Formata a chave em grupos de 4 dígitos
            chave_formatada = ' '.join([chave[i:i+4] for i in range(0, len(chave), 4)])
            c.drawString(2*cm, y - 0.6*cm, chave_formatada)
            
            # Código do Evento
            y -= 2*cm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2*cm, y, "Código do Evento:")
            c.setFont("Helvetica", 10)
            c.drawString(2*cm, y - 0.6*cm, tipo_evento)
            
            # Protocolo
            c.setFont("Helvetica-Bold", 12)
            c.drawString(10*cm, y, "Protocolo:")
            c.setFont("Helvetica", 10)
            c.drawString(10*cm, y - 0.6*cm, protocolo)
            
            # Status
            y -= 2*cm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2*cm, y, "Status:")
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.HexColor("#27ae60"))
            c.drawString(2*cm, y - 0.6*cm, "✓ Manifestação registrada com sucesso")
            
            # Data e Hora
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(10*cm, y, "Data/Hora:")
            c.setFont("Helvetica", 10)
            data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            c.drawString(10*cm, y - 0.6*cm, data_hora)
            
            # Justificativa (se houver)
            if justificativa:
                y -= 2.5*cm
                c.setFillColor(colors.black)
                c.setFont("Helvetica-Bold", 12)
                c.drawString(2*cm, y, "Justificativa:")
                
                # Quebra a justificativa em linhas se for muito longa
                c.setFont("Helvetica", 10)
                max_width = largura - 4*cm
                words = justificativa.split()
                lines = []
                current_line = []
                
                for word in words:
                    test_line = ' '.join(current_line + [word])
                    if c.stringWidth(test_line, "Helvetica", 10) <= max_width:
                        current_line.append(word)
                    else:
                        if current_line:
                            lines.append(' '.join(current_line))
                        current_line = [word]
                
                if current_line:
                    lines.append(' '.join(current_line))
                
                y_just = y - 0.6*cm
                for line in lines[:5]:  # Máximo 5 linhas
                    c.drawString(2*cm, y_just, line)
                    y_just -= 0.5*cm
                
                y = y_just - 0.5*cm
            
            # Caixa de informações técnicas
            y -= 2*cm
            c.setStrokeColor(colors.HexColor("#95a5a6"))
            c.setLineWidth(1)
            c.rect(2*cm, y - 3*cm, largura - 4*cm, 3*cm)
            
            c.setFillColor(colors.HexColor("#7f8c8d"))
            c.setFont("Helvetica-Bold", 10)
            c.drawString(2.5*cm, y - 0.6*cm, "Informações Técnicas")
            
            c.setFont("Helvetica", 8)
            c.drawString(2.5*cm, y - 1.2*cm, f"Evento: {tipo_evento} - {nome_evento}")
            c.drawString(2.5*cm, y - 1.7*cm, f"Documento: CT-e" if tipo_evento.startswith('6') else "Documento: NF-e")
            c.drawString(2.5*cm, y - 2.2*cm, f"cStat: 135 - Evento registrado e vinculado ao documento")
            c.drawString(2.5*cm, y - 2.7*cm, f"Gerado em: {data_hora}")
            
            # Rodapé
            c.setFillColor(colors.HexColor("#95a5a6"))
            c.setFont("Helvetica", 8)
            c.drawCentredString(largura/2, 1.5*cm, "Este documento é um comprovante de manifestação eletrônica")
            c.drawCentredString(largura/2, 1*cm, "Gerado automaticamente pelo sistema")
            
            # Finaliza o PDF
            c.save()
            
            print(f"[PDF] PDF gerado com sucesso: {pdf_path}")
            
        except Exception as e:
            print(f"[PDF] Erro ao gerar PDF: {e}")
            import traceback
            traceback.print_exc()
            # Se falhar, cria um PDF mínimo válido
            try:
                from reportlab.pdfgen import canvas
                c = canvas.Canvas(pdf_path, pagesize=A4)
                c.setFont("Helvetica", 12)
                c.drawString(2*cm, 28*cm, f"Manifestação: {nome_evento}")
                c.drawString(2*cm, 27*cm, f"Chave: {chave}")
                c.drawString(2*cm, 26*cm, f"Protocolo: {protocolo}")
                if justificativa:
                    c.drawString(2*cm, 25*cm, f"Justificativa: {justificativa[:80]}")
                c.save()
            except:
                pass

    def _lighten_color(self, hex_color):
        """Clareia uma cor hexadecimal."""
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = min(255, int(r * 1.2))
        g = min(255, int(g * 1.2))
        b = min(255, int(b * 1.2))
        return f'#{r:02x}{g:02x}{b:02x}'

    def _darken_color(self, hex_color):
        """Escurece uma cor hexadecimal."""
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = int(r * 0.8)
        g = int(g * 0.8)
        b = int(b * 0.8)
        return f'#{r:02x}{g:02x}{b:02x}'

    def importar_xmls_pasta(self):
        """Importa XMLs de uma pasta, organiza como se tivesse buscado na SEFAZ e separa emitidos/recebidos."""
        from datetime import datetime
        from lxml import etree
        import shutil
        
        try:
            # Seleciona pasta com XMLs
            pasta = QFileDialog.getExistingDirectory(self, "Selecionar pasta com XMLs para importar")
            if not pasta:
                return
            
            pasta_path = Path(pasta)
            
            # Busca todos os XMLs recursivamente
            xmls_encontrados = list(pasta_path.rglob("*.xml"))
            
            if not xmls_encontrados:
                QMessageBox.warning(self, "Importar XMLs", "Nenhum arquivo XML encontrado na pasta selecionada!")
                return
            
            # Confirma importação
            reply = QMessageBox.question(
                self,
                "Importar XMLs",
                f"Encontrados {len(xmls_encontrados)} arquivo(s) XML.\n\n"
                "A importação irá:\n"
                "• Identificar se são NFe, CTe ou NFS-e\n"
                "• Separar entre emitidos pela empresa ou recebidos\n"
                "• Organizar na estrutura correta\n"
                "• Registrar no banco de dados\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Carrega CNPJs da empresa para identificar emitidos
            certificados = self.db.load_certificates()
            cnpjs_empresa = set()
            for cert in certificados:
                cnpj = cert.get('cnpj_cpf', '') or cert.get('informante', '')
                cnpj_limpo = ''.join(c for c in str(cnpj) if c.isdigit())
                if cnpj_limpo:
                    cnpjs_empresa.add(cnpj_limpo)
            
            print(f"[IMPORTAR] CNPJs da empresa: {cnpjs_empresa}")
            
            if not cnpjs_empresa:
                QMessageBox.warning(self, "Importar XMLs", "Nenhum certificado cadastrado! Cadastre pelo menos um certificado primeiro.")
                return
            
            # Progress dialog
            progress = QProgressDialog("Importando XMLs...", "Cancelar", 0, len(xmls_encontrados), self)
            progress.setWindowTitle("Importando")
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            # Namespaces
            NS = {
                'nfe': 'http://www.portalfiscal.inf.br/nfe',
                'cte': 'http://www.portalfiscal.inf.br/cte'
            }
            
            importados = 0
            emitidos = 0
            recebidos = 0
            erros = []
            
            for idx, xml_file in enumerate(xmls_encontrados):
                if progress.wasCanceled():
                    break
                
                progress.setValue(idx)
                progress.setLabelText(f"Processando {idx+1}/{len(xmls_encontrados)}...\n{xml_file.name}")
                QApplication.processEvents()
                
                try:
                    # Lê o XML
                    xml_text = xml_file.read_text(encoding='utf-8')
                    tree = etree.fromstring(xml_text.encode('utf-8'))
                    
                    # Identifica o tipo
                    tipo = None
                    chave = None
                    cnpj_emitente = None
                    nome_emitente = None
                    numero = None
                    data_emissao = None
                    valor = None
                    cnpj_destinatario = None
                    nome_destinatario = None
                    ie_destinatario = None
                    cfop = None
                    ncm = None
                    uf = None
                    natureza = None
                    base_icms = None
                    valor_icms = None
                    status_nfe = None
                    
                    # Tenta NFe
                    if tree.find('.//nfe:infNFe', namespaces=NS) is not None:
                        tipo = 'NFe'
                        chave_elem = tree.find('.//nfe:infNFe', namespaces=NS)
                        chave = chave_elem.get('Id', '').replace('NFe', '') if chave_elem is not None else None
                        
                        # Emitente
                        cnpj_emitente = tree.findtext('.//nfe:emit/nfe:CNPJ', namespaces=NS) or ''
                        nome_emitente = tree.findtext('.//nfe:emit/nfe:xNome', namespaces=NS) or ''
                        uf = tree.findtext('.//nfe:emit/nfe:enderEmit/nfe:UF', namespaces=NS) or ''
                        
                        # Destinatário
                        cnpj_destinatario = tree.findtext('.//nfe:dest/nfe:CNPJ', namespaces=NS) or tree.findtext('.//nfe:dest/nfe:CPF', namespaces=NS) or ''
                        nome_destinatario = tree.findtext('.//nfe:dest/nfe:xNome', namespaces=NS) or ''
                        ie_destinatario = tree.findtext('.//nfe:dest/nfe:IE', namespaces=NS) or ''
                        
                        # Dados da nota
                        numero = tree.findtext('.//nfe:ide/nfe:nNF', namespaces=NS) or ''
                        data_emissao = tree.findtext('.//nfe:ide/nfe:dhEmi', namespaces=NS) or tree.findtext('.//nfe:ide/nfe:dEmi', namespaces=NS) or ''
                        natureza = tree.findtext('.//nfe:ide/nfe:natOp', namespaces=NS) or ''
                        
                        # Valores e impostos
                        valor = tree.findtext('.//nfe:total/nfe:ICMSTot/nfe:vNF', namespaces=NS) or '0'
                        base_icms = tree.findtext('.//nfe:total/nfe:ICMSTot/nfe:vBC', namespaces=NS) or '0'
                        valor_icms = tree.findtext('.//nfe:total/nfe:ICMSTot/nfe:vICMS', namespaces=NS) or '0'
                        
                        # CFOP e NCM do primeiro produto
                        cfop = tree.findtext('.//nfe:det[1]/nfe:prod/nfe:CFOP', namespaces=NS) or ''
                        ncm = tree.findtext('.//nfe:det[1]/nfe:prod/nfe:NCM', namespaces=NS) or ''
                        
                        # Status da nota (da autorização)
                        status_nfe = tree.findtext('.//nfe:protNFe/nfe:infProt/nfe:cStat', namespaces=NS) or ''
                    
                    # Tenta CTe
                    elif tree.find('.//cte:infCte', namespaces=NS) is not None:
                        tipo = 'CTe'
                        chave_elem = tree.find('.//cte:infCte', namespaces=NS)
                        chave = chave_elem.get('Id', '').replace('CTe', '') if chave_elem is not None else None
                        
                        # Emitente
                        cnpj_emitente = tree.findtext('.//cte:emit/cte:CNPJ', namespaces=NS) or ''
                        nome_emitente = tree.findtext('.//cte:emit/cte:xNome', namespaces=NS) or ''
                        uf = tree.findtext('.//cte:emit/cte:enderEmit/cte:UF', namespaces=NS) or ''
                        
                        # Destinatário (tomador)
                        cnpj_destinatario = tree.findtext('.//cte:dest/cte:CNPJ', namespaces=NS) or tree.findtext('.//cte:dest/cte:CPF', namespaces=NS) or ''
                        nome_destinatario = tree.findtext('.//cte:dest/cte:xNome', namespaces=NS) or ''
                        
                        # Dados do CT-e
                        numero = tree.findtext('.//cte:ide/cte:nCT', namespaces=NS) or ''
                        data_emissao = tree.findtext('.//cte:ide/cte:dhEmi', namespaces=NS) or ''
                        natureza = tree.findtext('.//cte:ide/cte:natOp', namespaces=NS) or ''
                        cfop = tree.findtext('.//cte:ide/cte:CFOP', namespaces=NS) or ''
                        
                        # Valor
                        valor = tree.findtext('.//cte:vPrest/cte:vTPrest', namespaces=NS) or '0'
                    
                    if not tipo or not chave:
                        erros.append(f"{xml_file.name}: Tipo não identificado")
                        continue
                    
                    # Remove caracteres não numéricos do CNPJ emitente e destinatário
                    cnpj_emitente_limpo = ''.join(c for c in cnpj_emitente if c.isdigit())
                    cnpj_destinatario_limpo = ''.join(c for c in cnpj_destinatario if c.isdigit())
                    
                    # Determina se foi emitido pela empresa ou recebido
                    emitido_pela_empresa = cnpj_emitente_limpo in cnpjs_empresa
                    
                    # Escolhe o informante (quem vai "organizar" o arquivo)
                    if emitido_pela_empresa:
                        # Emitido pela empresa - usa o CNPJ emitente
                        informante = cnpj_emitente_limpo
                        emitidos += 1
                    else:
                        # Recebido de terceiros - usa o primeiro CNPJ da empresa
                        informante = list(cnpjs_empresa)[0]
                        recebidos += 1
                    
                    # Extrai ano-mês da data de emissão
                    year_month = datetime.now().strftime("%Y-%m")
                    if data_emissao:
                        if len(data_emissao) >= 7:
                            year_month = data_emissao[:7]
                    
                    # Define pasta destino
                    dest_folder = DATA_DIR / "xmls" / informante / tipo / year_month
                    dest_folder.mkdir(parents=True, exist_ok=True)
                    
                    # Nome do arquivo destino
                    if numero:
                        # Usa padrão número-nome
                        nome_limpo = ''.join(c for c in nome_emitente if c.isalnum() or c in (' ', '-', '_'))[:50]
                        dest_file = dest_folder / f"{numero}-{nome_limpo}.xml"
                    else:
                        # Usa apenas a chave
                        dest_file = dest_folder / f"{chave}.xml"
                    
                    # Copia o arquivo
                    shutil.copy2(xml_file, dest_file)
                    
                    # Registra no banco de dados
                    self.db.register_xml_download(chave, str(dest_file), informante)
                    
                    # Salva dados da nota no banco (com todos os campos extraídos)
                    nota_data = {
                        'chave': chave,
                        'numero': numero,
                        'data_emissao': data_emissao[:10] if len(data_emissao) >= 10 else data_emissao,
                        'tipo': tipo,
                        'valor': valor,
                        'cnpj_emitente': cnpj_emitente_limpo,
                        'nome_emitente': nome_emitente,
                        'nome_destinatario': nome_destinatario,
                        'cnpj_destinatario': cnpj_destinatario_limpo,
                        'ie_tomador': ie_destinatario,
                        'cfop': cfop,
                        'ncm': ncm,
                        'natureza': natureza,
                        'uf': uf,
                        'base_icms': base_icms,
                        'valor_icms': valor_icms,
                        'status': 'Autorizado' if status_nfe == '100' else status_nfe,
                        'informante': informante,
                        'xml_status': 'COMPLETO'
                    }
                    self.db.save_note(nota_data)
                    
                    importados += 1
                    print(f"[IMPORTAR] {tipo} {numero} - {'EMITIDO' if emitido_pela_empresa else 'RECEBIDO'} - {dest_file}")
                    
                except Exception as e:
                    erros.append(f"{xml_file.name}: {str(e)}")
                    print(f"[ERRO IMPORTAR] {xml_file.name}: {e}")
                    import traceback
                    traceback.print_exc()
            
            progress.setValue(len(xmls_encontrados))
            
            # Atualiza interface
            self.refresh_all()
            
            # Mostra resultado
            mensagem = f"Importação concluída!\n\n"
            mensagem += f"✅ Arquivos importados: {importados}\n"
            mensagem += f"📤 Emitidos pela empresa: {emitidos}\n"
            mensagem += f"📥 Recebidos de terceiros: {recebidos}\n"
            
            if erros:
                mensagem += f"\n❌ Erros: {len(erros)}"
                if len(erros) <= 5:
                    mensagem += "\n\n" + "\n".join(erros[:5])
                else:
                    mensagem += "\n(Veja o console para detalhes)"
            
            QMessageBox.information(self, "Importar XMLs", mensagem)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro ao Importar", f"Erro: {e}")
            import traceback
            traceback.print_exc()

    def abrir_exportacao(self):
        """Abre o diálogo de exportação de arquivos."""
        try:
            # Verifica se há documentos selecionados na tabela
            selected_rows = self.table.selectionModel().selectedRows()
            
            if not selected_rows:
                QMessageBox.warning(
                    self,
                    "Exportar",
                    "Selecione pelo menos um documento na tabela para exportar!"
                )
                return
            
            # Abre diálogo de exportação
            dialog = ExportDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                self._executar_exportacao(dialog.get_opcoes())
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir exportação: {e}")
            import traceback
            traceback.print_exc()
    
    def _executar_exportacao(self, opcoes):
        """Executa a exportação com as opções selecionadas."""
        from datetime import datetime
        import shutil
        
        try:
            print("\n" + "="*60)
            print("🔍 INICIANDO EXPORTAÇÃO - DEBUG")
            print("="*60)
            print(f"Opções selecionadas: {opcoes}")
            
            # Seleciona pasta de destino
            pasta_destino = QFileDialog.getExistingDirectory(
                self,
                "Selecionar pasta de destino para exportação"
            )
            
            if not pasta_destino:
                print("❌ Usuário cancelou seleção de pasta")
                return
            
            pasta_destino = Path(pasta_destino)
            print(f"📁 Pasta destino: {pasta_destino}")
            
            # Obtém documentos selecionados
            selected_rows = self.table.selectionModel().selectedRows()
            total = len(selected_rows)
            print(f"📋 Total de documentos selecionados: {total}")
            
            # Progress dialog
            progress = QProgressDialog("Exportando arquivos...", "Cancelar", 0, total, self)
            progress.setWindowTitle("Exportar")
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            exportados = 0
            pulados = 0  # NFSe sem arquivos disponíveis
            erros = []
            
            for idx, row_index in enumerate(selected_rows):
                if progress.wasCanceled():
                    print("⚠️ Exportação cancelada pelo usuário")
                    break
                
                row = row_index.row()
                # Coluna 18 é "Chave" (última coluna do header)
                # Headers: XML(0), Num(1), D/Emit(2), Tipo(3), Valor(4), Venc(5), 
                #          Emissor CNPJ(6), Emissor Nome(7), Natureza(8), UF(9), 
                #          Base ICMS(10), Valor ICMS(11), IBS(12), CBS(13), 
                #          Status(14), CFOP(15), NCM(16), Tomador IE(17), Chave(18)
                chave = self.table.item(row, 18).text() if self.table.item(row, 18) else None
                
                if not chave:
                    print(f"⚠️ Linha {row}: Chave não encontrada na tabela")
                    continue
                
                print(f"\n📄 [{idx+1}/{total}] Processando chave: {chave}")
                
                progress.setLabelText(f"Exportando {idx+1}/{total}...")
                progress.setValue(idx)
                
                try:
                    # Busca informações do documento no banco
                    print(f"  🔍 Buscando documento no banco...")
                    doc = self.db.get_documento_por_chave(chave)
                    
                    if not doc:
                        erro_msg = f"Documento não encontrado no banco: {chave}"
                        print(f"  ❌ {erro_msg}")
                        erros.append(erro_msg)
                        continue
                    
                    print(f"  ✅ Documento encontrado: Número={doc.get('numero')}, Emitente={doc.get('nome_emitente')}")
                    
                    # Define nome do arquivo
                    if opcoes['nome_personalizado']:
                        # Usa número e nome do documento
                        numero = doc.get('numero', 'SN')
                        nome_emit = doc.get('nome_emitente', 'Desconhecido')
                        # Remove caracteres inválidos do nome
                        nome_emit_limpo = "".join(c for c in nome_emit if c.isalnum() or c in (' ', '-', '_')).strip()
                        nome_base = f"{numero}_{nome_emit_limpo}"
                        print(f"  📝 Nome personalizado: {nome_base}")
                    else:
                        # Nome padrão (chave de acesso)
                        nome_base = chave
                        print(f"  📝 Nome padrão (chave): {nome_base}")
                    
                    sucesso_xml = False
                    sucesso_pdf = False
                    
                    # Exporta XML
                    if opcoes['exportar_xml']:
                        print(f"  🔍 Procurando arquivo XML...")
                        xml_origem = self._encontrar_arquivo_xml(chave)
                        
                        if xml_origem and xml_origem.exists():
                            xml_destino = pasta_destino / f"{nome_base}.xml"
                            print(f"  ✅ XML encontrado: {xml_origem}")
                            print(f"  📤 Copiando para: {xml_destino}")
                            shutil.copy2(xml_origem, xml_destino)
                            sucesso_xml = True
                        else:
                            erro_msg = f"XML não encontrado: {chave}"
                            print(f"  ❌ {erro_msg}")
                            erros.append(erro_msg)
                    
                    # Exporta PDF
                    if opcoes['exportar_pdf']:
                        print(f"  🔍 Procurando arquivo PDF...")
                        pdf_origem = self._encontrar_arquivo_pdf(chave)
                        
                        if pdf_origem and pdf_origem.exists():
                            pdf_destino = pasta_destino / f"{nome_base}.pdf"
                            print(f"  ✅ PDF encontrado: {pdf_origem}")
                            print(f"  📤 Copiando para: {pdf_destino}")
                            shutil.copy2(pdf_origem, pdf_destino)
                            sucesso_pdf = True
                        else:
                            # Se PDF não existe, tenta gerar automaticamente
                            print(f"  ⚠️  PDF não encontrado, tentando gerar...")
                            
                            # Primeiro encontra o XML
                            xml_origem = self._encontrar_arquivo_xml(chave)
                            
                            if xml_origem and xml_origem.exists():
                                try:
                                    # Lê o conteúdo do XML
                                    with open(xml_origem, 'r', encoding='utf-8') as f:
                                        xml_content = f.read()
                                    
                                    # Cria PDF temporário
                                    import tempfile
                                    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
                                        pdf_temp = Path(tmp_pdf.name)
                                    
                                    print(f"  🔄 Gerando PDF a partir do XML: {xml_origem.name}")
                                    
                                    # Gera PDF usando módulo pdf_simple
                                    from modules.pdf_simple import generate_danfe_pdf
                                    
                                    tipo = doc.get('tipo', 'NFE')
                                    if generate_danfe_pdf(xml_content, str(pdf_temp), tipo):
                                        # PDF gerado com sucesso, copia para destino
                                        pdf_destino = pasta_destino / f"{nome_base}.pdf"
                                        print(f"  ✅ PDF gerado com sucesso!")
                                        print(f"  📤 Copiando para: {pdf_destino}")
                                        shutil.copy2(pdf_temp, pdf_destino)
                                        sucesso_pdf = True
                                        
                                        # Remove arquivo temporário
                                        pdf_temp.unlink()
                                    else:
                                        erro_msg = f"Falha ao gerar PDF: {chave}"
                                        print(f"  ❌ {erro_msg}")
                                        erros.append(erro_msg)
                                        if pdf_temp.exists():
                                            pdf_temp.unlink()
                                except Exception as e_gen:
                                    erro_msg = f"Erro ao gerar PDF {chave}: {str(e_gen)}"
                                    print(f"  ❌ {erro_msg}")
                                    erros.append(erro_msg)
                            else:
                                # XML não encontrado - comum para NFSe consultadas via API
                                tipo_doc = doc.get('tipo', 'NFE')
                                if tipo_doc == 'NFS-e':
                                    # NFSe sem XML é esperado (consultadas via API sem XML completo)
                                    print(f"  ⚠️  NFS-e sem arquivo XML disponível (dados apenas no banco)")
                                    print(f"  ℹ️  Para gerar PDF, é necessário ter o XML completo salvo")
                                    # Não adiciona erro para NFSe sem XML
                                else:
                                    erro_msg = f"XML não encontrado para gerar PDF: {chave}"
                                    print(f"  ❌ {erro_msg}")
                                    erros.append(erro_msg)
                    
                    if sucesso_xml or sucesso_pdf:
                        exportados += 1
                        print(f"  ✅ Exportado com sucesso!")
                    else:
                        # Documento não pôde ser exportado (sem XML e sem PDF)
                        if doc.get('tipo') == 'NFS-e':
                            print(f"  ⏭️  Pulando NFS-e sem arquivos disponíveis")
                            pulados += 1
                        else:
                            print(f"  ❌ Nenhum arquivo exportado")

                    
                except Exception as e:
                    erro_msg = f"{chave}: {str(e)}"
                    print(f"  ❌ ERRO: {erro_msg}")
                    import traceback
                    traceback.print_exc()
                    erros.append(erro_msg)
            
            progress.setValue(total)
            
            print("\n" + "="*60)
            print("📊 RESUMO DA EXPORTAÇÃO")
            print("="*60)
            if pulados > 0:
                print(f"⏭️  NFS-e puladas (sem arquivos): {pulados}")
            print(f"❌ Erros: {len(erros)}")
            print("="*60 + "\n")
            
            # Resultado
            mensagem = f"Exportação concluída!\n\n"
            mensagem += f"✅ Arquivos exportados: {exportados}\n"
            if pulados > 0:
                mensagem += f"⏭️  NFS-e puladas (sem arquivos): {pulados}\n"
                mensagem += f"    (Estas NFS-e foram consultadas via API mas não possuem\n"
                mensagem += f"     arquivo XML/PDF disponível para exportação)\n"
            mensagem += f"📁 Destino: {pasta_destino}\n"
            
            if erros:
                mensagem += f"\n❌ Erros: {len(erros)}"
                if len(erros) <= 5:
                    mensagem += "\n\n" + "\n".join(erros[:5])
                else:
                    mensagem += f"\n\nPrimeiros 5 erros:\n" + "\n".join(erros[:5])
                    mensagem += f"\n\n(Veja o console para lista completa)"
            
            QMessageBox.information(self, "Exportar", mensagem)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro na exportação: {e}")
            import traceback
            traceback.print_exc()
    
    def _encontrar_arquivo_xml(self, chave):
        """Encontra o arquivo XML de uma chave de acesso."""
        print(f"    🔍 Procurando XML para chave: {chave}")
        
        # PRIORIDADE 1: Consulta o banco de dados onde o caminho está registrado
        print(f"    💾 Consultando banco de dados...")
        try:
            with self.db._connect() as conn:
                cursor = conn.execute(
                    "SELECT caminho_arquivo FROM xmls_baixados WHERE chave = ?",
                    (chave,)
                )
                row = cursor.fetchone()
                if row and row[0]:
                    xml_path = Path(row[0])
                    print(f"    💾 Caminho encontrado no banco: {xml_path}")
                    if xml_path.exists():
                        print(f"    ✅ XML encontrado no banco: {xml_path}")
                        return xml_path
                    else:
                        print(f"    ⚠️ Caminho do banco não existe mais: {xml_path}")
                else:
                    print(f"    ⚠️ Chave não encontrada no banco de xmls_baixados")
        except Exception as e:
            print(f"    ❌ Erro ao consultar banco: {e}")
            import traceback
            traceback.print_exc()
        
        # PRIORIDADE 2: Busca em diretórios estruturados por informante
        # Formato: DATA_DIR/xmls/{informante}/{ano-mes}/{tipo}/{numero}-{nome}.xml
        print(f"    📂 DATA_DIR: {DATA_DIR}")
        xmls_dir = DATA_DIR / 'xmls'
        if xmls_dir.exists():
            print(f"    📂 Buscando em estrutura: {xmls_dir}")
            
            # Primeiro tenta busca rápida por nome de arquivo com chave
            print(f"    📂 Procurando arquivo: {chave}.xml")
            for xml_file in xmls_dir.rglob(f"{chave}.xml"):
                print(f"    ✅ Arquivo encontrado: {xml_file}")
                return xml_file
            
            # Se não encontrar, busca recursiva por XMLs e verifica conteúdo
            # (os XMLs são salvos como numero-nome.xml, então precisamos buscar no conteúdo)
            print(f"    🔍 Chave não encontrada no nome, buscando no conteúdo dos XMLs...")
            xml_files = list(xmls_dir.rglob("*.xml"))
            
            # Filtra arquivos de debug/protocolo
            xml_files = [
                f for f in xml_files 
                if not any(x in f.name.lower() for x in ['debug', 'protocolo', 'request', 'response'])
            ]
            
            print(f"    📊 Total de XMLs para verificar: {len(xml_files)}")
            
            # Busca a chave no conteúdo (limitado aos primeiros 1000 arquivos para não travar)
            for xml_file in xml_files[:1000]:
                try:
                    with open(xml_file, 'r', encoding='utf-8') as f:
                        content = f.read(2000)  # Lê apenas início do arquivo
                        if chave in content:
                            print(f"    ✅ Arquivo encontrado por conteúdo: {xml_file}")
                            return xml_file
                except Exception:
                    continue
            
            print(f"    ⚠️ Chave {chave} não encontrada em nenhum XML")
        else:
            print(f"    ⚠️ Diretório não existe: {xmls_dir}")
        
        # PRIORIDADE 3: Busca em diretórios legados
        diretorios = [
            BASE_DIR / 'xmls_chave',
            BASE_DIR / 'xml_extraidos', 
            BASE_DIR / 'xml_NFs',
        ]
        
        for diretorio in diretorios:
            if diretorio.exists():
                print(f"    📂 Verificando diretório legado: {diretorio}")
                for xml_file in diretorio.rglob(f"*{chave}*.xml"):
                    # FILTRO: Ignora arquivos de debug/protocolo
                    nome_arquivo = xml_file.name.lower()
                    if any(x in nome_arquivo for x in ['debug', 'protocolo', 'request', 'response']):
                        print(f"    ⚠️ Arquivo ignorado (debug/protocolo): {xml_file.name}")
                        continue
                    
                    print(f"    ✅ Arquivo encontrado: {xml_file}")
                    return xml_file
        
        print(f"    ❌ XML não encontrado em nenhum local")
        return None
    
    def _encontrar_arquivo_pdf(self, chave):
        """Encontra o arquivo PDF de uma chave de acesso."""
        print(f"    🔍 Procurando PDF para chave: {chave}")
        
        # PRIORIDADE 1: PDF ao lado do XML registrado no banco
        try:
            with self.db._connect() as conn:
                cursor = conn.execute(
                    "SELECT caminho_arquivo FROM xmls_baixados WHERE chave = ?",
                    (chave,)
                )
                row = cursor.fetchone()
                if row and row[0]:
                    xml_path = Path(row[0])
                    pdf_path = xml_path.with_suffix('.pdf')
                    if pdf_path.exists():
                        print(f"    ✅ PDF encontrado ao lado do XML: {pdf_path}")
                        return pdf_path
                    else:
                        print(f"    ⚠️ PDF não existe ao lado do XML: {pdf_path}")
        except Exception as e:
            print(f"    ⚠️ Erro ao consultar banco: {e}")
        
        # PRIORIDADE 2: Busca em estrutura por informante  
        xmls_dir = DATA_DIR / 'xmls'
        if xmls_dir.exists():
            print(f"    📂 Buscando em estrutura: {xmls_dir}")
            
            # Primeiro tenta busca rápida por nome de arquivo com chave
            for pdf_file in xmls_dir.rglob(f"{chave}.pdf"):
                print(f"    ✅ Arquivo encontrado: {pdf_file}")
                return pdf_file
            
            # Se não encontrar, procura PDF que corresponda a XML com a chave
            # (busca arquivos .xml e verifica se existe .pdf correspondente)
            print(f"    🔍 Chave não encontrada no nome, buscando PDF correspondente ao XML...")
            xml_files = list(xmls_dir.rglob("*.xml"))
            
            # Filtra arquivos de debug/protocolo
            xml_files = [
                f for f in xml_files 
                if not any(x in f.name.lower() for x in ['debug', 'protocolo', 'request', 'response'])
            ]
            
            # Busca a chave no conteúdo do XML e verifica se existe PDF
            for xml_file in xml_files[:1000]:
                try:
                    with open(xml_file, 'r', encoding='utf-8') as f:
                        content = f.read(2000)  # Lê apenas início do arquivo
                        if chave in content:
                            # Verifica se existe PDF com mesmo nome
                            pdf_file = xml_file.with_suffix('.pdf')
                            if pdf_file.exists():
                                print(f"    ✅ PDF encontrado correspondente ao XML: {pdf_file}")
                                return pdf_file
                            else:
                                print(f"    ⚠️ XML encontrado mas PDF não existe: {xml_file.name}")
                except Exception:
                    continue
        
        # PRIORIDADE 3: Busca em diretórios legados
        diretorios = [
            BASE_DIR / 'xmls_chave',
            BASE_DIR / 'xml_extraidos',
            BASE_DIR / 'xml_NFs',
        ]
        
        for diretorio in diretorios:
            if diretorio.exists():
                print(f"    📂 Verificando diretório legado: {diretorio}")
                for pdf_file in diretorio.rglob(f"*{chave}*.pdf"):
                    print(f"    ✅ Arquivo encontrado: {pdf_file}")
                    return pdf_file
        
        print(f"    ❌ PDF não encontrado em nenhum local")
        return None

    def abrir_relatorio(self):
        """Abre diálogo de relatório analítico IBS/CBS com filtros de período e empresa."""
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                                     QDateEdit, QComboBox, QTableWidget, QTableWidgetItem,
                                     QPushButton, QHeaderView, QFileDialog, QMessageBox,
                                     QProgressDialog)
        from PyQt5.QtCore import QDate, Qt
        from datetime import datetime
        
        try:
            # Cria dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Relatório IBS/CBS")
            dialog.resize(1400, 800)
            
            # Layout principal
            layout = QVBoxLayout(dialog)
            
            # === FILTROS ===
            filter_layout = QHBoxLayout()
            
            # Período
            filter_layout.addWidget(QLabel("Período:"))
            date_inicio = QDateEdit(QDate.currentDate().addMonths(-1))
            date_inicio.setDisplayFormat("dd/MM/yyyy")
            date_inicio.setCalendarPopup(True)
            date_fim = QDateEdit(QDate.currentDate())
            date_fim.setDisplayFormat("dd/MM/yyyy")
            date_fim.setCalendarPopup(True)
            filter_layout.addWidget(date_inicio)
            filter_layout.addWidget(QLabel("até"))
            filter_layout.addWidget(date_fim)
            
            # Empresa
            filter_layout.addWidget(QLabel("Empresa:"))
            combo_empresa = QComboBox()
            combo_empresa.setMinimumWidth(250)
            filter_layout.addWidget(combo_empresa)
            
            # Botões
            btn_atualizar = QPushButton("🔄 Atualizar")
            btn_exportar_excel = QPushButton("📊 Exportar Excel")
            btn_exportar_excel.setEnabled(False)
            filter_layout.addWidget(btn_atualizar)
            filter_layout.addWidget(btn_exportar_excel)
            filter_layout.addStretch()
            
            layout.addLayout(filter_layout)
            
            # === TABELA ===
            table = QTableWidget()
            headers = ["Data", "Tipo", "Número", "Emitente", "Destinatário", 
                      "Valor Total", "IBS", "CBS", "Chave"]
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            
            # Adiciona tooltips aos headers IBS e CBS
            header_item_ibs = table.horizontalHeaderItem(6)
            if header_item_ibs:
                header_item_ibs.setToolTip("Imposto sobre Bens e Serviços (Reforma Tributária)")
            
            header_item_cbs = table.horizontalHeaderItem(7)
            if header_item_cbs:
                header_item_cbs.setToolTip("Contribuição sobre Bens e Serviços (Reforma Tributária)")
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(True)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.setSelectionMode(QTableWidget.ExtendedSelection)  # Permite múltipla seleção
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            table.setSortingEnabled(True)  # Permite ordenar por clique nos cabeçalhos
            layout.addWidget(table)
            
            # Label de status
            lbl_status = QLabel("Selecione o período e clique em Atualizar")
            layout.addWidget(lbl_status)
            
            # Label de totais (rodapé)
            lbl_totais = QLabel("")
            lbl_totais.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 8px; font-weight: bold; border: 1px solid #ccc; }")
            layout.addWidget(lbl_totais)
            
            # === FUNÇÕES AUXILIARES ===
            
            def converter_moeda_para_float(valor_str):
                """Converte string formatada (1.234,56) para float."""
                try:
                    if not valor_str or valor_str == "0,00":
                        return 0.0
                    # Remove pontos de milhar e troca vírgula por ponto
                    valor_limpo = valor_str.replace('.', '').replace(',', '.')
                    return float(valor_limpo)
                except:
                    return 0.0
            
            def atualizar_totais():
                """Atualiza label de totais com base nas linhas selecionadas."""
                try:
                    selected_rows = table.selectionModel().selectedRows()
                    
                    if not selected_rows:
                        lbl_totais.setText("")
                        return
                    
                    # Calcula totais das colunas 5 (Valor Total), 6 (IBS), 7 (CBS)
                    total_valor = 0.0
                    total_ibs = 0.0
                    total_cbs = 0.0
                    
                    for row_index in selected_rows:
                        row = row_index.row()
                        
                        # Valor Total (coluna 5)
                        item_valor = table.item(row, 5)
                        if item_valor:
                            total_valor += converter_moeda_para_float(item_valor.text())
                        
                        # IBS (coluna 6)
                        item_ibs = table.item(row, 6)
                        if item_ibs:
                            total_ibs += converter_moeda_para_float(item_ibs.text())
                        
                        # CBS (coluna 7)
                        item_cbs = table.item(row, 7)
                        if item_cbs:
                            total_cbs += converter_moeda_para_float(item_cbs.text())
                    
                    # Formata totais
                    texto_totais = f"Selecionados: {len(selected_rows)} documento(s)  |  "
                    texto_totais += f"Total Valor: {formatar_moeda(total_valor)}  |  "
                    texto_totais += f"Total IBS: {formatar_moeda(total_ibs)}  |  "
                    texto_totais += f"Total CBS: {formatar_moeda(total_cbs)}"
                    
                    lbl_totais.setText(texto_totais)
                    
                except Exception as e:
                    print(f"Erro ao calcular totais: {e}")
            
            def formatar_moeda(valor):
                """Formata valor no padrão monetário brasileiro (1.234,56)."""
                try:
                    if valor is None or valor == "":
                        return "0,00"
                    
                    # Se já está formatado (contém "R$"), limpa e reconverte
                    if isinstance(valor, str):
                        # Remove "R$", espaços e outros caracteres não numéricos (exceto dígitos, vírgula e ponto)
                        valor_limpo = valor.replace('R$', '').replace(' ', '').strip()
                        
                        # Se tem vírgula, está no formato brasileiro (1.234,56)
                        if ',' in valor_limpo:
                            # Remove pontos de milhar e troca vírgula por ponto
                            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
                        
                        # Tenta converter para float
                        try:
                            valor_float = float(valor_limpo)
                        except ValueError:
                            return "0,00"
                    else:
                        valor_float = float(valor)
                    
                    # Formata no padrão brasileiro manualmente (independente de locale)
                    # Separa parte inteira e decimal
                    valor_abs = abs(valor_float)
                    parte_inteira = int(valor_abs)
                    parte_decimal = int(round((valor_abs - parte_inteira) * 100))
                    
                    # Formata parte inteira com separador de milhar (.)
                    parte_int_str = f"{parte_inteira:,}".replace(',', '.')
                    
                    # Monta valor final
                    valor_formatado = f"{'-' if valor_float < 0 else ''}{parte_int_str},{parte_decimal:02d}"
                    
                    return valor_formatado
                except:
                    return "0,00"
            
            def extrair_ibs_cbs_do_xml(chave):
                """
                Extrai valores de IBS e CBS do XML da nota.
                
                ⚠️ FUNÇÃO LEGADA - Usada apenas como fallback para notas antigas sem IBS/CBS no banco.
                Para melhor performance, execute: Menu > Configurações > Atualizar IBS/CBS das Notas (Ctrl+Shift+U)
                Isso populará o banco e tornará o relatório muito mais rápido.
                """
                try:
                    from lxml import etree
                    from pathlib import Path
                    
                    # Busca rápida do XML (sem busca pesada em conteúdo)
                    xml_path = None
                    
                    # PRIORIDADE 1: Consulta banco de dados
                    try:
                        with self.db._connect() as conn:
                            cursor = conn.execute(
                                "SELECT caminho_arquivo FROM xmls_baixados WHERE chave = ?",
                                (chave,)
                            )
                            row = cursor.fetchone()
                            if row and row[0]:
                                xml_path = Path(row[0])
                                if not xml_path.exists():
                                    xml_path = None
                    except:
                        pass
                    
                    # PRIORIDADE 2: Busca rápida por nome de arquivo apenas
                    if not xml_path:
                        xmls_dir = DATA_DIR / 'xmls'
                        if xmls_dir.exists():
                            # Busca apenas por nome de arquivo (rápido)
                            matches = list(xmls_dir.rglob(f"{chave}.xml"))
                            if matches:
                                xml_path = matches[0]
                    
                    # Se não encontrou, retorna zero (não faz busca pesada)
                    if not xml_path or not xml_path.exists():
                        return "0,00", "0,00"
                    
                    # Lê e parseia o XML
                    with open(xml_path, 'r', encoding='utf-8') as f:
                        xml_content = f.read()
                    
                    # NFS-e não tem IBS/CBS (são impostos de NF-e)
                    if 'nfse' in xml_content.lower() or 'servico' in xml_content.lower():
                        return "0,00", "0,00"
                    
                    tree = etree.fromstring(xml_content.encode('utf-8'))
                    
                    # Namespaces comuns para NF-e
                    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                    
                    ibs_value = None
                    cbs_value = None
                    
                    # ESTRATÉGIA 1: Busca com namespace específico
                    ibs_tags = tree.xpath('.//nfe:vIBS', namespaces=ns)
                    if ibs_tags and ibs_tags[0].text:
                        ibs_value = ibs_tags[0].text
                    
                    cbs_tags = tree.xpath('.//nfe:vCBS', namespaces=ns)
                    if cbs_tags and cbs_tags[0].text:
                        cbs_value = cbs_tags[0].text
                    
                    # ESTRATÉGIA 2: Busca em ICMSTot com namespace
                    if not ibs_value:
                        ibs_total = tree.xpath('.//nfe:ICMSTot/nfe:vIBS', namespaces=ns)
                        if ibs_total and ibs_total[0].text:
                            ibs_value = ibs_total[0].text
                    
                    if not cbs_value:
                        cbs_total = tree.xpath('.//nfe:ICMSTot/nfe:vCBS', namespaces=ns)
                        if cbs_total and cbs_total[0].text:
                            cbs_value = cbs_total[0].text
                    
                    # ESTRATÉGIA 3: Busca sem namespace (local-name)
                    if not ibs_value:
                        ibs_no_ns = tree.xpath(".//*[local-name()='vIBS']")
                        if ibs_no_ns and ibs_no_ns[0].text:
                            ibs_value = ibs_no_ns[0].text
                    
                    if not cbs_value:
                        cbs_no_ns = tree.xpath(".//*[local-name()='vCBS']")
                        if cbs_no_ns and cbs_no_ns[0].text:
                            cbs_value = cbs_no_ns[0].text
                    
                    # ESTRATÉGIA 4: Busca texto no XML (fallback)
                    if not ibs_value and '<vIBS>' in xml_content:
                        import re
                        match = re.search(r'<vIBS>([\d.,]+)</vIBS>', xml_content)
                        if match:
                            ibs_value = match.group(1)
                    
                    if not cbs_value and '<vCBS>' in xml_content:
                        import re
                        match = re.search(r'<vCBS>([\d.,]+)</vCBS>', xml_content)
                        if match:
                            cbs_value = match.group(1)
                    
                    return ibs_value or "0.00", cbs_value or "0.00"
                    
                except Exception as e:
                    # Erro silencioso - retorna zero
                    return "0,00", "0,00"
            
            def popular_empresas():
                """Popula combo com empresas (CNPJs dos certificados - destinatários)."""
                try:
                    combo_empresa.clear()
                    combo_empresa.addItem("Todas as empresas", None)
                    
                    # Busca empresas dos certificados configurados
                    with self.db._connect() as conn:
                        # Verifica se as colunas existem
                        cursor = conn.execute("PRAGMA table_info(certificados)")
                        columns = {col[1] for col in cursor.fetchall()}
                        
                        # Monta query baseada nas colunas disponíveis
                        if 'razao_social' in columns and 'ativo' in columns:
                            cursor = conn.execute("""
                                SELECT cnpj_cpf, razao_social 
                                FROM certificados 
                                WHERE ativo = 1 
                                ORDER BY razao_social
                            """)
                        elif 'razao_social' in columns:
                            cursor = conn.execute("""
                                SELECT cnpj_cpf, razao_social 
                                FROM certificados 
                                ORDER BY razao_social
                            """)
                        elif 'nome_certificado' in columns:
                            cursor = conn.execute("""
                                SELECT cnpj_cpf, nome_certificado 
                                FROM certificados 
                                ORDER BY nome_certificado
                            """)
                        else:
                            cursor = conn.execute("""
                                SELECT cnpj_cpf, informante 
                                FROM certificados 
                                ORDER BY informante
                            """)
                        
                        empresas = cursor.fetchall()
                        if empresas:
                            for cnpj, nome in empresas:
                                if cnpj:
                                    combo_empresa.addItem(f"{cnpj} - {nome or 'Sem Nome'}", cnpj)
                        else:
                            # Fallback: busca destinatários únicos das notas
                            cursor = conn.execute("""
                                SELECT DISTINCT cnpj_destinatario 
                                FROM notas_detalhadas 
                                WHERE cnpj_destinatario IS NOT NULL 
                                ORDER BY cnpj_destinatario
                            """)
                            destinatarios = cursor.fetchall()
                            for (cnpj,) in destinatarios:
                                if cnpj:
                                    combo_empresa.addItem(cnpj, cnpj)
                
                except Exception as e:
                    QMessageBox.warning(dialog, "Erro", f"Erro ao carregar empresas: {e}")
            
            def atualizar_relatorio():
                """Atualiza tabela com dados do período e empresa selecionados."""
                try:
                    # Limpa tabela
                    table.setRowCount(0)
                    btn_exportar_excel.setEnabled(False)
                    lbl_totais.setText("")  # Limpa totais
                    
                    # Pega filtros
                    data_ini = date_inicio.date().toString("yyyy-MM-dd")
                    data_fim = date_fim.date().toString("yyyy-MM-dd")
                    cnpj_filtro = combo_empresa.currentData()
                    
                    lbl_status.setText(f"Buscando dados de {date_inicio.date().toString('dd/MM/yyyy')} até {date_fim.date().toString('dd/MM/yyyy')}...")
                    dialog.repaint()
                    
                    # Query SQL - Filtra por DESTINATÁRIO (emitidas por terceiros)
                    # Inclui NFS-e usando informante OU cnpj_destinatario
                    # 🚀 PERFORMANCE: Inclui v_ibs e v_cbs na query (evita ler XMLs)
                    query = """
                        SELECT chave, data_emissao, tipo, numero, nome_emitente, 
                               cnpj_destinatario, valor, informante, v_ibs, v_cbs
                        FROM notas_detalhadas
                        WHERE data_emissao BETWEEN ? AND ?
                    """
                    params = [data_ini, data_fim]
                    
                    if cnpj_filtro:
                        # Para NFS-e, verifica tanto cnpj_destinatario quanto informante
                        query += " AND (cnpj_destinatario = ? OR (tipo LIKE '%NFS%' AND informante = ?))"
                        params.extend([cnpj_filtro, cnpj_filtro])
                    
                    query += " ORDER BY data_emissao DESC"
                    
                    with self.db._connect() as conn:
                        cursor = conn.execute(query, params)
                        rows = cursor.fetchall()
                    
                    if not rows:
                        lbl_status.setText("Nenhum documento encontrado no período selecionado")
                        return
                    
                    # Progress dialog
                    progress = QProgressDialog("Processando documentos...", "Cancelar", 0, len(rows), dialog)
                    progress.setWindowTitle("Gerando Relatório")
                    progress.setWindowModality(Qt.WindowModal)
                    progress.show()
                    
                    # Desabilita ordenação durante preenchimento (performance)
                    table.setSortingEnabled(False)
                    
                    # Popula tabela
                    table.setRowCount(len(rows))
                    
                    # Contador para rastrear extrações de XML (fallback para notas antigas)
                    extraidos_xml = 0
                    
                    for idx, row in enumerate(rows):
                        if progress.wasCanceled():
                            break
                        
                        chave, data_emissao, tipo, numero, nome_emit, cnpj_dest, valor, informante, v_ibs_db, v_cbs_db = row
                        
                        # Garante que valor seja float para formatação correta
                        try:
                            valor = float(valor) if valor else 0.0
                        except (ValueError, TypeError):
                            valor = 0.0
                        
                        # 🚀 PERFORMANCE OTIMIZADA: Sempre usa valores do banco quando disponíveis
                        # Só extrai do XML se valores estiverem NULL (notas antigas nunca atualizadas)
                        if v_ibs_db is not None and v_cbs_db is not None:
                            # Valores já estão no banco (podem ser zero, mas isso é correto)
                            # Garante conversão para float para formatação correta
                            try:
                                ibs = float(v_ibs_db) if v_ibs_db else 0.0
                                cbs = float(v_cbs_db) if v_cbs_db else 0.0
                            except (ValueError, TypeError):
                                ibs = 0.0
                                cbs = 0.0
                        else:
                            # Fallback: Extrai do XML (apenas para notas antigas sem IBS/CBS no banco)
                            # Sugestão: Execute "Atualizar IBS/CBS" no menu para popular o banco
                            ibs_str, cbs_str = extrair_ibs_cbs_do_xml(chave)
                            # Converte strings retornadas para float
                            try:
                                ibs = float(ibs_str) if ibs_str else 0.0
                                cbs = float(cbs_str) if cbs_str else 0.0
                            except (ValueError, TypeError):
                                ibs = 0.0
                                cbs = 0.0
                            extraidos_xml += 1
                        
                        # Atualiza progresso apenas a cada 50 documentos (performance)
                        if idx % 50 == 0 or idx == len(rows) - 1:
                            progress.setLabelText(f"Processando {idx+1}/{len(rows)} documentos...")
                            progress.setValue(idx)
                        
                        # Formata data
                        try:
                            data_fmt = datetime.strptime(data_emissao, "%Y-%m-%d").strftime("%d/%m/%Y")
                        except:
                            data_fmt = data_emissao
                        
                        # Preenche linha
                        table.setItem(idx, 0, QTableWidgetItem(data_fmt))
                        table.setItem(idx, 1, QTableWidgetItem(tipo or ""))
                        table.setItem(idx, 2, QTableWidgetItem(numero or ""))
                        table.setItem(idx, 3, QTableWidgetItem(nome_emit or ""))
                        table.setItem(idx, 4, QTableWidgetItem(cnpj_dest or ""))
                        table.setItem(idx, 5, QTableWidgetItem(formatar_moeda(valor)))
                        table.setItem(idx, 6, QTableWidgetItem(formatar_moeda(ibs)))
                        table.setItem(idx, 7, QTableWidgetItem(formatar_moeda(cbs)))
                        table.setItem(idx, 8, QTableWidgetItem(chave or ""))
                    
                    progress.setValue(len(rows))
                    
                    # Reabilita ordenação após preencher
                    table.setSortingEnabled(True)
                    
                    # Monta mensagem de status com informações de performance
                    status_msg = f"✅ Relatório gerado: {len(rows)} documento(s)"
                    
                    if extraidos_xml > 0:
                        # Há notas antigas sem IBS/CBS no banco
                        status_msg += f" | ⚠️ {extraidos_xml} nota(s) extraída(s) do XML (lento)"
                        status_msg += " | 💡 Execute 'Atualizar IBS/CBS' no menu para acelerar relatórios futuros"
                    else:
                        # Todos os dados vieram do banco (performance máxima)
                        status_msg += " | 🚀 IBS/CBS do banco (rápido)"
                    
                    lbl_status.setText(status_msg)
                    btn_exportar_excel.setEnabled(True)
                    
                except Exception as e:
                    QMessageBox.critical(dialog, "Erro", f"Erro ao gerar relatório: {e}")
                    import traceback
                    traceback.print_exc()
            
            def exportar_excel():
                """Exporta dados da tabela para Excel."""
                try:
                    # Verifica se há dados
                    if table.rowCount() == 0:
                        QMessageBox.warning(dialog, "Aviso", "Não há dados para exportar")
                        return
                    
                    # Seleciona arquivo destino
                    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
                    arquivo_sugerido = f"Relatorio_IBS_CBS_{data_hora}.xlsx"
                    
                    arquivo, _ = QFileDialog.getSaveFileName(
                        dialog,
                        "Salvar Relatório Excel",
                        arquivo_sugerido,
                        "Excel Files (*.xlsx)"
                    )
                    
                    if not arquivo:
                        return
                    
                    # Importa openpyxl
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, PatternFill
                    except ImportError:
                        QMessageBox.critical(
                            dialog,
                            "Erro",
                            "Biblioteca openpyxl não instalada.\n\n"
                            "Execute: pip install openpyxl"
                        )
                        return
                    
                    # Cria workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Relatório IBS CBS"
                    
                    # Cabeçalhos
                    headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    ws.append(headers)
                    
                    # Formata cabeçalho
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Dados
                    for row_idx in range(table.rowCount()):
                        row_data = []
                        for col_idx in range(table.columnCount()):
                            item = table.item(row_idx, col_idx)
                            row_data.append(item.text() if item else "")
                        ws.append(row_data)
                    
                    # Formata colunas de valores monetários (Valor Total, IBS, CBS) - alinhamento à direita
                    for row_idx in range(2, ws.max_row + 1):
                        # Coluna 6 = Valor Total, 7 = IBS, 8 = CBS (índice 1-based)
                        for col_idx in [6, 7, 8]:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Ajusta largura das colunas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Salva arquivo
                    wb.save(arquivo)
                    
                    QMessageBox.information(
                        dialog,
                        "Sucesso",
                        f"Relatório exportado com sucesso!\n\n{arquivo}"
                    )
                    
                except Exception as e:
                    QMessageBox.critical(dialog, "Erro", f"Erro ao exportar Excel: {e}")
                    import traceback
                    traceback.print_exc()
            
            # === CONECTA SINAIS ===
            btn_atualizar.clicked.connect(atualizar_relatorio)
            btn_exportar_excel.clicked.connect(exportar_excel)
            table.itemSelectionChanged.connect(atualizar_totais)  # Atualiza totais ao selecionar linhas
            
            # Inicializa
            popular_empresas()
            
            # Mostra dialog
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir relatório: {e}")
            import traceback
            traceback.print_exc()

    def do_busca_completa(self):
        """Busca completa: reseta NSU para 0 e busca todos os XMLs da SEFAZ."""
        from datetime import datetime, timedelta
        
        try:
            # Confirma operação
            reply = QMessageBox.question(
                self,
                "Busca Completa",
                "Esta operação irá:\n\n"
                "• Resetar o NSU para 0 (zero) - NFe, CTe e NFS-e\n"
                "• Limpar todos os bloqueios de erro 656\n"
                "• Buscar TODOS os XMLs desde o início\n"
                "• Pode demorar muito tempo dependendo da quantidade\n\n"
                "⚠️ Use 'Busca na SEFAZ' para buscar apenas documentos novos.\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Marca busca em andamento
            self._search_in_progress = True
            self._next_search_time = None
            
            # ✅ Reseta NSU para 0 no banco de dados e limpa bloqueios de erro 656
            # BUSCA COMPLETA = buscar TODOS os documentos desde o início
            try:
                with sqlite3.connect(str(DB_PATH)) as conn:
                    # Busca todos os certificados (informantes)
                    informantes = conn.execute("SELECT DISTINCT informante FROM certificados").fetchall()
                    total_informantes = len(informantes)
                    
                    # ✅ Reseta NSU individualmente para cada certificado (NFe, CTe E NFS-e)
                    for (informante,) in informantes:
                        conn.execute(
                            "INSERT OR REPLACE INTO nsu (informante, ult_nsu) VALUES (?, ?)",
                            (informante, '000000000000000')
                        )
                        # Reseta também NSU do CTe
                        conn.execute(
                            "INSERT OR REPLACE INTO nsu_cte (informante, ult_nsu) VALUES (?, ?)",
                            (informante, '000000000000000')
                        )
                        # Reseta também NSU do NFS-e
                        conn.execute(
                            "INSERT OR REPLACE INTO nsu_nfse (informante, ult_nsu) VALUES (?, ?)",
                            (informante, '000000000000000')
                        )
                    
                    # Limpa todos os bloqueios de erro 656
                    conn.execute("DELETE FROM erro_656")
                    conn.commit()
                
                # Atualiza timestamp da última busca
                self.db.set_last_search_time(datetime.now().isoformat())
                
                self.set_status(f"NSU resetado para {total_informantes} certificado(s) - iniciando busca completa", 2000)
            except Exception as e:
                QMessageBox.critical(self, "Busca Completa", f"Erro ao resetar NSU: {e}")
                self._search_in_progress = False
                return
            
            # Reseta estatísticas para busca completa
            self._search_stats = {
                'nfes_found': 0,
                'ctes_found': 0,
                'nfses_found': 0,
                'start_time': datetime.now(),
                'last_cert': '',
                'total_docs': 0,
                'current_cert': 0,
                'total_certs': total_informantes
            }
            
            # Mostra progress bar com range determinado
            self.search_progress.setVisible(True)
            self.search_progress.setRange(0, total_informantes)
            self.search_progress.setValue(0)
            self.search_summary_label.setText(f"🔄 Busca Completa: 0/{total_informantes} certificados | NFes: 0 | CTes: 0 | NFSes: 0")
            
            # Inicia busca na SEFAZ
            self.set_status("🔄 Busca Completa iniciada - aguarde...", 0)

            def on_progress(line: str):
                if not line:
                    return
                
                # Atualiza progresso baseado nos logs
                try:
                    import re
                    
                    # Detecta processamento de certificado - várias variações
                    if "Processando certificado" in line or "processando certificado" in line.lower():
                        match = re.search(r'CNPJ[=:\s]+(\d+)', line, re.IGNORECASE)
                        if match:
                            cnpj = match.group(1)
                            self._search_stats['current_cert'] += 1
                            self._search_stats['last_cert'] = cnpj[-4:]
                            
                            # Atualiza progress bar
                            current = min(self._search_stats['current_cert'], total_informantes)
                            self.search_progress.setValue(current)
                            
                            # Atualiza resumo
                            elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                            self.search_summary_label.setText(
                                f"🔄 Busca Completa: {current}/{total_informantes} certificados | "
                                f"NFes: {self._search_stats['nfes_found']} | "
                                f"CTes: {self._search_stats['ctes_found']} | "
                                f"NFSes: {self._search_stats['nfses_found']} | "
                                f"Cert: ...{self._search_stats['last_cert']} | "
                                f"{elapsed:.0f}s"
                            )
                    
                    # Detecta NFe encontrada
                    if "registrar_xml" in line.lower() or "infnfe" in line.lower():
                        self._search_stats['nfes_found'] += 1
                        elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                        current = min(self._search_stats['current_cert'], total_informantes)
                        self.search_summary_label.setText(
                            f"🔄 Busca Completa: {current}/{total_informantes} certificados | "
                            f"NFes: {self._search_stats['nfes_found']} | "
                            f"CTes: {self._search_stats['ctes_found']} | "
                            f"NFSes: {self._search_stats['nfses_found']} | "
                            f"Cert: ...{self._search_stats['last_cert']} | "
                            f"{elapsed:.0f}s"
                        )
                    
                    # Detecta CTe encontrado
                    if "processar_cte" in line.lower() or "🚛" in line:
                        self._search_stats['ctes_found'] += 1
                        elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                        current = min(self._search_stats['current_cert'], total_informantes)
                        self.search_summary_label.setText(
                            f"🔄 Busca Completa: {current}/{total_informantes} certificados | "
                            f"NFes: {self._search_stats['nfes_found']} | "
                            f"CTes: {self._search_stats['ctes_found']} | "
                            f"NFSes: {self._search_stats['nfses_found']} | "
                            f"Cert: ...{self._search_stats['last_cert']} | "
                            f"{elapsed:.0f}s"
                        )
                        
                except Exception:
                    pass  # Silencioso para evitar recursão
                
                # Detecta se a busca foi finalizada
                if "Busca de NSU finalizada" in line or "Busca concluída" in line or "=== Busca concluída:" in line:
                    # Marca que a busca finalizou
                    self._search_in_progress = False
                    
                    # Oculta progress bar
                    self.search_progress.setVisible(False)
                    
                    # Mostra resumo final
                    elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                    minutos = int(elapsed / 60)
                    segundos = int(elapsed % 60)
                    
                    tempo_str = f"{minutos}min {segundos}s" if minutos > 0 else f"{segundos}s"
                    
                    self.search_summary_label.setText(
                        f"✅ Busca Completa finalizada! NFes: {self._search_stats['nfes_found']} | "
                        f"CTes: {self._search_stats['ctes_found']} | "
                        f"NFSes: {self._search_stats['nfses_found']} | "
                        f"Tempo: {tempo_str}"
                    )
                    
                    # Extrai tempo de espera (em minutos)
                    import re
                    match = re.search(r'(\d+)\s*minutos', line)
                    if match:
                        minutos_espera = int(match.group(1))
                        self._next_search_time = datetime.now() + timedelta(minutes=minutos_espera)
                        self.set_status(f"✅ Busca completa finalizada. Próxima em {minutos_espera} minutos", 3000)
                    else:
                        self.set_status("✅ Busca completa finalizada", 3000)
                    return
                
                # Logs no console
                print(line)

            # Worker thread para não travar a interface
            class SearchWorker(QThread):
                finished_search = pyqtSignal(dict)
                progress_line = pyqtSignal(str)
                
                def run(self):
                    res = run_search(progress_cb=lambda line: self.progress_line.emit(line))
                    self.finished_search.emit(res)
            
            def on_finished(res: Dict[str, Any]):
                # Força finalização da busca
                self._search_in_progress = False
                
                # Oculta progress bar
                self.search_progress.setVisible(False)
                
                if not res.get("ok"):
                    error = res.get('error') or res.get('message')
                    print(f"Erro na busca completa: {error}")
                    self.set_status(f"❌ Erro: {error[:50]}...", 5000)
                    self.search_summary_label.setText(f"❌ Erro na busca completa")
                else:
                    # Busca finalizada com sucesso
                    elapsed = (datetime.now() - self._search_stats['start_time']).total_seconds()
                    minutos = int(elapsed / 60)
                    segundos = int(elapsed % 60)
                    tempo_str = f"{minutos}min {segundos}s" if minutos > 0 else f"{segundos}s"
                    
                    self.search_summary_label.setText(
                        f"✅ Busca Completa finalizada! NFes: {self._search_stats['nfes_found']} | "
                        f"CTes: {self._search_stats['ctes_found']} | "
                        f"NFSes: {self._search_stats['nfses_found']} | "
                        f"Tempo: {tempo_str}"
                    )
                    self.set_status("✅ Busca completa finalizada", 3000)
                    
                    # Atualiza a interface com os novos dados
                    self.refresh_all()
                    
                    # 🆕 CONSULTA DE EVENTOS após Busca Completa
                    print("[PÓS-BUSCA COMPLETA] Iniciando consulta de eventos dos documentos baixados...")
                    QTimer.singleShot(3000, lambda: self._atualizar_status_apos_busca())
                    
                    # 🆕 AUTO-VERIFICAÇÃO INTELIGENTE após Busca Completa
                    print("[PÓS-BUSCA COMPLETA] Agendando auto-verificação inteligente de XMLs RESUMO...")
                    # Proteção contra AttributeError - captura método antes do timer
                    try:
                        metodo_verificacao = self._iniciar_auto_verificacao_inteligente
                        QTimer.singleShot(8000, metodo_verificacao)
                    except AttributeError:
                        print("[PÓS-BUSCA COMPLETA] ⚠️ Método _iniciar_auto_verificacao_inteligente não disponível")
                    
                    # 🆕 BUSCA DE NFS-e após Busca Completa concluir com sucesso
                    print("\n" + "="*70)
                    print("[PÓS-BUSCA COMPLETA] ✅ NF-e e CT-e concluídos!")
                    print("[PÓS-BUSCA COMPLETA] 📋 NFS-e COMPLETA será processada em 10 segundos...")
                    print("[PÓS-BUSCA COMPLETA] ℹ️  Busca desde NSU=0 (todos documentos)")
                    print("="*70 + "\n")
                    QTimer.singleShot(10000, lambda: self._buscar_nfse_automatico(busca_completa=True))
            
            # Conecta sinais e inicia worker
            self._search_worker = SearchWorker()
            self._search_worker.progress_line.connect(on_progress)
            self._search_worker.finished_search.connect(on_finished)
            self._search_worker.start()
        except Exception as e:
            QMessageBox.critical(self, "PDFs em lote", f"Erro: {e}")

    def open_logs_folder(self):
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            path = str(LOGS_DIR)
            if sys.platform == "win32":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "Logs", f"Erro ao abrir pasta de logs: {e}")

    def check_updates(self):
        """Verifica e aplica atualizações do GitHub com auto-update TRUE."""
        from modules.updater import GitHubUpdater
        from PyQt5.QtWidgets import QProgressDialog, QMessageBox
        from PyQt5.QtCore import Qt
        
        try:
            # BASE_DIR: onde estão os arquivos .py (para atualizar)
            # DATA_DIR: onde criar backups (tem permissão de escrita)
            updater = GitHubUpdater("W4lterBr/NF-e", BASE_DIR, backup_dir=DATA_DIR / "backups")
            
            # Verifica se há atualizações
            has_update, current, remote = updater.check_for_updates()
            
            if remote == "Erro ao conectar":
                QMessageBox.warning(
                    self, 
                    "Atualizações",
                    "❌ Não foi possível conectar ao servidor de atualizações.\nVerifique sua conexão com a internet."
                )
                return
            
            if not has_update:
                QMessageBox.information(
                    self,
                    "Atualizações",
                    f"✅ Você já está na versão mais recente!\n\nVersão atual: {current}"
                )
                return
            
            # Pergunta se deseja atualizar
            reply = QMessageBox.question(
                self,
                "Atualização Disponível",
                f"📦 Nova versão disponível!\n\n"
                f"Versão atual: {current}\n"
                f"Nova versão: {remote}\n\n"
                f"Deseja atualizar agora?\n\n"
                f"⚠️ O aplicativo será fechado e atualizado automaticamente.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Cria diálogo de progresso
            progress = QProgressDialog("Baixando atualizações...", "Cancelar", 0, 0, self)
            progress.setWindowTitle("Atualizando")
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(False)
            progress.setAutoReset(False)
            progress.setCancelButton(None)  # Não permitir cancelar
            progress.show()
            
            def update_progress(msg):
                progress.setLabelText(msg)
                QApplication.processEvents()
            
            # MODO EXECUTÁVEL: Usa auto-update TRUE (substitui o .exe)
            if getattr(sys, 'frozen', False):
                logger.info("🚀 Modo executável: usando auto-update TRUE")
                result = updater.update_executable(progress_callback=update_progress)
                
                progress.close()
                
                if result['success']:
                    if result.get('restart_required'):
                        # Mostra mensagem e fecha o app para o launcher substituir
                        QMessageBox.information(
                            self,
                            "Atualização Automática",
                            result['message'] + "\n\n🔄 O aplicativo será fechado agora..."
                        )
                        # Fecha o aplicativo - o launcher fará o resto
                        QApplication.quit()
                        sys.exit(0)
                    else:
                        QMessageBox.information(
                            self,
                            "Atualização",
                            result['message']
                        )
                else:
                    QMessageBox.warning(
                        self,
                        "Erro na Atualização",
                        f"❌ {result['message']}\n\n💡 Tente baixar o instalador manualmente do GitHub."
                    )
            else:
                # MODO DESENVOLVIMENTO: Atualiza arquivos .py individuais
                logger.info("🔧 Modo desenvolvimento: atualizando arquivos Python")
                result = updater.apply_update(progress_callback=update_progress)
                
                progress.close()
                
                if result['success']:
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Information)
                    msg_box.setWindowTitle("Atualização Concluída")
                    msg_box.setText(result['message'])
                    
                    if result['updated_files']:
                        details = "Arquivos atualizados:\n" + "\n".join(f"• {f}" for f in result['updated_files'])
                        msg_box.setDetailedText(details)
                    
                    msg_box.setStandardButtons(QMessageBox.Ok)
                    msg_box.exec_()
                    
                    # Pergunta se deseja reiniciar
                    reply = QMessageBox.question(
                        self,
                        "Reiniciar Aplicativo",
                        "✅ Atualização concluída!\n\nDeseja reiniciar o aplicativo agora?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    )
                    
                    if reply == QMessageBox.Yes:
                        # Atualiza o título antes de reiniciar (caso usuário cancele)
                        self._update_window_title()
                        QApplication.quit()
                        # Desenvolvimento: reinicia com Python
                        os.execl(sys.executable, sys.executable, *sys.argv)
                    else:
                        # Usuário não quer reiniciar agora - atualiza título mesmo assim
                        self._update_window_title()
                else:
                    QMessageBox.warning(
                        self,
                        "Erro na Atualização",
                        f"❌ Erro ao aplicar atualização:\n\n{result['message']}"
                    )
        
        except Exception as e:
            logger.exception("Erro ao verificar atualizações")
            QMessageBox.critical(
                self,
                "Erro",
                f"❌ Erro inesperado ao verificar atualizações:\n\n{str(e)}"
            )

    def open_xmls_folder(self):
        try:
            xmls_dir = DATA_DIR / 'xmls'
            xmls_dir.mkdir(parents=True, exist_ok=True)
            base = str(xmls_dir)
            if sys.platform == "win32":
                os.startfile(base)  # type: ignore[attr-defined]
            else:
                subprocess.Popen(["xdg-open", base])
        except Exception as e:
            QMessageBox.critical(self, "XMLs", f"Erro ao abrir pasta xmls: {e}")

    def open_certificates(self):
        try:
            dlg = CertificateDialog(self.db, self)
            dlg.exec_()
            # Após fechar, recarrega dados (caso certificados impactem buscas futuras)
            try:
                self._populate_certs_tree()
            except Exception:
                pass
            self.set_status("Certificados atualizados", 2000)
        except Exception as e:
            QMessageBox.critical(self, "Certificados", f"Erro: {e}")
    
    def open_brasilnfe_config(self):
        """Abre configuração da API BrasilNFe para manifestação."""
        try:
            dlg = BrasilNFeConfigDialog(self.db, self)
            if dlg.exec_() == QDialog.Accepted:
                self.set_status("Configuração BrasilNFe atualizada", 2000)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir configuração: {e}")

    def limpar_dados(self):
        """Limpa interface e deleta XMLs baixados da SEFAZ."""
        try:
            # Confirmar ação
            reply = QMessageBox.question(
                self, 
                "Limpar Dados",
                "Esta ação irá:\n\n"
                "• Limpar a interface (tabela)\n"
                "• Deletar TODOS os XMLs baixados da SEFAZ\n"
                "• Limpar registros do banco de dados\n\n"
                "Esta operação NÃO pode ser desfeita!\n\n"
                "Deseja continuar?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            self.set_status("Limpando dados…")
            QApplication.processEvents()
            
            # 1. Limpar tabela da interface
            self.notes = []
            self.table.clearContents()
            self.table.setRowCount(0)
            
            # 2. Limpar banco de dados
            try:
                with sqlite3.connect(str(DB_PATH)) as conn:
                    # Limpar tabelas principais
                    conn.execute("DELETE FROM notas_detalhadas")
                    conn.execute("DELETE FROM xmls_baixados")
                    conn.execute("DELETE FROM nf_status")
                    # ⚠️ IMPORTANTE: NÃO deletar NSU! É o histórico de sincronização com SEFAZ
                    # Se deletar, sistema volta para NSU=0 e baixa TUDO novamente
                    # conn.execute("DELETE FROM nsu")  # ❌ REMOVIDO - não deve ser zerado
                    # conn.execute("DELETE FROM nsu_cte")  # ❌ REMOVIDO - não deve ser zerado
                    conn.commit()
            except Exception as e:
                QMessageBox.warning(self, "Limpar", f"Erro ao limpar banco de dados: {e}")
            
            # 3. Deletar XMLs da pasta xmls/
            xmls_folder = DATA_DIR / "xmls"
            deleted_count = 0
            try:
                if xmls_folder.exists():
                    import shutil
                    # Percorre todas as subpastas e deleta XMLs
                    for item in xmls_folder.iterdir():
                        if item.is_dir():
                            try:
                                shutil.rmtree(item)
                                deleted_count += 1
                            except Exception:
                                pass
                        elif item.suffix.lower() == '.xml':
                            try:
                                item.unlink()
                                deleted_count += 1
                            except Exception:
                                pass
            except Exception as e:
                QMessageBox.warning(self, "Limpar", f"Erro ao deletar XMLs: {e}")
            
            self.set_status(f"Dados limpos com sucesso ({deleted_count} itens removidos)", 3000)
            QMessageBox.information(
                self, 
                "Limpar Dados",
                f"Operação concluída!\n\n"
                f"• Interface limpa\n"
                f"• Banco de dados resetado\n"
                f"• {deleted_count} XMLs/pastas removidos"
            )
            
        except Exception as e:
            self.set_status("")
            QMessageBox.critical(self, "Limpar", f"Erro ao limpar dados: {e}")

    def open_storage_config(self):
        """Abre diálogo de configuração de armazenamento"""
        try:
            dlg = StorageConfigDialog(self.db, self)
            dlg.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Armazenamento", f"Erro: {e}")
    
    def _abrir_gerenciador_trabalhos(self):
        """Abre o Gerenciador de Trabalhos"""
        try:
            # Verifica se já existe uma instância aberta
            if hasattr(self, '_gerenciador_dialog') and self._gerenciador_dialog and self._gerenciador_dialog.isVisible():
                # Traz para frente se já está aberta
                self._gerenciador_dialog.raise_()
                self._gerenciador_dialog.activateWindow()
                return
            
            # Cria nova instância NÃO-MODAL para não bloquear a interface
            self._gerenciador_dialog = GerenciadorTrabalhosDialog(self)
            self._gerenciador_dialog.show()  # show() em vez de exec_() para não bloquear
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao abrir gerenciador: {e}")
    
    def _check_inatividade(self):
        """Verifica inatividade e inicia sincronização automática"""
        try:
            # ⛔ DESABILITADO: Consulta de eventos não deve ser automática
            # A consulta de eventos só deve ocorrer:
            # 1. Após busca na SEFAZ (distribuição DFe)
            # 2. Ao clicar no botão "🔄 Atualizar Status"
            # 3. Ao clicar no botão "Sincronizar Agora"
            return
            
            if not hasattr(self, '_ultimo_evento_usuario') or not hasattr(self, '_sync_worker'):
                return
            
            # Se usuário cancelou a sincronização, não inicia automaticamente
            if self._sync_cancelada_pelo_usuario:
                return
            
            tempo_inativo = (datetime.now() - self._ultimo_evento_usuario).total_seconds()
            
            if tempo_inativo > 30 and not self._sync_worker:
                print("[AUTO-SYNC] Usuário inativo há 30s, iniciando sincronização automática...")
                self._iniciar_sync_background()
        except Exception as e:
            print(f"[AUTO-SYNC] Erro ao verificar inatividade: {e}")
    
    def _iniciar_sync_background(self):
        """Inicia sincronização de eventos em background"""
        print("[SYNC BACKGROUND] Função _iniciar_sync_background() chamada")
        if self._sync_worker:
            print("[SYNC BACKGROUND] Worker já existe, abortando")
            return
        
        try:
            from PyQt5.QtCore import QThread, pyqtSignal, QObject
        except ImportError:
            print("[SYNC BACKGROUND] ERRO: Não foi possível importar PyQt5")
            return
        
        # Usa TODOS os documentos do banco, não apenas os filtrados
        try:
            docs = self.notes  # TODOS os documentos, não apenas os visíveis
            print(f"[SYNC BACKGROUND] Total de documentos carregados: {len(docs)}")
            if not docs:
                print("[SYNC BACKGROUND] Nenhum documento encontrado, abortando")
                return
        except Exception:
            return
        
        class SyncWorker(QObject):
            progress = pyqtSignal(str, int, int)
            finished = pyqtSignal()
            error = pyqtSignal(str)
            
            def __init__(self, parent_window):
                super().__init__()
                self.parent = parent_window
                self._pausado = False
                self._cancelado = False
            
            def pausar(self):
                self._pausado = True
            
            def retomar(self):
                self._pausado = False
            
            def cancelar(self):
                self._cancelado = True
            
            def run(self):
                try:
                    # Usa TODOS os documentos do banco
                    docs = self.parent.notes
                    if not docs:
                        self.finished.emit()
                        return
                    
                    print(f"[SYNC] Total de documentos no banco: {len(docs)}")
                    
                    total = len(docs)
                    self.progress.emit(f"Iniciando sincronização de {total} documentos...", 0, total)
                    
                    # Salva estado inicial
                    primeira_chave = docs[0].get('chave', '') if docs else ''
                    self.parent.db.save_sync_state(primeira_chave, total, 0)
                    
                    # Extrai chaves dos documentos (pula eventos e já cancelados)
                    chaves_para_consultar = []
                    eventos_pulados = 0
                    cancelados_pulados = 0
                    
                    for item in docs:
                        chave = item.get('chave', '')
                        xml_status = (item.get('xml_status') or '').upper()
                        
                        # Debug: conta eventos
                        if xml_status == 'EVENTO':
                            eventos_pulados += 1
                            continue
                        
                        # Pula eventos e documentos já cancelados
                        if chave and len(chave) == 44:
                            if self.parent.db.is_chave_cancelada(chave):
                                cancelados_pulados += 1
                            else:
                                chaves_para_consultar.append(chave)
                    
                    print(f"[SYNC] Eventos pulados: {eventos_pulados}")
                    print(f"[SYNC] Já cancelados pulados: {cancelados_pulados}")
                    print(f"[SYNC] Chaves para consultar: {len(chaves_para_consultar)}")
                    
                    if not chaves_para_consultar:
                        self.progress.emit("Nenhum documento para sincronizar", total, total)
                        self.finished.emit()
                        return
                    
                    self.progress.emit(f"Sincronizando {len(chaves_para_consultar)} documentos em paralelo...", 0, len(chaves_para_consultar))
                    
                    # Usa função paralela otimizada
                    def progress_callback(atual, total_chaves, chave):
                        if self._cancelado:
                            raise Exception("Cancelado pelo usuário")
                        
                        while self._pausado and not self._cancelado:
                            import time
                            time.sleep(0.5)
                        
                        numero = chave[:10]
                        self.progress.emit(f"Doc {numero}...", atual, total_chaves)
                        
                        # Salva progresso a cada 5 documentos
                        if atual % 5 == 0:
                            self.parent.db.save_sync_state(chave, total_chaves, atual)
                    
                    # Chama função otimizada com paralelização (5 workers)
                    from nfe_search import atualizar_status_notas_lote
                    certs = self.parent.db.load_certificates()
                    
                    stats = atualizar_status_notas_lote(
                        self.parent.db,
                        certs,
                        chaves_para_consultar,
                        progress_callback,
                        max_workers=5  # 5 consultas simultâneas
                    )
                    
                    # Limpa estado ao concluir
                    self.parent.db.clear_sync_state()
                    
                    # Mensagem final com estatísticas detalhadas
                    consultadas = stats.get('consultadas', 0)
                    canceladas = stats.get('canceladas', 0)
                    
                    # Conta quantos documentos de cada tipo foram consultados
                    nfes_count = sum(1 for ch in chaves_para_consultar if len(ch) == 44 and ch[20:22] == '55')
                    ctes_count = sum(1 for ch in chaves_para_consultar if len(ch) == 44 and ch[20:22] == '57')
                    
                    msg = f"✅ Sincronização concluída! Consultadas: {consultadas} (NF-e: {nfes_count}, CT-e: {ctes_count}) | Canceladas: {canceladas}"
                    self.progress.emit(msg, len(chaves_para_consultar), len(chaves_para_consultar))
                    
                    self.finished.emit()
                except Exception as e:
                    print(f"[SYNC] Erro: {e}")
                    self.error.emit(str(e))
        
        self._sync_thread = QThread()
        self._sync_worker = SyncWorker(self)
        self._sync_worker.moveToThread(self._sync_thread)
        
        self._sync_thread.started.connect(self._sync_worker.run)
        self._sync_worker.finished.connect(self._on_sync_finished)
        self._sync_worker.error.connect(self._on_sync_error)
        self._sync_worker.progress.connect(self._on_sync_progress)
        
        trabalho = {
            'id': datetime.now().timestamp(),
            'nome': 'Sincronização de Eventos',
            'tipo': 'sync_eventos',
            'status': 'Em execução',
            'progresso': 0,
            'total': 0,
            'mensagem': 'Iniciando...',
            'worker': self._sync_worker
        }
        self._trabalhos_ativos.append(trabalho)
        self._sync_thread.start()
    
    def _on_sync_progress(self, mensagem, atual, total):
        """Atualiza progresso da sincronização"""
        try:
            for trabalho in self._trabalhos_ativos:
                if trabalho.get('tipo') == 'sync_eventos' and trabalho.get('status') == 'Em execução':
                    trabalho['progresso'] = atual
                    trabalho['total'] = total
                    trabalho['mensagem'] = mensagem
                    break
            self.statusBar().showMessage(f"Sincronizando: {mensagem} ({atual}/{total})", 2000)
        except Exception:
            pass
    
    def _on_sync_finished(self):
        """Finaliza sincronização"""
        try:
            if self._sync_thread:
                self._sync_thread.quit()
                self._sync_thread.wait()
            self._sync_worker = None
            self._sync_thread = None
            self._trabalhos_ativos = [t for t in self._trabalhos_ativos if t.get('tipo') != 'sync_eventos']
            
            # FORÇA recarregar dados do banco
            print("[SYNC] Recarregando dados do banco...")
            self.notes = self.db.load_notes(limit=5000)
            print(f"[SYNC] {len(self.notes)} notas carregadas")
            
            # Atualiza tabela para mostrar status atualizados (SEM recarregar dados)
            self._refresh_table_only()
            
            self.statusBar().showMessage("✅ Sincronização concluída!", 5000)
        except Exception:
            pass
    
    def _on_sync_error(self, erro):
        """Trata erros da sincronização"""
        try:
            if self._sync_thread:
                self._sync_thread.quit()
                self._sync_thread.wait()
            self._sync_worker = None
            self._sync_thread = None
            self._trabalhos_ativos = [t for t in self._trabalhos_ativos if t.get('tipo') != 'sync_eventos']
            # Não limpa o estado em caso de erro - permite retomar
            self.statusBar().showMessage(f"Erro: {erro}", 5000)
        except Exception:
            pass
    
    def _retomar_sync_background(self, estado: dict):
        """Retoma sincronização de onde parou."""
        if self._sync_worker:
            return
        
        try:
            from PyQt5.QtCore import QThread, pyqtSignal, QObject
        except ImportError:
            return
        
        try:
            docs = self.filtered() if self.tabs.currentIndex() == 0 else self.filtered_emitidos()
            if not docs:
                self.db.clear_sync_state()
                return
            
            # Encontra o índice da última chave processada
            ultima_chave = estado.get('ultima_chave', '')
            idx_inicio = 0
            
            if ultima_chave:
                for idx, doc in enumerate(docs):
                    if doc.get('chave') == ultima_chave:
                        idx_inicio = idx + 1  # Começa no próximo
                        break
            
            # Se já processou todos, limpa estado
            if idx_inicio >= len(docs):
                self.db.clear_sync_state()
                QMessageBox.information(
                    self,
                    "Sincronização Concluída",
                    "Todos os documentos já foram processados!"
                )
                return
            
            # Cria slice dos docs restantes
            docs_restantes = docs[idx_inicio:]
            processados_anteriormente = estado.get('docs_processados', 0)
            total_original = estado.get('total_docs', len(docs))
            
            print(f"[SYNC-RETOMAR] Retomando do documento {idx_inicio+1}/{len(docs)}")
            
        except Exception as e:
            print(f"[SYNC-RETOMAR] Erro ao preparar retomada: {e}")
            return
        
        class SyncWorkerRetomar(QObject):
            progress = pyqtSignal(str, int, int)
            finished = pyqtSignal()
            error = pyqtSignal(str)
            
            def __init__(self, parent_window, docs_restantes, idx_inicio, processados_ant, total_orig):
                super().__init__()
                self.parent = parent_window
                self.docs = docs_restantes
                self.idx_inicio = idx_inicio
                self.processados_anteriormente = processados_ant
                self.total_original = total_orig
                self._pausado = False
                self._cancelado = False
            
            def pausar(self):
                self._pausado = True
            
            def retomar(self):
                self._pausado = False
            
            def cancelar(self):
                self._cancelado = True
            
            def run(self):
                try:
                    if not self.docs:
                        self.finished.emit()
                        return
                    
                    total_docs = self.total_original
                    self.progress.emit(f"Retomando... {len(self.docs)} docs restantes", 
                                     self.processados_anteriormente, total_docs)
                    
                    for idx, item in enumerate(self.docs):
                        if self._cancelado:
                            break
                        
                        while self._pausado and not self._cancelado:
                            import time
                            time.sleep(0.5)
                        
                        if self._cancelado:
                            break
                        
                        chave = item.get('chave', '')
                        numero = item.get('numero', chave[:10])
                        
                        # Pula documentos cancelados
                        if chave and self.parent.db.is_chave_cancelada(chave):
                            atual_global = self.processados_anteriormente + idx + 1
                            self.progress.emit(f"Doc {numero} (cancelado - pulado)", atual_global, total_docs)
                            continue
                        
                        atual_global = self.processados_anteriormente + idx + 1
                        self.progress.emit(f"Doc {numero}", atual_global, total_docs)
                        
                        try:
                            informante = item.get('informante', '')
                            certs = self.parent.db.load_certificates()
                            cert_uf = next((c for c in certs if c.get('informante') == informante), certs[0] if certs else None)
                            
                            if cert_uf and chave and len(chave) == 44:
                                from nfe_search import NFeService, salvar_xml_por_certificado
                                from lxml import etree
                                
                                service = NFeService(
                                    cert_path=cert_uf.get('caminho', ''),
                                    senha=cert_uf.get('senha', ''),
                                    informante=cert_uf.get('informante', ''),
                                    cuf=cert_uf.get('cUF_autor', '50')
                                )
                                
                                resposta_xml = service.consultar_eventos_chave(chave)
                                if resposta_xml:
                                    root = etree.fromstring(resposta_xml.encode('utf-8'))
                                    eventos = root.findall('.//{http://www.portalfiscal.inf.br/nfe}retEvento')
                                    
                                    for evento in eventos:
                                        tp_evento = evento.findtext('.//{http://www.portalfiscal.inf.br/nfe}tpEvento')
                                        cstat = evento.findtext('.//{http://www.portalfiscal.inf.br/nfe}cStat')
                                        
                                        if cstat in ['135', '136'] and tp_evento:
                                            evento_xml_str = etree.tostring(evento, encoding='utf-8').decode('utf-8')
                                            
                                            # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local + TODOS os perfis ativos
                                            cnpj_informante = cert_uf.get('informante')
                                            nome_cert = cert_uf.get('nome_certificado')
                                            
                                            # 1. Salva em backup local (xmls/)
                                            salvar_xml_por_certificado(evento_xml_str, cnpj_informante, pasta_base="xmls")
                                            
                                            # 2. Salva em TODOS os perfis ativos (pasta_base=None)
                                            salvar_xml_por_certificado(evento_xml_str, cnpj_informante, pasta_base=None, nome_certificado=nome_cert)
                                            
                                            if tp_evento == '110111' and cstat == '135':
                                                self.parent.db.marcar_chave_cancelada(chave, 'Cancelamento de NF-e')
                                                print(f"[SYNC] Chave {chave[:10]}... marcada como cancelada")
                                            
                                            if tp_evento in ['210200', '210210', '210220', '210240']:
                                                self.parent.db.register_manifestacao(
                                                    chave, tp_evento, cert_uf.get('informante'),
                                                    datetime.now().isoformat()
                                                )
                        except Exception as e:
                            print(f"[SYNC] Erro: {e}")
                        
                        # Salva progresso a cada 5 documentos
                        if (idx + 1) % 5 == 0:
                            self.parent.db.save_sync_state(chave, total_docs, atual_global)
                        
                        import time
                        time.sleep(1.5)
                    
                    # Limpa estado ao concluir
                    self.parent.db.clear_sync_state()
                    self.finished.emit()
                except Exception as e:
                    self.error.emit(str(e))
        
        self._sync_thread = QThread()
        self._sync_worker = SyncWorkerRetomar(
            self, docs_restantes, idx_inicio, 
            processados_anteriormente, total_original
        )
        self._sync_worker.moveToThread(self._sync_thread)
        
        self._sync_thread.started.connect(self._sync_worker.run)
        self._sync_worker.finished.connect(self._on_sync_finished)
        self._sync_worker.error.connect(self._on_sync_error)
        self._sync_worker.progress.connect(self._on_sync_progress)
        
        trabalho = {
            'id': datetime.now().timestamp(),
            'nome': 'Sincronização de Eventos (Retomada)',
            'tipo': 'sync_eventos',
            'status': 'Em execução',
            'progresso': processados_anteriormente,
            'total': total_original,
            'mensagem': f'Retomando do doc {idx_inicio+1}...',
            'worker': self._sync_worker
        }
        self._trabalhos_ativos.append(trabalho)
        self._sync_thread.start()


class GerenciadorTrabalhosDialog(QDialog):
    """Dialog estilo Windows Task Manager para gerenciar trabalhos em background"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("Gerenciador de Trabalhos")
        self.resize(1000, 600)
        self.setModal(False)
        
        # Estilo global do dialog
        self.setStyleSheet("""
            QDialog {
                background-color: #fafafa;
            }
        """)
        
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header estilo Windows moderno
        header = QWidget()
        header.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                          stop:0 #0078d4, stop:1 #005a9e);
                border-bottom: 3px solid #004578;
            }
        """)
        header_layout = QVBoxLayout()
        header_layout.setContentsMargins(25, 20, 25, 20)
        header_layout.setSpacing(8)
        
        # Título com ícone
        titulo_layout = QHBoxLayout()
        titulo_layout.setSpacing(10)
        
        icone_label = QLabel("⚙️")
        icone_label.setTextFormat(Qt.PlainText)
        icone_label.setTextInteractionFlags(Qt.NoTextInteraction)
        icone_label.setOpenExternalLinks(False)
        icone_label.setWordWrap(False)
        icone_label.setFocusPolicy(Qt.NoFocus)
        icone_label.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        icone_font = QFont()
        icone_font.setFamily("Segoe UI Emoji")
        icone_font.setPointSize(56)
        icone_font.setUnderline(False)
        icone_font.setStyleStrategy(QFont.PreferAntialias)
        icone_label.setFont(icone_font)
        icone_label.setAlignment(Qt.AlignCenter)
        icone_label.setStyleSheet("""
            QLabel {
                color: white;
                background: transparent;
                text-decoration: none;
                border: none;
                border-bottom: none;
                outline: none;
            }
        """)
        titulo_layout.addWidget(icone_label)
        
        titulo = QLabel("Gerenciador de Trabalhos")
        titulo.setTextFormat(Qt.PlainText)
        titulo.setTextInteractionFlags(Qt.NoTextInteraction)
        titulo.setOpenExternalLinks(False)
        titulo.setWordWrap(False)
        titulo.setFocusPolicy(Qt.NoFocus)
        titulo.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        # Configura fonte explicitamente sem underline
        titulo_font = QFont()
        titulo_font.setFamily("Segoe UI")
        titulo_font.setPointSize(20)
        titulo_font.setBold(True)
        titulo_font.setUnderline(False)
        titulo_font.setStyleStrategy(QFont.PreferAntialias)
        titulo.setFont(titulo_font)
        titulo.setStyleSheet("""
            QLabel {
                color: white;
                background: transparent;
                padding-left: 10px;
                border: none;
                text-decoration: none;
                border-bottom: none;
                outline: none;
            }
        """)
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        
        header_layout.addLayout(titulo_layout)
        
        subtitulo = QLabel("Acompanhe e controle todas as tarefas em segundo plano")
        subtitulo.setTextFormat(Qt.PlainText)
        subtitulo.setTextInteractionFlags(Qt.NoTextInteraction)
        subtitulo.setOpenExternalLinks(False)
        subtitulo.setWordWrap(False)
        subtitulo.setFocusPolicy(Qt.NoFocus)
        subtitulo.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        # Configura fonte explicitamente sem underline
        subtitulo_font = QFont()
        subtitulo_font.setFamily("Segoe UI")
        subtitulo_font.setPointSize(11)
        subtitulo_font.setUnderline(False)
        subtitulo_font.setStyleStrategy(QFont.PreferAntialias)
        subtitulo.setFont(subtitulo_font)
        subtitulo.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.9);
                background: transparent;
                padding-left: 52px;
                border: none;
                text-decoration: none;
                border-bottom: none;
                outline: none;
            }
        """)
        header_layout.addWidget(subtitulo)
        
        header.setLayout(header_layout)
        main_layout.addWidget(header)
        
        # Área de conteúdo com barra de ferramentas
        content = QWidget()
        content.setStyleSheet("background-color: #fafafa;")
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(15)
        
        # Barra de ferramentas
        toolbar = QWidget()
        toolbar.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setContentsMargins(10, 5, 10, 5)
        
        btn_atualizar = QPushButton("🔄 Atualizar")
        btn_atualizar.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #005a9e;
            }
            QPushButton:pressed {
                background-color: #004578;
            }
        """)
        btn_atualizar.clicked.connect(self._atualizar_lista)
        toolbar_layout.addWidget(btn_atualizar)
        
        # Botão Agendar Tarefa
        btn_agendar = QPushButton("⏰ Agendar Tarefa")
        btn_agendar.setToolTip("Configure tarefas automáticas")
        btn_agendar.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)
        btn_agendar.clicked.connect(self._abrir_agendador)
        toolbar_layout.addWidget(btn_agendar)
        
        toolbar_layout.addStretch()
        
        info_label = QLabel("⏱ Atualização automática a cada 1 segundo")
        info_label.setStyleSheet("""
            color: #666;
            font-size: 10px;
            background: transparent;
        """)
        toolbar_layout.addWidget(info_label)
        
        toolbar.setLayout(toolbar_layout)
        content_layout.addWidget(toolbar)
        
        # Tabela de trabalhos com sombra
        table_container = QWidget()
        table_container.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 6px;
                border: 1px solid #e0e0e0;
            }
        """)
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["📋 Tarefa", "📊 Status", "📈 Progresso", "🎮 Ações"])
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.table.setColumnWidth(1, 140)
        self.table.setColumnWidth(3, 280)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #f0f0f0;
                border: none;
                border-radius: 6px;
                selection-background-color: #e3f2fd;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: #0078d4;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                          stop:0 #f8f8f8, stop:1 #ececec);
                padding: 12px 8px;
                border: none;
                border-right: 1px solid #e0e0e0;
                border-bottom: 2px solid #0078d4;
                font-weight: bold;
                font-size: 11px;
                color: #333;
            }
            QHeaderView::section:first {
                border-top-left-radius: 6px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 6px;
                border-right: none;
            }
        """)
        
        table_layout.addWidget(self.table)
        table_container.setLayout(table_layout)
        content_layout.addWidget(table_container)
        
        # Rodapé com informações detalhadas
        footer = QWidget()
        footer.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 12px;
            }
        """)
        footer_layout = QHBoxLayout()
        footer_layout.setContentsMargins(15, 10, 15, 10)
        
        self.status_label = QLabel("ℹ Nenhum trabalho em execução")
        self.status_label.setStyleSheet("""
            QLabel {
                background-color: transparent;
                color: #666;
                font-size: 12px;
                font-weight: bold;
            }
        """)
        footer_layout.addWidget(self.status_label)
        
        footer_layout.addStretch()
        
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("""
            QLabel {
                background-color: transparent;
                color: #999;
                font-size: 10px;
            }
        """)
        footer_layout.addWidget(self.info_label)
        
        footer.setLayout(footer_layout)
        content_layout.addWidget(footer)
        
        content.setLayout(content_layout)
        main_layout.addWidget(content)
        
        self.setLayout(main_layout)
        
        # Lista de workers ativos
        self.workers = []
        
        # Timer para atualizar a lista
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self._atualizar_lista)
        self.update_timer.start(1000)
        
        # Atualização inicial
        self._atualizar_lista()
    
    def _atualizar_status_lote(self):
        """Atualiza status de todas as notas consultando eventos na SEFAZ"""
        # Verifica se já há atualização em andamento
        if hasattr(self, '_update_worker') and self._update_worker and self._update_worker.isRunning():
            QMessageBox.warning(self, "Aviso", "Já há uma atualização de status em andamento!")
            return
        
        # Confirma ação
        reply = QMessageBox.question(
            self,
            "Atualizar Status",
            "Deseja consultar o status atual de TODAS as notas na SEFAZ?\n\n"
            "⚠️ Esta operação pode demorar alguns minutos dependendo da quantidade de notas.\n\n"
            "Isso irá detectar:\n"
            "• Notas canceladas\n"
            "• Cartas de correção\n"
            "• Outros eventos",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        # Obtém certificados
        certs = self.parent_window.db.load_certificates() if self.parent_window else []
        if not certs:
            QMessageBox.warning(self, "Erro", "Nenhum certificado configurado!")
            return
        
        # Obtém lista de chaves (apenas notas com status "Autorizado" para otimizar)
        notas = self.parent_window.notes if self.parent_window else []
        chaves = []
        for nota in notas:
            status = (nota.get('status') or '').lower()  # CORRIGIDO: campo é 'status'
            chave = nota.get('chave')
            # Consulta apenas notas "autorizadas" (não consulta já canceladas)
            if chave and len(chave) == 44 and 'autoriza' in status:
                chaves.append(chave)
        
        if not chaves:
            QMessageBox.information(self, "Info", "Nenhuma nota autorizada para atualizar.")
            return
        
        # Cria diálogo de progresso NÃO-MODAL
        progress = QProgressDialog("Consultando status das notas...", "Cancelar", 0, len(chaves), self)
        progress.setWindowTitle("Atualizando Status")
        progress.setWindowModality(Qt.NonModal)  # NÃO-MODAL para não bloquear
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()  # Mostra explicitamente
        
        # Variável para controlar cancelamento
        cancelado = [False]
        
        # Função de callback para atualizar progresso
        def progress_callback(current, total, chave):
            if progress.wasCanceled() or cancelado[0]:
                cancelado[0] = True
                raise Exception("Cancelado pelo usuário")
            progress.setLabelText(f"Consultando {current}/{total}...\nChave: {chave[:15]}...")
            progress.setValue(current)
            # Removido QApplication.processEvents() que pode causar travamentos
        
        # Executa atualização em thread
        from PyQt5.QtCore import QThread, pyqtSignal
        
        class UpdateStatusWorker(QThread):
            finished = pyqtSignal(dict)
            error = pyqtSignal(str)
            
            def __init__(self, db, certs, chaves, callback):
                super().__init__()
                self.db = db
                self.certs = certs
                self.chaves = chaves
                self.callback = callback
            
            def run(self):
                try:
                    from nfe_search import atualizar_status_notas_lote
                    stats = atualizar_status_notas_lote(
                        self.db,
                        self.certs,
                        self.chaves,
                        self.callback
                    )
                    self.finished.emit(stats)
                except Exception as e:
                    self.error.emit(str(e))
        
        def on_finished(stats):
            progress.close()
            
            msg = f"✅ Atualização concluída!\n\n"
            msg += f"📊 Estatísticas:\n"
            msg += f"• Notas consultadas: {stats.get('consultadas', 0)}\n"
            msg += f"• Notas canceladas encontradas: {stats.get('canceladas', 0)}\n"
            msg += f"• Status atualizados: {stats.get('atualizadas', 0)}\n"
            msg += f"• Erros: {stats.get('erros', 0)}\n\n"
            msg += f"A tabela será atualizada automaticamente."
            
            QMessageBox.information(self, "Concluído", msg)
            
            # Limpa referência ao worker
            self._update_worker = None
            
            # FORÇA recarregar dados do banco
            if self.parent_window:
                print("[UPDATE-STATUS] Recarregando dados do banco...")
                self.parent_window.notes = self.parent_window.db.load_notes(limit=5000)
                print(f"[UPDATE-STATUS] {len(self.parent_window.notes)} notas carregadas")
                
                # Atualiza visualização (SEM recarregar dados novamente)
                self.parent_window._refresh_table_only()
        
        def on_error(error_msg):
            progress.close()
            
            # Limpa referência ao worker
            self._update_worker = None
            
            if "Cancelado" not in error_msg:
                QMessageBox.warning(self, "Erro", f"Erro ao atualizar status:\n{error_msg}")
        
        worker = UpdateStatusWorker(
            self.parent_window.db,
            certs,
            chaves,
            progress_callback
        )
        worker.finished.connect(on_finished)
        worker.error.connect(on_error)
        worker.start()
        
        # Mantém referência ao worker
        self._update_worker = worker
    
    def _iniciar_sync_manual(self):
        """Inicia sincronização manual"""
        print("[SYNC MANUAL] Botão Sincronizar Agora clicado")
        if self.parent_window and hasattr(self.parent_window, '_iniciar_sync_background'):
            print("[SYNC MANUAL] Chamando _iniciar_sync_background()")
            # Reabilita auto-sync quando usuário inicia manualmente
            self.parent_window._sync_cancelada_pelo_usuario = False
            self.parent_window._iniciar_sync_background()
            QMessageBox.information(
                self,
                "Sincronização Iniciada",
                "A sincronização de eventos foi iniciada em segundo plano.\n\n"
                "Você pode acompanhar o progresso nesta janela."
            )
        else:
            print("[SYNC MANUAL] ERRO: parent_window ou _iniciar_sync_background não disponível")
    
    def _reprocessar_resumos(self):
        """Reprocessa notas com status RESUMO (resNFe) e baixa XMLs completos"""
        if not self.parent_window:
            return
        
        # Confirma ação
        reply = QMessageBox.question(
            self,
            "Reprocessar Resumos (resNFe)",
            "Esta função irá:\n\n"
            "1. Buscar todas as notas com status RESUMO\n"
            "2. Para cada nota, buscar o XML completo na SEFAZ por chave\n"
            "3. Salvar os XMLs completos e atualizar o banco\n\n"
            "⚠️ Esta operação pode demorar alguns minutos.\n\n"
            "Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # Cria worker para reprocessar
        worker = ReprocessarResumosWorker(self.parent_window)
        
        # Registra worker
        self.workers.append(worker)
        
        # Conecta signals
        def on_finished_reprocessar(encontrados, total):
            QMessageBox.information(
                self,
                "Reprocessamento Concluído",
                f"✅ Processo concluído!\n\n"
                f"• Total processado: {total}\n"
                f"• XMLs completos encontrados: {encontrados}\n\n"
                f"A interface será atualizada."
            )
            # Remove worker da lista
            if worker in self.workers:
                self.workers.remove(worker)
            # Atualiza interface
            if self.parent_window:
                self.parent_window.refresh_all()
        
        def on_error_reprocessar(msg):
            QMessageBox.critical(self, "Erro", f"Erro no reprocessamento:\n{msg}")
            # Remove worker da lista
            if worker in self.workers:
                self.workers.remove(worker)
        
        worker.finished.connect(on_finished_reprocessar)
        worker.error.connect(on_error_reprocessar)
        
        # Inicia thread
        worker.start()
        
        QMessageBox.information(
            self,
            "Reprocessamento Iniciado",
            "🔄 O reprocessamento de resumos foi iniciado em segundo plano.\n\n"
            "Você pode continuar usando o sistema normalmente."
        )
    
    def _iniciar_auto_verificacao(self):
        """Inicia auto-verificação de notas com status RESUMO em background"""
        print("[DEBUG AUTO-VERIFICAÇÃO] Método _iniciar_auto_verificacao chamado")
        
        if not self.parent_window:
            print("[DEBUG AUTO-VERIFICAÇÃO] ERRO: parent_window não definido!")
            QMessageBox.warning(
                self,
                "Erro",
                "Erro interno: parent_window não definido.\n\n"
                "Tente fechar e reabrir o Gerenciador de Trabalhos."
            )
            return
        
        print("[DEBUG AUTO-VERIFICAÇÃO] parent_window OK, verificando database...")
        
        if not hasattr(self.parent_window, 'db') or not self.parent_window.db:
            print("[DEBUG AUTO-VERIFICAÇÃO] ERRO: Database não disponível!")
            QMessageBox.warning(
                self,
                "Erro",
                "Erro interno: Database não disponível."
            )
            return
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Database OK, iniciando contagem...")
        
        # Define AutoVerificacaoWorker ANTES de usar
        from PyQt5.QtCore import QThread, pyqtSignal
        
        class AutoVerificacaoWorker(QThread):
            progress = pyqtSignal(str, int, int)  # mensagem, atual, total
            finished = pyqtSignal(int, int)  # encontrados, nao_encontrados
            error = pyqtSignal(str)
            log_message = pyqtSignal(str)  # Para enviar logs ao terminal
            
            def __init__(self, parent_window, notas_resumo):
                super().__init__()
                self.parent_window = parent_window
                self.notas_resumo = notas_resumo
                self._cancelado = False
                self._pausado = False
                self.buscar_notas_na_thread = False  # Flag para buscar notas na thread
            
            def cancelar(self):
                self._cancelado = True
            
            def pausar(self):
                self._pausado = True
                self.log_message.emit("[AUTO-VERIFICAÇÃO] ⏸️ Pausando...")
            
            def retomar(self):
                self._pausado = False
                self.log_message.emit("[AUTO-VERIFICAÇÃO] ▶️ Retomando...")
            
            def log(self, msg: str):
                """Envia log para o terminal"""
                self.log_message.emit(msg)
            
            def run(self):
                print("[DEBUG WORKER] ========== AutoVerificacaoWorker.run() INICIADO ==========")
                try:
                    print("[DEBUG WORKER] Importando módulos...")
                    from nfe_search import DatabaseManager, NFeService, salvar_xml_por_certificado
                    from lxml import etree
                    
                    print(f"[DEBUG WORKER] parent_window: {self.parent_window}")
                    print(f"[DEBUG WORKER] parent_window.db: {self.parent_window.db if self.parent_window else 'N/A'}")
                    print(f"[DEBUG WORKER] parent_window.db.db_path: {self.parent_window.db.db_path if self.parent_window and self.parent_window.db else 'N/A'}")
                    
                    print("[DEBUG WORKER] Criando DatabaseManager...")
                    db_nfe = DatabaseManager(str(self.parent_window.db.db_path))
                    print("[DEBUG WORKER] DatabaseManager criado!")
                    
                    print(f"[DEBUG WORKER] buscar_notas_na_thread: {self.buscar_notas_na_thread}")
                    
                    # Se flag ativada, busca notas dentro da thread (não trava UI)
                    if self.buscar_notas_na_thread:
                        print("[DEBUG WORKER] Entrando no bloco buscar_notas_na_thread...")
                        self.log("[AUTO-VERIFICAÇÃO] 🔍 Buscando notas com status RESUMO no banco de dados...")
                        self.progress.emit("Buscando notas RESUMO...", 0, 100)
                        
                        try:
                            # Busca direto do banco (muito mais rápido)
                            with db_nfe._connect() as conn:
                                cursor = conn.execute("""
                                    SELECT chave, informante, tipo, data_emissao, nome_emitente, numero, xml_status
                                    FROM notas_detalhadas 
                                    WHERE xml_status = 'RESUMO' 
                                    AND tipo NOT IN ('CTE', 'CT-e', 'CTe')
                                    ORDER BY data_emissao DESC
                                """)
                                
                                self.notas_resumo = []
                                for row in cursor:
                                    nota = {
                                        'chave': row[0],
                                        'informante': row[1],
                                        'tipo': row[2],
                                        'data_emissao': row[3],
                                        'nome_emitente': row[4] if len(row) > 4 else '',
                                        'numero': row[5] if len(row) > 5 else '',
                                        'xml_status': row[6] if len(row) > 6 else 'RESUMO'
                                    }
                                    self.notas_resumo.append(nota)
                            
                            self.log(f"[AUTO-VERIFICAÇÃO] ✅ Encontradas {len(self.notas_resumo)} notas com status RESUMO")
                        except Exception as e:
                            self.log(f"[AUTO-VERIFICAÇÃO] ❌ Erro ao buscar notas: {e}")
                            self.error.emit(f"Erro ao buscar notas RESUMO:\n{e}")
                            return
                    
                    total = len(self.notas_resumo)
                    if total == 0:
                        self.log("[AUTO-VERIFICAÇÃO] ℹ️ Nenhuma nota RESUMO encontrada")
                        self.finished.emit(0, 0)
                        return
                    
                    # Carrega certificados
                    certs = self.parent_window.db.load_certificates()
                    if not certs:
                        self.error.emit("Nenhum certificado configurado!")
                        return
                    
                    encontrados = 0
                    nao_encontrados = 0
                    chaves_invalidas = 0
                    
                    for idx, nota in enumerate(self.notas_resumo, 1):
                        if self._cancelado:
                            self.log("[AUTO-VERIFICAÇÃO] ❌ Cancelado pelo usuário")
                            break
                        
                        while self._pausado:
                            import time
                            time.sleep(0.5)
                        
                        chave = nota.get('chave', '')
                        informante = nota.get('informante', '')
                        nome = nota.get('nome_emitente', 'N/A')
                        
                        # ======= VALIDAÇÃO DE CHAVE =======
                        # Filtrar NFS-e e chaves inválidas (deve ter 44 dígitos numéricos)
                        if not chave or len(chave) != 44 or not chave.isdigit():
                            chaves_invalidas += 1
                            tipo_invalido = "NFS-e" if "NFSE" in chave.upper() or "NSU" in chave.upper() else "Chave inválida"
                            self.log(f"\n[AUTO-VERIFICAÇÃO] [{idx}/{total}] ⚠️ {tipo_invalido} - IGNORADO")
                            self.log(f"[AUTO-VERIFICAÇÃO]    Chave: {chave[:50]}...")
                            self.progress.emit(f"[{idx}/{total}] Ignorando {tipo_invalido}...", idx, total)
                            continue
                        # ==================================
                        
                        self.log(f"\n[AUTO-VERIFICAÇÃO] [{idx}/{total}] {nome[:40]} - {chave[:20]}...")
                        self.progress.emit(f"[{idx}/{total}] Buscando XML completo...", idx, total)
                        
                        # ⏱️ Delay de 2 segundos entre consultas para respeitar limite da SEFAZ
                        if idx > 2:  # Pula delay nas primeiras 2 notas
                            import time
                            time.sleep(2)
                        
                        # Tenta buscar XML completo
                        xml_completo = None
                        cert_usado = None  # 🔧 Guarda qual certificado conseguiu buscar o XML
                        motivo_rejeicao = None
                        consumo_indevido = False  # Flag para detectar limite de consultas
                        
                        for cert in certs:
                            # Se já detectou consumo indevido, para IMEDIATAMENTE
                            if consumo_indevido:
                                break
                            
                            try:
                                svc = NFeService(
                                    cert.get('caminho'), 
                                    cert.get('senha'), 
                                    cert.get('cnpj_cpf'), 
                                    cert.get('cUF_autor')
                                )
                                # Tenta buscar pelo método de distribuição (mais correto)
                                xml_resp = svc.fetch_by_chave_dist(chave)
                                
                                if not xml_resp:
                                    continue
                                
                                # Verifica se é XML completo (nfeProc) ou resumo (resNFe)
                                if '<nfeProc' in xml_resp or '<procNFe' in xml_resp:
                                    xml_completo = xml_resp
                                    cert_usado = cert  # 🔧 Guarda certificado que teve sucesso
                                    break  # Encontrou XML completo!
                                
                                # Verifica código de status
                                import re
                                cstat_match = re.search(r'<cStat>(\d+)</cStat>', xml_resp)
                                if cstat_match:
                                    cstat = cstat_match.group(1)
                                    motivo_match = re.search(r'<xMotivo>(.+?)</xMotivo>', xml_resp)
                                    motivo_texto = motivo_match.group(1) if motivo_match else ""
                                    
                                    # 656 pode ser "Documento localizado" OU "Consumo Indevido"
                                    if cstat == '656':
                                        if 'Consumo Indevido' in motivo_texto or 'Ultrapassou' in motivo_texto:
                                            # Consumo Indevido (limite de consultas por hora)
                                            motivo_rejeicao = motivo_texto
                                            consumo_indevido = True  # Marca flag
                                            break  # Para IMEDIATAMENTE
                                        else:
                                            # Documento localizado (mas só resumo disponível)
                                            motivo_rejeicao = motivo_texto if motivo_texto else "Documento localizado (apenas resumo)"
                                            # Continua tentando outros certificados
                                            continue
                                    
                                    # 217/231 = Não possui permissão
                                    elif cstat in ['217', '231']:
                                        motivo_rejeicao = motivo_texto if motivo_texto else "Sem permissão"
                                        # Para de tentar (nenhum certificado terá permissão)
                                        break
                                
                            except Exception as e:
                                # Silencia erros individuais de certificado
                                continue
                        
                        if xml_completo and (('<nfeProc' in xml_completo) or ('<procNFe' in xml_completo)):
                            encontrados += 1
                            self.log(f"[AUTO-VERIFICAÇÃO]    ✅ XML completo encontrado!")
                            # Salva XML
                            try:
                                # 🔧 Usa o certificado que CONSEGUIU buscar o XML, não o informante da nota
                                cert_cnpj = cert_usado.get('cnpj_cpf') if cert_usado else informante
                                nome_cert = self.parent_window.db.get_cert_nome_by_informante(cert_cnpj)
                                
                                # 🆕 ARMAZENAMENTO AUTOMÁTICO: Salva em backup local + TODOS os perfis ativos
                                # 1. Salva em backup local (xmls/)
                                salvar_xml_por_certificado(xml_completo, cert_cnpj, pasta_base="xmls")
                                
                                # 2. Salva em TODOS os perfis ativos (pasta_base=None)
                                salvar_xml_por_certificado(xml_completo, cert_cnpj, pasta_base=None, nome_certificado=nome_cert)
                                
                                # Atualiza status da nota no banco
                                nota_update = {
                                    'chave': chave,
                                    'xml_status': 'COMPLETO',
                                    'informante': informante
                                }
                                # Carrega dados existentes da nota
                                with self.parent_window.db._connect() as conn:
                                    existing = conn.execute(
                                        "SELECT * FROM notas_detalhadas WHERE chave = ?",
                                        (chave,)
                                    ).fetchone()
                                    if existing:
                                        # Mescla dados existentes com atualização
                                        columns = [desc[0] for desc in conn.execute("SELECT * FROM notas_detalhadas LIMIT 0").description]
                                        for col, val in zip(columns, existing):
                                            if col not in nota_update:
                                                nota_update[col] = val
                                # Salva nota completa com xml_status atualizado
                                self.parent_window.db.save_note(nota_update)
                            except Exception as e:
                                self.log(f"[AUTO-VERIFICAÇÃO]    ⚠️ Erro ao salvar: {e}")
                                pass
                        else:
                            nao_encontrados += 1
                            # Usa motivo de rejeição capturado durante a tentativa
                            if not motivo_rejeicao:
                                motivo_rejeicao = "Não disponível"
                            self.log(f"[AUTO-VERIFICAÇÃO]    ❌ XML não disponível - Motivo: {motivo_rejeicao}")
                            
                            # Se detectou consumo indevido, PARA o processo
                            if consumo_indevido:
                                self.log(f"\n[AUTO-VERIFICAÇÃO] ⚠️ LIMITE DE CONSULTAS ATINGIDO!")
                                self.log(f"[AUTO-VERIFICAÇÃO] A SEFAZ bloqueou temporariamente as consultas.")
                                self.log(f"[AUTO-VERIFICAÇÃO] Aguarde 1 hora antes de tentar novamente.")
                                self.log(f"[AUTO-VERIFICAÇÃO] Processadas: {idx}/{total} notas")
                                break  # SAI DO LOOP de notas
                    
                    self.log(f"\n[AUTO-VERIFICAÇÃO] ========================================")
                    self.log(f"[AUTO-VERIFICAÇÃO] ✅ Processo concluído!")
                    self.log(f"[AUTO-VERIFICAÇÃO]    • Total de registros RESUMO: {total}")
                    self.log(f"[AUTO-VERIFICAÇÃO]    • Chaves inválidas (NFS-e, etc): {chaves_invalidas}")
                    self.log(f"[AUTO-VERIFICAÇÃO]    • Chaves válidas consultadas: {total - chaves_invalidas}")
                    self.log(f"[AUTO-VERIFICAÇÃO]    • XMLs completos encontrados: {encontrados}")
                    self.log(f"[AUTO-VERIFICAÇÃO]    • XMLs não disponíveis: {nao_encontrados}")
                    if encontrados > 0:
                        taxa_sucesso = (encontrados / (total - chaves_invalidas)) * 100 if (total - chaves_invalidas) > 0 else 0
                        self.log(f"[AUTO-VERIFICAÇÃO]    • Taxa de sucesso: {taxa_sucesso:.1f}%")
                    self.log(f"[AUTO-VERIFICAÇÃO] ========================================")
                    
                    print("[DEBUG WORKER] Emitindo signal finished...")
                    self.finished.emit(encontrados, nao_encontrados)
                    print("[DEBUG WORKER] Signal finished emitido!")
                    
                except Exception as e:
                    import traceback
                    print(f"[DEBUG WORKER] ========== ERRO NO WORKER ==========")
                    print(f"[DEBUG WORKER] Tipo do erro: {type(e)}")
                    print(f"[DEBUG WORKER] Mensagem: {str(e)}")
                    print(f"[DEBUG WORKER] Traceback:")
                    traceback.print_exc()
                    print(f"[DEBUG WORKER] ====================================")
                    self.log(f"[AUTO-VERIFICAÇÃO] ❌ ERRO: {str(e)}")
                    self.log(traceback.format_exc())
                    self.error.emit(f"Erro na auto-verificação:\n{e}")
        
        # Primeiro, conta quantas notas RESUMO existem
        print("[DEBUG AUTO-VERIFICAÇÃO] Iniciando query de contagem...")
        try:
            with self.parent_window.db._connect() as conn:
                print("[DEBUG AUTO-VERIFICAÇÃO] Conexão estabelecida, executando query...")
                count = conn.execute("""
                    SELECT COUNT(*) FROM notas_detalhadas 
                    WHERE xml_status = 'RESUMO' 
                    AND tipo NOT IN ('CTE', 'CT-e', 'CTe')
                """).fetchone()[0]
            
            print(f"[DEBUG AUTO-VERIFICAÇÃO] Query executada! Encontradas {count} notas RESUMO")
            
            if count == 0:
                print("[DEBUG AUTO-VERIFICAÇÃO] Nenhuma nota RESUMO, exibindo mensagem...")
                QMessageBox.information(
                    self,
                    "Auto-Verificação",
                    "✅ Não há notas com status RESUMO para processar!"
                )
                return
            
            print("[DEBUG AUTO-VERIFICAÇÃO] Exibindo diálogo de confirmação...")
            # Confirma ação
            reply = QMessageBox.question(
                self,
                "Auto-Verificação",
                f"🔍 Encontradas {count} notas com status RESUMO.\n\n"
                f"Deseja buscar os XMLs completos na SEFAZ?\n\n"
                f"⚠️ Esta operação pode demorar alguns minutos.\n"
                f"A tarefa rodará em segundo plano.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            print(f"[DEBUG AUTO-VERIFICAÇÃO] Resposta do usuário: {'SIM' if reply == QMessageBox.Yes else 'NÃO'}")
            
            if reply == QMessageBox.No:
                print("[DEBUG AUTO-VERIFICAÇÃO] Usuário cancelou")
                return
                
        except Exception as e:
            print(f"[DEBUG AUTO-VERIFICAÇÃO] ERRO na contagem: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erro", f"Erro ao verificar notas RESUMO:\n{e}")
            return
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Criando worker...")
        # Agora cria o worker
        worker = AutoVerificacaoWorker(self.parent_window, [])  # Lista vazia temporária
        worker.buscar_notas_na_thread = True  # Flag para buscar notas dentro da thread
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Worker criado, registrando na lista...")
        # Registra worker na lista
        self.workers.append(worker)
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Conectando signals...")
        # Conecta signals básicos
        def on_finished_worker(encontrados, nao_encontrados):
            print(f"[DEBUG AUTO-VERIFICAÇÃO] Worker finalizado! Encontrados: {encontrados}, Não encontrados: {nao_encontrados}")
            QMessageBox.information(
                self,
                "Auto-Verificação Concluída",
                f"✅ Processo concluído!\n\n"
                f"• XMLs completos encontrados: {encontrados}\n"
                f"• XMLs não disponíveis: {nao_encontrados}\n\n"
                f"A interface será atualizada."
            )
            # Remove worker da lista
            if worker in self.workers:
                self.workers.remove(worker)
            # Atualiza interface
            if self.parent_window:
                self.parent_window.refresh_all()
        
        def on_error_worker(msg):
            print(f"[DEBUG AUTO-VERIFICAÇÃO] ERRO no worker: {msg}")
            QMessageBox.critical(self, "Erro", f"Erro na auto-verificação:\n{msg}")
            # Remove worker da lista
            if worker in self.workers:
                self.workers.remove(worker)
        
        def on_progress_worker(msg, current, total):
            print(f"[DEBUG AUTO-VERIFICAÇÃO] Progresso: {msg} ({current}/{total})")
        
        def on_log_worker(msg):
            print(f"[AUTO-VERIFICAÇÃO LOG] {msg}")
        
        worker.finished.connect(on_finished_worker)
        worker.error.connect(on_error_worker)
        worker.progress.connect(on_progress_worker)
        worker.log_message.connect(on_log_worker)
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Iniciando thread...")
        # Inicia thread
        worker.start()
        
        print("[DEBUG AUTO-VERIFICAÇÃO] Thread iniciada, exibindo mensagem final...")
        QMessageBox.information(
            self,
            "Auto-Verificação Iniciada",
            f"🔍 A busca de XMLs completos foi iniciada em segundo plano.\n\n"
            f"Total de notas: {count}\n\n"
            f"Você pode continuar usando o sistema normalmente."
        )
        print("[DEBUG AUTO-VERIFICAÇÃO] Método _iniciar_auto_verificacao finalizado!")
    
    def _iniciar_auto_verificacao_inteligente(self):
        """
        Inicia auto-verificação INTELIGENTE com controle de quota SEFAZ
        - Respeita limite de 20 consultas/hora/certificado
        - Prioriza notas mais recentes
        - Exibe saldo de consultas disponíveis
        - Para automaticamente ao atingir limites
        """
        print("[AUTO-VERIFICAÇÃO INTELIGENTE] Iniciando verificação com controle de quota...")
        
        # Carrega gerenciador de quotas
        from modules.quota_manager import QuotaManager
        quota_mgr = QuotaManager()
        
        # Obtém certificados
        certs = self.db.load_certificates()
        if not certs:
            print("[AUTO-VERIFICAÇÃO INTELIGENTE] Nenhum certificado configurado")
            return
        
        # Verifica saldo de quotas
        status_quotas = quota_mgr.get_status_todos_certificados(certs)
        
        # Conta total de consultas disponíveis
        total_disponiveis = sum(s['disponiveis'] for s in status_quotas.values())
        
        print(f"[AUTO-VERIFICAÇÃO INTELIGENTE] Saldo de consultas disponíveis: {total_disponiveis}")
        
        # Se não tem consultas disponíveis, não inicia
        if total_disponiveis == 0:
            print("[AUTO-VERIFICAÇÃO INTELIGENTE] ⚠️ Nenhuma consulta disponível. Aguarde 1 hora.")
            # Mostra notificação
            self.set_status("⚠️ Limite de consultas SEFAZ atingido. Aguarde 1 hora.", 5000)
            return
        
        # Conta quantas notas RESUMO existem
        with self.db._connect() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT COUNT(*) FROM notas_detalhadas 
                WHERE xml_status = 'RESUMO'
                AND LENGTH(chave) = 44
                AND chave NOT LIKE '%NFSE%'
                AND chave NOT LIKE '%NSU%'
            """)
            count = cursor.fetchone()[0]
        
        if count == 0:
            print("[AUTO-VERIFICAÇÃO INTELIGENTE] ✅ Não há notas RESUMO para verificar")
            self.set_status("✅ Todos os XMLs já foram baixados", 3000)
            return
        
        # Calcula quantas notas processar (mínimo entre disponível e total)
        notas_processar = min(total_disponiveis, count)
        
        msg = (
            f"🔍 Auto-Verificação Inteligente\n\n"
            f"📊 Status de Quotas SEFAZ:\n"
        )
        
        for cnpj, status in status_quotas.items():
            razao = next((c.get('razao_social', cnpj[:8]) for c in certs if c.get('cnpj_cpf') == cnpj), cnpj[:8])
            msg += f"  • {razao}: {status['disponiveis']}/{status['limite']} consultas disponíveis\n"
        
        msg += (
            f"\n📋 Notas RESUMO encontradas: {count}\n"
            f"🎯 Serão processadas: {notas_processar} (limite SEFAZ)\n\n"
            f"Deseja iniciar a verificação?"
        )
        
        from PyQt5.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "Auto-Verificação Inteligente",
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            print("[AUTO-VERIFICAÇÃO INTELIGENTE] Cancelado pelo usuário")
            return
        
        # Inicia a verificação com limite
        self._executar_auto_verificacao_com_limite(notas_processar, quota_mgr)
    
    def _executar_auto_verificacao_com_limite(self, limite: int, quota_mgr):
        """Executa auto-verificação respeitando limite de consultas"""
        print(f"[AUTO-VERIFICAÇÃO INTELIGENTE] Iniciando com limite de {limite} consultas")
        
        from PyQt5.QtCore import QThread, pyqtSignal
        
        class AutoVerificacaoInteligenteWorker(QThread):
            """Worker para auto-verificação com controle de quota"""
            progress = pyqtSignal(str, int, int)
            finished = pyqtSignal(int, int, int)  # encontrados, nao_encontrados, quota_esgotada
            error = pyqtSignal(str)
            log_message = pyqtSignal(str)
            
            def __init__(self, parent_window, limite_consultas, quota_manager):
                super().__init__()
                self.parent_window = parent_window
                self.limite = limite_consultas
                self.quota_mgr = quota_manager
                self._cancelado = False
            
            def log(self, msg: str):
                self.log_message.emit(msg)
            
            def run(self):
                try:
                    from nfe_search import NFeService
                    import re
                    import time
                    
                    # Carrega certificados
                    certs = self.parent_window.db.load_certificates()
                    
                    # Busca notas RESUMO (priorizando mais recentes)
                    with self.parent_window.db._connect() as conn:
                        cursor = conn.cursor()
                        cursor.execute("""
                            SELECT chave, informante, tipo, data_emissao, nome_emitente
                            FROM notas_detalhadas 
                            WHERE xml_status = 'RESUMO'
                            AND LENGTH(chave) = 44
                            AND chave NOT LIKE '%NFSE%'
                            AND chave NOT LIKE '%NSU%'
                            ORDER BY data_emissao DESC
                            LIMIT ?
                        """, (self.limite,))
                        notas = cursor.fetchall()
                    
                    total = len(notas)
                    encontrados = 0
                    nao_encontrados = 0
                    consultas_realizadas = 0
                    quota_esgotada = 0
                    
                    self.log(f"[AUTO-VERIFICAÇÃO] 🎯 Processando {total} notas (mais recentes)")
                    
                    for idx, (chave, informante, tipo, data_emissao, nome) in enumerate(notas, 1):
                        if self._cancelado:
                            break
                        
                        self.progress.emit(f"[{idx}/{total}] Verificando...", idx, total)
                        
                        # Busca certificado com quota disponível
                        cert_com_quota = None
                        for cert in certs:
                            cnpj = cert.get('cnpj_cpf', '')
                            if self.quota_mgr.pode_consultar(cnpj):
                                cert_com_quota = cert
                                break
                        
                        if not cert_com_quota:
                            self.log(f"[AUTO-VERIFICAÇÃO] ⚠️ Quota esgotada! Parando em {idx}/{total}")
                            quota_esgotada = 1
                            break
                        
                        # Delay entre consultas (2 segundos)
                        if idx > 1:
                            time.sleep(2)
                        
                        # Realiza consulta
                        try:
                            svc = NFeService(
                                cert_com_quota.get('caminho'),
                                cert_com_quota.get('senha'),
                                cert_com_quota.get('cnpj_cpf'),
                                cert_com_quota.get('cUF_autor')
                            )
                            
                            xml_resp = svc.fetch_by_chave_dist(chave)
                            
                            # Registra consulta na quota
                            self.quota_mgr.registrar_consulta(cert_com_quota.get('cnpj_cpf'))
                            consultas_realizadas += 1
                            
                            if xml_resp and ('<nfeProc' in xml_resp or '<procNFe' in xml_resp):
                                encontrados += 1
                                self.log(f"[{idx}/{total}] ✅ {nome[:30]} - XML encontrado!")
                                # Salva XML (código de salvamento aqui)
                            else:
                                nao_encontrados += 1
                                # Verifica motivo
                                cstat_match = re.search(r'<cStat>(\d+)</cStat>', xml_resp or '')
                                if cstat_match:
                                    cstat = cstat_match.group(1)
                                    if cstat == '656' and ('Consumo Indevido' in (xml_resp or '')):
                                        self.log(f"[AUTO-VERIFICAÇÃO] ⚠️ Limite SEFAZ atingido!")
                                        quota_esgotada = 1
                                        break
                        
                        except Exception as e:
                            self.log(f"[{idx}/{total}] ⚠️ Erro: {str(e)[:50]}")
                            nao_encontrados += 1
                    
                    self.finished.emit(encontrados, nao_encontrados, quota_esgotada)
                
                except Exception as e:
                    import traceback
                    self.error.emit(f"{str(e)}\n{traceback.format_exc()}")
        
        # Cria e inicia worker
        worker = AutoVerificacaoInteligenteWorker(self.parent_window, limite, quota_mgr)
        
        def on_finished(encontrados, nao_encontrados, quota_esgotada):
            total = encontrados + nao_encontrados
            msg = f"✅ Verificação concluída!\n\n"
            msg += f"XMLs encontrados: {encontrados}\n"
            msg += f"Não disponíveis: {nao_encontrados}\n"
            msg += f"Total consultado: {total}\n"
            
            if quota_esgotada:
                msg += "\n⚠️ Quota SEFAZ esgotada. Aguarde 1 hora para continuar."
            
            QMessageBox.information(self, "Verificação Concluída", msg)
            
            if self.parent_window:
                self.parent_window.refresh_all()
        
        def on_error(msg):
            QMessageBox.critical(self, "Erro", f"Erro na verificação:\n{msg}")
        
        def on_log(msg):
            print(msg)
        
        worker.finished.connect(on_finished)
        worker.error.connect(on_error)
        worker.log_message.connect(on_log)
        worker.start()
        
        # Guarda referência
        self.workers.append(worker)
        
        self.set_status(f"🔍 Verificando {limite} notas com controle de quota...", 0)
    
    def _exibir_status_quotas(self):
        """Exibe status detalhado das quotas SEFAZ"""
        from modules.quota_manager import QuotaManager
        from PyQt5.QtWidgets import QMessageBox
        
        quota_mgr = QuotaManager()
        certs = self.parent_window.db.load_certificates() if self.parent_window else []
        
        if not certs:
            QMessageBox.warning(self, "Status de Quotas", "Nenhum certificado configurado.")
            return
        
        status_quotas = quota_mgr.get_status_todos_certificados(certs)
        
        msg = "📊 STATUS DE QUOTAS SEFAZ\n"
        msg += "=" * 50 + "\n\n"
        msg += "Limite: 20 consultas por chave por hora por certificado\n\n"
        
        total_disponiveis = 0
        total_usadas = 0
        
        for cnpj, status in status_quotas.items():
            razao = next((c.get('razao_social', 'Sem razão') for c in certs if c.get('cnpj_cpf') == cnpj), 'Desconhecido')
            
            # Barra de progresso visual
            disponiveis = status['disponiveis']
            usadas = status['usadas']
            percentual = status['percentual']
            
            # Cria barra visual
            total_barras = 20
            barras_cheias = int((disponiveis / 20) * total_barras)
            barras_vazias = total_barras - barras_cheias
            barra = "█" * barras_cheias + "░" * barras_vazias
            
            # Emoji baseado no status
            if percentual >= 80:
                emoji = "🟢"
            elif percentual >= 40:
                emoji = "🟡"
            elif percentual >= 20:
                emoji = "🟠"
            else:
                emoji = "🔴"
            
            msg += f"{emoji} {razao[:30]}\n"
            msg += f"   CNPJ: {cnpj}\n"
            msg += f"   [{barra}] {disponiveis}/20\n"
            msg += f"   Disponíveis: {disponiveis} | Usadas: {usadas}\n\n"
            
            total_disponiveis += disponiveis
            total_usadas += usadas
        
        msg += "=" * 50 + "\n"
        msg += f"📈 TOTAL GERAL:\n"
        msg += f"   Consultas disponíveis: {total_disponiveis}\n"
        msg += f"   Consultas usadas: {total_usadas}\n"
        
        # Recomendação
        if total_disponiveis == 0:
            msg += "\n⚠️ ATENÇÃO: Quota esgotada!\n"
            msg += "Aguarde até 1 hora para novas consultas.\n"
        elif total_disponiveis < 10:
            msg += "\n⚠️ Quota baixa. Use com cuidado.\n"
        else:
            msg += "\n✅ Quota disponível para consultas.\n"
        
        QMessageBox.information(self, "Status de Quotas SEFAZ", msg)
    
    def _abrir_agendador(self):
        """Abre diálogo de agendamento de tarefas"""
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                                     QComboBox, QSpinBox, QCheckBox, QGroupBox,
                                     QPushButton, QTimeEdit, QMessageBox)
        from PyQt5.QtCore import QTime, QSettings
        
        dialog = QDialog(self)
        dialog.setWindowTitle("⏰ Agendador de Tarefas")
        dialog.setMinimumWidth(500)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #fafafa;
            }
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        titulo = QLabel("Configure quando e quais tarefas serão executadas automaticamente")
        titulo.setStyleSheet("font-size: 11px; color: #666; margin-bottom: 10px;")
        titulo.setWordWrap(True)
        layout.addWidget(titulo)
        
        # Grupo 1: Quando executar
        grupo_quando = QGroupBox("⏰ Quando Executar?")
        layout_quando = QVBoxLayout()
        
        # Opção: Ao iniciar
        check_ao_iniciar = QCheckBox("Executar ao iniciar o sistema")
        check_ao_iniciar.setStyleSheet("padding: 5px;")
        layout_quando.addWidget(check_ao_iniciar)
        
        # Opção: Horário específico
        layout_horario = QHBoxLayout()
        check_horario = QCheckBox("Executar em horário específico:")
        time_edit = QTimeEdit()
        time_edit.setDisplayFormat("HH:mm")
        time_edit.setTime(QTime(8, 0))  # Padrão: 08:00
        time_edit.setEnabled(False)
        check_horario.toggled.connect(time_edit.setEnabled)
        layout_horario.addWidget(check_horario)
        layout_horario.addWidget(time_edit)
        layout_horario.addStretch()
        layout_quando.addLayout(layout_horario)
        
        # Opção: Intervalo periódico
        layout_intervalo = QHBoxLayout()
        check_intervalo = QCheckBox("Repetir a cada:")
        spin_intervalo = QSpinBox()
        spin_intervalo.setMinimum(1)
        spin_intervalo.setMaximum(24)
        spin_intervalo.setValue(2)
        spin_intervalo.setSuffix(" horas")
        spin_intervalo.setEnabled(False)
        check_intervalo.toggled.connect(spin_intervalo.setEnabled)
        layout_intervalo.addWidget(check_intervalo)
        layout_intervalo.addWidget(spin_intervalo)
        layout_intervalo.addStretch()
        layout_quando.addLayout(layout_intervalo)
        
        grupo_quando.setLayout(layout_quando)
        layout.addWidget(grupo_quando)
        
        # Grupo 2: O que executar
        grupo_oque = QGroupBox("🎯 Qual Tarefa Executar?")
        layout_oque = QVBoxLayout()
        
        combo_tarefa = QComboBox()
        combo_tarefa.addItems([
            "Buscar Notas na SEFAZ",
            "Busca Completa (NSU)",
            "Atualizar Status de Notas",
            "Baixar XMLs Faltantes",
            "Gerar PDFs Pendentes",
            "Manifestação Automática",
            "Sincronizar Documentos"
        ])
        combo_tarefa.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
            }
        """)
        layout_oque.addWidget(combo_tarefa)
        
        # Descrição da tarefa
        label_descricao = QLabel()
        label_descricao.setWordWrap(True)
        label_descricao.setStyleSheet("""
            background-color: #f8f9fa;
            padding: 10px;
            border-radius: 4px;
            color: #555;
            font-size: 10px;
        """)
        
        descricoes = {
            0: "Busca novos documentos fiscais na SEFAZ usando distribuição DFe.",
            1: "Busca completa usando NSU (últimos 90 dias).",
            2: "Consulta status de notas autorizadas (cancelamentos, etc).",
            3: "Baixa XMLs completos para notas que estão como RESUMO.",
            4: "Gera arquivos PDF para XMLs que ainda não possuem PDF.",
            5: "Manifesta automaticamente notas pendentes (Ciência da Operação).",
            6: "Sincroniza todos os documentos e atualiza banco de dados."
        }
        
        def atualizar_descricao(index):
            label_descricao.setText(f"ℹ️ {descricoes.get(index, '')}")
        
        combo_tarefa.currentIndexChanged.connect(atualizar_descricao)
        atualizar_descricao(0)  # Mostra descrição inicial
        
        layout_oque.addWidget(label_descricao)
        grupo_oque.setLayout(layout_oque)
        layout.addWidget(grupo_oque)
        
        # Carregar configurações salvas
        settings = QSettings('NFE_System', 'BOT_NFE')
        check_ao_iniciar.setChecked(settings.value('agendador/ao_iniciar', False, type=bool))
        check_horario.setChecked(settings.value('agendador/horario_ativo', False, type=bool))
        time_edit.setTime(QTime.fromString(settings.value('agendador/horario', '08:00'), 'HH:mm'))
        check_intervalo.setChecked(settings.value('agendador/intervalo_ativo', False, type=bool))
        spin_intervalo.setValue(settings.value('agendador/intervalo_horas', 2, type=int))
        combo_tarefa.setCurrentIndex(settings.value('agendador/tarefa_index', 0, type=int))
        
        # Botões
        layout_botoes = QHBoxLayout()
        
        btn_salvar = QPushButton("💾 Salvar Configuração")
        btn_salvar.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 10px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        
        def salvar_configuracao():
            # Valida se ao menos uma opção foi marcada
            if not (check_ao_iniciar.isChecked() or check_horario.isChecked() or check_intervalo.isChecked()):
                QMessageBox.warning(
                    dialog,
                    "Configuração Incompleta",
                    "Por favor, selecione ao menos uma opção de quando executar a tarefa."
                )
                return
            
            # Salva configurações
            settings.setValue('agendador/ao_iniciar', check_ao_iniciar.isChecked())
            settings.setValue('agendador/horario_ativo', check_horario.isChecked())
            settings.setValue('agendador/horario', time_edit.time().toString('HH:mm'))
            settings.setValue('agendador/intervalo_ativo', check_intervalo.isChecked())
            settings.setValue('agendador/intervalo_horas', spin_intervalo.value())
            settings.setValue('agendador/tarefa_index', combo_tarefa.currentIndex())
            settings.setValue('agendador/tarefa_nome', combo_tarefa.currentText())
            
            QMessageBox.information(
                dialog,
                "Configuração Salva",
                f"✅ Tarefa agendada com sucesso!\n\n"
                f"Tarefa: {combo_tarefa.currentText()}\n"
                f"Ao iniciar: {'Sim' if check_ao_iniciar.isChecked() else 'Não'}\n"
                f"Horário: {time_edit.time().toString('HH:mm') if check_horario.isChecked() else 'Não'}\n"
                f"Intervalo: {f'{spin_intervalo.value()}h' if check_intervalo.isChecked() else 'Não'}"
            )
            
            dialog.accept()
        
        btn_salvar.clicked.connect(salvar_configuracao)
        btn_cancelar.clicked.connect(dialog.reject)
        
        layout_botoes.addWidget(btn_salvar)
        layout_botoes.addWidget(btn_cancelar)
        layout.addLayout(layout_botoes)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def _reprocessar_resumos(self):
        """Reprocessa notas RESUMO em background"""
        from PyQt5.QtCore import QThread, pyqtSignal
        
        class ReprocessarResumosWorker(QThread):
            """Worker para reprocessar notas com status RESUMO"""
            progress = pyqtSignal(str, int, int)  # mensagem, atual, total
            finished = pyqtSignal(int, int)  # encontrados, total
            error = pyqtSignal(str)
            log_message = pyqtSignal(str)
            
            def __init__(self, parent_window):
                super().__init__()
                self.parent_window = parent_window
                self._cancelado = False
                self._pausado = False
            
            def cancelar(self):
                self._cancelado = True
            
            def pausar(self):
                """Pausa a execução do worker"""
                self._pausado = not self._pausado
                if self._pausado:
                    self.log("[REPROCESSAR] ⏸️ Pausado")
                else:
                    self.log("[REPROCESSAR] ▶️ Retomado")
            
            def log(self, msg: str):
                """Envia log para o terminal"""
                self.log_message.emit(msg)
            
            def run(self):
                try:
                    from nfe_search import DatabaseManager, NFeService, salvar_xml_por_certificado, extrair_nota_detalhada
                    from lxml import etree
                    
                    self.log("[REPROCESSAR] ========================================")
                    self.log("[REPROCESSAR] Iniciando reprocessamento de resumos (resNFe)")
                    self.log("[REPROCESSAR] ========================================")
                    
                    db_nfe = DatabaseManager(str(self.parent_window.db.db_path))
                    
                    # Busca notas com RESUMO direto do banco
                    self.log("[REPROCESSAR] 🔍 Buscando notas RESUMO no banco...")
                    self.progress.emit("Buscando notas RESUMO...", 0, 100)
                    
                    with db_nfe._connect() as conn:
                        cursor = conn.execute("""
                            SELECT chave, informante, tipo, data_emissao, nome_emitente
                            FROM notas_detalhadas 
                            WHERE xml_status = 'RESUMO'
                            ORDER BY data_emissao DESC
                        """)
                        notas_resumo = [
                            {'chave': row[0], 'informante': row[1], 'tipo': row[2], 
                             'data_emissao': row[3], 'nome_emitente': row[4]}
                            for row in cursor
                        ]
                    
                    total = len(notas_resumo)
                    self.log(f"[REPROCESSAR] ✅ Encontradas {total} notas RESUMO")
                    
                    if total == 0:
                        self.error.emit("Nenhuma nota com status RESUMO encontrada!")
                        return
                    
                    # Carrega certificados
                    certs = self.parent_window.db.load_certificates()
                    if not certs:
                        self.error.emit("Nenhum certificado configurado!")
                        return
                    
                    encontrados = 0
                    
                    for idx, nota in enumerate(notas_resumo, 1):
                        # Verifica pausa
                        while self._pausado and not self._cancelado:
                            import time
                            time.sleep(0.5)
                        
                        if self._cancelado:
                            self.log("[REPROCESSAR] ❌ Cancelado pelo usuário")
                            break
                        
                        chave = nota['chave']
                        informante = nota['informante']
                        
                        self.log(f"\n[REPROCESSAR] [{idx}/{total}] Processando chave: {chave}")
                        self.progress.emit(f"[{idx}/{total}] Buscando {chave[:20]}...", idx, total)
                        
                        # Tenta buscar XML completo com cada certificado
                        xml_completo = None
                        for cert in certs:
                            try:
                                cnpj_cert = cert.get('cnpj_cpf')
                                senha_cert = cert.get('senha')
                                caminho_cert = cert.get('caminho')
                                cuf_cert = cert.get('cUF_autor')
                                
                                # Cria serviço
                                svc = NFeService(caminho_cert, senha_cert, cnpj_cert, cuf_cert)
                                
                                # Detecta tipo (NF-e ou CT-e)
                                modelo = chave[20:22] if len(chave) >= 22 else '55'
                                is_cte = modelo == '57'
                                
                                # Busca XML
                                if is_cte:
                                    xml_completo = svc.fetch_prot_cte(chave)
                                else:
                                    # fetch_by_chave_dist retorna resposta SOAP, precisa extrair documentos
                                    resp_soap = svc.fetch_by_chave_dist(chave)
                                    if resp_soap:
                                        # Verifica código de retorno SEFAZ
                                        from lxml import etree
                                        try:
                                            root = etree.fromstring(resp_soap.encode('utf-8'))
                                            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                                            cstat_elem = root.find('.//nfe:cStat', ns)
                                            xmotivo_elem = root.find('.//nfe:xMotivo', ns)
                                            if cstat_elem is not None:
                                                cstat = cstat_elem.text
                                                xmotivo = xmotivo_elem.text if xmotivo_elem is not None else "N/A"
                                                if cstat != "138":  # 138 = Documento localizado
                                                    self.log(f"[REPROCESSAR]    ⚠️ SEFAZ retornou cStat={cstat}: {xmotivo}")
                                                    
                                                    # Se atingiu limite, avisa e sugere aguardar
                                                    if cstat == "656" and "20 consultas por hora" in xmotivo:
                                                        self.log(f"[REPROCESSAR]    ⏳ Limite de 20 consultas/hora atingido!")
                                                        self.log(f"[REPROCESSAR]    💡 Sugestão: Pause e aguarde 1 hora para continuar")
                                                        # Não para automaticamente, deixa usuário decidir
                                        except:
                                            pass
                                        
                                        # Extrai documentos da resposta (descompacta docZip)
                                        from nfe_search import XMLProcessor
                                        processor = XMLProcessor(informante=cnpj_cert)
                                        docs = processor.extract_docs(resp_soap)
                                        
                                        if not docs:
                                            continue  # Pula para próximo certificado
                                        
                                        # Procura por XML completo (nfeProc ou procNFe)
                                        for nsu, xml in docs:
                                            if '<nfeProc' in xml or '<procNFe' in xml:
                                                xml_completo = xml
                                                self.log(f"[REPROCESSAR]    ✅ XML completo encontrado (NSU {nsu})! Tamanho: {len(xml)} bytes")
                                                break
                                        
                                        if xml_completo:
                                            break
                                
                                if xml_completo and ('<nfeProc' in xml_completo or '<procNFe' in xml_completo):
                                    break
                                    
                            except Exception as e:
                                self.log(f"[REPROCESSAR]    ⚠️ Erro com cert {cnpj_cert}: {str(e)[:80]}")
                                continue
                        
                        if not xml_completo or ('<nfeProc' not in xml_completo and '<procNFe' not in xml_completo):
                            self.log(f"[REPROCESSAR]    ❌ XML completo não encontrado")
                            continue
                        
                        # Salva XML e atualiza banco
                        try:
                            # Busca nome do certificado
                            nome_cert = db_nfe.get_cert_nome_by_informante(informante)
                            
                            # Salva em xmls/
                            salvar_xml_por_certificado(xml_completo, informante, pasta_base="xmls", nome_certificado=nome_cert)
                            
                            # Salva em storage se configurado
                            pasta_storage = db_nfe.get_config('storage_pasta_base', 'xmls')
                            if pasta_storage and pasta_storage != 'xmls':
                                salvar_xml_por_certificado(xml_completo, informante, pasta_base=pasta_storage, nome_certificado=nome_cert)
                            
                            # Cria parser temporário para extrair nota
                            from nfe_search import XMLProcessor
                            parser_temp = XMLProcessor(informante=informante)
                            
                            # Extrai e salva nota
                            nota_detalhada = extrair_nota_detalhada(
                                xml_txt=xml_completo,
                                parser=parser_temp,
                                db=db_nfe,
                                chave=chave,
                                informante=informante,
                                nsu_documento=None
                            )
                            nota_detalhada['informante'] = informante
                            nota_detalhada['xml_status'] = 'COMPLETO'
                            db_nfe.salvar_nota_detalhada(nota_detalhada)
                            
                            self.log(f"[REPROCESSAR]    💾 Nota salva: {nota_detalhada.get('numero', 'N/A')}")
                            encontrados += 1
                            
                        except Exception as e:
                            self.log(f"[REPROCESSAR]    ❌ Erro ao salvar: {str(e)[:100]}")
                    
                    self.log(f"\n[REPROCESSAR] ========================================")
                    self.log(f"[REPROCESSAR] ✅ Concluído: {encontrados}/{total} XMLs baixados")
                    self.log(f"[REPROCESSAR] ========================================")
                    
                    self.finished.emit(encontrados, total)
                    
                except Exception as e:
                    self.log(f"[REPROCESSAR] ❌ ERRO: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    self.error.emit(f"Erro no reprocessamento:\n\n{str(e)}")
        
        # Cria worker para reprocessar resumos
        worker_repro = ReprocessarResumosWorker(self.parent_window)
        
        # Adiciona trabalho à lista (total será atualizado quando worker iniciar)
        trabalho = {
            'tipo': 'reprocessar_resumos',
            'nome': 'Reprocessamento de Resumos',
            'status': 'Em execução',
            'progresso': 0,
            'total': 0,  # Será atualizado pelo worker
            'mensagem': 'Iniciando...',
            'worker': worker_repro
        }
        self.parent_window._trabalhos_ativos.append(trabalho)
        
        # Conecta sinais do worker de reprocessamento
        def on_progress_repro(msg, atual, total):
            for t in self.parent_window._trabalhos_ativos:
                if t.get('tipo') == 'reprocessar_resumos':
                    t['progresso'] = atual
                    t['total'] = total
                    t['mensagem'] = msg
                    break
        
        def on_finished_repro(encontrados, total):
            # Remove da lista de trabalhos
            self.parent_window._trabalhos_ativos = [
                t for t in self.parent_window._trabalhos_ativos
                if t.get('tipo') != 'reprocessar_resumos'
            ]
            
            QMessageBox.information(
                self,
                "Reprocessamento Concluído",
                f"Reprocessamento de resumos concluído!\n\n"
                f"✅ XMLs completos encontrados: {encontrados}\n"
                f"📊 Total processado: {total}\n\n"
                f"A interface será atualizada."
            )
            
            # Atualiza interface
            self.parent_window.refresh_all()
        
        def on_error_repro(error_msg):
            QMessageBox.critical(self, "Erro", error_msg)
        
        def on_log_repro(msg):
            print(msg)
        
        worker_repro.progress.connect(on_progress_repro)
        worker_repro.finished.connect(on_finished_repro)
        worker_repro.error.connect(on_error_repro)
        worker_repro.log_message.connect(on_log_repro)
        
        # Inicia worker
        worker_repro.start()
    
    def _iniciar_reprocessamento_resumos(self):
        """Reprocessa notas que estão como RESUMO, buscando XML completo"""
        # Busca notas RESUMO
        with self.db._connect() as conn:
            cursor = conn.execute("""
                SELECT chave, informante, nome_emitente 
                FROM notas_detalhadas 
                WHERE xml_status = 'RESUMO'
                ORDER BY data_emissao DESC
            """)
            notas_resumo = [{'chave': r[0], 'informante': r[1], 'nome_emitente': r[2]} for r in cursor]
        
        if not notas_resumo:
            QMessageBox.information(
                self,
                "Nenhuma nota pendente",
                "Não há notas com status RESUMO para reprocessar."
            )
            return
        
        # Cria e inicia o worker
        worker = AutoVerificacaoWorker(self.parent_window, notas_resumo)
        
        # Adiciona à lista de trabalhos ativos
        trabalho = {
            'tipo': 'auto_verificacao',
            'nome': 'Auto-Verificação de XMLs',
            'status': 'Em execução',
            'progresso': 0,
            'total': len(notas_resumo),
            'mensagem': 'Iniciando...',
            'worker': worker
        }
        self.parent_window._trabalhos_ativos.append(trabalho)
        
        # Conecta sinais
        def on_progress(msg, atual, total):
            for t in self.parent_window._trabalhos_ativos:
                if t.get('tipo') == 'auto_verificacao':
                    t['progresso'] = atual
                    t['total'] = total
                    t['mensagem'] = msg
                    break
        
        def on_finished(encontrados, nao_encontrados):
            # Remove da lista de trabalhos
            self.parent_window._trabalhos_ativos = [
                t for t in self.parent_window._trabalhos_ativos
                if t.get('tipo') != 'auto_verificacao'
            ]
            
            # Atualiza interface de forma OTIMIZADA (usa QTimer para garantir execução na thread principal)
            from PyQt5.QtCore import QTimer
            def atualizar_interface():
                try:
                    print("[AUTO-VERIFICAÇÃO] Atualizando interface...")
                    
                    # 1. Apenas corrige xml_status (rápido - já está otimizado)
                    self.parent_window._corrigir_xml_status_automatico()
                    
                    # 2. Apenas atualiza a tabela SEM recarregar dados
                    # (os dados em self.notes já estão corretos, só precisa re-renderizar)
                    self.parent_window._refresh_table_only()
                    
                    print("[AUTO-VERIFICAÇÃO] Interface atualizada!")
                except Exception as e:
                    print(f"[AUTO-VERIFICAÇÃO] Erro ao atualizar interface: {e}")
                    import traceback
                    traceback.print_exc()
            
            # Executa atualização com delay pequeno (evita travamento)
            QTimer.singleShot(200, atualizar_interface)
            
            # Mostra resultado IMEDIATAMENTE (não espera atualização)
            print(f"[AUTO-VERIFICAÇÃO] Exibindo resultado final...")
            QMessageBox.information(
                self,
                "Auto-Verificação Concluída",
                f"Verificação concluída!\n\n"
                f"✅ XMLs encontrados: {encontrados}\n"
                f"❌ XMLs não encontrados: {nao_encontrados}\n\n"
                f"A interface será atualizada em instantes."
            )
        
        def on_error(error_msg):
            # Remove da lista de trabalhos
            self.parent_window._trabalhos_ativos = [
                t for t in self.parent_window._trabalhos_ativos
                if t.get('tipo') != 'auto_verificacao'
            ]
            
            if "Cancelado" not in error_msg:
                QMessageBox.warning(self, "Erro", f"Erro na auto-verificação:\n{error_msg}")
        
        worker.progress.connect(on_progress)
        worker.finished.connect(on_finished)
        worker.error.connect(on_error)
        
        # Conecta sinal de log para imprimir no terminal
        worker.log_message.connect(lambda msg: print(msg, flush=True))
        
        worker.start()
        
        QMessageBox.information(
            self,
            "Auto-Verificação Iniciada",
            f"A auto-verificação foi iniciada em segundo plano.\n\n"
            f"Total de notas: {len(notas_resumo)}\n\n"
            f"Você pode acompanhar o progresso nesta janela."
        )
    
    def _atualizar_lista(self):
        """Atualiza a lista de trabalhos"""
        # Remove workers que já terminaram
        self.workers = [w for w in self.workers if w.isRunning()]
        
        # Coleta trabalhos de múltiplas fontes
        trabalhos = []
        
        # 1. Trabalhos do parent_window (antigo sistema)
        if self.parent_window and hasattr(self.parent_window, '_trabalhos_ativos'):
            trabalhos.extend(self.parent_window._trabalhos_ativos)
        
        # 2. Workers ativos (novo sistema)
        for worker in self.workers:
            trabalhos.append({
                'tipo': 'auto_verificacao',
                'nome': 'Auto-Verificação de XMLs',
                'status': 'Em execução',
                'progresso': 0,
                'mensagem': 'Buscando XMLs completos...',
                'total': 100,
                'worker': worker
            })
        
        self.table.setRowCount(len(trabalhos))
        
        # Controla visibilidade do botão "Sincronizar Agora"
        # Esconde se houver sincronização em andamento, mostra caso contrário
        tem_sync_ativa = any(t.get('tipo') == 'sync_eventos' and t.get('status') == 'Em execução' 
                             for t in trabalhos)
        
        if hasattr(self, 'btn_sync'):
            self.btn_sync.setEnabled(not tem_sync_ativa)
            if tem_sync_ativa:
                self.btn_sync.setStyleSheet("""
                    QPushButton {
                        background-color: #cccccc;
                        color: #666666;
                        border: none;
                        padding: 8px 20px;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 11px;
                    }
                """)
                self.btn_sync.setText("⚡ Sincronizando...")
            else:
                self.btn_sync.setStyleSheet("""
                    QPushButton {
                        background-color: #16c60c;
                        color: white;
                        border: none;
                        padding: 8px 20px;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 11px;
                    }
                    QPushButton:hover {
                        background-color: #13a10e;
                    }
                    QPushButton:pressed {
                        background-color: #0e7c0a;
                    }
                """)
                self.btn_sync.setText("⚡ Sincronizar Agora")
        
        # Configura altura das linhas
        for i in range(len(trabalhos)):
            self.table.setRowHeight(i, 80)
        
        for idx, trabalho in enumerate(trabalhos):
            # Coluna 0: Nome da tarefa com estilo
            nome_widget = QWidget()
            nome_layout = QVBoxLayout()
            nome_layout.setContentsMargins(12, 10, 12, 10)
            nome_layout.setSpacing(5)
            
            nome_label = QLabel(f"🔄 {trabalho.get('nome', 'Tarefa')}")
            nome_label.setWordWrap(True)
            nome_label.setStyleSheet("""
                font-size: 13px;
                font-weight: bold;
                color: #333;
            """)
            nome_layout.addWidget(nome_label)
            
            tipo_label = QLabel(f"Tipo: {trabalho.get('tipo', 'desconhecido')}")
            tipo_label.setStyleSheet("""
                font-size: 10px;
                color: #999;
            """)
            nome_layout.addWidget(tipo_label)
            
            nome_widget.setLayout(nome_layout)
            self.table.setCellWidget(idx, 0, nome_widget)
            
            # Coluna 1: Status com badge colorido
            status = trabalho.get('status', 'Desconhecido')
            status_widget = QWidget()
            status_layout = QHBoxLayout()
            status_layout.setContentsMargins(8, 8, 8, 8)
            
            status_label = QLabel(status)
            status_label.setAlignment(Qt.AlignCenter)
            
            if status == 'Em execução':
                status_label.setStyleSheet("""
                    background-color: #0078d4;
                    color: white;
                    padding: 6px 12px;
                    border-radius: 12px;
                    font-weight: bold;
                    font-size: 11px;
                """)
            elif status == 'Pausado':
                status_label.setStyleSheet("""
                    background-color: #ff8c00;
                    color: white;
                    padding: 6px 12px;
                    border-radius: 12px;
                    font-weight: bold;
                    font-size: 11px;
                """)
            elif status == 'Concluído':
                status_label.setStyleSheet("""
                    background-color: #16c60c;
                    color: white;
                    padding: 6px 12px;
                    border-radius: 12px;
                    font-weight: bold;
                    font-size: 11px;
                """)
            else:
                status_label.setStyleSheet("""
                    background-color: #d13438;
                    color: white;
                    padding: 6px 12px;
                    border-radius: 12px;
                    font-weight: bold;
                    font-size: 11px;
                """)
            
            status_layout.addWidget(status_label)
            status_widget.setLayout(status_layout)
            self.table.setCellWidget(idx, 1, status_widget)
            
            # Coluna 2: Progresso com barra e informações
            progresso_widget = QWidget()
            progresso_layout = QVBoxLayout()
            progresso_layout.setContentsMargins(12, 10, 12, 10)
            progresso_layout.setSpacing(5)
            
            # Mensagem acima da barra
            mensagem = trabalho.get('mensagem', '')
            total = trabalho.get('total', 0)
            atual = trabalho.get('progresso', 0)
            
            msg_label = QLabel(mensagem)
            msg_label.setWordWrap(True)
            msg_label.setStyleSheet("""
                font-size: 11px;
                color: #000;
                font-weight: bold;
            """)
            progresso_layout.addWidget(msg_label)
            
            # Barra de progresso
            progresso_bar = QProgressBar()
            progresso_bar.setMaximum(trabalho.get('total', 100))
            progresso_bar.setValue(trabalho.get('progresso', 0))
            progresso_bar.setTextVisible(True)
            progresso_bar.setFixedHeight(26)
            progresso_bar.setStyleSheet("""
                QProgressBar {
                    border: 2px solid #e0e0e0;
                    border-radius: 5px;
                    text-align: center;
                    background-color: #f5f5f5;
                    color: #000;
                    font-weight: bold;
                    font-size: 11px;
                }
                QProgressBar::chunk {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                              stop:0 #0078d4, stop:1 #00bcf2);
                    border-radius: 3px;
                }
            """)
            
            if total > 0:
                percentual = int((atual / total) * 100)
                progresso_bar.setFormat(f"{atual}/{total} ({percentual}%)")
            else:
                progresso_bar.setFormat("Aguardando...")
            
            progresso_layout.addWidget(progresso_bar)
            progresso_widget.setLayout(progresso_layout)
            self.table.setCellWidget(idx, 2, progresso_widget)
            
            # Coluna 3: Ações com botões estilizados
            acoes_widget = QWidget()
            acoes_layout = QHBoxLayout()
            acoes_layout.setContentsMargins(10, 8, 10, 8)
            acoes_layout.setSpacing(10)
            
            worker = trabalho.get('worker')
            
            if status == 'Em execução':
                btn_pausar = QPushButton("⏸ Pausar")
                btn_pausar.setFixedHeight(36)
                btn_pausar.setCursor(Qt.PointingHandCursor)
                btn_pausar.setStyleSheet("""
                    QPushButton {
                        background-color: #ff8c00;
                        color: white;
                        border: none;
                        padding: 6px 16px;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background-color: #ff7700;
                    }
                    QPushButton:pressed {
                        background-color: #e67700;
                    }
                """)
                btn_pausar.clicked.connect(lambda checked, w=worker, t=trabalho: self._pausar(w, t))
                acoes_layout.addWidget(btn_pausar)
            elif status == 'Pausado':
                btn_retomar = QPushButton("▶ Retomar")
                btn_retomar.setFixedHeight(36)
                btn_retomar.setCursor(Qt.PointingHandCursor)
                btn_retomar.setStyleSheet("""
                    QPushButton {
                        background-color: #16c60c;
                        color: white;
                        border: none;
                        padding: 6px 16px;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background-color: #13a10e;
                    }
                    QPushButton:pressed {
                        background-color: #0e7c0a;
                    }
                """)
                btn_retomar.clicked.connect(lambda checked, w=worker, t=trabalho: self._retomar(w, t))
                acoes_layout.addWidget(btn_retomar)
            
            btn_cancelar = QPushButton("✖ Cancelar")
            btn_cancelar.setFixedHeight(36)
            btn_cancelar.setCursor(Qt.PointingHandCursor)
            btn_cancelar.setStyleSheet("""
                QPushButton {
                    background-color: #d13438;
                    color: white;
                    border: none;
                    padding: 6px 16px;
                    border-radius: 5px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #b52d30;
                }
                QPushButton:pressed {
                    background-color: #992628;
                }
            """)
            btn_cancelar.clicked.connect(lambda checked, w=worker, t=trabalho: self._cancelar(w, t))
            acoes_layout.addWidget(btn_cancelar)
            
            acoes_layout.addStretch()
            acoes_widget.setLayout(acoes_layout)
            self.table.setCellWidget(idx, 3, acoes_widget)
        
        # Atualizar rodapé com informações detalhadas
        if trabalhos:
            em_execucao = sum(1 for t in trabalhos if t.get('status') == 'Em execução')
            pausados = sum(1 for t in trabalhos if t.get('status') == 'Pausado')
            concluidos = sum(1 for t in trabalhos if t.get('status') == 'Concluído')
            
            if em_execucao > 0:
                self.status_label.setText(f"✅ {em_execucao} trabalho(s) em execução")
                self.status_label.setStyleSheet("""
                    QLabel {
                        background-color: transparent;
                        color: #16c60c;
                        font-size: 12px;
                        font-weight: bold;
                    }
                """)
            elif pausados > 0:
                self.status_label.setText(f"⏸ {pausados} trabalho(s) pausado(s)")
                self.status_label.setStyleSheet("""
                    QLabel {
                        background-color: transparent;
                        color: #ff8c00;
                        font-size: 12px;
                        font-weight: bold;
                    }
                """)
            else:
                self.status_label.setText(f"✅ Todos os trabalhos concluídos")
                self.status_label.setStyleSheet("""
                    QLabel {
                        background-color: transparent;
                        color: #16c60c;
                        font-size: 12px;
                        font-weight: bold;
                    }
                """)
            
            # Info adicional
            from datetime import datetime
            agora = datetime.now().strftime("%H:%M:%S")
            self.info_label.setText(f"Total: {len(trabalhos)} | Ativos: {em_execucao} | Pausados: {pausados} | Concluídos: {concluidos} | Atualizado: {agora}")
        else:
            self.status_label.setText("ℹ Nenhum trabalho em execução")
            self.status_label.setStyleSheet("""
                QLabel {
                    background-color: transparent;
                    color: #666;
                    font-size: 12px;
                    font-weight: bold;
                }
            """)
            self.info_label.setText("Clique em 'Sincronizar Agora' para iniciar uma nova tarefa")
    
    def _pausar(self, worker, trabalho):
        """Pausa um trabalho"""
        print(f"[DEBUG GERENCIADOR] _pausar chamado - worker: {worker}, trabalho: {trabalho.get('nome')}")
        if worker:
            print(f"[DEBUG GERENCIADOR] Chamando worker.pausar()...")
            worker.pausar()
            trabalho['status'] = 'Pausado'
            self._atualizar_lista()
            print(f"[DEBUG GERENCIADOR] Worker pausado!")
        else:
            print(f"[DEBUG GERENCIADOR] ERRO: worker é None!")
    
    def _retomar(self, worker, trabalho):
        """Retoma um trabalho pausado"""
        print(f"[DEBUG GERENCIADOR] _retomar chamado - worker: {worker}")
        if worker:
            print(f"[DEBUG GERENCIADOR] Chamando worker.retomar()...")
            worker.retomar()
            trabalho['status'] = 'Em execução'
            self._atualizar_lista()
            print(f"[DEBUG GERENCIADOR] Worker retomado!")
        else:
            print(f"[DEBUG GERENCIADOR] ERRO: worker é None!")
    
    def _cancelar(self, worker, trabalho):
        """Cancela um trabalho"""
        print(f"[DEBUG GERENCIADOR] _cancelar chamado - worker: {worker}, trabalho: {trabalho.get('nome')}")
        resposta = QMessageBox.question(
            self,
            "Cancelar Trabalho",
            f"Deseja realmente cancelar a tarefa '{trabalho.get('nome', 'Tarefa')}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if resposta == QMessageBox.Yes:
            print(f"[DEBUG GERENCIADOR] Usuário confirmou cancelamento")
            if worker:
                print(f"[DEBUG GERENCIADOR] Chamando worker.cancelar()...")
                worker.cancelar()
                print(f"[DEBUG GERENCIADOR] Worker cancelado!")
            else:
                print(f"[DEBUG GERENCIADOR] ERRO: worker é None!")
            trabalho['status'] = 'Cancelado'
            self._atualizar_lista()
        else:
            print(f"[DEBUG GERENCIADOR] Usuário cancelou o cancelamento")
    
    def closeEvent(self, event):
        """Para o timer ao fechar"""
        self.update_timer.stop()
        super().closeEvent(event)


class CertificateDialog(QDialog):
    def __init__(self, db: UIDB, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Gerenciar Certificados Digitais")
        self.resize(1100, 500)
        
        # Layout principal
        v = QVBoxLayout(self)
        v.setSpacing(15)
        v.setContentsMargins(20, 20, 20, 20)
        
        # Cabeçalho com título e estatísticas
        header = QHBoxLayout()
        title_label = QLabel("📜 Certificados Cadastrados")
        title_font = title_label.font()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        header.addWidget(title_label)
        
        self.stats_label = QLabel()
        self.stats_label.setStyleSheet("color: #666; font-size: 11px;")
        header.addWidget(self.stats_label)
        header.addStretch(1)
        v.addLayout(header)

        # Tabela estilizada
        self.table = QTableWidget()
        headers = ["Informante", "CNPJ/CPF", "Nome do Certificado", "UF", "Status", "Vencimento"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        # Habilita ordenação clicável nos cabeçalhos
        self.table.setSortingEnabled(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                background-color: white;
                gridline-color: #f0f0f0;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: black;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #ddd;
                font-weight: bold;
            }
        """)
        
        # Ajustar larguras das colunas
        self.table.setColumnWidth(0, 180)  # Informante
        self.table.setColumnWidth(1, 180)  # CNPJ
        self.table.setColumnWidth(2, 350)  # Nome do Certificado
        self.table.setColumnWidth(3, 60)   # UF
        self.table.setColumnWidth(4, 120)  # Status
        self.table.setColumnWidth(5, 180)  # Vencimento
        
        v.addWidget(self.table)

        # Botões estilizados
        h = QHBoxLayout()
        h.setSpacing(10)
        
        btn_add = QPushButton("➕ Adicionar Certificado")
        btn_add.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        
        btn_edit = QPushButton("✏️ Editar Selecionado")
        btn_edit.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
        """)
        
        btn_replace = QPushButton("🔄 Substituir Certificado")
        btn_replace.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E64A19;
            }
        """)
        
        btn_del = QPushButton("🗑️ Remover Selecionado")
        btn_del.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:pressed {
                background-color: #c4180a;
            }
        """)
        
        btn_close = QPushButton("✖️ Fechar")
        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #546E7A;
            }
            QPushButton:pressed {
                background-color: #455A64;
            }
        """)
        
        btn_add.clicked.connect(self._on_add)
        btn_edit.clicked.connect(self._on_edit)
        btn_replace.clicked.connect(self._on_replace)
        btn_del.clicked.connect(self._on_delete)
        btn_close.clicked.connect(self.accept)
        
        h.addStretch(1)
        h.addWidget(btn_add)
        h.addWidget(btn_edit)
        h.addWidget(btn_replace)
        h.addWidget(btn_del)
        h.addWidget(btn_close)
        v.addLayout(h)

        self.reload()

    def _extract_cert_name(self, cert_path: str) -> str:
        """Extrai o nome do certificado (CN) do arquivo .pfx"""
        try:
            from cryptography import x509
            from cryptography.hazmat.backends import default_backend
            import os
            
            if not os.path.exists(cert_path):
                return "❌ Arquivo não encontrado"
            
            # Tenta extrair informações do certificado
            # Nota: cryptography requer senha, então vamos apenas mostrar o nome do arquivo
            filename = os.path.basename(cert_path)
            return filename.replace('.pfx', '').replace('.p12', '')
        except Exception:
            # Fallback: retorna apenas o nome do arquivo
            try:
                import os
                filename = os.path.basename(cert_path)
                return filename.replace('.pfx', '').replace('.p12', '')
            except Exception:
                return "N/D"
    
    def _extract_cert_expiry(self, cert_path: str, senha: str = "") -> str:
        """Extrai a data de vencimento do certificado .pfx"""
        try:
            from cryptography.hazmat.primitives.serialization import pkcs12
            from cryptography.hazmat.backends import default_backend
            import os
            from datetime import datetime
            
            if not os.path.exists(cert_path):
                return "❌ Não encontrado"
            
            # Lê o arquivo do certificado
            with open(cert_path, 'rb') as f:
                pfx_data = f.read()
            
            # Tenta carregar o certificado
            try:
                # Tenta sem senha primeiro
                private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                    pfx_data, None, default_backend()
                )
            except Exception:
                try:
                    # Tenta com senha se fornecida
                    if senha:
                        private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                            pfx_data, senha.encode(), default_backend()
                        )
                    else:
                        return "N/D"
                except Exception:
                    return "N/D"
            
            if certificate:
                # Extrai data de validade
                expiry = certificate.not_valid_after_utc if hasattr(certificate, 'not_valid_after_utc') else certificate.not_valid_after
                expiry_str = expiry.strftime("%d/%m/%Y")
                
                # Verifica se está vencido
                hoje = datetime.now()
                if expiry.replace(tzinfo=None) < hoje:
                    return f"🔴 {expiry_str} (Vencido)"
                else:
                    # Verifica se vence em menos de 30 dias
                    dias_restantes = (expiry.replace(tzinfo=None) - hoje).days
                    if dias_restantes <= 30:
                        return f"🟡 {expiry_str} ({dias_restantes}d)"
                    else:
                        return f"🟢 {expiry_str}"
            
            return "N/D"
        except Exception as e:
            return "N/D"

    def reload(self):
        try:
            certs = self.db.load_certificates()
        except Exception:
            certs = []
        
        # Atualizar estatísticas
        total = len(certs)
        ativos = sum(1 for c in certs if c.get('ativo', 1) == 1)
        inativos = total - ativos
        self.stats_label.setText(f"Total: {total} | Ativos: {ativos} | Inativos: {inativos}")
        
        self.table.setRowCount(len(certs))
        for r, c in enumerate(certs):
            # Armazena ID como dado oculto na primeira coluna
            cert_id = c.get('id')
            
            def cell(x: Any, alignment=None):
                item = QTableWidgetItem(str(x or ""))
                if alignment:
                    item.setTextAlignment(alignment)
                # Armazena ID como UserRole em cada célula para referência
                item.setData(Qt.UserRole, cert_id)
                return item
            
            # Informante (CNPJ formatado)
            informante = c.get('informante', '')
            if informante and len(informante) >= 11:
                if len(informante) == 14:  # CNPJ
                    informante_fmt = f"{informante[:2]}.{informante[2:5]}.{informante[5:8]}/{informante[8:12]}-{informante[12:]}"
                else:  # CPF
                    informante_fmt = f"{informante[:3]}.{informante[3:6]}.{informante[6:9]}-{informante[9:]}"
            else:
                informante_fmt = informante
            self.table.setItem(r, 0, cell(informante_fmt))
            
            # CNPJ/CPF do certificado (formatado)
            cnpj_cpf = c.get('cnpj_cpf', '')
            if cnpj_cpf and len(cnpj_cpf) >= 11:
                if len(cnpj_cpf) == 14:
                    cnpj_fmt = f"{cnpj_cpf[:2]}.{cnpj_cpf[2:5]}.{cnpj_cpf[5:8]}/{cnpj_cpf[8:12]}-{cnpj_cpf[12:]}"
                else:
                    cnpj_fmt = f"{cnpj_cpf[:3]}.{cnpj_cpf[3:6]}.{cnpj_cpf[6:9]}-{cnpj_cpf[9:]}"
            else:
                cnpj_fmt = cnpj_cpf
            self.table.setItem(r, 1, cell(cnpj_fmt))
            
            # Nome do Certificado
            # Prioriza o nome personalizado, senão extrai do arquivo
            nome_cert = c.get('nome_certificado')
            if nome_cert:
                cert_name = nome_cert
            else:
                cert_name = self._extract_cert_name(c.get('caminho', ''))
            
            name_cell = cell(cert_name)
            name_cell.setForeground(QBrush(QColor("#2196F3")))
            name_cell.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.table.setItem(r, 2, name_cell)
            
            # UF
            uf_cell = cell(c.get('cUF_autor', ''), Qt.AlignCenter)
            self.table.setItem(r, 3, uf_cell)
            
            # Status (Ativo/Inativo com ícones)
            ativo = c.get('ativo', 1)
            status_text = "🟢 Ativo" if ativo == 1 else "🔴 Inativo"
            status_cell = cell(status_text, Qt.AlignCenter)
            if ativo == 1:
                status_cell.setForeground(QBrush(QColor("#4CAF50")))
            else:
                status_cell.setForeground(QBrush(QColor("#f44336")))
            status_cell.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.table.setItem(r, 4, status_cell)
            
            # Data de vencimento do certificado
            senha = c.get('senha', '')
            vencimento = self._extract_cert_expiry(c.get('caminho', ''), senha)
            venc_cell = cell(vencimento, Qt.AlignCenter)
            
            # Colorir conforme status do vencimento
            if "🔴" in vencimento:  # Vencido
                venc_cell.setForeground(QBrush(QColor("#f44336")))
                venc_cell.setFont(QFont("Segoe UI", 9, QFont.Bold))
            elif "🟡" in vencimento:  # Próximo ao vencimento
                venc_cell.setForeground(QBrush(QColor("#FF9800")))
                venc_cell.setFont(QFont("Segoe UI", 9, QFont.Bold))
            elif "🟢" in vencimento:  # Válido
                venc_cell.setForeground(QBrush(QColor("#4CAF50")))
            else:
                venc_cell.setForeground(QBrush(QColor("#999")))
            
            self.table.setItem(r, 5, venc_cell)

    def _on_add(self):
        dlg = AddCertificateDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            if not data:
                return
            
            success, error_msg = self.db.save_certificate(data)
            
            if success:
                QMessageBox.information(
                    self,
                    "✅ Sucesso",
                    f"Certificado adicionado com sucesso!\n\n"
                    f"Informante: {data.get('informante', 'N/D')}\n\n"
                    f"Observação: Se havia um registro antigo com o mesmo CNPJ/CPF,\n"
                    f"ele foi substituído automaticamente."
                )
            else:
                error_details = error_msg or "Erro desconhecido"
                QMessageBox.critical(
                    self, 
                    "❌ Erro ao Salvar Certificado", 
                    f"Não foi possível salvar o certificado.\n\n"
                    f"Detalhes do erro:\n{error_details}\n\n"
                    f"Verifique também os logs no terminal para mais informações."
                )
            self.reload()

    def _on_edit(self):
        """Abre diálogo para editar certificado selecionado"""
        idxs = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not idxs:
            QMessageBox.warning(self, "Editar", "Selecione um certificado para editar!")
            return
        
        row = idxs[0].row()
        # Recupera ID armazenado no UserRole da primeira coluna
        first_item = self.table.item(row, 0)
        if not first_item:
            return
        cert_id = first_item.data(Qt.UserRole)
        if not cert_id:
            return
        
        # Busca dados do certificado no banco
        try:
            with sqlite3.connect(self.db.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute("SELECT * FROM certificados WHERE id = ?", (cert_id,))
                cert_data = cursor.fetchone()
                
                if not cert_data:
                    QMessageBox.warning(self, "Editar", "Certificado não encontrado!")
                    return
                
                # Converte para dict
                cert_dict = dict(cert_data)
                
        except Exception as e:
            QMessageBox.critical(self, "Editar", f"Erro ao carregar certificado: {e}")
            return
        
        # Abre diálogo de edição
        dlg = EditCertificateDialog(self, cert_dict)
        if dlg.exec_() == QDialog.Accepted:
            updated_data = dlg.get_data()
            if updated_data:
                # Atualiza no banco
                try:
                    with sqlite3.connect(self.db.db_path) as conn:
                        conn.execute("""
                            UPDATE certificados 
                            SET nome_certificado = ?, cUF_autor = ?
                            WHERE id = ?
                        """, (updated_data.get('nome_certificado'), 
                              updated_data.get('cUF_autor'),
                              cert_id))
                        conn.commit()
                    
                    QMessageBox.information(
                        self,
                        "✅ Sucesso",
                        "Certificado atualizado com sucesso!"
                    )
                    self.reload()
                    
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao atualizar certificado: {e}")

    def _on_replace(self):
        """Substitui o arquivo e senha do certificado mantendo o histórico"""
        idxs = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not idxs:
            QMessageBox.warning(self, "Substituir", "Selecione um certificado para substituir!")
            return
        
        row = idxs[0].row()
        # Recupera ID armazenado no UserRole da primeira coluna
        first_item = self.table.item(row, 0)
        if not first_item:
            return
        cert_id = first_item.data(Qt.UserRole)
        if not cert_id:
            return
        
        # Busca dados do certificado no banco
        try:
            with sqlite3.connect(self.db.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute("SELECT * FROM certificados WHERE id = ?", (cert_id,))
                cert_data = cursor.fetchone()
                
                if not cert_data:
                    QMessageBox.warning(self, "Substituir", "Certificado não encontrado!")
                    return
                
                # Converte para dict
                cert_dict = dict(cert_data)
                
        except Exception as e:
            QMessageBox.critical(self, "Substituir", f"Erro ao carregar certificado: {e}")
            return
        
        # Abre diálogo de substituição
        dlg = ReplaceCertificateDialog(self, cert_dict)
        if dlg.exec_() == QDialog.Accepted:
            replacement_data = dlg.get_data()
            if replacement_data:
                # Atualiza no banco (apenas caminho e senha)
                try:
                    # Criptografa a senha antes de salvar
                    from modules.crypto_portable import get_portable_crypto as get_crypto
                    senha_to_save = replacement_data.get('senha', '')
                    if senha_to_save:
                        crypto = get_crypto()
                        senha_to_save = crypto.encrypt(senha_to_save)
                    
                    with sqlite3.connect(self.db.db_path) as conn:
                        conn.execute("""
                            UPDATE certificados 
                            SET caminho = ?, senha = ?
                            WHERE id = ?
                        """, (replacement_data.get('caminho'), 
                              senha_to_save,
                              cert_id))
                        conn.commit()
                    
                    QMessageBox.information(
                        self,
                        "✅ Sucesso",
                        f"Certificado substituído com sucesso!\n\n"
                        f"✔️ Novo arquivo: {replacement_data.get('caminho')}\n"
                        f"✔️ Nova validade: {replacement_data.get('validade', 'N/D')}\n\n"
                        f"📂 Todo o histórico de notas foi mantido."
                    )
                    self.reload()
                    
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao substituir certificado: {e}")

    def _on_delete(self):
        idxs = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not idxs:
            return
        row = idxs[0].row()
        # Recupera ID armazenado no UserRole da primeira coluna
        first_item = self.table.item(row, 0)
        if not first_item:
            return
        cert_id = first_item.data(Qt.UserRole)
        if not cert_id:
            return
        if QMessageBox.question(self, "Remover", f"Remover certificado selecionado?") != QMessageBox.Yes:
            return
        try:
            with sqlite3.connect(self.db.db_path) as conn:
                conn.execute("DELETE FROM certificados WHERE id = ?", (cert_id,))
                conn.commit()
        except Exception as e:
            QMessageBox.critical(self, "Remover", f"Erro ao remover: {e}")
            return
        self.reload()


class ReplaceCertificateDialog(QDialog):
    """Diálogo para substituir arquivo e senha do certificado"""
    def __init__(self, parent=None, cert_data: dict = None):
        super().__init__(parent)
        self.cert_data = cert_data or {}
        self.setWindowTitle("🔄 Substituir Certificado")
        self.resize(700, 550)
        
        # Estilo moderno
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                font-size: 12px;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 2px solid #FF9800;
            }
            QLineEdit:disabled {
                background-color: #f5f5f5;
                color: #666;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
        """)
        
        v = QVBoxLayout(self)
        v.setSpacing(20)
        v.setContentsMargins(30, 30, 30, 30)
        
        # Cabeçalho
        header_label = QLabel("🔄 Substituir Certificado Digital")
        header_font = header_label.font()
        header_font.setPointSize(13)
        header_font.setBold(True)
        header_label.setFont(header_font)
        v.addWidget(header_label)
        
        # Aviso importante
        warning_box = QLabel(
            "⚠️ <b>Importante:</b> Esta operação substitui apenas o arquivo .pfx e a senha.<br>"
            "Todo o histórico de notas e configurações serão mantidos."
        )
        warning_box.setStyleSheet("""
            QLabel {
                background-color: #FFF3CD;
                border: 2px solid #FFA726;
                border-radius: 5px;
                padding: 10px;
                color: #856404;
            }
        """)
        warning_box.setWordWrap(True)
        v.addWidget(warning_box)
        
        # Info do certificado atual (somente leitura)
        current_group = QGroupBox("📋 Certificado Atual")
        current_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        
        current_layout = QVBoxLayout()
        current_layout.setSpacing(8)
        
        # Informações do certificado atual
        cnpj_label = QLabel(f"🆔 CNPJ/CPF: {self.cert_data.get('cnpj_cpf', 'N/D')}")
        cnpj_label.setStyleSheet("color: #333; font-weight: normal;")
        current_layout.addWidget(cnpj_label)
        
        inf_label = QLabel(f"👤 Informante: {self.cert_data.get('informante', 'N/D')}")
        inf_label.setStyleSheet("color: #333; font-weight: normal;")
        current_layout.addWidget(inf_label)
        
        razao = self.cert_data.get('razao_social', 'N/D')
        if razao and razao != 'N/D' and len(razao) > 50:
            razao = razao[:50] + "..."
        razao_label = QLabel(f"🏢 Razão Social: {razao}")
        razao_label.setStyleSheet("color: #333; font-weight: normal;")
        current_layout.addWidget(razao_label)
        
        # Arquivo atual
        arquivo_atual = self.cert_data.get('caminho', 'N/D')
        if len(arquivo_atual) > 60:
            arquivo_atual = "..." + arquivo_atual[-60:]
        arquivo_label = QLabel(f"📄 Arquivo atual: {arquivo_atual}")
        arquivo_label.setStyleSheet("color: #666; font-style: italic; font-weight: normal; font-size: 10px;")
        arquivo_label.setWordWrap(True)
        current_layout.addWidget(arquivo_label)
        
        current_group.setLayout(current_layout)
        v.addWidget(current_group)
        
        # Novo certificado
        new_group = QGroupBox("🆕 Novo Certificado")
        new_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #FF9800;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        
        new_layout = QVBoxLayout()
        new_layout.setSpacing(15)
        
        # Campo de arquivo
        arquivo_label = QLabel("📁 Novo Arquivo do Certificado:")
        arquivo_label.setStyleSheet("font-weight: bold; color: #333;")
        new_layout.addWidget(arquivo_label)
        
        cert_layout = QHBoxLayout()
        self.cert_edit = QLineEdit()
        self.cert_edit.setPlaceholderText("Selecione o novo arquivo .pfx...")
        self.cert_edit.setReadOnly(True)
        
        btn_browse = QPushButton("🔍 Procurar...")
        btn_browse.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        btn_browse.clicked.connect(self._browse_cert)
        
        cert_layout.addWidget(self.cert_edit, 1)
        cert_layout.addWidget(btn_browse)
        new_layout.addLayout(cert_layout)
        
        # Campo de senha
        senha_label = QLabel("🔐 Nova Senha do Certificado:")
        senha_label.setStyleSheet("font-weight: bold; color: #333;")
        new_layout.addWidget(senha_label)
        
        self.senha_edit = QLineEdit()
        self.senha_edit.setPlaceholderText("Digite a senha do novo certificado...")
        self.senha_edit.setEchoMode(QLineEdit.Password)
        new_layout.addWidget(self.senha_edit)
        
        # Botão para validar
        validate_layout = QHBoxLayout()
        validate_layout.addStretch()
        btn_validate = QPushButton("🔎 Validar Novo Certificado")
        btn_validate.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        btn_validate.clicked.connect(self._validate_cert)
        validate_layout.addWidget(btn_validate)
        validate_layout.addStretch()
        new_layout.addLayout(validate_layout)
        
        # Label de validação
        self.validation_label = QLabel("")
        self.validation_label.setStyleSheet("padding: 5px; font-weight: normal;")
        self.validation_label.setWordWrap(True)
        new_layout.addWidget(self.validation_label)
        
        new_group.setLayout(new_layout)
        v.addWidget(new_group)
        
        v.addStretch()
        
        # Botões finais
        h = QHBoxLayout()
        h.addStretch()
        
        self.btn_save = QPushButton("🔄 Substituir")
        self.btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #ccc;
                color: #666;
            }
        """)
        self.btn_save.clicked.connect(self.accept)
        self.btn_save.setEnabled(False)  # Desabilitado até validar
        
        btn_cancel = QPushButton("❌ Cancelar")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        h.addWidget(self.btn_save)
        h.addWidget(btn_cancel)
        v.addLayout(h)
        
        # Armazena dados de validação
        self.validated_data = None
    
    def _browse_cert(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Novo Certificado Digital",
            "",
            "Certificados (*.pfx *.p12);;Todos os arquivos (*.*)"
        )
        if path:
            self.cert_edit.setText(path)
            self.validation_label.setText("")
            self.btn_save.setEnabled(False)
            self.validated_data = None
    
    def _validate_cert(self):
        """Valida o novo certificado"""
        cert_path = self.cert_edit.text().strip()
        senha = self.senha_edit.text().strip()
        
        if not cert_path:
            self.validation_label.setText("⚠️ Selecione o arquivo do certificado!")
            self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
            return
        
        try:
            from cryptography.hazmat.primitives.serialization import pkcs12
            from cryptography.hazmat.backends import default_backend
            from cryptography import x509
            from datetime import datetime
            import os
            
            if not os.path.exists(cert_path):
                self.validation_label.setText("❌ Arquivo não encontrado!")
                self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
                return
            
            # Lê o arquivo
            with open(cert_path, 'rb') as f:
                pfx_data = f.read()
            
            # Tenta carregar o certificado
            try:
                private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                    pfx_data,
                    senha.encode() if senha else None,
                    default_backend()
                )
            except Exception as e:
                self.validation_label.setText(f"❌ Erro ao carregar certificado!\n{str(e)[:100]}")
                self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
                return
            
            if not certificate:
                self.validation_label.setText("❌ Certificado não encontrado no arquivo!")
                self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
                return
            
            # Extrai CNPJ do novo certificado
            subject = certificate.subject
            cn = None
            for attr in subject:
                if attr.oid._name == 'commonName':
                    cn = attr.value
                    break
            
            # Extrai CNPJ do CN
            import re
            novo_cnpj = None
            if cn:
                match = re.search(r'\d{14}', cn)
                if match:
                    novo_cnpj = match.group(0)
            
            # Valida se é o mesmo CNPJ
            cnpj_atual = self.cert_data.get('informante', '')
            if novo_cnpj != cnpj_atual:
                self.validation_label.setText(
                    f"❌ CNPJ incompatível!\n"
                    f"Atual: {cnpj_atual}\n"
                    f"Novo: {novo_cnpj or 'Não identificado'}"
                )
                self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
                return
            
            # Extrai validade
            validade = certificate.not_valid_after_utc
            validade_str = validade.strftime("%d/%m/%Y")
            
            # Verifica se não está vencido
            hoje = datetime.now(validade.tzinfo)
            if validade < hoje:
                self.validation_label.setText(
                    f"⚠️ Certificado já vencido!\n"
                    f"Vencimento: {validade_str}"
                )
                self.validation_label.setStyleSheet("color: #FF9800; padding: 5px;")
                self.btn_save.setEnabled(False)
                return
            
            # Sucesso!
            dias_restantes = (validade.replace(tzinfo=None) - datetime.now()).days
            self.validation_label.setText(
                f"✅ Certificado válido!\n"
                f"CNPJ: {novo_cnpj}\n"
                f"Validade: {validade_str} ({dias_restantes} dias)"
            )
            self.validation_label.setStyleSheet("color: #4CAF50; padding: 5px; font-weight: bold;")
            self.btn_save.setEnabled(True)
            
            # Armazena dados validados
            self.validated_data = {
                'caminho': cert_path,
                'senha': senha,
                'cnpj': novo_cnpj,
                'validade': validade_str
            }
            
        except Exception as e:
            self.validation_label.setText(f"❌ Erro na validação: {str(e)[:100]}")
            self.validation_label.setStyleSheet("color: #f44336; padding: 5px;")
            import traceback
            traceback.print_exc()
    
    def get_data(self) -> Optional[Dict[str, Any]]:
        """Retorna os dados validados"""
        return self.validated_data


class EditCertificateDialog(QDialog):
    """Diálogo para editar dados do certificado"""
    def __init__(self, parent=None, cert_data: dict = None):
        super().__init__(parent)
        self.cert_data = cert_data or {}
        self.setWindowTitle("✏️ Editar Certificado")
        self.resize(600, 400)
        
        # Estilo moderno
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                font-size: 12px;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 2px solid #2196F3;
            }
            QLineEdit:disabled {
                background-color: #f5f5f5;
                color: #666;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
        """)
        
        v = QVBoxLayout(self)
        v.setSpacing(20)
        v.setContentsMargins(30, 30, 30, 30)
        
        # Cabeçalho
        header_label = QLabel("Editar dados do certificado")
        header_font = header_label.font()
        header_font.setPointSize(13)
        header_font.setBold(True)
        header_label.setFont(header_font)
        v.addWidget(header_label)
        
        # Info do certificado (somente leitura)
        info_group = QGroupBox("📋 Informações do Certificado")
        info_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        
        info_layout = QVBoxLayout()
        info_layout.setSpacing(10)
        
        # CNPJ/CPF
        cnpj_label = QLabel(f"🆔 CNPJ/CPF: {self.cert_data.get('cnpj_cpf', 'N/D')}")
        cnpj_label.setStyleSheet("color: #333; font-weight: normal;")
        info_layout.addWidget(cnpj_label)
        
        # Informante
        inf_label = QLabel(f"👤 Informante: {self.cert_data.get('informante', 'N/D')}")
        inf_label.setStyleSheet("color: #333; font-weight: normal;")
        info_layout.addWidget(inf_label)
        
        # Razão Social
        razao = self.cert_data.get('razao_social', 'N/D')
        if razao and razao != 'N/D' and len(razao) > 50:
            razao = razao[:50] + "..."
        razao_label = QLabel(f"🏢 Razão Social: {razao}")
        razao_label.setStyleSheet("color: #333; font-weight: normal;")
        info_layout.addWidget(razao_label)
        
        info_group.setLayout(info_layout)
        v.addWidget(info_group)
        
        # Campos editáveis
        edit_group = QGroupBox("✏️ Campos Editáveis")
        edit_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #2196F3;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        
        edit_layout = QVBoxLayout()
        edit_layout.setSpacing(15)
        
        # Nome do Certificado
        nome_label = QLabel("📝 Nome do Certificado:")
        nome_label.setStyleSheet("font-weight: bold; color: #333;")
        edit_layout.addWidget(nome_label)
        
        self.nome_edit = QLineEdit()
        self.nome_edit.setText(self.cert_data.get('nome_certificado', ''))
        self.nome_edit.setPlaceholderText("Ex: Walter Transportes, Empresa Filial SP...")
        self.nome_edit.setStyleSheet("""
            QLineEdit {
                background-color: #fffef0;
                border: 2px solid #FFA726;
            }
            QLineEdit:focus {
                border: 2px solid #FF9800;
            }
        """)
        edit_layout.addWidget(self.nome_edit)
        
        # Nota sobre nome
        nota_nome = QLabel("💡 Este nome será usado ao salvar arquivos em vez do CNPJ")
        nota_nome.setStyleSheet("color: #666; font-style: italic; font-size: 10px;")
        nota_nome.setWordWrap(True)
        edit_layout.addWidget(nota_nome)
        
        # UF Autor
        uf_label = QLabel("📍 UF Autor:")
        uf_label.setStyleSheet("font-weight: bold; color: #333;")
        edit_layout.addWidget(uf_label)
        
        self.uf_edit = QLineEdit()
        self.uf_edit.setText(str(self.cert_data.get('cUF_autor', '')))
        self.uf_edit.setPlaceholderText("Ex: 33 (Rio de Janeiro)")
        edit_layout.addWidget(self.uf_edit)
        
        edit_group.setLayout(edit_layout)
        v.addWidget(edit_group)
        
        v.addStretch()
        
        # Botões
        h = QHBoxLayout()
        h.addStretch()
        
        btn_save = QPushButton("💾 Salvar")
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_save.clicked.connect(self.accept)
        
        btn_cancel = QPushButton("❌ Cancelar")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        h.addWidget(btn_save)
        h.addWidget(btn_cancel)
        v.addLayout(h)
    
    def get_data(self) -> Optional[Dict[str, Any]]:
        """Retorna os dados editados"""
        nome = self.nome_edit.text().strip()
        uf = self.uf_edit.text().strip()
        
        return {
            "nome_certificado": nome if nome else None,
            "cUF_autor": uf if uf else self.cert_data.get('cUF_autor')
        }


class AddCertificateDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📜 Adicionar Certificado Digital")
        self.resize(700, 600)
        
        # Estilo moderno para o diálogo
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                font-size: 12px;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 2px solid #2196F3;
            }
            QLineEdit:disabled {
                background-color: #f5f5f5;
                color: #666;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
        """)
        
        v = QVBoxLayout(self)
        v.setSpacing(20)
        v.setContentsMargins(30, 30, 30, 30)
        
        # Cabeçalho
        header_label = QLabel("Selecione o arquivo do certificado digital (.pfx)")
        header_font = header_label.font()
        header_font.setPointSize(13)
        header_font.setBold(True)
        header_label.setFont(header_font)
        v.addWidget(header_label)
        
        # Seção: Arquivo do certificado
        cert_section = QVBoxLayout()
        cert_section.setSpacing(10)
        
        cert_label = QLabel("📁 Arquivo do Certificado:")
        cert_label.setStyleSheet("font-weight: bold; color: #333;")
        cert_section.addWidget(cert_label)
        
        cert_layout = QHBoxLayout()
        self.cert_edit = QLineEdit()
        self.cert_edit.setPlaceholderText("Selecione o arquivo .pfx do certificado...")
        self.cert_edit.setReadOnly(True)
        
        btn_browse = QPushButton("🔍 Procurar...")
        btn_browse.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        btn_browse.clicked.connect(self._browse_cert)
        
        cert_layout.addWidget(self.cert_edit, 1)
        cert_layout.addWidget(btn_browse)
        cert_section.addLayout(cert_layout)
        v.addLayout(cert_section)
        
        # Seção: Senha
        senha_section = QVBoxLayout()
        senha_section.setSpacing(10)
        
        senha_label = QLabel("🔐 Senha do Certificado:")
        senha_label.setStyleSheet("font-weight: bold; color: #333;")
        senha_section.addWidget(senha_label)
        
        self.senha_edit = QLineEdit()
        self.senha_edit.setPlaceholderText("Digite a senha do certificado...")
        self.senha_edit.setEchoMode(QLineEdit.Password)
        senha_section.addWidget(self.senha_edit)
        v.addLayout(senha_section)
        
        # Botão para extrair informações
        extract_layout = QHBoxLayout()
        extract_layout.addStretch()
        btn_extract = QPushButton("🔎 Extrair Informações do Certificado")
        btn_extract.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        btn_extract.clicked.connect(self._extract_cert_info)
        extract_layout.addWidget(btn_extract)
        extract_layout.addStretch()
        v.addLayout(extract_layout)
        
        # Linha separadora
        separator = QLabel()
        separator.setFrameStyle(QLabel.HLine | QLabel.Sunken)
        separator.setStyleSheet("background-color: #ddd; max-height: 2px;")
        v.addWidget(separator)
        
        # Informações extraídas (desabilitadas por padrão)
        info_label = QLabel("ℹ️ Informações Extraídas:")
        info_label.setStyleSheet("font-weight: bold; color: #333; font-size: 13px;")
        v.addWidget(info_label)
        
        # Grid com informações
        from PyQt5.QtWidgets import QGridLayout
        grid = QGridLayout()
        grid.setSpacing(15)
        grid.setColumnStretch(1, 1)
        
        # Informante
        grid.addWidget(QLabel("👤 Informante (CNPJ/CPF):"), 0, 0)
        self.informante_edit = QLineEdit()
        self.informante_edit.setPlaceholderText("Será preenchido automaticamente...")
        self.informante_edit.setReadOnly(True)
        grid.addWidget(self.informante_edit, 0, 1)
        
        # CNPJ/CPF do certificado
        grid.addWidget(QLabel("🆔 CNPJ/CPF do Certificado:"), 1, 0)
        self.cnpj_edit = QLineEdit()
        self.cnpj_edit.setPlaceholderText("Será preenchido automaticamente...")
        self.cnpj_edit.setReadOnly(True)
        grid.addWidget(self.cnpj_edit, 1, 1)
        
        # Razão Social
        grid.addWidget(QLabel("🏢 Razão Social:"), 2, 0)
        self.razao_social_edit = QLineEdit()
        self.razao_social_edit.setPlaceholderText("Será preenchido automaticamente...")
        self.razao_social_edit.setReadOnly(True)
        grid.addWidget(self.razao_social_edit, 2, 1)
        
        # UF
        grid.addWidget(QLabel("📍 UF Autor:"), 3, 0)
        self.uf_edit = QLineEdit()
        self.uf_edit.setPlaceholderText("Ex: 33 (Rio de Janeiro)")
        grid.addWidget(self.uf_edit, 3, 1)
        
        # Titular
        grid.addWidget(QLabel("📋 Titular:"), 4, 0)
        self.titular_edit = QLineEdit()
        self.titular_edit.setPlaceholderText("Será preenchido automaticamente...")
        self.titular_edit.setReadOnly(True)
        grid.addWidget(self.titular_edit, 4, 1)
        
        # Validade
        grid.addWidget(QLabel("📅 Válido até:"), 5, 0)
        self.validade_edit = QLineEdit()
        self.validade_edit.setPlaceholderText("Será preenchido automaticamente...")
        self.validade_edit.setReadOnly(True)
        grid.addWidget(self.validade_edit, 5, 1)
        
        # Nome do Certificado (campo editável)
        grid.addWidget(QLabel("📝 Nome do Certificado:"), 6, 0)
        self.nome_cert_edit = QLineEdit()
        self.nome_cert_edit.setPlaceholderText("Ex: Walter Transportes, Empresa Filial SP...")
        self.nome_cert_edit.setReadOnly(False)
        self.nome_cert_edit.setStyleSheet("""
            QLineEdit {
                background-color: #fffef0;
                border: 2px solid #FFA726;
            }
            QLineEdit:focus {
                border: 2px solid #FF9800;
            }
        """)
        grid.addWidget(self.nome_cert_edit, 6, 1)
        
        # Nota explicativa
        nota_label = QLabel("💡 O nome do certificado será usado ao salvar arquivos em vez do CNPJ")
        nota_label.setStyleSheet("color: #666; font-style: italic; font-size: 10px; padding-top: 5px;")
        nota_label.setWordWrap(True)
        grid.addWidget(nota_label, 7, 0, 1, 2)
        
        v.addLayout(grid)
        
        v.addStretch()
        
        # Botões finais
        h = QHBoxLayout()
        h.addStretch()
        
        btn_save = QPushButton("💾 Salvar")
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_save.clicked.connect(self.accept)
        
        btn_cancel = QPushButton("❌ Cancelar")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        h.addWidget(btn_save)
        h.addWidget(btn_cancel)
        v.addLayout(h)
    
    def _browse_cert(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Certificado Digital",
            "",
            "Certificados (*.pfx *.p12);;Todos os arquivos (*.*)"
        )
        if path:
            self.cert_edit.setText(path)
    
    def _consultar_uf_cnpj(self, cnpj: str) -> tuple[Optional[str], Optional[str]]:
        """Consulta UF e Razão Social do CNPJ via API Brasil.
        
        Returns:
            tuple: (uf, razao_social) ou (None, None) em caso de erro
        """
        try:
            import requests
            url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
            print(f"[DEBUG] Consultando dados do CNPJ {cnpj} via API Brasil...")
            
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                uf = data.get('uf')
                razao_social = data.get('razao_social')
                print(f"[DEBUG] UF encontrada: {uf}, Razão Social: {razao_social}")
                return (uf, razao_social)
            else:
                print(f"[DEBUG] Erro ao consultar CNPJ: status {response.status_code}")
                return (None, None)
        except Exception as e:
            print(f"[DEBUG] Erro ao consultar API Brasil: {e}")
            return (None, None)
    
    def _extract_cert_info(self):
        """Extrai informações do certificado automaticamente."""
        cert_path = self.cert_edit.text().strip()
        senha = self.senha_edit.text().strip()
        
        if not cert_path:
            QMessageBox.warning(self, "Atenção", "Selecione primeiro o arquivo do certificado!")
            return
        
        if not senha:
            reply = QMessageBox.question(
                self,
                "Senha",
                "Nenhuma senha foi fornecida. Deseja tentar extrair sem senha?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
        
        try:
            from cryptography.hazmat.primitives.serialization import pkcs12
            from cryptography.hazmat.backends import default_backend
            from cryptography import x509
            from datetime import datetime
            import os
            
            if not os.path.exists(cert_path):
                QMessageBox.critical(self, "Erro", "Arquivo do certificado não encontrado!")
                return
            
            # Lê o arquivo
            with open(cert_path, 'rb') as f:
                pfx_data = f.read()
            
            # Tenta carregar o certificado
            try:
                private_key, certificate, additional_certs = pkcs12.load_key_and_certificates(
                    pfx_data,
                    senha.encode() if senha else None,
                    default_backend()
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Erro ao Carregar Certificado",
                    f"Não foi possível carregar o certificado.\n\n"
                    f"Verifique se a senha está correta.\n\n"
                    f"Erro: {e}"
                )
                return
            
            if not certificate:
                QMessageBox.warning(self, "Aviso", "Certificado não encontrado no arquivo!")
                return
            
            # Extrai informações do certificado
            subject = certificate.subject
            
            # Nome comum (CN)
            cn = None
            for attr in subject:
                if attr.oid == x509.NameOID.COMMON_NAME:
                    cn = attr.value
                    break
            
            # CNPJ ou CPF (procura em diferentes campos)
            documento = None
            
            # Tenta extrair do campo serialNumber ou do CN
            for attr in subject:
                if attr.oid == x509.NameOID.SERIAL_NUMBER:
                    # Remove caracteres não numéricos
                    serial = ''.join(c for c in attr.value if c.isdigit())
                    if len(serial) in [11, 14]:  # CPF ou CNPJ
                        documento = serial
                        break
            
            # Se não encontrou no serialNumber, tenta extrair do CN
            if not documento and cn:
                # Procura padrão de CNPJ/CPF no CN
                import re
                # CNPJ: XX.XXX.XXX/XXXX-XX ou 14 dígitos
                cnpj_match = re.search(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', cn)
                if cnpj_match:
                    documento = ''.join(c for c in cnpj_match.group(1) if c.isdigit())
                else:
                    # CPF: XXX.XXX.XXX-XX ou 11 dígitos
                    cpf_match = re.search(r'(\d{3}\.?\d{3}\.?\d{3}-?\d{2})', cn)
                    if cpf_match:
                        documento = ''.join(c for c in cpf_match.group(1) if c.isdigit())
            
            # Data de validade
            expiry = certificate.not_valid_after_utc if hasattr(certificate, 'not_valid_after_utc') else certificate.not_valid_after
            validade_str = expiry.strftime("%d/%m/%Y")
            
            # Verifica se está vencido
            hoje = datetime.now()
            if expiry.replace(tzinfo=None) < hoje:
                status_validade = f"❌ VENCIDO em {validade_str}"
            else:
                dias_restantes = (expiry.replace(tzinfo=None) - hoje).days
                if dias_restantes <= 30:
                    status_validade = f"⚠️ {validade_str} (Vence em {dias_restantes} dias)"
                else:
                    status_validade = f"✅ {validade_str}"
            
            # Preenche os campos
            if documento:
                self.informante_edit.setText(documento)
                self.cnpj_edit.setText(documento)
                
                # Consulta UF e Razão Social automaticamente via API Brasil
                if len(documento) == 14:  # É CNPJ
                    uf_encontrada, razao_social = self._consultar_uf_cnpj(documento)
                    
                    # Preenche razão social
                    if razao_social:
                        self.razao_social_edit.setText(razao_social)
                        print(f"[DEBUG] Razão Social preenchida: {razao_social}")
                    
                    # Preenche UF
                    if uf_encontrada:
                        # Mapeia UF para código
                        uf_to_codigo = {
                            'AC': '12', 'AL': '27', 'AP': '16', 'AM': '13', 'BA': '29',
                            'CE': '23', 'DF': '53', 'ES': '32', 'GO': '52', 'MA': '21',
                            'MT': '51', 'MS': '50', 'MG': '31', 'PA': '15', 'PB': '25',
                            'PR': '41', 'PE': '26', 'PI': '22', 'RJ': '33', 'RN': '24',
                            'RS': '43', 'RO': '11', 'RR': '14', 'SC': '42', 'SP': '35',
                            'SE': '28', 'TO': '17'
                        }
                        codigo_uf = uf_to_codigo.get(uf_encontrada)
                        if codigo_uf:
                            # Preenche no campo de texto
                            self.uf_edit.setText(codigo_uf)
                            print(f"[DEBUG] UF preenchida automaticamente: {codigo_uf} ({uf_encontrada})")
            else:
                QMessageBox.warning(
                    self,
                    "Atenção",
                    "Não foi possível extrair o CNPJ/CPF do certificado.\n"
                    "Você precisará preencher manualmente."
                )
            
            self.titular_edit.setText(cn or "N/D")
            self.validade_edit.setText(status_validade)
            
            uf_msg = ""
            razao_msg = ""
            if len(documento or '') == 14:
                uf_msg = f"\nUF: Preenchida automaticamente"
                if self.razao_social_edit.text():
                    razao_msg = f"\nRazão Social: {self.razao_social_edit.text()}"
            
            QMessageBox.information(
                self,
                "Sucesso",
                f"Informações extraídas com sucesso!\n\n"
                f"Titular: {cn or 'N/D'}\n"
                f"CNPJ/CPF: {documento or 'Não encontrado'}{razao_msg}\n"
                f"Validade: {status_validade}{uf_msg}"
            )
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "Erro",
                f"Erro ao extrair informações do certificado:\n\n{e}"
            )

    def get_data(self) -> Optional[Dict[str, Any]]:
        """Retorna os dados preenchidos no formulário."""
        cert_path = self.cert_edit.text().strip()
        senha = self.senha_edit.text().strip()
        informante = self.informante_edit.text().strip()
        cnpj_cpf = self.cnpj_edit.text().strip()
        cuf = self.uf_edit.text().strip()
        razao_social = self.razao_social_edit.text().strip()
        nome_certificado = self.nome_cert_edit.text().strip()
        
        if not cert_path:
            QMessageBox.warning(self, "Atenção", "Selecione o arquivo do certificado!")
            return None
        
        # Verifica se o arquivo existe
        import os
        if not os.path.exists(cert_path):
            QMessageBox.critical(
                self, 
                "Erro",
                f"Arquivo do certificado não encontrado:\n\n{cert_path}\n\n"
                "Verifique se o arquivo ainda existe no local especificado."
            )
            return None
        
        if not informante:
            QMessageBox.warning(self, "Atenção", "Informante não foi preenchido!\n\nClique em 'Extrair Informações' primeiro.")
            return None
        
        if not cnpj_cpf:
            QMessageBox.warning(self, "Atenção", "CNPJ/CPF do certificado não foi preenchido!\n\nClique em 'Extrair Informações' primeiro.")
            return None
        
        if not cuf:
            QMessageBox.warning(self, "Atenção", "Preencha o campo 'UF Autor' (código da UF, ex: 33 para RJ)")
            return None
        
        # Validação do código UF (deve ser numérico entre 11 e 53)
        try:
            cuf_int = int(cuf)
            if cuf_int < 11 or cuf_int > 53:
                QMessageBox.warning(
                    self, 
                    "Atenção",
                    f"Código UF inválido: {cuf}\n\n"
                    "Use um código entre 11 e 53.\n"
                    "Ex: 33 (Rio de Janeiro), 35 (São Paulo)"
                )
                return None
        except ValueError:
            QMessageBox.warning(
                self,
                "Atenção", 
                f"Código UF deve ser numérico!\n\n"
                f"Valor informado: '{cuf}'\n\n"
                "Ex: 33 (Rio de Janeiro), 35 (São Paulo)"
            )
            return None
        
        print(f"[DEBUG] Dados do certificado validados:")
        print(f"  - Informante: {informante}")
        print(f"  - CNPJ/CPF: {cnpj_cpf}")
        print(f"  - Caminho: {cert_path}")
        print(f"  - UF Autor: {cuf}")
        print(f"  - Razão Social: {razao_social or '(não preenchido)'}")
        print(f"  - Nome Certificado: {nome_certificado or '(não preenchido)'}")
        
        return {
            "informante": informante,
            "cnpj_cpf": cnpj_cpf,
            "caminho": cert_path,
            "senha": senha,
            "cUF_autor": cuf,
            "ativo": 1,
            "razao_social": razao_social if razao_social else None,
            "nome_certificado": nome_certificado if nome_certificado else None
        }


class AutofillWorker(QThread):
    log_line = pyqtSignal(str)
    percent = pyqtSignal(int)

    def __init__(self, db: UIDB, items: List[Dict[str, Any]]):
        super().__init__()
        self.db = db
        self.items = items

    def _emit(self, txt: str):
        try:
            self.log_line.emit(txt)
        except Exception:
            pass

    def run(self):
        try:
            total = max(1, len(self.items))
            # Pré-carrega certificados
            try:
                certs = self.db.load_certificates()
            except Exception:
                certs = []
            for idx, item in enumerate(self.items, start=1):
                chave = (item.get('chave') or '').strip()
                tipo = (item.get('tipo') or 'NFe').strip()
                informante = (item.get('informante') or '').strip()
                if not chave:
                    continue
                self.percent.emit(int((idx-1)/total*100))
                self._emit(f"[Autofill] {idx}/{len(self.items)} chave={chave}")

                # 1) Tenta XML local
                xml_text = resolve_xml_text(item)
                # 2) Se não encontrou, tenta SEFAZ via sandbox com os certificados
                if not xml_text:
                    # Prioriza certificado do informante quando possível
                    ordered = list(certs)
                    try:
                        ordered.sort(key=lambda c: 0 if (str(c.get('informante') or '').strip() == informante) else 1)
                    except Exception:
                        pass
                    prefer = ('nfeProc', 'NFe') if tipo.upper().replace('-', '') in ('NFE', 'NFE') else ('procCTe', 'CTe')
                    for c in ordered:
                        try:
                            payload = {
                                "cert": {
                                    "path": c.get('caminho') or '',
                                    "senha": c.get('senha') or '',
                                    "cnpj": c.get('cnpj_cpf') or '',
                                    "cuf": c.get('cUF_autor') or ''
                                },
                                "chave": chave,
                                "prefer": list(prefer)
                            }
                            res = sandbox.run_task("fetch_by_chave", payload, timeout=240)
                            if res.get("ok") and res.get("data", {}).get("xml"):
                                xml_text = res["data"]["xml"]
                                break
                        except Exception:
                            continue
                if not xml_text:
                    self._emit(f"  - XML não encontrado (local/SEFAZ): {chave}")
                    continue
                # 3) Extrai número/data e atualiza banco (e xml_status COMPLETO)
                numero = ''
                data_emissao = ''
                try:
                    import xml.etree.ElementTree as ET
                    tree = ET.fromstring(xml_text.encode('utf-8'))
                    nfe_ns = '{http://www.portalfiscal.inf.br/nfe}'
                    cte_ns = '{http://www.portalfiscal.inf.br/cte}'
                    if tree.find(f'.//{nfe_ns}infNFe') is not None:
                        numero = tree.findtext(f'.//{nfe_ns}ide/{nfe_ns}nNF') or ''
                        data_emissao = tree.findtext(f'.//{nfe_ns}ide/{nfe_ns}dhEmi') or tree.findtext(f'.//{nfe_ns}ide/{nfe_ns}dEmi') or ''
                    elif tree.find(f'.//{cte_ns}infCte') is not None:
                        numero = tree.findtext(f'.//{cte_ns}ide/{cte_ns}nCT') or ''
                        data_emissao = tree.findtext(f'.//{cte_ns}ide/{cte_ns}dhEmi') or tree.findtext(f'.//{cte_ns}ide/{cte_ns}dEmi') or ''
                    if data_emissao:
                        data_emissao = data_emissao[:19]
                except Exception:
                    pass
                upd: Dict[str, Any] = {'chave': chave}
                if numero:
                    upd['numero'] = numero
                if data_emissao:
                    upd['data_emissao'] = data_emissao
                if xml_text:
                    upd['xml_status'] = 'COMPLETO'
                if len(upd) > 1:
                    ok = self.db.save_note(upd)
                    if ok:
                        self._emit(f"  - Atualizado: numero={numero or '-'} data={data_emissao or '-'} status={'COMPLETO' if xml_text else 'RESUMO'}")
                self.percent.emit(int(idx/total*100))
        except Exception as e:
            try:
                self._emit(f"[Autofill] Erro: {e}")
            except Exception:
                pass


class ExportDialog(QDialog):
    """Diálogo para configurar opções de exportação de arquivos"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📤 Exportar Arquivos")
        self.resize(500, 400)
        
        # Estilo moderno
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #ddd;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
                font-size: 13px;
            }
            QLabel {
                font-size: 11px;
            }
            QRadioButton, QCheckBox {
                font-size: 11px;
                padding: 5px;
            }
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton#btn_exportar {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton#btn_exportar:hover {
                background-color: #45a049;
            }
            QPushButton#btn_cancelar {
                background-color: #f44336;
                color: white;
            }
            QPushButton#btn_cancelar:hover {
                background-color: #da190b;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        titulo = QLabel("Selecione as opções de exportação:")
        titulo.setStyleSheet("font-size: 14px; font-weight: bold; color: #333; margin-bottom: 10px;")
        layout.addWidget(titulo)
        
        # Grupo: Tipo de arquivo
        grupo_tipo = QGroupBox("📁 Tipo de Arquivo")
        layout_tipo = QVBoxLayout()
        
        self.radio_xml = QRadioButton("Exportar apenas XML")
        self.radio_pdf = QRadioButton("Exportar apenas PDF")
        self.radio_ambos = QRadioButton("Exportar PDF e XML")
        self.radio_ambos.setChecked(True)  # Padrão
        
        layout_tipo.addWidget(self.radio_xml)
        layout_tipo.addWidget(self.radio_pdf)
        layout_tipo.addWidget(self.radio_ambos)
        grupo_tipo.setLayout(layout_tipo)
        layout.addWidget(grupo_tipo)
        
        # Grupo: Nomenclatura
        grupo_nome = QGroupBox("📝 Nomenclatura dos Arquivos")
        layout_nome = QVBoxLayout()
        
        self.radio_nome_padrao = QRadioButton("Padrão (Chave de Acesso)")
        self.radio_nome_personalizado = QRadioButton("Personalizado (Número + Nome do Documento)")
        self.radio_nome_padrao.setChecked(True)  # Padrão
        
        # Explicação
        label_explicacao = QLabel(
            "• Padrão: Arquivo será salvo com a chave de acesso completa\n"
            "  Exemplo: 35210112345678000190550010000123451234567890.xml\n\n"
            "• Personalizado: Arquivo será salvo com número e nome\n"
            "  Exemplo: 123456_Nome_da_Empresa.xml"
        )
        label_explicacao.setStyleSheet("font-size: 10px; color: #666; padding: 10px; background: #f0f0f0; border-radius: 5px;")
        label_explicacao.setWordWrap(True)
        
        layout_nome.addWidget(self.radio_nome_padrao)
        layout_nome.addWidget(self.radio_nome_personalizado)
        layout_nome.addWidget(label_explicacao)
        grupo_nome.setLayout(layout_nome)
        layout.addWidget(grupo_nome)
        
        # Espaçador
        layout.addStretch()
        
        # Botões
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_exportar = QPushButton("✅ Exportar")
        btn_exportar.setObjectName("btn_exportar")
        btn_exportar.clicked.connect(self.accept)
        btn_exportar.setMinimumWidth(120)
        
        btn_cancelar = QPushButton("❌ Cancelar")
        btn_cancelar.setObjectName("btn_cancelar")
        btn_cancelar.clicked.connect(self.reject)
        btn_cancelar.setMinimumWidth(120)
        
        btn_layout.addWidget(btn_exportar)
        btn_layout.addWidget(btn_cancelar)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def get_opcoes(self):
        """Retorna as opções selecionadas pelo usuário"""
        return {
            'exportar_xml': self.radio_xml.isChecked() or self.radio_ambos.isChecked(),
            'exportar_pdf': self.radio_pdf.isChecked() or self.radio_ambos.isChecked(),
            'nome_personalizado': self.radio_nome_personalizado.isChecked()
        }


class StorageConfigDialog(QDialog):
    """Diálogo para configurar perfis de armazenamento múltiplos"""
    
    def __init__(self, db: UIDB, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_profile_id = None
        self.setWindowTitle("⚙️ Perfis de Armazenamento")
        self.resize(900, 650)
        
        # Estilo moderno
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #ddd;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
                font-size: 13px;
            }
            QLabel {
                font-size: 11px;
            }
            QLineEdit, QComboBox {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
            }
            QLineEdit:focus, QComboBox:focus {
                border: 2px solid #2196F3;
            }
            QRadioButton {
                font-size: 11px;
                spacing: 8px;
            }
            QPushButton {
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 11px;
                font-weight: bold;
            }
            QListWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #eee;
            }
            QListWidget::item:selected {
                background-color: #2196F3;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #e3f2fd;
            }
        """)
        
        # Layout principal horizontal (lista + configurações)
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # === LADO ESQUERDO: Lista de Perfis ===
        left_panel = QVBoxLayout()
        left_panel.setSpacing(10)
        
        # Título da lista
        profiles_title = QLabel("📋 Perfis de Armazenamento")
        profiles_title_font = profiles_title.font()
        profiles_title_font.setPointSize(12)
        profiles_title_font.setBold(True)
        profiles_title.setFont(profiles_title_font)
        profiles_title.setStyleSheet("color: #333;")
        left_panel.addWidget(profiles_title)
        
        # Lista de perfis
        self.profiles_list = QListWidget()
        self.profiles_list.setMaximumWidth(250)
        self.profiles_list.setMinimumHeight(400)
        self.profiles_list.currentItemChanged.connect(self._on_profile_selected)
        left_panel.addWidget(self.profiles_list)
        
        # Botões de gerenciamento de perfis
        profiles_buttons_layout = QVBoxLayout()
        profiles_buttons_layout.setSpacing(5)
        
        btn_add_profile = QPushButton("➕ Novo Perfil")
        btn_add_profile.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_add_profile.clicked.connect(self._add_profile)
        profiles_buttons_layout.addWidget(btn_add_profile)
        
        btn_delete_profile = QPushButton("🗑️ Excluir Perfil")
        btn_delete_profile.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_delete_profile.clicked.connect(self._delete_profile)
        profiles_buttons_layout.addWidget(btn_delete_profile)
        
        left_panel.addLayout(profiles_buttons_layout)
        
        # Info sobre perfis ativos
        info_profiles = QLabel(
            "💡 <b>Dica:</b><br>"
            "• Arquivos são salvos em TODOS<br>"
            "  os perfis ativos (✅)<br>"
            "• Desmarque para desativar<br>"
            "• Excluir NÃO apaga arquivos"
        )
        info_profiles.setStyleSheet("""
            background-color: #fff3cd;
            border: 1px solid #ffc107;
            border-radius: 6px;
            padding: 10px;
            color: #856404;
            font-size: 9px;
        """)
        info_profiles.setWordWrap(True)
        left_panel.addWidget(info_profiles)
        
        main_layout.addLayout(left_panel)
        
        # === LADO DIREITO: Configurações do Perfil Selecionado ===
        right_panel = QVBoxLayout()
        right_panel.setSpacing(15)
        
        # Título do perfil selecionado
        self.profile_title = QLabel("📝 Selecione um perfil")
        profile_title_font = self.profile_title.font()
        profile_title_font.setPointSize(12)
        profile_title_font.setBold(True)
        self.profile_title.setFont(profile_title_font)
        self.profile_title.setStyleSheet("color: #333;")
        right_panel.addWidget(self.profile_title)
        
        # Scroll area para as configurações
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        
        scroll_content = QWidget()
        config_layout = QVBoxLayout(scroll_content)
        config_layout.setSpacing(15)
        
        # === GRUPO 1: Nome do Perfil ===
        group_nome = QGroupBox("📝 Nome do Perfil")
        nome_layout = QVBoxLayout()
        nome_layout.setSpacing(10)
        
        nome_label = QLabel("Identificação deste perfil de armazenamento:")
        nome_label.setStyleSheet("font-weight: normal; color: #666;")
        nome_layout.addWidget(nome_label)
        
        self.nome_edit = QLineEdit()
        self.nome_edit.setPlaceholderText("Ex: Pasta do Contador, Backup Nuvem, etc.")
        nome_layout.addWidget(self.nome_edit)
        
        group_nome.setLayout(nome_layout)
        config_layout.addWidget(group_nome)
        
        # === GRUPO 2: Pasta Base ===
        group_pasta = QGroupBox("📂 Pasta Base de Armazenamento")
        pasta_layout = QVBoxLayout()
        pasta_layout.setSpacing(10)
        
        pasta_label = QLabel("Caminho completo da pasta onde os arquivos serão salvos:")
        pasta_label.setStyleSheet("font-weight: normal; color: #666;")
        pasta_layout.addWidget(pasta_label)
        
        # Layout horizontal para campo + botão
        pasta_h_layout = QHBoxLayout()
        pasta_h_layout.setSpacing(8)
        
        self.pasta_edit = QLineEdit()
        self.pasta_edit.setPlaceholderText("Ex: C:/Arquivo Walter - Empresas/Notas NFe")
        pasta_h_layout.addWidget(self.pasta_edit, 1)
        
        btn_browse = QPushButton("📁 Procurar...")
        btn_browse.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        btn_browse.clicked.connect(self._browse_folder)
        pasta_h_layout.addWidget(btn_browse)
        
        pasta_layout.addLayout(pasta_h_layout)
        
        pasta_obs = QLabel("💡 Os arquivos serão salvos em: [PASTA_BASE]/[CNPJ]/[MÊS]/[TIPO]/")
        pasta_obs.setStyleSheet("color: #888; font-size: 10px; font-style: italic;")
        pasta_layout.addWidget(pasta_obs)
        
        group_pasta.setLayout(pasta_layout)
        config_layout.addWidget(group_pasta)
        
        # === GRUPO 3: Formato do Mês ===
        group_mes = QGroupBox("📅 Formato da Pasta de Mês")
        mes_layout = QVBoxLayout()
        mes_layout.setSpacing(10)
        
        mes_label = QLabel("Como deseja organizar as pastas por mês:")
        mes_label.setStyleSheet("font-weight: normal; color: #666;")
        mes_layout.addWidget(mes_label)
        
        self.formato_combo = QComboBox()
        self.formato_combo.addItem("📅 AAAA-MM  (2025-01, 2025-02...)", "AAAA-MM")
        self.formato_combo.addItem("📅 MM-AAAA  (01-2025, 02-2025...)", "MM-AAAA")
        self.formato_combo.addItem("📅 MMAAAA  (012025, 022025...)", "MMAAAA")
        self.formato_combo.addItem("📅 AAAA/MM  (2025/01, 2025/02...)", "AAAA/MM")
        self.formato_combo.addItem("📅 MM/AAAA  (01/2025, 02/2025...)", "MM/AAAA")
        mes_layout.addWidget(self.formato_combo)
        
        self.mes_exemplo = QLabel("📁 Exemplo: xmls/33251845000109/2025-01/NFe/")
        self.mes_exemplo.setStyleSheet("color: #888; font-size: 10px; font-style: italic; margin-top: 5px;")
        mes_layout.addWidget(self.mes_exemplo)
        
        group_mes.setLayout(mes_layout)
        config_layout.addWidget(group_mes)
        
        # === GRUPO 3.5: Hierarquia de Pastas ===
        group_hierarquia = QGroupBox("📊 Hierarquia de Pastas (definido na criação)")
        hierarquia_layout = QVBoxLayout()
        hierarquia_layout.setSpacing(8)
        
        self.lbl_tipo_org = QLabel()
        self.lbl_tipo_org.setStyleSheet("""
            background-color: #e8f5e9;
            border: 1px solid #4CAF50;
            border-radius: 4px;
            padding: 8px;
            font-size: 11px;
        """)
        self.lbl_tipo_org.setWordWrap(True)
        hierarquia_layout.addWidget(self.lbl_tipo_org)
        
        hierarquia_info = QLabel("💡 A hierarquia é definida na criação do perfil e não pode ser alterada")
        hierarquia_info.setStyleSheet("color: #888; font-size: 9px; font-style: italic;")
        hierarquia_layout.addWidget(hierarquia_info)
        
        group_hierarquia.setLayout(hierarquia_layout)
        config_layout.addWidget(group_hierarquia)
        
        # === GRUPO 4: Organização XML/PDF ===
        group_org = QGroupBox("🗂️ Organização de XML e PDF")
        org_layout = QVBoxLayout()
        org_layout.setSpacing(12)
        
        org_label = QLabel("Como deseja organizar XML e PDF:")
        org_label.setStyleSheet("font-weight: normal; color: #666;")
        org_layout.addWidget(org_label)
        
        self.radio_juntos = QRadioButton("📄 XMLs e PDFs na mesma pasta")
        self.radio_juntos.setToolTip("Exemplo: xmls/33251845000109/2025-01/NFe/ (contém .xml e .pdf)")
        org_layout.addWidget(self.radio_juntos)
        
        juntos_exemplo = QLabel("      └─ Exemplo: NFe/nota12345.xml + nota12345.pdf na mesma pasta")
        juntos_exemplo.setStyleSheet("color: #888; font-size: 10px; margin-left: 30px;")
        org_layout.addWidget(juntos_exemplo)
        
        self.radio_separados = QRadioButton("📁 XMLs e PDFs em pastas separadas")
        self.radio_separados.setToolTip("Exemplo: xmls/ e pdfs/ em pastas diferentes")
        self.radio_separados.setChecked(True)
        org_layout.addWidget(self.radio_separados)
        
        sep_exemplo = QLabel("      └─ Exemplo: xmls/33251845000109/2025-01/NFe/ e pdfs/33251845000109/2025-01/NFe/")
        sep_exemplo.setStyleSheet("color: #888; font-size: 10px; margin-left: 30px;")
        org_layout.addWidget(sep_exemplo)
        
        group_org.setLayout(org_layout)
        config_layout.addWidget(group_org)
        
        # === GRUPO 5: Status do Perfil ===
        group_status = QGroupBox("⚡ Status do Perfil")
        status_layout = QVBoxLayout()
        status_layout.setSpacing(10)
        
        self.checkbox_ativo = QCheckBox("✅ Perfil ativo (arquivos serão salvos neste perfil)")
        self.checkbox_ativo.setChecked(True)
        self.checkbox_ativo.setStyleSheet("font-weight: normal; font-size: 11px;")
        status_layout.addWidget(self.checkbox_ativo)
        
        status_info = QLabel("💡 Desmarque para desativar este perfil temporariamente")
        status_info.setStyleSheet("color: #888; font-size: 10px; font-style: italic;")
        status_layout.addWidget(status_info)
        
        group_status.setLayout(status_layout)
        config_layout.addWidget(group_status)
        
        # === INFORMAÇÃO ADICIONAL ===
        info_box = QLabel(
            "ℹ️ <b>Estrutura de Pastas:</b><br>"
            "   • <b>NFe</b>: Notas Fiscais Eletrônicas<br>"
            "   • <b>CTe</b>: Conhecimentos de Transporte<br>"
            "   • <b>NFe/Eventos</b>: Eventos de NF-e (cancelamento, carta de correção...)<br>"
            "   • <b>CTe/Eventos</b>: Eventos de CT-e"
        )
        info_box.setStyleSheet("""
            background-color: #e3f2fd;
            border: 1px solid #90caf9;
            border-radius: 6px;
            padding: 12px;
            color: #1565c0;
            font-size: 10px;
        """)
        info_box.setWordWrap(True)
        config_layout.addWidget(info_box)
        
        config_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        right_panel.addWidget(scroll)
        
        # === BOTÕES DE AÇÃO ===
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        btn_apply = QPushButton("🔄 Aplicar (Copiar XMLs)")
        btn_apply.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                min-width: 140px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        btn_apply.setToolTip("Copia todos os XMLs e PDFs existentes para este perfil com a estrutura configurada")
        btn_apply.clicked.connect(self._apply_profile)
        
        btn_save = QPushButton("💾 Salvar Perfil")
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_save.clicked.connect(self._save_profile)
        
        btn_close = QPushButton("✖ Fechar")
        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_close.clicked.connect(self.accept)
        
        button_layout.addWidget(btn_apply)
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_close)
        right_panel.addLayout(button_layout)
        
        main_layout.addLayout(right_panel, 1)
        
        # Conecta mudanças para atualizar exemplo
        self.pasta_edit.textChanged.connect(self._update_example)
        self.formato_combo.currentIndexChanged.connect(self._update_example)
        
        # Carrega perfis
        self._ensure_table_exists()
        self._load_profiles()
    
    def _ensure_table_exists(self):
        """Garante que a tabela de perfis existe"""
        try:
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            # Verifica se tabela existe
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='perfis_armazenamento'")
            if not cursor.fetchone():
                # Cria tabela e migra config atual
                cursor.execute("""
                    CREATE TABLE perfis_armazenamento (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        nome TEXT NOT NULL,
                        pasta_base TEXT NOT NULL,
                        formato_pasta_mes TEXT DEFAULT 'AAAA-MM',
                        xml_pdf_separado INTEGER DEFAULT 1,
                        organizacao_tipo TEXT DEFAULT 'CERTIFICADO_TIPO',
                        ativo INTEGER DEFAULT 1,
                        is_default INTEGER DEFAULT 0,
                        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Cria índices
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_perfil_ativo ON perfis_armazenamento(ativo)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_perfil_default ON perfis_armazenamento(is_default)")
                
                # Migra config atual para Perfil 1
                pasta_base = self.db.get_config('storage_pasta_base', str(DATA_DIR / 'xmls'))
                formato_mes = self.db.get_config('storage_formato_mes', 'AAAA-MM')
                xml_pdf_separado = int(self.db.get_config('storage_xml_pdf_separado', '1'))
                
                cursor.execute("""
                    INSERT INTO perfis_armazenamento 
                    (nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo, ativo, is_default)
                    VALUES (?, ?, ?, ?, ?, 1, 1)
                """, ("Perfil 1", pasta_base, formato_mes, xml_pdf_separado, 'CERTIFICADO_TIPO'))
            else:
                # Tabela já existe - verifica se tem coluna organizacao_tipo
                cursor.execute("PRAGMA table_info(perfis_armazenamento)")
                columns = {col[1] for col in cursor.fetchall()}
                if 'organizacao_tipo' not in columns:
                    # Adiciona coluna para bancos antigos
                    cursor.execute("ALTER TABLE perfis_armazenamento ADD COLUMN organizacao_tipo TEXT DEFAULT 'CERTIFICADO_TIPO'")
                    print("[INFO] Coluna 'organizacao_tipo' adicionada aos perfis existentes")
                
                conn.commit()
                print("[INFO] Tabela de perfis criada e Perfil 1 migrado")
            
            conn.close()
        except Exception as e:
            print(f"[ERRO] Ao garantir tabela de perfis: {e}")
    
    def _load_profiles(self):
        """Carrega todos os perfis do banco"""
        try:
            self.profiles_list.clear()
            
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo, ativo, is_default
                FROM perfis_armazenamento
                ORDER BY is_default DESC, id ASC
            """)
            
            profiles = cursor.fetchall()
            conn.close()
            
            if not profiles:
                # Nenhum perfil encontrado, cria Perfil 1 padrão
                self._create_default_profile()
                self._load_profiles()  # Recarrega
                return
            
            for profile in profiles:
                profile_id, nome, pasta_base, formato, separado, organizacao_tipo, ativo, is_default = profile
                
                # Monta texto do item
                status_icon = "✅" if ativo else "⭕"
                default_icon = "⭐" if is_default else ""
                item_text = f"{status_icon} {nome} {default_icon}"
                
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, profile_id)
                
                # Cor diferente se inativo
                if not ativo:
                    item.setForeground(QBrush(QColor("#999")))
                
                self.profiles_list.addItem(item)
            
            # Seleciona o primeiro perfil
            if self.profiles_list.count() > 0:
                self.profiles_list.setCurrentRow(0)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar perfis:\n{e}")
            print(f"[ERRO] _load_profiles: {e}")
    
    def _create_default_profile(self):
        """Cria perfil padrão (Perfil 1)"""
        try:
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            pasta_base = self.db.get_config('storage_pasta_base', str(DATA_DIR / 'xmls'))
            formato_mes = self.db.get_config('storage_formato_mes', 'AAAA-MM')
            xml_pdf_separado = int(self.db.get_config('storage_xml_pdf_separado', '1'))
            
            cursor.execute("""
                INSERT INTO perfis_armazenamento 
                (nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo, ativo, is_default)
                VALUES (?, ?, ?, ?, ?, 1, 1)
            """, ("Perfil 1", pasta_base, formato_mes, xml_pdf_separado, 'CERTIFICADO_TIPO'))
            
            conn.commit()
            conn.close()
            
            print("[INFO] Perfil 1 (padrão) criado")
        except Exception as e:
            print(f"[ERRO] _create_default_profile: {e}")
    
    def _on_profile_selected(self, current, previous):
        """Quando um perfil é selecionado na lista"""
        if not current:
            return
        
        profile_id = current.data(Qt.UserRole)
        self._load_profile_config(profile_id)
    
    def _load_profile_config(self, profile_id):
        """Carrega configurações de um perfil específico"""
        try:
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo, ativo
                FROM perfis_armazenamento
                WHERE id = ?
            """, (profile_id,))
            
            profile = cursor.fetchone()
            conn.close()
            
            if not profile:
                return
            
            nome, pasta_base, formato, separado, organizacao_tipo, ativo = profile
            
            # Guarda organizacao_tipo para exibir
            self.current_organizacao_tipo = organizacao_tipo or 'CERTIFICADO_TIPO'
            
            # Atualiza título
            self.profile_title.setText(f"📝 Editando: {nome}")
            
            # Atualiza label de hierarquia
            if self.current_organizacao_tipo == 'TIPO_CERTIFICADO':
                self.lbl_tipo_org.setText(
                    "📂 <b>Tipo → Certificado → Mês</b><br>"
                    "<span style='color: #666;'>Exemplo: NFe/61-MATPARCG/012026/</span>"
                )
            else:
                self.lbl_tipo_org.setText(
                    "📂 <b>Certificado → Tipo → Mês</b> (padrão)<br>"
                    "<span style='color: #666;'>Exemplo: 61-MATPARCG/012026/NFe/</span>"
                )
            
            # Preenche campos
            self.nome_edit.setText(nome)
            self.pasta_edit.setText(pasta_base)
            
            # Formato do mês
            for i in range(self.formato_combo.count()):
                if self.formato_combo.itemData(i) == formato:
                    self.formato_combo.setCurrentIndex(i)
                    break
            
            # XML/PDF separado
            if separado:
                self.radio_separados.setChecked(True)
            else:
                self.radio_juntos.setChecked(True)
            
            # Status
            self.checkbox_ativo.setChecked(bool(ativo))
            
            # Atualiza exemplo
            self._update_example()
            
            # Salva ID atual
            self.current_profile_id = profile_id
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar perfil:\n{e}")
            print(f"[ERRO] _load_profile_config: {e}")
    
    def _add_profile(self):
        """Adiciona novo perfil"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox, QPushButton
            
            # Cria dialog customizado
            dialog = QDialog(self)
            dialog.setWindowTitle("Novo Perfil de Armazenamento")
            dialog.setMinimumWidth(500)
            
            layout = QVBoxLayout(dialog)
            
            # Nome do perfil
            layout.addWidget(QLabel("Nome do perfil:"))
            txt_nome = QLineEdit(f"Perfil {self.profiles_list.count() + 1}")
            layout.addWidget(txt_nome)
            
            # Tipo de organização
            layout.addWidget(QLabel("Organização de pastas:"))
            combo_org = QComboBox()
            combo_org.addItem("Certificado → Tipo → Mês (padrão)", "CERTIFICADO_TIPO")
            combo_org.addItem("Tipo → Certificado → Mês (novo)", "TIPO_CERTIFICADO")
            combo_org.setToolTip(
                "CERTIFICADO_TIPO: 61-MATPARCG/012026/NFe/\n"
                "TIPO_CERTIFICADO: NFe/61-MATPARCG/012026/"
            )
            layout.addWidget(combo_org)
            
            # Exemplo visual
            lbl_exemplo = QLabel()
            lbl_exemplo.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 8px; border: 1px solid #ccc; }")
            layout.addWidget(lbl_exemplo)
            
            def atualizar_exemplo():
                org_tipo = combo_org.currentData()
                if org_tipo == "TIPO_CERTIFICADO":
                    exemplo = "📁 Exemplo:\nNFe/\n  └─ 61-MATPARCG/\n      └─ 012026/\n          └─ arquivo.xml"
                else:
                    exemplo = "📁 Exemplo:\n61-MATPARCG/\n  └─ 012026/\n      └─ NFe/\n          └─ arquivo.xml"
                lbl_exemplo.setText(exemplo)
            
            combo_org.currentIndexChanged.connect(atualizar_exemplo)
            atualizar_exemplo()
            
            # Botões
            btn_layout = QHBoxLayout()
            btn_ok = QPushButton("Criar")
            btn_cancel = QPushButton("Cancelar")
            btn_layout.addWidget(btn_ok)
            btn_layout.addWidget(btn_cancel)
            layout.addLayout(btn_layout)
            
            btn_ok.clicked.connect(dialog.accept)
            btn_cancel.clicked.connect(dialog.reject)
            
            if dialog.exec_() != QDialog.Accepted:
                return
            
            nome = txt_nome.text().strip()
            if not nome:
                QMessageBox.warning(self, "Aviso", "Nome do perfil não pode estar vazio!")
                return
            
            organizacao_tipo = combo_org.currentData()
            
            # Usa configurações do perfil atual ou padrão
            if self.current_profile_id:
                conn = sqlite3.connect(self.db.db_path)
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT pasta_base, formato_pasta_mes, xml_pdf_separado
                    FROM perfis_armazenamento
                    WHERE id = ?
                """, (self.current_profile_id,))
                
                result = cursor.fetchone()
                if result:
                    pasta_base, formato, separado = result
                else:
                    pasta_base = str(DATA_DIR / 'xmls')
                    formato = 'AAAA-MM'
                    separado = 1
                
                conn.close()
            else:
                pasta_base = str(DATA_DIR / 'xmls')
                formato = 'AAAA-MM'
                separado = 1
            
            # Insere novo perfil
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO perfis_armazenamento 
                (nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo, ativo, is_default)
                VALUES (?, ?, ?, ?, ?, 1, 0)
            """, (nome, pasta_base, formato, separado, organizacao_tipo))
            
            new_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            # Recarrega lista
            self._load_profiles()
            
            # Seleciona novo perfil
            for i in range(self.profiles_list.count()):
                item = self.profiles_list.item(i)
                if item.data(Qt.UserRole) == new_id:
                    self.profiles_list.setCurrentRow(i)
                    break
            
            QMessageBox.information(self, "Sucesso", f"Perfil '{nome}' criado com sucesso!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao criar perfil:\n{e}")
            print(f"[ERRO] _add_profile: {e}")
    
    def _delete_profile(self):
        """Exclui perfil selecionado"""
        try:
            current_item = self.profiles_list.currentItem()
            if not current_item:
                QMessageBox.warning(self, "Atenção", "Selecione um perfil para excluir!")
                return
            
            profile_id = current_item.data(Qt.UserRole)
            
            # Verifica se é o perfil padrão
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT nome, is_default FROM perfis_armazenamento WHERE id = ?", (profile_id,))
            result = cursor.fetchone()
            
            if not result:
                conn.close()
                return
            
            nome, is_default = result
            
            # Conta quantos perfis existem
            cursor.execute("SELECT COUNT(*) FROM perfis_armazenamento")
            total_perfis = cursor.fetchone()[0]
            
            if total_perfis <= 1:
                QMessageBox.warning(
                    self,
                    "Atenção",
                    "Não é possível excluir o último perfil!\n\n"
                    "Pelo menos um perfil deve existir para salvar arquivos."
                )
                conn.close()
                return
            
            # Confirma exclusão
            reply = QMessageBox.question(
                self,
                "Confirmar Exclusão",
                f"Deseja realmente excluir o perfil '{nome}'?\n\n"
                f"⚠️ IMPORTANTE:\n"
                f"• O perfil será removido do sistema\n"
                f"• Arquivos já salvos NÃO serão apagados\n"
                f"• Novos arquivos não serão mais salvos neste perfil",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                conn.close()
                return
            
            # Exclui perfil
            cursor.execute("DELETE FROM perfis_armazenamento WHERE id = ?", (profile_id,))
            conn.commit()
            
            # Se era o perfil padrão, define outro como padrão
            if is_default:
                cursor.execute("""
                    UPDATE perfis_armazenamento 
                    SET is_default = 1 
                    WHERE id = (SELECT MIN(id) FROM perfis_armazenamento)
                """)
                conn.commit()
            
            conn.close()
            
            # Recarrega lista
            self._load_profiles()
            
            QMessageBox.information(self, "Sucesso", f"Perfil '{nome}' excluído com sucesso!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao excluir perfil:\n{e}")
            print(f"[ERRO] _delete_profile: {e}")
    
    def _save_profile(self):
        """Salva configurações do perfil atual"""
        try:
            if not self.current_profile_id:
                QMessageBox.warning(self, "Atenção", "Selecione um perfil para salvar!")
                return
            
            nome = self.nome_edit.text().strip()
            pasta = self.pasta_edit.text().strip()
            
            if not nome:
                QMessageBox.warning(self, "Atenção", "Informe um nome para o perfil!")
                return
            
            if not pasta:
                QMessageBox.warning(self, "Atenção", "Informe o caminho da pasta base!")
                return
            
            # Valida pasta
            try:
                pasta_path = Path(pasta)
                
                if not pasta_path.exists():
                    reply = QMessageBox.question(
                        self,
                        "Pasta não existe",
                        f"A pasta:\n{pasta}\n\nNão existe. Deseja criá-la?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    
                    if reply == QMessageBox.Yes:
                        pasta_path.mkdir(parents=True, exist_ok=True)
                    else:
                        return
                
                if not pasta_path.is_dir():
                    QMessageBox.warning(self, "Atenção", "O caminho informado não é um diretório válido!")
                    return
                    
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Caminho inválido:\n{e}")
                return
            
            # Salva no banco
            formato = self.formato_combo.currentData()
            separado = 1 if self.radio_separados.isChecked() else 0
            ativo = 1 if self.checkbox_ativo.isChecked() else 0
            
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE perfis_armazenamento
                SET nome = ?,
                    pasta_base = ?,
                    formato_pasta_mes = ?,
                    xml_pdf_separado = ?,
                    ativo = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (nome, str(pasta_path), formato, separado, ativo, self.current_profile_id))
            
            conn.commit()
            conn.close()
            
            # Recarrega lista para atualizar ícones
            self._load_profiles()
            
            # Reseleciona o perfil atual
            for i in range(self.profiles_list.count()):
                item = self.profiles_list.item(i)
                if item.data(Qt.UserRole) == self.current_profile_id:
                    self.profiles_list.setCurrentRow(i)
                    break
            
            QMessageBox.information(
                self,
                "✅ Sucesso",
                f"Perfil '{nome}' salvo com sucesso!\n\n"
                f"📁 Pasta: {pasta_path}\n"
                f"📅 Formato: {formato}\n"
                f"📄 XML/PDF: {'Separados' if separado else 'Juntos'}\n"
                f"⚡ Status: {'Ativo ✅' if ativo else 'Inativo ⭕'}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar perfil:\n{e}")
            print(f"[ERRO] _save_profile: {e}")
    
    def _browse_folder(self):
        """Abre diálogo para selecionar pasta"""
        current_path = self.pasta_edit.text().strip()
        
        # Se já tem um caminho, usa como inicial
        if current_path and os.path.exists(current_path):
            initial_dir = current_path
        else:
            initial_dir = str(Path.home())

        
        folder = QFileDialog.getExistingDirectory(
            self,
            "Selecionar Pasta de Armazenamento",
            initial_dir,
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if folder:
            self.pasta_edit.setText(folder)
        
        folder = QFileDialog.getExistingDirectory(
            self,
            "Selecionar Pasta de Armazenamento",
            initial_dir,
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if folder:
            self.pasta_edit.setText(folder)
    
    def _update_example(self):
        """Atualiza o exemplo de caminho conforme as configurações"""
        try:
            pasta = self.pasta_edit.text().strip()
            if not pasta:
                pasta = str(DATA_DIR / 'xmls')
            
            formato = self.formato_combo.currentData() or 'AAAA-MM'
            
            # Gera exemplo de mês
            from datetime import datetime
            now = datetime.now()
            if formato == 'AAAA-MM':
                mes_exemplo = f"{now.year}-{now.month:02d}"
            elif formato == 'MM-AAAA':
                mes_exemplo = f"{now.month:02d}-{now.year}"
            elif formato == 'MMAAAA':
                mes_exemplo = f"{now.month:02d}{now.year}"
            elif formato == 'AAAA/MM':
                mes_exemplo = f"{now.year}/{now.month:02d}"
            else:  # MM/AAAA
                mes_exemplo = f"{now.month:02d}/{now.year}"
            
            # Atualiza label de exemplo
            exemplo = f"📁 Exemplo: {pasta}/33251845000109/{mes_exemplo}/NFe/"
            
            if hasattr(self, 'mes_exemplo'):
                self.mes_exemplo.setText(exemplo)
        except Exception:
            pass
    
    def _apply_profile(self):
        """Aplica o perfil copiando todos os XMLs e PDFs existentes para o local configurado"""
        try:
            if not self.current_profile_id:
                QMessageBox.warning(self, "Atenção", "Selecione um perfil para aplicar!")
                return
            
            # Lê configurações do perfil do banco
            conn = sqlite3.connect(self.db.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT nome, pasta_base, formato_pasta_mes, xml_pdf_separado, organizacao_tipo
                FROM perfis_armazenamento
                WHERE id = ?
            """, (self.current_profile_id,))
            
            result = cursor.fetchone()
            conn.close()
            
            if not result:
                QMessageBox.warning(self, "Erro", "Perfil não encontrado no banco de dados!")
                return
            
            nome_perfil, pasta_destino, formato_mes, xml_pdf_separado, organizacao_tipo = result
            organizacao_tipo = organizacao_tipo or 'CERTIFICADO_TIPO'
            
            print(f"[INFO PERFIL] Aplicando perfil ID={self.current_profile_id}: '{nome_perfil}'")
            print(f"[INFO PERFIL] organizacao_tipo={organizacao_tipo}, formato={formato_mes}")
            
            if not pasta_destino:
                QMessageBox.warning(self, "Atenção", "Configure a pasta base antes de aplicar!")
                return
            
            # Define texto de organização para exibição
            if organizacao_tipo == 'TIPO_CERTIFICADO':
                texto_org = "Tipo → Certificado → Mês (NFe/Certificado/mmaaaa)"
            else:
                texto_org = "Certificado → Tipo → Mês (Certificado/mmaaaa/NFe)"
            
            # Valida pasta destino
            try:
                pasta_destino_path = Path(pasta_destino)
                
                if not pasta_destino_path.exists():
                    reply = QMessageBox.question(
                        self,
                        "Pasta não existe",
                        f"A pasta de destino:\n{pasta_destino}\n\nNão existe. Deseja criá-la?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    
                    if reply == QMessageBox.Yes:
                        pasta_destino_path.mkdir(parents=True, exist_ok=True)
                    else:
                        return
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Caminho inválido:\n{e}")
                return
            
            # Pergunta confirmação
            # Converte formato_mes de código para texto legível
            formatos_map = {
                'AAAA-MM': 'Ano-Mês (2025-01)',
                'MM-AAAA': 'Mês-Ano (01-2025)',
                'MMAAAA': 'MêsAno (012025)',
                'AAAA/MM': 'Ano/Mês (2025/01)',
                'MM/AAAA': 'Mês/Ano (01/2025)'
            }
            texto_formato = formatos_map.get(formato_mes, formato_mes)
            
            reply = QMessageBox.question(
                self,
                "Confirmar Aplicação",
                f"Deseja copiar todos os XMLs e PDFs existentes para:\n\n"
                f"📁 {pasta_destino}\n\n"
                f"📅 Formato: {texto_formato}\n"
                f"📊 Hierarquia: {texto_org}\n"
                f"🗂️ XML/PDF: {'Pastas separadas' if xml_pdf_separado else 'Mesma pasta'}\n\n"
                f"⚠️ Esta operação pode levar alguns minutos dependendo da quantidade de arquivos.\n\n"
                f"Os arquivos originais NÃO serão modificados ou removidos.",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Define pasta de origem (busca na pasta xmls local)
            pasta_origem = DATA_DIR / 'xmls'
            
            if not pasta_origem.exists():
                QMessageBox.information(self, "Informação", "Nenhuma pasta 'xmls' encontrada para copiar.")
                return
            
            # Executa cópia com as configurações do perfil
            self._copiar_arquivos_para_perfil(
                pasta_origem, 
                pasta_destino_path,
                formato_mes,
                xml_pdf_separado,
                organizacao_tipo
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao aplicar perfil:\n{e}")
            print(f"[ERRO] _apply_profile: {e}")
            import traceback
            traceback.print_exc()
    
    def _copiar_arquivos_para_perfil(self, origem: Path, destino: Path, formato_mes: str, xml_pdf_separado: int, organizacao_tipo: str):
        """Copia arquivos XML/PDF da pasta de origem para o perfil - MODO RÁPIDO (sem reprocessar)
        
        Args:
            origem: Pasta de origem (xmls local)
            destino: Pasta de destino do perfil
            formato_mes: Formato de mês (AAAA-MM, MM-AAAA, etc)
            xml_pdf_separado: 1 = pastas separadas, 0 = mesma pasta
            organizacao_tipo: 'CERTIFICADO_TIPO' ou 'TIPO_CERTIFICADO'
        """
        try:
            import shutil
            import re
            from lxml import etree
            from datetime import datetime
            
            # Cria diálogo de progresso
            progress = QProgressDialog("Preparando cópia de arquivos...", "Cancelar", 0, 100, self)
            progress.setWindowTitle("Aplicando Perfil - Copiando Arquivos")
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            
            # Pastas a ignorar
            pastas_ignorar = ['Debug de notas', 'Resumos', 'Eventos', 'Outros', 'debug', 'resumos', 'eventos', 'outros']
            
            # Lista arquivos XML
            arquivos_xml = []
            for arquivo in origem.rglob('*.xml'):
                deve_ignorar = False
                for parte in arquivo.parts:
                    if parte in pastas_ignorar:
                        deve_ignorar = True
                        break
                
                if not deve_ignorar:
                    arquivos_xml.append(arquivo)
            
            total = len(arquivos_xml)
            if total == 0:
                QMessageBox.information(self, "Informação", "Nenhum arquivo XML encontrado para copiar.")
                return
            
            progress.setMaximum(total)
            progress.setLabelText(f"Copiando {total} arquivo(s)...")
            
            # Carrega mapeamento de CNPJ -> Nome do Certificado
            mapeamento_nomes = {}
            try:
                certs = self.db.load_certificates()
                for cert in certs:
                    informante = cert.get('informante', '')
                    nome_cert = cert.get('nome_certificado', '')
                    if informante and nome_cert:
                        nome_limpo = re.sub(r'[\\/*?:"<>|]', "_", nome_cert).strip()
                        mapeamento_nomes[informante] = nome_limpo
            except Exception as e:
                print(f"[ERRO] Ao carregar certificados: {e}")
            
            # Usa parâmetros recebidos (já lidos do perfil do banco)
            copiados = 0
            erros = 0
            
            for idx, arquivo_xml in enumerate(arquivos_xml):
                if progress.wasCanceled():
                    QMessageBox.information(self, "Cancelado", f"Cópia cancelada. {copiados} arquivo(s) copiado(s).")
                    return
                
                try:
                    # 🚫 FILTRO CRÍTICO: Verifica nome do arquivo ANTES de processar
                    # APENAS NF-e, CT-e e NFS-e completas devem ser copiadas
                    nome_arquivo = arquivo_xml.name.upper()
                    palavras_evento = ['EVENTO', 'CIENCIA', 'CONFIRMACAO', 'DESCONHECIMENTO', 
                                      'NAO_REALIZADA', 'CANCELAMENTO', 'CARTA_CORRECAO']
                    
                    if any(palavra in nome_arquivo for palavra in palavras_evento):
                        print(f"[IGNORADO] Evento não copiado para perfil: {arquivo_xml.name}")
                        continue
                    
                    # Extrai CNPJ da estrutura de pastas
                    caminho_relativo = arquivo_xml.relative_to(origem)
                    partes = list(caminho_relativo.parts)
                    
                    if len(partes) < 3:  # Precisa ter pelo menos: CNPJ/AAAA-MM/TIPO/arquivo.xml
                        continue
                    
                    cnpj_pasta = partes[0]
                    tipo_pasta = partes[2]  # NFe, CTe, etc.
                    
                    # 🚫 FILTRO ADICIONAL: Verifica se a pasta é de Eventos
                    if "Eventos" in tipo_pasta or "Eventos" in str(arquivo_xml):
                        print(f"[IGNORADO] Arquivo em pasta de eventos: {arquivo_xml.name}")
                        continue
                    
                    # Normaliza CNPJ (remove caracteres especiais)
                    cnpj_normalizado = ''.join(c for c in cnpj_pasta if c.isdigit())
                    
                    # Busca nome do certificado (SEMPRE usa nome, nunca CNPJ)
                    nome_cert = mapeamento_nomes.get(cnpj_normalizado)
                    if not nome_cert:
                        print(f"[AVISO] Nome do certificado não encontrado para CNPJ {cnpj_normalizado}")
                        erros += 1
                        continue  # Pula se não encontrar o nome
                    
                    pasta_cert = nome_cert
                    
                    # Lê apenas a data do XML (parse mínimo)
                    ano = None
                    mes = None
                    
                    try:
                        tree = etree.parse(str(arquivo_xml))
                        root = tree.getroot()
                        
                        # Detecta tipo de documento
                        root_tag = root.tag.split('}')[-1] if '}' in root.tag else root.tag
                        
                        # Define namespace baseado no tipo
                        if 'cte' in root_tag.lower():
                            ns = '{http://www.portalfiscal.inf.br/cte}'
                        else:
                            ns = '{http://www.portalfiscal.inf.br/nfe}'
                        
                        # Busca data de emissão (tenta várias tags)
                        data_elem = root.find(f'.//{ns}dhEmi')
                        if data_elem is None:
                            data_elem = root.find(f'.//{ns}dEmi')
                        if data_elem is None:
                            data_elem = root.find(f'.//{ns}dhRecbto')
                        
                        if data_elem is not None and data_elem.text:
                            data_str = data_elem.text.split('T')[0]  # Remove hora se tiver
                            if len(data_str) >= 7:  # AAAA-MM-DD
                                ano = data_str[:4]
                                mes = data_str[5:7]
                                print(f"[DEBUG DATA] {arquivo_xml.name}: data extraída do XML = {ano}-{mes}")
                    except Exception as e:
                        print(f"[AVISO] Erro ao ler data do XML {arquivo_xml.name}: {e}")
                    
                    # Fallback 1: Tenta extrair da estrutura de pastas (AAAA-MM ou MMAAAA)
                    if not ano or not mes:
                        if len(partes) >= 2:
                            data_pasta = partes[1]  # pode ser "2025-12" ou "122025"
                            if '-' in data_pasta and len(data_pasta) == 7:
                                # Formato AAAA-MM
                                ano = data_pasta[:4]
                                mes = data_pasta[5:7]
                                print(f"[DEBUG DATA] {arquivo_xml.name}: data extraída da pasta = {ano}-{mes}")
                            elif len(data_pasta) == 6 and data_pasta.isdigit():
                                # Formato MMAAAA
                                mes = data_pasta[:2]
                                ano = data_pasta[2:6]
                                print(f"[DEBUG DATA] {arquivo_xml.name}: data extraída da pasta MMAAAA = {ano}-{mes}")
                    
                    # Fallback 2: usa data atual
                    if not ano or not mes:
                        now = datetime.now()
                        ano = str(now.year)
                        mes = f"{now.month:02d}"
                        print(f"[AVISO] {arquivo_xml.name}: usando data atual = {ano}-{mes}")
                    
                    # Valida ano e mês extraídos
                    if len(ano) != 4 or len(mes) != 2 or not ano.isdigit() or not mes.isdigit():
                        print(f"[ERRO] Data inválida para {arquivo_xml.name}: ano={ano}, mes={mes}")
                        erros += 1
                        continue
                    
                    # Aplica formato de mês configurado
                    if formato_mes == 'MM-AAAA':
                        ano_mes = f"{mes}-{ano}"
                    elif formato_mes == 'MMAAAA':
                        ano_mes = f"{mes}{ano}"
                    elif formato_mes == 'AAAA/MM':
                        ano_mes = f"{ano}/{mes}"
                    elif formato_mes == 'MM/AAAA':
                        ano_mes = f"{mes}/{ano}"
                    else:  # AAAA-MM (padrão)
                        ano_mes = f"{ano}-{mes}"
                    
                    print(f"[DEBUG PASTA] {arquivo_xml.name}: ano_mes final = {ano_mes}")
                    
                    # 🆕 Monta caminho de destino baseado em organizacao_tipo
                    if organizacao_tipo == 'TIPO_CERTIFICADO':
                        # Novo formato: Tipo/Certificado/mmaaaa
                        # Exemplo: NFe/61-MATPARCG/012026/
                        pasta_dest = destino / tipo_pasta / pasta_cert / ano_mes
                    else:
                        # Formato padrão: Certificado/mmaaaa/Tipo
                        # Exemplo: 61-MATPARCG/012026/NFe/
                        if xml_pdf_separado:
                            pasta_dest = destino / pasta_cert / ano_mes / tipo_pasta
                        else:
                            # XML e PDF na mesma pasta: NOME/MES/
                            pasta_dest = destino / pasta_cert / ano_mes
                    
                    pasta_dest.mkdir(parents=True, exist_ok=True)
                    
                    # Copia XML
                    xml_destino = pasta_dest / arquivo_xml.name
                    if not xml_destino.exists():
                        shutil.copy2(arquivo_xml, xml_destino)
                        copiados += 1
                        
                        # Copia PDF se existir
                        pdf_original = arquivo_xml.with_suffix('.pdf')
                        if pdf_original.exists():
                            pdf_destino = xml_destino.with_suffix('.pdf')
                            if not pdf_destino.exists():
                                shutil.copy2(pdf_original, pdf_destino)
                    
                    progress.setValue(idx + 1)
                    progress.setLabelText(f"Copiando {idx + 1}/{total}: {arquivo_xml.name}")
                    QApplication.processEvents()
                    
                except Exception as e:
                    print(f"[ERRO] Ao copiar {arquivo_xml.name}: {e}")
                    erros += 1
            
            progress.close()
            
            # Mensagem final
            formatos_map = {
                'AAAA-MM': 'Ano-Mês (2025-01)',
                'MM-AAAA': 'Mês-Ano (01-2025)',
                'MMAAAA': 'MêsAno (012025)',
                'AAAA/MM': 'Ano/Mês (2025/01)',
                'MM/AAAA': 'Mês/Ano (01/2025)'
            }
            texto_formato = formatos_map.get(formato_mes, formato_mes)
            texto_org = "Tipo → Certificado → Mês" if organizacao_tipo == 'TIPO_CERTIFICADO' else "Certificado → Tipo → Mês"
            
            msg = f"✅ Perfil aplicado com sucesso!\n\n"
            msg += f"📊 Estatísticas:\n"
            msg += f"   • Arquivos copiados: {copiados}\n"
            if erros > 0:
                msg += f"   • Erros: {erros}\n"
            msg += f"\n📁 Destino: {destino}\n"
            msg += f"📅 Formato aplicado: {texto_formato}\n"
            msg += f"📊 Hierarquia: {texto_org}\n"
            msg += f"\n💡 Os arquivos originais foram mantidos em:\n   {origem}"
            
            QMessageBox.information(self, "Aplicação Concluída", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao copiar arquivos:\n{e}")
            import traceback
            traceback.print_exc()

def main():
    # Parse argumentos de linha de comando
    parser = argparse.ArgumentParser(description='BOT Busca NFE')
    parser.add_argument('--startup', action='store_true', 
                       help='Inicia em modo startup (minimizado)')
    parser.add_argument('--minimized', action='store_true',
                       help='Inicia minimizado')
    args = parser.parse_args()
    
    # Keep console open when run under VS Code for visibility
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    os.environ.setdefault("PYTHONUTF8", "1")
    
    # ===== PROTEÇÃO CONTRA MÚLTIPLAS INSTÂNCIAS =====
    # Cria um mutex único para o sistema "Busca XML"
    # Se já existir, significa que outra instância está rodando
    if sys.platform == "win32":
        kernel32 = ctypes.windll.kernel32
        ERROR_ALREADY_EXISTS = 183
        
        # Nome único do mutex (pode ser qualquer string única)
        mutex_name = "Global\\BuscaXML_SingleInstance_Mutex_9A8B7C6D"
        
        # Tenta criar o mutex
        mutex = kernel32.CreateMutexW(None, False, mutex_name)
        last_error = kernel32.GetLastError()
        
        # Se o mutex já existe, outra instância está rodando
        if last_error == ERROR_ALREADY_EXISTS:
            # Mostra mensagem de erro usando MessageBox do Windows (mais confiável que QMessageBox antes do QApplication)
            user32 = ctypes.windll.user32
            MB_OK = 0x00000000
            MB_ICONWARNING = 0x00000030
            MB_TOPMOST = 0x00040000
            
            mensagem = (
                "O sistema 'Busca XML' já está em execução!\n\n"
                "Não é permitido abrir múltiplas instâncias do programa.\n\n"
                "Por favor, use a instância que já está aberta."
            )
            user32.MessageBoxW(None, mensagem, "Busca XML - Já em Execução", MB_OK | MB_ICONWARNING | MB_TOPMOST)
            sys.exit(1)
        
        # Mantém o mutex aberto durante toda a execução do programa
        # Ele será automaticamente liberado quando o processo terminar
    # ===== FIM DA PROTEÇÃO =====
    
    app = QApplication(sys.argv)
    
    # Define o ícone do aplicativo (aparece na barra de tarefas do Windows)
    icon_path = BASE_DIR / 'Logo.ico'
    if not icon_path.exists():
        icon_path = BASE_DIR / 'Logo.png'
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))
    
    # Não encerra o app quando a janela é fechada (vai para bandeja)
    app.setQuitOnLastWindowClosed(False)
    
    w = MainWindow()
    
    # Agenda busca automática se iniciado via startup
    if args.startup or args.minimized:
        print("[STARTUP] Modo startup detectado - agendando busca automática em 10 minutos")
        
        # Agenda busca para 10 minutos (600 segundos)
        def executar_busca_automatica():
            print("[STARTUP] Executando busca automática agendada...")
            try:
                w.refresh_all()
                if w.tray_icon:
                    w.tray_icon.showMessage(
                        "Busca Automática Concluída",
                        "A busca por novas notas foi executada com sucesso.",
                        QSystemTrayIcon.Information,
                        5000
                    )
            except Exception as e:
                print(f"[STARTUP] Erro na busca automática: {e}")
                if w.tray_icon:
                    w.tray_icon.showMessage(
                        "Erro na Busca Automática",
                        f"Ocorreu um erro: {str(e)[:100]}",
                        QSystemTrayIcon.Warning,
                        5000
                    )
        
        w.task_scheduler.schedule_task(
            "Busca Automática SEFAZ",
            executar_busca_automatica,
            600  # 10 minutos = 600 segundos
        )
        
        # Inicia minimizado (não mostra janela)
        print("[STARTUP] Iniciando em segundo plano...")
    else:
        # Modo normal - mostra janela
        w.show()
        w._center_window()  # Centraliza depois de mostrar
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

